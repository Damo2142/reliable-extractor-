"""
Export all point names used in the SBS Composition Engine to a single Excel
workbook, one sheet per equipment family.

For static families (AHU/VAV/FCU/UV/RAD/RHC): the module set is
  cores + required_modules + every module whose category is in the
  family's available_categories.  All points from those modules are unioned.

For wizard plant families (HW-PLANT, CHW-PLANT-*): the module set is built by
  calling the real wizard assemblers across every mutually-exclusive option
  combination, so all possible point names appear.

Run: cd /srv/reliable-generator-dev && PYTHONPATH=. venv/bin/python composition/export_point_names.py
"""
import composition.module_registry as r
from composition.module_registry import (
    EQUIPMENT_FAMILIES, get_core_modules, get_module, list_modules,
    hwp_assemble, chwp_assemble,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/srv/dfa/drops/SBS-Composition-Engine-Point-Names.xlsx"

# ---------------------------------------------------------------------------
# Build the module list for a family
# ---------------------------------------------------------------------------

def static_module_ids(fam, info):
    """Module IDs for a non-wizard family: cores + required + category matches."""
    ids = set(get_core_modules(fam))
    ids |= set(info.get("required_modules", []))
    avail = set(info.get("available_categories", []))
    for mid in list_modules():
        try:
            mod = get_module(mid)
        except Exception:
            continue
        if mod.category in avail:
            ids.add(mid)
    # Only keep buildable ids
    return [m for m in ids if m in r._BUILDERS]


# HW plant: cover every mutually-exclusive option so all names appear
_HW_PARAM_SETS = [
    dict(boiler_type="cascade", num_boilers=4, spt_output="analog",
         monitor_boiler_temps=True, pump_type="cs", num_pumps=4,
         mixing_valve=True, iso_valves=True, comb_damper=True,
         heat_exchanger=True, hx_valve_type="single_mod",
         ahu_integration=True, num_ahus=8, makeup_water=True),
    dict(boiler_type="cascade", num_boilers=4, spt_output="bacnet",
         monitor_boiler_temps=True, pump_type="vfd", num_pumps=4),
    dict(boiler_type="full", num_boilers=4, monitor_boiler_temps=True,
         pump_type="pri-sec", num_primary=4, num_secondary=4),
    dict(boiler_type="full", num_boilers=4, heat_exchanger=True,
         hx_valve_type="single_onoff", pump_type="cs", num_pumps=2),
    dict(boiler_type="full", num_boilers=4, heat_exchanger=True,
         hx_valve_type="third_twothird", pump_type="cs", num_pumps=2),
]

# CHW air-cooled: all optionals, no tower
_CHW_AIR_PARAM_SETS = [
    dict(num_chillers=4, num_pri_pumps=4, num_sec_pumps=4, num_dp_sensors=2,
         bypass_valve=True, iso_valves=True, makeup_water=True,
         ahu_integration=True, num_ahus=8),
]

# CHW tower: add condenser water + towers
_CHW_TOWER_PARAM_SETS = [
    dict(num_chillers=4, num_pri_pumps=4, num_sec_pumps=4, num_dp_sensors=2,
         num_cw_pumps=4, num_towers=4, tower_bypass=True,
         bypass_valve=True, iso_valves=True, makeup_water=True,
         ahu_integration=True, num_ahus=8),
]


def plant_modules(fam):
    if fam == "HW-PLANT":
        sets, fn = _HW_PARAM_SETS, hwp_assemble
    elif fam == "CHW-PLANT-TOWER":
        sets, fn = _CHW_TOWER_PARAM_SETS, chwp_assemble
    else:  # CHW-PLANT-AIR
        sets, fn = _CHW_AIR_PARAM_SETS, chwp_assemble
    mods = []
    seen = set()
    for p in sets:
        for m in fn(p):
            key = id(m)
            if key in seen:
                continue
            seen.add(key)
            mods.append(m)
    return mods


# ---------------------------------------------------------------------------
# Flatten a module's objects into uniform rows
# ---------------------------------------------------------------------------

def rows_from_module(mod):
    """Yield (obj_type, name, ref, detail, description, source_module)."""
    src = getattr(mod, "id", "") or ""
    for p in mod.inputs:
        yield (p.point_type, p.name, p.row, p.range_code, p.description, p.module or src)
    for p in mod.outputs:
        detail = p.range_code
        if getattr(p, "reverse", False):
            detail += " (reverse)"
        yield (p.point_type, p.name, p.row, detail, p.description, p.module or src)
    for p in mod.values:
        if p.point_type == "MV" and p.states:
            detail = "states: " + ", ".join(str(s) for s in (
                p.states.values() if isinstance(p.states, dict) else p.states))
        else:
            d = f"default={p.default}"
            if p.units:
                d += f" {p.units}"
            if p.min_val is not None or p.max_val is not None:
                d += f" [{p.min_val}..{p.max_val}]"
            detail = d
        yield (p.point_type, p.name, p.instance, detail, p.description, p.module or src)
    for lp in mod.loops:
        detail = f"in={lp.input_ref} sp={lp.setpoint_ref} out={lp.output_ref} {lp.action}"
        yield ("LOOP", lp.name, lp.instance, detail, lp.description, lp.module or src)
    for t in mod.tables:
        yield ("TABLE", t.name, t.instance, f"{t.input_units}->{t.output_units}",
               t.description, t.module or src)
    for a in mod.arrays:
        yield ("ARRAY", a.name, a.instance, f"size={a.size}", a.description, a.module or src)
    for pr in mod.programs:
        yield ("PRG", pr.name, pr.instance, pr.filename, pr.description, pr.module or src)
    for s in mod.schedules:
        yield ("SCHED", s.name, s.instance, f"default={s.default_state}",
               s.description, s.module or src)
    for tr in mod.trends:
        yield ("TREND", tr.name, tr.instance, f"monitors {tr.monitored_point} ({tr.trend_type})",
               "", tr.module or src)


# Sort order for object types
_TYPE_ORDER = {t: i for i, t in enumerate(
    ["AI", "BI", "AO", "BO", "AV", "BV", "MV", "LOOP", "TABLE", "ARRAY",
     "PRG", "SCHED", "TREND"])}


_PLANT_OPTIONAL = ("mix", "iso", "comb", "damper", "exchang", "hx",
                   "ahu", "makeup", "bypass")


def _usage_for(mid, standard_ids, wizard):
    """Classify a module's points as Standard (variant-defining/always present)
    or Optional (toggle-able add-on)."""
    if wizard:
        if "core" in mid:
            return "Standard"
        if any(k in mid for k in _PLANT_OPTIONAL):
            return "Optional"
        return "Standard"   # boiler/chiller/pump modules
    return "Standard" if mid in standard_ids else "Optional"


def family_rows(fam, info):
    wizard = bool(info.get("wizard"))
    if wizard:
        mods = plant_modules(fam)
        standard_ids = set()
    else:
        standard_ids = set(get_core_modules(fam)) | set(info.get("required_modules", []))
        mods = [get_module(m) for m in static_module_ids(fam, info)]
    seen = {}   # (type, name) -> row list
    order = []
    for mod in mods:
        usage = _usage_for(getattr(mod, "id", "") or "", standard_ids, wizard)
        for row in rows_from_module(mod):
            otype, name = row[0], row[1]
            key = (otype, name)
            lst = list(row) + [usage]   # append Usage
            if key in seen:
                existing = seen[key]
                # merge source module note if different
                if row[5] and row[5] not in existing[5]:
                    existing[5] = existing[5] + ", " + row[5]
                # Standard wins over Optional if the name appears in both
                if usage == "Standard":
                    existing[6] = "Standard"
                continue
            seen[key] = lst
            order.append(lst)
    _u = {"Standard": 0, "Optional": 1}
    order.sort(key=lambda x: (_u.get(x[6], 9), _TYPE_ORDER.get(x[0], 99),
                              (x[2] if isinstance(x[2], int) else 0), x[1]))
    return order


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COLS = ["Object Type", "Usage", "Point Name", "Row / Instance",
        "Range / Default / Detail", "Description", "Source Module(s)"]
WIDTHS = [12, 10, 30, 14, 42, 50, 26]
OPT_FILL = PatternFill("solid", fgColor="FCE4D6")   # light orange for Optional


def write_sheet(ws, fam, info, rows):
    ws.cell(row=1, column=1, value=f"{fam}  —  {info.get('name','')}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=info.get("description", "")).font = Font(
        italic=True, size=9, color="595959")
    n_std = sum(1 for r in rows if r[6] == "Standard")
    n_opt = len(rows) - n_std
    ws.cell(row=3, column=1,
            value=f"{len(rows)} unique points  ({n_std} standard, {n_opt} optional add-on)").font = Font(
        size=9, color="595959")
    hr = 5
    for c, name in enumerate(COLS, 1):
        cell = ws.cell(row=hr, column=c, value=name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = WIDTHS[c - 1]
    for i, row in enumerate(rows):
        otype, name, ref, detail, desc, source, usage = row
        ordered = [otype, usage, name, ref, detail, desc, source]
        for c, val in enumerate(ordered, 1):
            cell = ws.cell(row=hr + 1 + i, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (5, 6, 7)))
            cell.font = Font(size=9)
            if usage == "Optional":
                cell.fill = OPT_FILL
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


def main():
    wb = Workbook()
    index = wb.active
    index.title = "INDEX"

    fam_data = []
    for fam, info in EQUIPMENT_FAMILIES.items():
        try:
            rows = family_rows(fam, info)
        except Exception as e:
            print(f"  !! {fam}: {e}")
            rows = []
        fam_data.append((fam, info, rows))

    # Index sheet
    index.cell(row=1, column=1, value="SBS Composition Engine — Point Names by Equipment Family").font = TITLE_FONT
    idx_cols = ["Family ID", "Name", "Prefix", "Total Points", "Standard", "Optional", "Wizard"]
    idx_w = [20, 42, 12, 13, 11, 11, 8]
    for c, name in enumerate(idx_cols, 1):
        cell = index.cell(row=3, column=c, value=name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        index.column_dimensions[cell.column_letter].width = idx_w[c - 1]
    for i, (fam, info, rows) in enumerate(fam_data):
        index.cell(row=4 + i, column=1, value=fam).font = Font(size=9, bold=True)
        index.cell(row=4 + i, column=2, value=info.get("name", "")).font = Font(size=9)
        index.cell(row=4 + i, column=3, value=info.get("prefix", "")).font = Font(size=9)
        n_std = sum(1 for r in rows if r[6] == "Standard")
        index.cell(row=4 + i, column=4, value=len(rows)).font = Font(size=9)
        index.cell(row=4 + i, column=5, value=n_std).font = Font(size=9)
        index.cell(row=4 + i, column=6, value=len(rows) - n_std).font = Font(size=9)
        index.cell(row=4 + i, column=7, value="Yes" if info.get("wizard") else "").font = Font(size=9)
    index.freeze_panes = "A4"

    # One sheet per family (Excel sheet-name limit 31 chars; family IDs fit)
    used = set()
    for fam, info, rows in fam_data:
        title = fam[:31]
        base = title
        n = 1
        while title in used:
            n += 1
            title = f"{base[:28]}_{n}"
        used.add(title)
        ws = wb.create_sheet(title)
        write_sheet(ws, fam, info, rows)

    wb.save(OUT)
    total = sum(len(r) for _, _, r in fam_data)
    print(f"Wrote {OUT}")
    print(f"  {len(fam_data)} family sheets + INDEX")
    print(f"  {total} point rows total")
    for fam, info, rows in fam_data:
        print(f"    {fam:18s} {len(rows):4d}")


if __name__ == "__main__":
    main()
