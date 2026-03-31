"""
SBS Composition Engine v2 — IO Schedule Export/Import

Provides:
  - Config store: in-memory cache of assembled configs keyed by short UUID
  - Export: generate IO schedule Excel (physical I/O only, one tab per controller)
  - Import: read modified terminal assignments, validate point names, update config
  - Generate from stored config: re-assemble with terminal overrides applied

Point Name column is LOCKED in the export — user can only edit Terminal column.
"""

import io
import uuid
import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter


# ─── Config Store ────────────────────────────────────────────────────────────

_config_store = {}  # config_id -> StoredConfig dict

MAX_STORED_CONFIGS = 50  # evict oldest when full


def _evict_oldest():
    """Remove oldest config when store is full."""
    if len(_config_store) >= MAX_STORED_CONFIGS:
        oldest_id = min(_config_store, key=lambda k: _config_store[k]["created"])
        del _config_store[oldest_id]


def store_config(inputs, outputs, family, controller, source, params):
    """
    Store an assembled config's I/O data and return a short config_id.

    Args:
        inputs:     list of InputPoint objects
        outputs:    list of OutputPoint objects
        family:     equipment family string (e.g. "AHU-VAV", "HW-PLANT")
        controller: controller model string (e.g. "MPS")
        source:     assembly source ("ahu", "hwp", "chwp")
        params:     original assembly parameters (for re-assembly)

    Returns:
        config_id (str, 8 chars)
    """
    _evict_oldest()
    config_id = uuid.uuid4().hex[:8]
    _config_store[config_id] = {
        "inputs": copy.deepcopy(inputs),
        "outputs": copy.deepcopy(outputs),
        "family": family,
        "controller": controller,
        "source": source,
        "params": params,
        "terminal_overrides": {},  # point_name -> new_row (populated by import)
        "created": datetime.now().isoformat(),
    }
    return config_id


def get_stored_config(config_id):
    """Get a stored config by ID. Raises ValueError if not found."""
    if config_id not in _config_store:
        raise ValueError(f"Config '{config_id}' not found. Assemble first to get a config_id.")
    return _config_store[config_id]


def get_terminal_overrides(config_id):
    """Get terminal overrides dict for a config. Returns empty dict if none."""
    cfg = get_stored_config(config_id)
    return dict(cfg["terminal_overrides"])


# ─── Excel Styles ────────────────────────────────────────────────────────────

_HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_DATA_FONT = Font(name="Calibri", size=10)
_LOCKED_FONT = Font(name="Calibri", size=10, color="333333")
_LOCKED_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
_EDITABLE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
_SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")


# ─── Export ──────────────────────────────────────────────────────────────────

def export_io_schedule(config_id):
    """
    Generate IO schedule Excel from a stored config.

    Returns bytes of the .xlsx file.

    Columns: Terminal | Point Name | Description | Type | Units | Controller | Notes
    Physical IO only (AI, AO, BI, BO). One tab per controller.
    Point Name column is locked (sheet protection enabled, Terminal column unlocked).
    """
    cfg = get_stored_config(config_id)
    inputs = cfg["inputs"]
    outputs = cfg["outputs"]
    controller = cfg["controller"] or "MPS"
    family = cfg["family"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = controller

    # Title
    ws.cell(1, 1, value=f"IO Schedule — {family}").font = _TITLE_FONT
    ws.cell(2, 1, value=f"Controller: {controller} | Config: {config_id} | "
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _SUBTITLE_FONT

    # Headers at row 4
    headers = ["Terminal", "Point Name", "Description", "Type", "Units", "Controller", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    row = 5

    # --- Inputs section ---
    for pt in sorted(inputs, key=lambda p: p.row):
        terminal = f"IN{pt.row}"
        _write_io_row(ws, row, terminal, pt.name, pt.description,
                      pt.point_type, getattr(pt, 'units', ''), controller,
                      getattr(pt, 'range_code', ''))
        row += 1

    # Spacer
    row += 1

    # --- Outputs section ---
    for pt in sorted(outputs, key=lambda p: p.row):
        terminal = f"OUT{pt.row}"
        _write_io_row(ws, row, terminal, pt.name, pt.description,
                      pt.point_type, getattr(pt, 'units', '%'), controller,
                      getattr(pt, 'range_code', ''))
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12   # Terminal (editable)
    ws.column_dimensions["B"].width = 30   # Point Name (locked)
    ws.column_dimensions["C"].width = 40   # Description
    ws.column_dimensions["D"].width = 6    # Type
    ws.column_dimensions["E"].width = 8    # Units
    ws.column_dimensions["F"].width = 14   # Controller
    ws.column_dimensions["G"].width = 30   # Notes

    # Enable sheet protection — Point Name column locked, Terminal column editable
    ws.protection = openpyxl.worksheet.protection.SheetProtection(
        sheet=True,
        objects=True,
        scenarios=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        sort=False,
        autoFilter=False,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_io_row(ws, row, terminal, point_name, description, point_type, units, controller, notes):
    """Write a single I/O row with proper protection."""
    # Terminal — EDITABLE (green background, unlocked)
    cell_term = ws.cell(row=row, column=1, value=terminal)
    cell_term.fill = _EDITABLE_FILL
    cell_term.font = _DATA_FONT
    cell_term.border = _THIN_BORDER
    cell_term.protection = Protection(locked=False)

    # Point Name — LOCKED (gray background)
    cell_name = ws.cell(row=row, column=2, value=point_name)
    cell_name.fill = _LOCKED_FILL
    cell_name.font = _LOCKED_FONT
    cell_name.border = _THIN_BORDER
    cell_name.protection = Protection(locked=True)

    # Description — locked
    cell_desc = ws.cell(row=row, column=3, value=description)
    cell_desc.font = _DATA_FONT
    cell_desc.border = _THIN_BORDER
    cell_desc.protection = Protection(locked=True)

    # Type — locked
    cell_type = ws.cell(row=row, column=4, value=point_type)
    cell_type.font = _DATA_FONT
    cell_type.border = _THIN_BORDER
    cell_type.alignment = Alignment(horizontal="center")
    cell_type.protection = Protection(locked=True)

    # Units — locked
    cell_units = ws.cell(row=row, column=5, value=units or "")
    cell_units.font = _DATA_FONT
    cell_units.border = _THIN_BORDER
    cell_units.protection = Protection(locked=True)

    # Controller — locked
    cell_ctrl = ws.cell(row=row, column=6, value=controller)
    cell_ctrl.font = _DATA_FONT
    cell_ctrl.border = _THIN_BORDER
    cell_ctrl.protection = Protection(locked=True)

    # Notes — editable
    cell_notes = ws.cell(row=row, column=7, value=notes or "")
    cell_notes.font = _DATA_FONT
    cell_notes.border = _THIN_BORDER
    cell_notes.protection = Protection(locked=False)


# ─── Import ──────────────────────────────────────────────────────────────────

def import_io_schedule(config_id, file_data):
    """
    Read modified terminal assignments from an uploaded IO schedule Excel.

    Validates:
      1. Point names unchanged (reject with error listing modified names)
      2. No terminal conflicts (two different points on same terminal)
      3. Terminal format valid (IN{n} or OUT{n})

    On success, stores terminal overrides in the config store.

    Returns dict with:
      - ok: True
      - changes: list of {point_name, old_terminal, new_terminal}
      - total_points: int

    Raises ValueError on validation failure.
    """
    cfg = get_stored_config(config_id)

    # Build original point name -> terminal mapping
    orig_input_map = {}   # point_name -> "IN{row}"
    orig_output_map = {}  # point_name -> "OUT{row}"
    for pt in cfg["inputs"]:
        orig_input_map[pt.name] = f"IN{pt.row}"
    for pt in cfg["outputs"]:
        orig_output_map[pt.name] = f"OUT{pt.row}"
    orig_all = {**orig_input_map, **orig_output_map}

    # Read uploaded Excel
    wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # First tab

    uploaded_points = []  # list of (terminal, point_name)
    for row in ws.iter_rows(min_row=5, values_only=True):  # Data starts at row 5
        if not row or not row[0] or not row[1]:
            continue
        terminal = str(row[0]).strip()
        point_name = str(row[1]).strip()
        if not terminal or not point_name:
            continue
        # Skip spacer rows or header repeats
        if terminal.lower() in ("terminal",) or point_name.lower() in ("point name",):
            continue
        uploaded_points.append((terminal, point_name))

    wb.close()

    # --- Validation 1: point names unchanged ---
    uploaded_names = set(p[1] for p in uploaded_points)
    orig_names = set(orig_all.keys())
    missing_names = orig_names - uploaded_names
    extra_names = uploaded_names - orig_names
    errors = []
    if missing_names:
        errors.append(f"Missing point names (removed or renamed): {sorted(missing_names)}")
    if extra_names:
        errors.append(f"Unknown point names (added or renamed): {sorted(extra_names)}")
    if errors:
        raise ValueError("Point Name validation failed — names must not be modified.\n" +
                         "\n".join(errors))

    # --- Validation 2: terminal format ---
    for terminal, name in uploaded_points:
        if not (terminal.startswith("IN") or terminal.startswith("OUT")):
            errors.append(f"Invalid terminal format '{terminal}' for point '{name}' "
                          f"— must be IN{{n}} or OUT{{n}}")
            continue
        suffix = terminal[2:] if terminal.startswith("IN") else terminal[3:]
        try:
            int(suffix)
        except ValueError:
            errors.append(f"Invalid terminal number '{terminal}' for point '{name}'")

    if errors:
        raise ValueError("Terminal format validation failed:\n" + "\n".join(errors))

    # --- Validation 3: no terminal conflicts ---
    input_terminals = {}   # terminal -> point_name
    output_terminals = {}  # terminal -> point_name
    for terminal, name in uploaded_points:
        bucket = input_terminals if terminal.startswith("IN") else output_terminals
        if terminal in bucket:
            errors.append(f"Terminal conflict: '{terminal}' assigned to both "
                          f"'{bucket[terminal]}' and '{name}'")
        else:
            bucket[terminal] = name

    if errors:
        raise ValueError("Terminal conflict validation failed:\n" + "\n".join(errors))

    # --- Build overrides ---
    changes = []
    terminal_overrides = {}
    for terminal, name in uploaded_points:
        old_terminal = orig_all.get(name)
        if old_terminal != terminal:
            # Extract row number
            if terminal.startswith("IN"):
                new_row = int(terminal[2:])
            else:
                new_row = int(terminal[3:])
            terminal_overrides[name] = new_row
            changes.append({
                "point_name": name,
                "old_terminal": old_terminal,
                "new_terminal": terminal,
            })

    # Store overrides
    cfg["terminal_overrides"] = terminal_overrides

    return {
        "ok": True,
        "changes": changes,
        "total_points": len(uploaded_points),
        "overrides_applied": len(changes),
    }


# ─── Apply Overrides ────────────────────────────────────────────────────────

def apply_terminal_overrides(config_id, inputs, outputs):
    """
    Apply stored terminal overrides to freshly assembled input/output lists.

    Modifies the .row attribute of matching InputPoint/OutputPoint objects in-place.
    Also updates highest_input_row / highest_output_row.

    Returns (highest_input_row, highest_output_row).
    """
    overrides = get_terminal_overrides(config_id)
    if not overrides:
        hi_in = max((p.row for p in inputs), default=0)
        hi_out = max((p.row for p in outputs), default=0)
        return hi_in, hi_out

    for pt in inputs:
        if pt.name in overrides:
            pt.row = overrides[pt.name]

    for pt in outputs:
        if pt.name in overrides:
            pt.row = overrides[pt.name]

    # Re-sort
    inputs.sort(key=lambda p: p.row)
    outputs.sort(key=lambda p: p.row)

    hi_in = max((p.row for p in inputs), default=0)
    hi_out = max((p.row for p in outputs), default=0)
    return hi_in, hi_out
