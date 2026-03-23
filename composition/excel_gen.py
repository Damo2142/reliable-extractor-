"""
SBS Composition Engine v2 — Excel Generator (Reliable Controls Output Tool)

Generates an Excel workbook with 11 tabs mirroring RC Studio pages exactly:
1. Inputs (AI/BI)
2. Outputs (AO/BO)
3. Values (AV/BV/MV)
4. Loops (PID)
5. Tables
6. Arrays
7. Schedules
8. Trends (STLs)
9. System Groups
10. Programs
11. Custom Units

Row number = physical terminal number. Empty rows = unused terminals (intentional).
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from composition.models import ControllerConfig


# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
UNUSED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=10, color="666666")


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER


def _add_title(ws, title, subtitle=""):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    return 4  # Start data at row 4


def generate_excel(config: ControllerConfig) -> bytes:
    """Generate the 11-tab Excel workbook from an assembled config."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _build_inputs_tab(wb, config)
    _build_outputs_tab(wb, config)
    _build_values_tab(wb, config)
    _build_loops_tab(wb, config)
    _build_tables_tab(wb, config)
    _build_arrays_tab(wb, config)
    _build_schedules_tab(wb, config)
    _build_trends_tab(wb, config)
    _build_system_groups_tab(wb, config)
    _build_programs_tab(wb, config)
    _build_custom_units_tab(wb, config)
    _build_alarms_tab(wb, config)
    _build_commissioning_tab(wb, config)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_inputs_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Inputs")
    start = _add_title(ws, "INPUTS", f"{config.device_name} — {config.equipment_family}")

    headers = ["Terminal", "Name", "Type", "Value", "Units", "Range", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    # Build lookup by row
    input_by_row = {pt.row: pt for pt in config.inputs}
    max_row = config.highest_input_row

    for row_num in range(1, max_row + 1):
        r = start + row_num
        if row_num in input_by_row:
            pt = input_by_row[row_num]
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"{config.device_name}-{pt.name}")
            ws.cell(row=r, column=3, value=pt.point_type)
            ws.cell(row=r, column=4, value="")  # Present value placeholder
            ws.cell(row=r, column=5, value=pt.units)
            ws.cell(row=r, column=6, value=pt.range_code)
            ws.cell(row=r, column=7, value=pt.description)
            ws.cell(row=r, column=8, value=pt.module)
        else:
            # Empty row — unused terminal (intentional, never fill)
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"--- IN{row_num} unused ---")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 16


def _build_outputs_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Outputs")
    start = _add_title(ws, "OUTPUTS", f"{config.device_name} — {config.equipment_family}")

    headers = ["Terminal", "Name", "Type", "Value", "Units", "Range", "Min V", "Max V", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    output_by_row = {pt.row: pt for pt in config.outputs}
    max_row = config.highest_output_row

    for row_num in range(1, max_row + 1):
        r = start + row_num
        if row_num in output_by_row:
            pt = output_by_row[row_num]
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"{config.device_name}-{pt.name}")
            ws.cell(row=r, column=3, value=pt.point_type)
            ws.cell(row=r, column=4, value="")
            ws.cell(row=r, column=5, value=pt.units)
            ws.cell(row=r, column=6, value=pt.range_code)
            ws.cell(row=r, column=7, value=pt.min_v)
            ws.cell(row=r, column=8, value=pt.max_v)
            ws.cell(row=r, column=9, value=pt.description)
            ws.cell(row=r, column=10, value=pt.module)
        else:
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"--- OUT{row_num} unused ---")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 16


def _build_values_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Values")
    start = _add_title(ws, "VALUES (AV / BV / MV)", f"{config.device_name}")

    headers = ["Instance", "Name", "Type", "Default Value", "Units", "Range / States", "Min", "Max", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    # Build single lookup by instance across ALL value types (AV/BV/MV share instance space)
    val_by_inst = {v.instance: v for v in config.values}
    max_inst = max(val_by_inst.keys()) if val_by_inst else 0

    for inst in range(1, max_inst + 1):
        r = start + inst
        if inst in val_by_inst:
            val = val_by_inst[inst]
            prefix = {"AV": "AV", "BV": "BV", "MV": "MV"}.get(val.point_type, "AV")
            ws.cell(row=r, column=1, value=f"{prefix}{inst}")
            ws.cell(row=r, column=2, value=f"{config.device_name}-{val.name}")
            ws.cell(row=r, column=3, value=val.point_type)
            if val.point_type == "MV" and val.states:
                state_str = "/".join([f"{k}-{v}" for k, v in sorted(val.states.items())])
                default_text = val.states.get(val.default, str(val.default)) if val.default else ""
                ws.cell(row=r, column=4, value=default_text)
                ws.cell(row=r, column=5, value=state_str)
                ws.cell(row=r, column=6, value=state_str)
            elif val.point_type == "BV":
                ws.cell(row=r, column=4, value="Off" if not val.default else "On")
                ws.cell(row=r, column=5, value="Off/On")
                ws.cell(row=r, column=6, value="Off/On")
            else:
                ws.cell(row=r, column=4, value=str(val.default))
                ws.cell(row=r, column=5, value=val.units or "")
                ws.cell(row=r, column=6, value="")
            ws.cell(row=r, column=7, value=val.min_val if val.min_val is not None else "")
            ws.cell(row=r, column=8, value=val.max_val if val.max_val is not None else "")
            if val.point_type == "MV" and val.states:
                ws.cell(row=r, column=9, value=state_str)
            else:
                ws.cell(row=r, column=9, value=val.description)
            ws.cell(row=r, column=10, value=val.module)
        else:
            # Empty filler row — use AV as default type label
            ws.cell(row=r, column=1, value=f"AV{inst}")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 16


def _build_loops_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Loops")
    start = _add_title(ws, "PID LOOPS", f"{config.device_name}")

    headers = ["Loop", "Name", "Input", "Setpoint", "Output", "Action", "P Band", "Integral", "D", "Bias", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    loop_by_inst = {lp.instance: lp for lp in config.loops}
    max_loop = max(loop_by_inst.keys()) if loop_by_inst else 0
    for inst in range(1, max_loop + 1):
        r = start + inst
        if inst in loop_by_inst:
            lp = loop_by_inst[inst]
            action_sym = "+" if lp.action == "direct" else "-"
            ws.cell(row=r, column=1, value=f"LOOP{inst}")
            ws.cell(row=r, column=2, value=lp.name)
            ws.cell(row=r, column=3, value=f"{config.device_name}-{lp.input_ref}")
            ws.cell(row=r, column=4, value=f"{config.device_name}-{lp.setpoint_ref}")
            ws.cell(row=r, column=5, value=f"{config.device_name}-{lp.output_ref}" if lp.output_ref else "")
            ws.cell(row=r, column=6, value=action_sym)
            ws.cell(row=r, column=7, value=lp.p_band)
            ws.cell(row=r, column=8, value=lp.integral)
            ws.cell(row=r, column=9, value=lp.derivative)
            ws.cell(row=r, column=10, value=lp.bias)
            ws.cell(row=r, column=11, value=lp.description)
            ws.cell(row=r, column=12, value=lp.module)
        else:
            ws.cell(row=r, column=1, value=f"LOOP{inst}")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    for c, w in enumerate([10, 20, 30, 30, 30, 8, 10, 10, 6, 6, 30, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_tables_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Tables")
    start = _add_title(ws, "TABLES", f"{config.device_name}")

    # Determine max data points across all tables for column count
    max_pts = max((len(tbl.data_points) for tbl in config.tables), default=0) if config.tables else 0
    headers = ["Table", "Name", "Input Units", "Output Units", "Description", "Module"]
    # Add X/Y column pairs for each data point
    for p in range(max_pts):
        headers.append(f"X{p+1}")
        headers.append(f"Y{p+1}")

    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    tbl_by_inst = {tbl.instance: tbl for tbl in config.tables}
    max_tbl = max(tbl_by_inst.keys()) if tbl_by_inst else 0
    for inst in range(1, max_tbl + 1):
        r = start + inst
        if inst in tbl_by_inst:
            tbl = tbl_by_inst[inst]
            ws.cell(row=r, column=1, value=f"TBL{inst}")
            ws.cell(row=r, column=2, value=tbl.name)
            ws.cell(row=r, column=3, value=tbl.input_units)
            ws.cell(row=r, column=4, value=tbl.output_units)
            ws.cell(row=r, column=5, value=tbl.description)
            ws.cell(row=r, column=6, value=tbl.module)
            for p, (x, y) in enumerate(tbl.data_points):
                ws.cell(row=r, column=7 + p * 2, value=x)
                ws.cell(row=r, column=8 + p * 2, value=y)
        else:
            ws.cell(row=r, column=1, value=f"TBL{inst}")
            for c2 in range(1, len(headers) + 1):
                ws.cell(row=r, column=c2).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 16


def _build_arrays_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Arrays")
    start = _add_title(ws, "ARRAYS", f"{config.device_name}")
    ws.cell(row=start, column=1, value="(No arrays for this configuration)")
    ws.cell(row=start, column=1).font = DATA_FONT


def _build_schedules_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Schedules")
    start = _add_title(ws, "SCHEDULES", f"{config.device_name}")

    headers = ["Schedule", "Name", "Default State", "States", "Priority", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    sched_by_inst = {sch.instance: sch for sch in config.schedules}
    max_sched = max(sched_by_inst.keys()) if sched_by_inst else 0
    for inst in range(1, max_sched + 1):
        r = start + inst
        if inst in sched_by_inst:
            sch = sched_by_inst[inst]
            ws.cell(row=r, column=1, value=f"SCHED{inst}")
            ws.cell(row=r, column=2, value=f"{config.device_name}-{sch.name}")
            ws.cell(row=r, column=3, value=sch.default_state)
            ws.cell(row=r, column=4, value=" / ".join(sch.states))
            ws.cell(row=r, column=5, value=sch.priority)
            ws.cell(row=r, column=6, value=sch.description)
            ws.cell(row=r, column=7, value=sch.module)
        else:
            ws.cell(row=r, column=1, value=f"SCHED{inst}")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    for c, w in enumerate([10, 35, 14, 25, 10, 35, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_trends_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Trends")
    start = _add_title(ws, "SINGLE POINT TREND LOGS (STL)", f"{config.device_name}")

    headers = ["STL", "Name", "Monitored Point", "Type", "Interval/COV Delta", "Buffer"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    stl_by_inst = {tr.instance: tr for tr in config.trends}
    max_stl = max(stl_by_inst.keys()) if stl_by_inst else 0
    for inst in range(1, max_stl + 1):
        r = start + inst
        if inst in stl_by_inst:
            tr = stl_by_inst[inst]
            ws.cell(row=r, column=1, value=f"STL{inst}")
            ws.cell(row=r, column=2, value=f"{config.device_name}-{tr.name}")
            ws.cell(row=r, column=3, value=f"{config.device_name}-{tr.monitored_point}")
            ws.cell(row=r, column=4, value=tr.trend_type.capitalize())
            if tr.trend_type == "polled":
                ws.cell(row=r, column=5, value=tr.interval)
            else:
                ws.cell(row=r, column=5, value=tr.cov_delta)
            ws.cell(row=r, column=6, value=tr.buffer_size)
        else:
            ws.cell(row=r, column=1, value=f"STL{inst}")
            for c2 in range(1, len(headers) + 1):
                ws.cell(row=r, column=c2).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    for c, w in enumerate([10, 35, 35, 10, 16, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_system_groups_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("System Groups")
    start = _add_title(ws, "SYSTEM GROUPS (Graphic Pages)", f"{config.device_name}")

    headers = ["Group Name", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    for i, sg in enumerate(config.system_groups):
        r = start + 1 + i
        ws.cell(row=r, column=1, value=sg.name)
        ws.cell(row=r, column=2, value=sg.description)
        ws.cell(row=r, column=3, value=sg.module)
        _style_data(ws, r, len(headers))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16


def _build_programs_tab(wb, config: ControllerConfig):
    ws = wb.create_sheet("Programs")
    start = _add_title(ws, "PROGRAMS", f"{config.device_name}")

    headers = ["PRG#", "Name", "Filename", "Enabled", "Exec Order", "Description", "Module"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    prg_by_inst = {prg.instance: prg for prg in config.programs}
    max_prg = max(prg_by_inst.keys()) if prg_by_inst else 0
    for inst in range(1, max_prg + 1):
        r = start + inst
        if inst in prg_by_inst:
            prg = prg_by_inst[inst]
            ws.cell(row=r, column=1, value=f"PRG{inst}")
            ws.cell(row=r, column=2, value=f"{config.device_name}-{prg.name}")
            ws.cell(row=r, column=3, value=prg.filename)
            ws.cell(row=r, column=4, value="Yes" if prg.enabled else "No")
            ws.cell(row=r, column=5, value=prg.exec_order)
            ws.cell(row=r, column=6, value=prg.description)
            ws.cell(row=r, column=7, value=prg.module)
        else:
            ws.cell(row=r, column=1, value=f"PRG{inst}")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        _style_data(ws, r, len(headers))

    for c, w in enumerate([8, 35, 28, 10, 12, 50, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_custom_units_tab(wb, config: ControllerConfig):
    """SBS Standard custom unit enumerations — ships with every build."""
    ws = wb.create_sheet("Custom Units")
    start = _add_title(ws, "CUSTOM UNITS", "SBS Standard Enumerations")

    headers = ["Index", "Analog Text", "Digital On", "Digital Off",
               "State 1", "State 2", "State 3", "State 4",
               "State 5", "State 6", "State 7", "State 8"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    # SBS Standard custom units — CUSTOM enumerations only
    # Standard BACnet units (°F, %, ppm, etc.) are NOT listed here —
    # they are built into the controller firmware.
    # Only custom digital and multi-state enumerations go in this table.
    custom_units = [
        (64,  "",       "Occupied",  "Unoccupied","", "", "", "", "", "", "", ""),
        (65,  "",       "Enabled",   "Disabled",  "", "", "", "", "", "", "", ""),
        (66,  "",       "On",        "Off",       "", "", "", "", "", "", "", ""),
        (67,  "",       "Open",      "Closed",    "", "", "", "", "", "", "", ""),
        (68,  "",       "Start",     "Stop",      "", "", "", "", "", "", "", ""),
        (69,  "",       "Run",       "Shutdown",  "", "", "", "", "", "", "", ""),
        (70,  "",       "Alarm",     "Normal",    "", "", "", "", "", "", "", ""),
        (71,  "",       "Dirty",     "Clean",     "", "", "", "", "", "", "", ""),
        (72,  "",       "Yes",       "No",        "", "", "", "", "", "", "", ""),
        (73,  "",       "Active",    "Inactive",  "", "", "", "", "", "", "", ""),
        (74,  "",       "Reset",     "Normal",    "", "", "", "", "", "", "", ""),
        (75,  "",       "Tripped",   "Normal",    "", "", "", "", "", "", "", ""),
        (76,  "",       "High",      "Low",       "", "", "", "", "", "", "", ""),
        (77,  "",       "",          "",          "Occupied", "Bypass", "Standby", "Unoccupied", "NSB", "NSF", "OSH", "OSC"),
        (78,  "",       "",          "",          "Vent", "Cool", "Reheat", "Heat", "Init", "", "", ""),
        (79,  "",       "",          "",          "Off", "Low", "Med", "High", "", "", "", ""),
    ]

    for i, cu in enumerate(custom_units):
        r = start + 1 + i
        for c, val in enumerate(cu, 1):
            ws.cell(row=r, column=c, value=val)
        _style_data(ws, r, len(headers))

    for c, w in enumerate([8, 12, 14, 14, 12, 12, 12, 14, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_alarms_tab(wb, config: ControllerConfig):
    """Alarms tab — auto-generated alarm definitions (from Dave's Alarm Builder format)."""
    from composition.alarm_gen import generate_alarm_excel_data, generate_alarm_bas
    ws = wb.create_sheet("Alarms")
    start = _add_title(ws, "ALARM DEFINITIONS", f"{config.device_name}")

    headers = ["Point", "Type", "AlarmType", "Priority", "Delay", "Deadband",
               "High", "Low", "Condition", "Message", "ReverseLogic"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    rows = generate_alarm_excel_data(config)
    for i, row in enumerate(rows):
        r = start + 1 + i
        for c, key in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))
        _style_data(ws, r, len(headers))

    # Add alarm BAS code preview at the bottom
    bas_code = generate_alarm_bas(config)
    if bas_code:
        gap = start + len(rows) + 3
        ws.cell(row=gap, column=1, value="ALARM PROGRAM (.BAS)")
        ws.cell(row=gap, column=1).font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        for j, line in enumerate(bas_code.split("\n")):
            ws.cell(row=gap + 1 + j, column=1, value=line)
            ws.cell(row=gap + 1 + j, column=1).font = Font(name="Consolas", size=9)

    for c, w in enumerate([35, 8, 14, 14, 8, 10, 8, 8, 12, 35, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_commissioning_tab(wb, config: ControllerConfig):
    """Commissioning checklist — field point checkout (from Dave's Report Generator concept)."""
    from composition.alarm_gen import generate_commissioning_checklist
    ws = wb.create_sheet("Commissioning")
    start = _add_title(ws, "COMMISSIONING POINT CHECKLIST", f"{config.device_name}")

    headers = ["Point", "Object", "Type", "Range", "Description",
               "Check", "Expected", "Actual", "Pass", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _style_header(ws, start, len(headers))

    GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    rows = generate_commissioning_checklist(config)
    for i, row in enumerate(rows):
        r = start + 1 + i
        for c, key in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            cell.font = DATA_FONT
            # Green fill for "Actual", "Pass", "Notes" columns (field entry)
            if key in ("Actual", "Pass", "Notes"):
                cell.fill = GREEN_FILL

    for c, w in enumerate([35, 10, 6, 20, 35, 30, 20, 15, 8, 25], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
