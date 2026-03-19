"""
SBS Composition Engine v2 — Vismation Deliverable Export

Generates a comprehensive Excel workbook for Vismation to build their
Visio stencil tool. Contains:
  Tab 1: All 55 AHU Add-On Options with I/O row assignments
  Tab 2: Standard I/O Row Map (master reference — AHU-VAV)
  Tab 3: Standard Configurations (common combos)
  Tab 4: Controller Selection Guide
  Tab 5: Expansion Board Rules
  Tab 6: Feature → I/O Lookup (select features → get exact points)
"""

import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from composition.module_registry import get_module, list_modules, STANDARD_CONFIGS
from composition.assembler import assemble
from composition.program_loader import inject_program_code


# Styles
HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CAT_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
CAT_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUB_FONT = Font(name="Calibri", bold=True, size=10, color="666666")
UNUSED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YES_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))


def _hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    if sub:
        ws.cell(row=2, column=1, value=sub).font = SUB_FONT
    return 4


# ── All 55 add-on options organized by category ──
ADD_ON_OPTIONS = [
    # (category, module_id, option_name, description)
    # COOLING (6)
    ("Cooling", "clg-chw",     "CHW",      "Chilled water cooling coil with modulating valve"),
    ("Cooling", "clg-dx",      "DX-1",     "Single stage DX compressor"),
    ("Cooling", "clg-dx-2",    "DX-2",     "Two stage DX compressor"),
    ("Cooling", "clg-dx-vfd",  "DX-VFD",   "Variable speed DX compressor"),
    ("Cooling", None,          "DX-3+",    "Three+ stage DX (use DX-2 + additional staging)"),
    ("Cooling", None,          "NONE",     "No mechanical cooling"),

    # HEATING (10)
    ("Heating", "htg-hw",       "HW",       "Hot water heating coil with modulating valve"),
    ("Heating", "htg-elec",     "ELEC-1",   "Single stage electric heat"),
    ("Heating", "htg-elec-2",   "ELEC-2",   "Two stage electric heat"),
    ("Heating", "htg-elec-3",   "ELEC-3",   "Three stage electric heat"),
    ("Heating", "htg-elec-scr", "ELEC-SCR", "Modulating electric heat (SCR controller)"),
    ("Heating", "htg-gas",      "GAS-1",    "Single stage gas-fired heat"),
    ("Heating", None,           "GAS-2",    "Two stage gas-fired heat"),
    ("Heating", "htg-gas-mod",  "GAS-MOD",  "Modulating gas-fired heat"),
    ("Heating", None,           "STEAM",    "Steam heating coil"),
    ("Heating", None,           "NONE",     "No heating"),

    # PREHEAT (5)
    ("Preheat", "ph-hw",     "PH-HW",     "Hot water preheat coil + pump"),
    ("Preheat", "ph-elec",   "PH-ELEC",   "Electric preheat coil"),
    ("Preheat", None,        "PH-STEAM",  "Steam preheat coil"),
    ("Preheat", "ph-glycol", "PH-GLYCOL", "Glycol preheat coil + pump"),
    ("Preheat", None,        "NONE",      "No preheat"),

    # SUPPLY FAN (4)
    ("Supply Fan", "fan-sf-vfd", "SF-VFD",  "Variable frequency drive"),
    ("Supply Fan", "fan-sf-cs",  "SF-CS",   "Constant speed (start/stop only)"),
    ("Supply Fan", "fan-sf-ecm", "SF-ECM",  "Electronically commutated motor"),
    ("Supply Fan", None,         "SF-2SPD", "Two speed motor"),

    # RETURN/RELIEF (5)
    ("Return/Relief Fan", "fan-rf-vfd",  "RF-VFD",  "Return fan VFD with building pressure"),
    ("Return/Relief Fan", "fan-rf-cs",   "RF-CS",   "Return fan constant speed"),
    ("Return/Relief Fan", "fan-rlf-vfd", "RLF-VFD", "Relief fan VFD"),
    ("Return/Relief Fan", "fan-rlf-cs",  "RLF-CS",  "Relief fan constant speed"),
    ("Return/Relief Fan", None,          "NONE",    "No return/relief fan"),

    # EXHAUST FAN (3)
    ("Exhaust Fan", "fan-ef-vfd", "EF-VFD", "Exhaust fan VFD with airflow tracking"),
    ("Exhaust Fan", "fan-ef-cs",  "EF-CS",  "Exhaust fan constant speed"),
    ("Exhaust Fan", None,         "NONE",   "No exhaust fan"),

    # ECONOMIZER (5)
    ("Economizer", "econ-db",      "ECON-DB",      "Dry bulb temperature lockout"),
    ("Economizer", "econ-enth",    "ECON-ENTH",    "Enthalpy-based changeover"),
    ("Economizer", "econ-diff",    "ECON-DIFF",    "Differential enthalpy (OA vs RA)"),
    ("Economizer", "econ-diff-db", "ECON-DIFF-DB", "Differential dry bulb (OAT vs RAT)"),
    ("Economizer", None,           "NONE",         "No economizer"),

    # ENERGY RECOVERY (2)
    ("Energy Recovery", "erw",  "ERV-WHL", "Energy recovery wheel (FIXED SPEED — start/stop + bypass)"),
    ("Energy Recovery", None,   "NONE",    "No energy recovery"),

    # VENTILATION (5)
    ("Ventilation", "vent-fix",  "VENT-FIX",  "Fixed minimum OA damper position"),
    ("Ventilation", "vent-ams",  "VENT-AMS",  "Airflow measuring station"),
    ("Ventilation", "dcv-co2",   "DCV-CO2",   "Demand controlled ventilation — CO2 sensor"),
    ("Ventilation", "dcv-occ",   "DCV-OCC",   "Demand controlled ventilation — occupancy sensor"),
    ("Ventilation", "vent-100",  "VENT-100",  "100% outside air (DOAS/MAU)"),

    # HUMIDITY (5)
    ("Humidity", "hum-stm",   "HUM-STM",   "Steam humidifier"),
    ("Humidity", "hum-elec",  "HUM-ELEC",  "Electric humidifier"),
    ("Humidity", "hum-ultra", "HUM-ULTRA", "Ultrasonic humidifier"),
    ("Humidity", "dehum-sc",  "DEHUM-SC",  "Dehumidification (subcooling + reheat)"),
    ("Humidity", None,        "NONE",      "No humidity control"),

    # DAMPERS (5)
    ("Dampers", None, "DMP-OA-RA",    "OA + RA dampers (standard with economizer)"),
    ("Dampers", None, "DMP-OA-RA-EA", "OA + RA + EA dampers (with exhaust)"),
    ("Dampers", None, "DMP-OA",       "OA damper only (100% OA unit)"),
    ("Dampers", None, "DMP-RLF",      "Relief damper"),
    ("Dampers", None, "DMP-ISO",      "Isolation damper"),

    # SAFETY (8)
    ("Safety", "safe-filter",       "FLTR-DP",   "Main filter DP switch (always present)"),
    ("Safety", "safe-smoke",        "SMOKE-SA",  "Supply air duct smoke detector (always)"),
    ("Safety", "safe-freeze",       "FREEZE-LL", "Low limit freeze stat (always)"),
    ("Safety", "safe-freeze-dp",    "FREEZE-DP", "Freeze stat DP across coil face"),
    ("Safety", "safe-hi-static",    "HI-STATIC", "Supply air high static pressure switch"),
    ("Safety", "safe-fire-sd",      "FIRE-SD",   "Fire shutdown contact"),
    ("Safety", "safe-cond-ovf",     "COND-OVF",  "Condensate overflow switch"),
    ("Safety", "safe-filter-oa",    "FLTR-OA",   "OA pre-filter DP switch"),
]


def generate_vismation_excel() -> bytes:
    """Generate the Vismation deliverable workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    _build_addon_options_tab(wb)
    _build_master_row_map_tab(wb)
    _build_standard_configs_tab(wb)
    _build_controller_guide_tab(wb)
    _build_expansion_rules_tab(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_addon_options_tab(wb):
    ws = wb.create_sheet("Add-On Options (55)")
    start = _title(ws, "SBS AHU ADD-ON OPTIONS — 55 FEATURES",
                   "Each option can be selected per category. Select one per mutually exclusive group.")

    headers = ["#", "Category", "Option Code", "Description",
               "Input Rows Used", "Output Rows Used",
               "Values Added", "Loops Added", "Programs Added"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _hdr(ws, start, len(headers))

    r = start + 1
    last_cat = ""
    num = 0
    for cat, mod_id, option, desc in ADD_ON_OPTIONS:
        num += 1

        # Category separator
        if cat != last_cat:
            ws.cell(row=r, column=1, value="")
            ws.cell(row=r, column=2, value=cat.upper())
            ws.cell(row=r, column=2).font = CAT_FONT
            ws.cell(row=r, column=2).fill = CAT_FILL
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = CAT_FILL
                ws.cell(row=r, column=c).border = THIN
            r += 1
            last_cat = cat

        ws.cell(row=r, column=1, value=num)
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=option)
        ws.cell(row=r, column=4, value=desc)

        if mod_id:
            try:
                mod = get_module(mod_id)
                in_rows = ", ".join([f"{p.row}" for p in mod.inputs]) or "—"
                out_rows = ", ".join([f"{p.row}" for p in mod.outputs]) or "—"
                ws.cell(row=r, column=5, value=in_rows)
                ws.cell(row=r, column=6, value=out_rows)
                ws.cell(row=r, column=7, value=len(mod.values))
                ws.cell(row=r, column=8, value=len(mod.loops))
                ws.cell(row=r, column=9, value=len(mod.programs))
            except Exception:
                ws.cell(row=r, column=5, value="—")
        else:
            ws.cell(row=r, column=5, value="—")
            ws.cell(row=r, column=6, value="—")

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    for c, w in enumerate([5, 18, 14, 55, 20, 20, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_master_row_map_tab(wb):
    ws = wb.create_sheet("Master I_O Row Map")
    start = _title(ws, "STANDARD AHU-VAV I/O ROW MAP",
                   "Row number = physical terminal number. Empty rows NEVER filled.")

    # INPUTS section
    ws.cell(row=start, column=1, value="INPUTS").font = Font(bold=True, size=12, color="1F4E79")
    start += 1
    headers = ["Row", "Point Name", "Type", "Range", "Units", "Description", "Feature Module", "Always Present"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _hdr(ws, start, len(headers))

    # Load row map JSON
    import os
    map_path = os.path.join(os.path.dirname(__file__), "row_maps", "ahu_vav.json")
    with open(map_path) as f:
        row_map = json.load(f)

    max_in = max(int(k) for k in row_map["inputs"].keys())
    r = start + 1
    for row_num in range(1, max_in + 1):
        key = str(row_num)
        if key in row_map["inputs"]:
            pt = row_map["inputs"][key]
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"{{device-name}}-{pt['name']}")
            ws.cell(row=r, column=3, value=pt["type"])
            ws.cell(row=r, column=4, value=pt["range"])
            ws.cell(row=r, column=5, value=pt.get("units", ""))
            ws.cell(row=r, column=6, value=pt["desc"])
            ws.cell(row=r, column=7, value=pt["module"])
            always = "YES" if pt.get("always") else ""
            ws.cell(row=r, column=8, value=always)
            if always:
                ws.cell(row=r, column=8).fill = YES_FILL
        else:
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"--- unused ---")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    # OUTPUTS section
    r += 2
    ws.cell(row=r, column=1, value="OUTPUTS").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    out_headers = ["Row", "Point Name", "Type", "Range", "Min V", "Max V", "Description", "Feature Module", "Reverse Acting"]
    for c, h in enumerate(out_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _hdr(ws, r, len(out_headers))
    r += 1

    max_out = max(int(k) for k in row_map["outputs"].keys())
    for row_num in range(1, max_out + 1):
        key = str(row_num)
        if key in row_map["outputs"]:
            pt = row_map["outputs"][key]
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"{{device-name}}-{pt['name']}")
            ws.cell(row=r, column=3, value=pt["type"])
            ws.cell(row=r, column=4, value=pt["range"])
            ws.cell(row=r, column=5, value=pt.get("min_v", ""))
            ws.cell(row=r, column=6, value=pt.get("max_v", ""))
            ws.cell(row=r, column=7, value=pt["desc"])
            ws.cell(row=r, column=8, value=pt["module"])
            rev = "YES" if pt.get("reverse") else ""
            ws.cell(row=r, column=9, value=rev)
        else:
            ws.cell(row=r, column=1, value=row_num)
            ws.cell(row=r, column=2, value=f"--- unused ---")
            for c in range(1, len(out_headers) + 1):
                ws.cell(row=r, column=c).fill = UNUSED_FILL
        for c in range(1, len(out_headers) + 1):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    for c, w in enumerate([6, 35, 6, 20, 8, 8, 45, 18, 12], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_standard_configs_tab(wb):
    ws = wb.create_sheet("Standard Configurations")
    start = _title(ws, "STANDARD AHU CONFIGURATIONS",
                   "Pre-built configs — select one and generate complete controller package")

    headers = ["Config Name", "Description",
               "Cooling", "Heating", "Preheat", "Supply Fan",
               "Return/Exhaust", "Economizer", "ERW", "Ventilation",
               "Inputs", "Outputs", "Controller", "Expansion"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _hdr(ws, start, len(headers))

    r = start + 1
    for cid, std in STANDARD_CONFIGS.items():
        config = assemble(std["modules"])
        exp = f"{config.expansion_count}x {config.expansion_model}" if config.expansion_count else "None"

        # Determine what's in each category
        mods = set(config.selected_modules)
        cooling = next((m for m in mods if m.startswith("clg-")), "NONE")
        heating = next((m for m in mods if m.startswith("htg-")), "NONE")
        preheat = next((m for m in mods if m.startswith("ph-")), "NONE")
        sf = next((m for m in mods if m.startswith("fan-sf-")), "NONE")
        rf_ef = next((m for m in mods if m.startswith(("fan-rf-", "fan-ef-", "fan-rlf-"))), "NONE")
        econ = next((m for m in mods if m.startswith("econ-")), "NONE")
        erw_mod = "ERW" if "erw" in mods else "NONE"
        vent = next((m for m in mods if m in ("vent-fix","vent-ams","dcv-co2","dcv-occ","vent-100")), "NONE")

        ws.cell(row=r, column=1, value=cid)
        ws.cell(row=r, column=2, value=std["name"])
        ws.cell(row=r, column=3, value=cooling)
        ws.cell(row=r, column=4, value=heating)
        ws.cell(row=r, column=5, value=preheat)
        ws.cell(row=r, column=6, value=sf)
        ws.cell(row=r, column=7, value=rf_ef)
        ws.cell(row=r, column=8, value=econ)
        ws.cell(row=r, column=9, value=erw_mod)
        ws.cell(row=r, column=10, value=vent)
        ws.cell(row=r, column=11, value=f"{len(config.inputs)} (row {config.highest_input_row})")
        ws.cell(row=r, column=12, value=f"{len(config.outputs)} (row {config.highest_output_row})")
        ws.cell(row=r, column=13, value=config.controller_model)
        ws.cell(row=r, column=14, value=exp)

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    for c, w in enumerate([28, 42, 12, 12, 12, 14, 14, 14, 8, 12, 16, 16, 10, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_controller_guide_tab(wb):
    ws = wb.create_sheet("Controller Selection")
    start = _title(ws, "RELIABLE CONTROLS CONTROLLER SELECTION GUIDE",
                   "Select based on highest occupied I/O row. AHUs use MPS family.")

    headers = ["Controller", "Family", "Base Inputs", "Base Outputs",
               "Max Expansion", "Expansion Module", "Exp Inputs", "Exp Outputs",
               "Max Total Inputs", "Max Total Outputs", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _hdr(ws, start, len(headers))

    controllers = [
        ("MPS",      "MACH-ProSys",     12, 8, 7, "MPP-IO-U", 12, 8, 96, 64, "Primary AHU controller — Ethernet + 2x RS-485"),
        ("MPWS",     "MACH-ProWebSys",  12, 8, 7, "MPP-IO-U", 12, 8, 96, 64, "MPS with built-in web server"),
        ("MPV-LCD",  "MACH-ProView",    12, 8, 7, "MPP-IO-U", 12, 8, 96, 64, "MPS with color touchscreen LCD"),
        ("MP2",      "MACH-Pro2",       12, 8, 3, "MPP-IO-U", 12, 8, 48, 32, "Midsize — max 3 expansion boards"),
        ("MP1",      "MACH-Pro1",       12, 8, 0, "—",        0,  0, 12,  8, "Midsize — no expansion"),
        ("MPC",      "MACH-ProCom",      0, 0, 8, "MPP-IO-U", 12, 8, 96, 64, "Network controller — expansion I/O only"),
        ("MPWC",     "MACH-ProWebCom",   0, 0, 8, "MPP-IO-U", 12, 8, 96, 64, "MPC with web server"),
        ("MPA-36",   "MACH-ProAir",      3, 6, 0, "—",        0,  0,  3,  6, "Small AHU — 3UI + 6UO + 8BI + 8BO"),
        ("MPA-35",   "MACH-ProAir",      3, 5, 0, "—",        0,  0,  3,  5, "Small AHU — 3UI + 5UO + 8BI + 8BO"),
        ("MPA-34",   "MACH-ProAir",      3, 4, 0, "—",        0,  0,  3,  4, "Small AHU — 3UI + 4UO + 8BI + 8BO"),
        ("MPZ-88",   "MACH-ProZone",     8, 8, 0, "—",        0,  0,  8,  8, "Zone — FCU, UH, EF"),
        ("MPZ-44",   "MACH-ProZone",     4, 4, 0, "—",        0,  0,  4,  4, "Zone — simple FCU, EF"),
    ]

    r = start + 1
    for ctrl in controllers:
        for c, val in enumerate(ctrl, 1):
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    # Selection logic
    r += 2
    ws.cell(row=r, column=1, value="SELECTION LOGIC").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    rules = [
        "1. Count highest occupied INPUT row and highest occupied OUTPUT row",
        "2. Base controller covers first N inputs and M outputs (see table above)",
        "3. Each expansion board adds more I/O (MPP-IO-U = 12 inputs + 8 outputs)",
        "4. Input 13 on MPS = expansion board 1, row 1",
        "5. Select the smallest controller + expansion count that covers all rows",
        "6. NEVER fill gaps — if row 34 is used but rows 28, 35 are empty, that's correct",
        "7. Standard AHU: MPS or MPWS (prefer MPWS for web access)",
        "8. Small AHU (≤3 inputs, ≤6 outputs, no expansion): MPA family",
        "9. Network/plant: MPC or MPWC",
    ]
    for rule in rules:
        ws.cell(row=r, column=1, value=rule).font = DATA_FONT
        r += 1

    for c, w in enumerate([12, 18, 12, 14, 14, 16, 12, 12, 16, 16, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _build_expansion_rules_tab(wb):
    ws = wb.create_sheet("Expansion Rules")
    start = _title(ws, "EXPANSION BOARD I/O MAPPING",
                   "MPS base: 12 inputs, 8 outputs. Each MPP-IO-U adds 12 in + 8 out.")

    headers = ["Board", "Input Rows", "Output Rows", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _hdr(ws, start, len(headers))

    boards = [
        ("MPS Base",      "1–12",   "1–8",    "Onboard I/O"),
        ("Expansion 1",   "13–24",  "9–16",   "First MPP-IO-U board"),
        ("Expansion 2",   "25–36",  "17–24",  "Second MPP-IO-U board"),
        ("Expansion 3",   "37–48",  "25–32",  "Third MPP-IO-U board"),
        ("Expansion 4",   "49–60",  "33–40",  "Fourth MPP-IO-U board"),
        ("Expansion 5",   "61–72",  "41–48",  "Fifth MPP-IO-U board"),
        ("Expansion 6",   "73–84",  "49–56",  "Sixth MPP-IO-U board"),
        ("Expansion 7",   "85–96",  "57–64",  "Seventh MPP-IO-U board (max for MPS)"),
    ]

    r = start + 1
    for board in boards:
        for c, val in enumerate(board, 1):
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="IMPORTANT RULES").font = Font(bold=True, size=12, color="B71C1C")
    r += 1
    rules = [
        "• Row number = physical terminal number on the controller/expansion board",
        "• ALL inputs are universal — range code in software determines AI vs BI",
        "• ALL outputs are universal — range code determines AO vs BO",
        "• If a feature is NOT selected, the row stays EMPTY — never shift points up",
        "• Empty rows in the middle are CORRECT and INTENTIONAL",
        "• One extra expansion board costs less than one service call on a non-standard layout",
        "• Every job of the same type has the same wiring layout — consistency is everything",
    ]
    for rule in rules:
        ws.cell(row=r, column=1, value=rule).font = DATA_FONT
        r += 1

    for c, w in enumerate([18, 14, 14, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


if __name__ == "__main__":
    data = generate_vismation_excel()
    out = "/srv/dfa/drops/SBS-Vismation-AHU-Standard-IO.xlsx"
    with open(out, "wb") as f:
        f.write(data)
    print(f"Vismation deliverable saved: {out} ({len(data):,} bytes)")
