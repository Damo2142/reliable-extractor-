"""Verify all FACTORY-flagged Excel points appear in the .pan, with range."""
import sys, glob, os, struct
from openpyxl import load_workbook
from composition.pan_compiler import crc16_kermit
from composition.pan_intake import _read_index_table, _parse_tlv_records

PKG = sys.argv[1]
xlsx = os.path.join(PKG, "RC-Studio-Output.xlsx")
pan = glob.glob(os.path.join(PKG, "*.pan"))[0]
raw = open(pan, "rb").read()

# Collect FACTORY rows from Inputs/Outputs/Values
wb = load_workbook(xlsx)
factory = []  # (name, objtype)
for tab, mod_col in (("Inputs", 7), ("Outputs", 9), ("Values", 9)):
    ws = wb[tab]
    hr = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] in ("Terminal", "Instance"):
            hr = i
            break
    if hr is None:
        continue
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        nm = row[1] if len(row) > 1 else None
        mod = row[mod_col] if len(row) > mod_col else None
        ty = row[2] if tab != "Values" else (str(row[0])[:2] if row[0] else "")
        if nm and str(mod).strip() == "FACTORY":
            factory.append((str(nm).strip(), str(ty).strip()))

# Walk .pan: name from _parse_tlv_records (reliable); range via byte-pattern
# search for the vendor-range record [00 04 1D 91 <val>] (position-independent).
def block_name_range(payload):
    props = _parse_tlv_records(payload)
    name = props.get(0x4D)
    if not isinstance(name, str):
        name = None
    rng = None
    i = payload.find(b'\x00\x04\x1d\x91')
    if i != -1 and i + 4 < len(payload):
        rng = payload[i + 4]
    return name, rng

pan_names = {}
ok = tot = 0
for tid, off, cnt in _read_index_table(raw):
    pos = off
    for _ in range(cnt):
        if pos + 12 > len(raw):
            break
        sz = struct.unpack_from('<H', raw, pos)[0]
        t = sz + 2
        if pos + t > len(raw):
            break
        payload = raw[pos + 12:pos + t]
        if crc16_kermit(payload) == struct.unpack_from('>H', raw, pos + 10)[0]:
            ok += 1
        tot += 1
        nm, rng = block_name_range(payload)
        if nm:
            pan_names[nm] = rng
        pos += t

print(f"{os.path.basename(pan)}")
print(f"  FACTORY points in Excel: {len(factory)}")
allpresent = True
for nm, ty in factory:
    present = nm in pan_names
    if not present:
        allpresent = False
    rng = pan_names.get(nm)
    rstr = f"range={rng}" if rng is not None else "range=?"
    print(f"    [{'OK ' if present else 'MISS'}] {ty:3s} {nm:28s} {rstr}")
print(f"  CRC: {ok}/{tot}  {'ALL PASS' if ok == tot else 'FAIL'}")
print(f"  ALL FACTORY PRESENT: {allpresent}")
