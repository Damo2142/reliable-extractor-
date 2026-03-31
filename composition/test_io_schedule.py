"""
End-to-end test for IO schedule export/import feature.

Path 1: Generate standard CHW air cooled 2+2 package, confirm it works.
Path 2: Export IO schedule, move 3 points to different terminals,
        import back, generate package, confirm .pan has updated terminals
        and all point names intact.
Validation: Reject modified point names, reject terminal conflicts.

Run: cd /srv/reliable-generator-dev && venv/bin/python composition/test_io_schedule.py
"""
import sys, json, io, os, tempfile, shutil
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pathlib import Path
from composition.module_registry import chwp_assemble
from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
from composition.io_schedule import (
    store_config, export_io_schedule, import_io_schedule,
    apply_terminal_overrides, get_stored_config,
)
import openpyxl


def _load_program_code(merged, prg_dir):
    for prg in merged['programs']:
        if prg.code and len(prg.code) > 50:
            continue
        bas_path = prg_dir / prg.filename
        if bas_path.exists():
            prg.code = bas_path.read_text()


def _compile_to_pan(merged, alarm_code, prg_dir, config_name):
    wb = write_excel(merged, generate_trends(merged), alarm_code, config_name)
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()
    tmp = tempfile.mkdtemp(prefix="sbs-test-")
    try:
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_bytes)
        tmp_prg = os.path.join(tmp, "programs")
        os.makedirs(tmp_prg, exist_ok=True)
        for prg in merged['programs']:
            bas_path = prg_dir / prg.filename
            code = prg.code if (prg.code and len(prg.code) > 50) else (
                bas_path.read_text() if bas_path.exists() else f"10 REM {prg.name}\n")
            with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                f.write(code)
        with open(os.path.join(tmp_prg, "PRG-ALARMS.bas"), "w") as f:
            f.write(alarm_code)
        from compile_from_excel import compile_package
        return compile_package(tmp, verbose=False)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_path1_standard_chw():
    """PATH 1: Standard CHW air cooled 2+2 generates correctly."""
    print("PATH 1: Standard CHW Air Cooled 2+2")
    params = {'num_chillers': 2, 'num_pri_pumps': 2, 'num_sec_pumps': 2,
              'num_dp_sensors': 2, 'config_name': 'CHW-Air-2+2'}
    prg_dir = Path(__file__).parent / 'programs' / 'chw_plant'

    modules = chwp_assemble(params)
    merged = merge_modules(modules)
    alarm_code = generate_alarm_bas(merged)
    _load_program_code(merged, prg_dir)

    no_code = [p.name for p in merged['programs'] if not p.code or len(p.code) < 50]
    assert not no_code, f"Programs without code: {no_code}"

    pan = _compile_to_pan(merged, alarm_code, prg_dir, 'CHW-Air-2+2')
    assert pan[:2] == b'\xc0\xba', "Bad .pan magic"
    assert len(pan) > 1000, "Pan too small"

    print(f"  {len(merged['inputs'])} inputs, {len(merged['outputs'])} outputs, "
          f"{len(merged['programs'])} programs, .pan={len(pan)} bytes")
    print("  PASS ✓")
    return merged, pan


def test_path2_export_import(merged_orig, pan_orig):
    """PATH 2: Export, modify 3 terminals, import, regenerate, verify."""
    print("\nPATH 2: IO Export → Modify → Import → Regenerate")
    params = {'num_chillers': 2, 'num_pri_pumps': 2, 'num_sec_pumps': 2,
              'num_dp_sensors': 2, 'config_name': 'CHW-Air-2+2'}
    prg_dir = Path(__file__).parent / 'programs' / 'chw_plant'

    # Store config
    config_id = store_config(
        merged_orig['inputs'], merged_orig['outputs'],
        "CHW-PLANT-AIR", "MPS", "chwp", {"params": params},
    )

    # Export
    io_excel = export_io_schedule(config_id)
    assert len(io_excel) > 1000

    # Modify 3 terminals
    wb = openpyxl.load_workbook(io.BytesIO(io_excel))
    ws = wb[wb.sheetnames[0]]
    for row_idx in range(5, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value
        if name == 'PCHWP1-STS':
            ws.cell(row=row_idx, column=1, value='IN30')
        elif name == 'PCHWP2-STS':
            ws.cell(row=row_idx, column=1, value='IN31')
        elif name == 'SCHWP1-S/S':
            ws.cell(row=row_idx, column=1, value='OUT20')
    mod_buf = io.BytesIO()
    wb.save(mod_buf)

    # Import
    result = import_io_schedule(config_id, mod_buf.getvalue())
    assert result['ok']
    assert result['overrides_applied'] == 3
    print(f"  Import: {result['overrides_applied']} overrides applied")

    # Re-assemble with overrides
    modules2 = chwp_assemble(params)
    merged2 = merge_modules(modules2)
    alarm_code2 = generate_alarm_bas(merged2)
    _load_program_code(merged2, prg_dir)
    apply_terminal_overrides(config_id, merged2['inputs'], merged2['outputs'])

    # Verify terminals changed
    in_map = {p.name: p.row for p in merged2['inputs']}
    out_map = {p.name: p.row for p in merged2['outputs']}
    assert in_map['PCHWP1-STS'] == 30
    assert in_map['PCHWP2-STS'] == 31
    assert out_map['SCHWP1-S/S'] == 20
    print("  Terminal overrides verified ✓")

    # Verify names intact
    assert sorted(p.name for p in merged_orig['inputs']) == sorted(p.name for p in merged2['inputs'])
    assert sorted(p.name for p in merged_orig['outputs']) == sorted(p.name for p in merged2['outputs'])
    print("  All point names intact ✓")

    # Regenerate .pan
    pan2 = _compile_to_pan(merged2, alarm_code2, prg_dir, 'CHW-Air-2+2-mod')
    assert pan2[:2] == b'\xc0\xba'
    assert pan_orig != pan2, ".pan should differ after terminal change"
    print(f"  .pan differs from original ✓ ({len(pan2)} bytes)")

    # Validation: reject modified names
    cfg = get_stored_config(config_id)
    cfg['terminal_overrides'] = {}
    wb_bad = openpyxl.load_workbook(io.BytesIO(io_excel))
    ws_bad = wb_bad[wb_bad.sheetnames[0]]
    for r in range(5, ws_bad.max_row + 1):
        if ws_bad.cell(r, 2).value == 'CHWS-T':
            ws_bad.cell(r, 2, value='CHWS-TEMP')
            break
    bad_buf = io.BytesIO()
    wb_bad.save(bad_buf)
    try:
        import_io_schedule(config_id, bad_buf.getvalue())
        assert False, "Should have rejected"
    except ValueError:
        print("  Rejects modified names ✓")

    # Validation: reject conflicts
    cfg['terminal_overrides'] = {}
    wb_c = openpyxl.load_workbook(io.BytesIO(io_excel))
    ws_c = wb_c[wb_c.sheetnames[0]]
    for r in range(5, ws_c.max_row + 1):
        if ws_c.cell(r, 2).value in ('CHWS-T', 'OAT'):
            ws_c.cell(r, 1, value='IN1')
    c_buf = io.BytesIO()
    wb_c.save(c_buf)
    try:
        import_io_schedule(config_id, c_buf.getvalue())
        assert False, "Should have rejected"
    except ValueError:
        print("  Rejects terminal conflicts ✓")

    print("  PASS ✓")


if __name__ == "__main__":
    print("=" * 60)
    merged, pan = test_path1_standard_chw()
    test_path2_export_import(merged, pan)
    print("\n" + "=" * 60)
    print("ALL TESTS PASS ✓")
    print("=" * 60)
