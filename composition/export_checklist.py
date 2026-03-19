"""
Generate the "What I Need" Excel checklist for Dave to fill.
One tab per system, each listing exactly what to export and where to drop it.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HDR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SEC_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
DATA_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=10, color="B71C1C", bold=True)
DONE_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

SYSTEMS = [
    {
        "tab": "Boiler Plant",
        "folder": "boiler-plant",
        "controller": "Cascade leader + follower (2 controllers or 1?)",
        "notes": "This is the #1 priority — gives me the plant control standard",
        "items": [
            ("Inputs list", "All physical inputs — same format as A201-inputs.txt (terminal, name, type, range, units)"),
            ("Outputs list", "All physical outputs — same format as A201-outputs.txt (terminal, name, type, range, min/max V)"),
            ("Values list", "All AV/BV/MV — same format as A201-values.txt (instance, name, type, default, units)"),
            ("Loops list", "All PID loops — instance, name, input, setpoint, P, I, D, action"),
            ("Tables list", "All scaling tables — number, name, data points (V→engineering units)"),
            ("Programs list", "All programs — PRG#, name, enabled, description"),
            (".bas files", "Export every .bas program file individually"),
            ("Schedules", "Any weekly schedules — instance, name, default state"),
            ("Trends", "List of trend logs (or just say 'trend everything')"),
            ("System Groups", "Graphic page names"),
            ("Controller model", "What controller? MPS? How many expansion boards?"),
            ("Equipment details", "How many boilers? Lead/lag? Cascade or direct fire control? Outdoor reset curve?"),
        ],
    },
    {
        "tab": "Chiller Plant",
        "folder": "chiller-plant",
        "controller": "Plant controller for chiller staging",
        "notes": "If not on this site, any site will do. Or I can build from SBS standard + PPS.",
        "items": [
            ("Inputs list", "All physical inputs"),
            ("Outputs list", "All physical outputs"),
            ("Values list", "All AV/BV/MV"),
            ("Loops list", "PID loops"),
            ("Programs list + .bas files", "All programs with code"),
            ("Controller model", "MPS? Expansion count?"),
            ("Equipment details", "How many chillers? Air-cooled or water-cooled? VFD compressors? Staging method?"),
        ],
    },
    {
        "tab": "Cooling Tower",
        "folder": "cooling-tower",
        "controller": "Usually same controller as chiller plant or separate?",
        "notes": "If on same controller as chiller, just note that. Otherwise export separately.",
        "items": [
            ("Inputs list", "All physical inputs"),
            ("Outputs list", "All physical outputs — fan staging/VFD, isolation valves"),
            ("Values list", "All AV/BV/MV"),
            ("Loops list", "Basin temp control, approach control"),
            ("Programs list + .bas files", "All programs with code"),
            ("Equipment details", "How many cells? Fan staging or VFD? Free cooling bypass?"),
        ],
    },
    {
        "tab": "HW Pump Station",
        "folder": "pump-station-hw",
        "controller": "Separate controller or on boiler plant controller?",
        "notes": "Primary/secondary? Variable primary? DP control?",
        "items": [
            ("Inputs list", "All physical inputs — DP sensor, flow, pump status"),
            ("Outputs list", "All physical outputs — pump commands, VFD speeds"),
            ("Values list", "All AV/BV/MV"),
            ("Loops list", "DP control loop, any others"),
            ("Programs list + .bas files", "All programs with code"),
            ("Equipment details", "How many pumps? Lead/lag? VFD? Primary/secondary or variable primary?"),
        ],
    },
    {
        "tab": "CHW Pump Station",
        "folder": "pump-station-chw",
        "controller": "Separate controller or on chiller plant controller?",
        "notes": "Same questions as HW pump station",
        "items": [
            ("Inputs list", "All physical inputs"),
            ("Outputs list", "All physical outputs"),
            ("Values list", "All AV/BV/MV"),
            ("Loops list", "DP control loop"),
            ("Programs list + .bas files", "All programs with code"),
            ("Equipment details", "How many pumps? Lead/lag? VFD? Primary/secondary?"),
        ],
    },
    {
        "tab": "VAV HW Reheat",
        "folder": "vav-hw-reheat",
        "controller": "RCFA model? Which variant?",
        "notes": "Get both floating and modulating if you have them. One is fine though.",
        "items": [
            ("Inputs list", "All physical inputs (flow sensor, space temp, etc.)"),
            ("Outputs list", "All physical outputs (damper, reheat valve — note floating vs modulating)"),
            ("Values list", "All AV/BV/MV (setpoints, flow min/max, etc.)"),
            ("Loops list", "Flow loop, space temp loop, reheat loop"),
            ("Programs list + .bas files", "All programs with code"),
            ("Controller model", "RCFA-M-34? RCFA-M-35?"),
            ("Equipment details", "Floating or modulating valve? Factory or field damper actuator?"),
        ],
    },
    {
        "tab": "VAV Cooling Only",
        "folder": "vav-cooling-only",
        "controller": "RCFA model?",
        "notes": "Simplest VAV — just damper, no reheat",
        "items": [
            ("Inputs list", "All physical inputs"),
            ("Outputs list", "All physical outputs (damper only)"),
            ("Values list", "All AV/BV/MV"),
            ("Loops list", "Flow loop, space temp"),
            ("Programs list + .bas files", "All programs with code"),
            ("Controller model", "RCFA-M-33? RCFA-M-34?"),
        ],
    },
    {
        "tab": "Fan Coil Unit",
        "folder": "fan-coil",
        "controller": "MPZ? Which model?",
        "notes": "2-pipe or 4-pipe? What fan speeds?",
        "items": [
            ("Inputs list", "All physical inputs (space temp, filter, etc.)"),
            ("Outputs list", "All physical outputs (fan, valves)"),
            ("Values list", "All AV/BV/MV"),
            ("Programs list + .bas files", "All programs with code"),
            ("Controller model", "MPZ-44? MPZ-88?"),
            ("Equipment details", "2-pipe or 4-pipe? How many fan speeds? HW/CHW/both?"),
        ],
    },
    {
        "tab": "Unit Ventilator",
        "folder": "unit-vent",
        "controller": "MPZ? MPA?",
        "notes": "Classroom style? With OA damper?",
        "items": [
            ("Inputs list", "All physical inputs"),
            ("Outputs list", "All physical outputs (fan, valves, OA damper)"),
            ("Values list", "All AV/BV/MV"),
            ("Programs list + .bas files", "All programs with code"),
            ("Controller model", "What model?"),
            ("Equipment details", "HW/CHW/DX? OA damper? How many fan speeds? CO2?"),
        ],
    },
    {
        "tab": "Exhaust Fan",
        "folder": "exhaust-fan",
        "controller": "MPZ-44? RC-SSC?",
        "notes": "Scheduled, interlock, VFD — whatever types you have on site",
        "items": [
            ("Inputs list", "Status, DP, any sensors"),
            ("Outputs list", "Start/stop, VFD speed if applicable"),
            ("Values list", "Any AV/BV"),
            ("Programs list + .bas files", "All programs with code"),
            ("Equipment details", "Scheduled? Interlock? VFD? Constant speed? Kitchen hood?"),
        ],
    },
    {
        "tab": "Unit Heater",
        "folder": "unit-heater",
        "controller": "MPZ-44? Simple thermostat?",
        "notes": "Cabinet, wall-mount, overhead — whatever you have",
        "items": [
            ("Inputs list", "Space temp, status"),
            ("Outputs list", "Fan, valve or contactor"),
            ("Values list", "Setpoints"),
            ("Programs list + .bas files", "Program code"),
            ("Equipment details", "HW or electric? Fan type? Space temp or duct mount?"),
        ],
    },
    {
        "tab": "Radiant Heat",
        "folder": "radiant-heat",
        "controller": "MPZ? Simple?",
        "notes": "In-floor, overhead panel — whatever you have",
        "items": [
            ("Inputs list", "Space temp, slab temp, supply water temp"),
            ("Outputs list", "Valve, pump"),
            ("Values list", "Setpoints"),
            ("Programs list + .bas files", "Program code"),
            ("Equipment details", "In-floor or overhead? HW valve modulating? Slab sensor?"),
        ],
    },
]


def generate():
    wb = Workbook()
    wb.remove(wb.active)

    # Summary tab first
    ws = wb.create_sheet("OVERVIEW")
    ws.cell(row=1, column=1, value="SBS COMPOSITION ENGINE — EXPORT CHECKLIST").font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=2, column=1, value="Drop files to: /srv/dfa/drops/plant-exports/{folder}/").font = Font(size=10, color="666666")
    ws.cell(row=3, column=1, value="Same format as the A201 AHU export — text files + individual .bas programs").font = Font(size=10, color="666666")

    headers = ["#", "System", "Drop Folder", "Controller", "Priority", "Status"]
    r = 5
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
        ws.cell(row=r, column=c).font = HDR_FONT
        ws.cell(row=r, column=c).fill = HDR_FILL
        ws.cell(row=r, column=c).border = THIN
    r += 1

    priorities = {
        "Boiler Plant": "HIGH",
        "VAV HW Reheat": "HIGH",
        "VAV Cooling Only": "HIGH",
        "Fan Coil Unit": "MEDIUM",
        "Unit Ventilator": "MEDIUM",
        "Exhaust Fan": "MEDIUM",
        "HW Pump Station": "MEDIUM",
        "Chiller Plant": "MEDIUM",
        "Cooling Tower": "LOW",
        "CHW Pump Station": "LOW",
        "Unit Heater": "LOW",
        "Radiant Heat": "LOW",
    }

    for i, sys in enumerate(SYSTEMS):
        ws.cell(row=r, column=1, value=i+1)
        ws.cell(row=r, column=2, value=sys["tab"])
        ws.cell(row=r, column=3, value=f"/srv/dfa/drops/plant-exports/{sys['folder']}/")
        ws.cell(row=r, column=4, value=sys["controller"])
        ws.cell(row=r, column=5, value=priorities.get(sys["tab"], "LOW"))
        ws.cell(row=r, column=6, value="")  # Status — Dave fills in
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN
        r += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    # One tab per system
    for sys in SYSTEMS:
        ws = wb.create_sheet(sys["tab"])
        ws.cell(row=1, column=1, value=sys["tab"].upper()).font = Font(bold=True, size=14, color="1F4E79")
        ws.cell(row=2, column=1, value=f"Drop folder: /srv/dfa/drops/plant-exports/{sys['folder']}/").font = Font(size=10, color="666666")
        ws.cell(row=3, column=1, value=f"Controller: {sys['controller']}").font = Font(size=10, color="666666")
        if sys.get("notes"):
            ws.cell(row=4, column=1, value=sys["notes"]).font = NOTE_FONT

        r = 6
        headers = ["#", "What I Need", "Details / Format", "Done?"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=h)
            ws.cell(row=r, column=c).font = HDR_FONT
            ws.cell(row=r, column=c).fill = HDR_FILL
            ws.cell(row=r, column=c).border = THIN
        r += 1

        for j, (item, detail) in enumerate(sys["items"]):
            ws.cell(row=r, column=1, value=j+1)
            ws.cell(row=r, column=2, value=item)
            ws.cell(row=r, column=3, value=detail)
            ws.cell(row=r, column=4, value="")  # Done checkbox
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = DATA_FONT
                ws.cell(row=r, column=c).border = THIN
            r += 1

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 70
        ws.column_dimensions["D"].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    data = generate()
    out = "/srv/dfa/drops/SBS-Export-Checklist-For-Dave.xlsx"
    with open(out, "wb") as f:
        f.write(data)
    print(f"Checklist saved: {out} ({len(data):,} bytes)")
    print(f"\nDrop folders created at /srv/dfa/drops/plant-exports/:")
    import os
    for d in sorted(os.listdir("/srv/dfa/drops/plant-exports/")):
        print(f"  {d}/")
