"""
Investigation tool (read-only): compare what the Excel IO schedule declares
against what the compiled .pan actually contains.

Usage: PYTHONPATH=. venv/bin/python composition/_audit_excel_vs_pan.py <pkgdir>
where <pkgdir> holds RC-Studio-Output.xlsx and a *.pan file.
"""
import sys, glob, os, struct
from openpyxl import load_workbook
from composition.pan_intake import decompile_pan, _read_index_table, _parse_tlv_records

PKG = sys.argv[1]
xlsx = os.path.join(PKG, "RC-Studio-Output.xlsx")
pan = glob.glob(os.path.join(PKG, "*.pan"))[0]

wb = load_workbook(xlsx)


def _hdr_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Terminal" or (row and row[0] == "Instance"):
            return i
    return None


def excel_points(tab, name_col=1):
    """Return list of (name, type) for real (non-unused, non-blank) rows."""
    if tab not in wb.sheetnames:
        return []
    ws = wb[tab]
    hr = _hdr_row(ws)
    out = []
    if hr is None:
        return out
    hdr = [c for c in next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    try:
        ni = hdr.index("Name")
        ti = hdr.index("Type")
    except ValueError:
        return out
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        nm = row[ni] if ni < len(row) else None
        ty = row[ti] if ti < len(row) else None
        if not nm:
            continue
        if isinstance(nm, str) and ("unused" in nm or nm.startswith("---")):
            continue
        out.append((str(nm).strip(), (str(ty).strip() if ty else "")))
    return out


# ---- Excel side ----
ex_in = excel_points("Inputs")
ex_out = excel_points("Outputs")
ex_val = excel_points("Values")
excel_all = {}
for nm, ty in ex_in + ex_out + ex_val:
    excel_all[nm] = ty

# ---- .pan side ----
with open(pan, "rb") as f:
    raw = f.read()
dec = decompile_pan(raw)

pan_names = set()
for sec in ("inputs", "outputs", "values", "loops", "schedules", "tables", "trends"):
    for o in dec.get(sec, []):
        if o.get("name"):
            pan_names.add(o["name"].strip())

# decompile_pan does not categorize every object type (e.g. MO=14).
# Walk the index directly and pull the 0x4D name from every block so the
# audit sees all named objects regardless of type.
for tid, off, cnt in _read_index_table(raw):
    pos = off
    for _ in range(cnt):
        if pos + 12 > len(raw):
            break
        sz = struct.unpack_from('<H', raw, pos)[0]
        tot = sz + 2
        if pos + tot > len(raw):
            break
        props = _parse_tlv_records(raw[pos + 12:pos + tot])
        nm = props.get(0x4D)
        if isinstance(nm, str) and nm.strip():
            pan_names.add(nm.strip())
        pos += tot

print(f"PACKAGE: {PKG}")
print(f".pan   : {os.path.basename(pan)}  ({dec['file_size']} bytes)")
print(f"block counts: {dec['block_counts']}")
print(f"Excel physical+value points: {len(excel_all)}  "
      f"(in={len(ex_in)} out={len(ex_out)} val={len(ex_val)})")
print(f".pan named objects: {len(pan_names)}")
print()

missing = [(n, t) for n, t in excel_all.items() if n not in pan_names]
print(f"==== IN EXCEL, MISSING FROM .pan: {len(missing)} ====")
for n, t in sorted(missing, key=lambda x: (x[1], x[0])):
    print(f"   [{t:3s}] {n}")

# Anything in .pan inputs/outputs not in excel (factory from blank?) — informational
ex_io_names = {n for n, _ in ex_in + ex_out}
pan_io = []
for sec in ("inputs", "outputs"):
    for o in dec.get(sec, []):
        if o.get("name"):
            pan_io.append((o["type"], o["name"].strip()))
extra_io = [(t, n) for t, n in pan_io if n not in ex_io_names]
print()
print(f"==== .pan I/O objects NOT in Excel I/O (factory-from-blank etc.): {len(extra_io)} ====")
for t, n in sorted(extra_io):
    print(f"   [{t:3s}] {n}")
