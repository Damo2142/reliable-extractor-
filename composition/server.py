"""
SBS Composition Engine v2 — FastAPI Standalone Server

Standalone web server for the composition engine UI.
Runs on port 8087 (separate from DFA platform and composer).
"""

import sys
import os
import io
import json
import zipfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional

from composition.assembler import assemble, CONTROLLER_SPECS, _select_controller
from composition.excel_gen import generate_excel
from composition.program_loader import inject_program_code, number_programs, prefix_local_points, format_program_commas
from composition.module_registry import (
    list_modules, list_by_category, get_module, STANDARD_CONFIGS, EQUIPMENT_FAMILIES,
    hwp_assemble, chwp_assemble
)
from composition.io_schedule import (
    store_config, get_stored_config, export_io_schedule,
    import_io_schedule, apply_terminal_overrides, get_terminal_overrides,
)

app = FastAPI(
    title="SBS Composition Engine v2",
    description="Vendor-neutral HVAC module composition + Reliable Controls output",
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class AssembleRequest(BaseModel):
    modules: List[str]
    controller_model: str = "auto"
    equipment_family: str = "VAV-AHU"


class GenerateRequest(BaseModel):
    modules: List[str]
    controller_model: str = "auto"
    equipment_family: str = "VAV-AHU"


# --- API Endpoints ---

@app.get("/api/modules")
async def api_list_modules():
    return list_by_category()


@app.get("/api/modules/{module_id}")
async def api_get_module(module_id: str):
    try:
        mod = get_module(module_id)
    except ValueError:
        raise HTTPException(404, f"Module not found: {module_id}")
    return {
        "id": mod.id, "name": mod.name, "category": mod.category,
        "description": mod.description, "is_core": mod.is_core,
        "requires": mod.requires, "conflicts": mod.conflicts,
        "mutually_exclusive_group": mod.mutually_exclusive_group,
        "inputs": len(mod.inputs), "outputs": len(mod.outputs),
        "values": len(mod.values), "loops": len(mod.loops),
        "program_count": len(mod.programs), "soo_paragraph": mod.soo_paragraph,
        "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename,
                      "description": p.description, "enabled": p.enabled,
                      "exec_order": p.exec_order} for p in mod.programs],
        "point_summary": {
            "AI": sum(1 for p in mod.inputs if p.point_type == "AI"),
            "AO": sum(1 for p in mod.outputs if p.point_type == "AO"),
            "DI": sum(1 for p in mod.inputs if p.point_type == "DI" or p.point_type == "BI"),
            "DO": sum(1 for p in mod.outputs if p.point_type == "DO" or p.point_type == "BO"),
            "AV": sum(1 for p in mod.values if p.point_type == "AV"),
            "BV": sum(1 for p in mod.values if p.point_type == "BV"),
            "MV": sum(1 for p in mod.values if p.point_type == "MV"),
        },
    }


@app.get("/api/families")
async def api_list_families():
    return EQUIPMENT_FAMILIES


@app.get("/api/standards")
async def api_list_standards():
    return STANDARD_CONFIGS


@app.get("/api/controllers")
async def api_list_controllers():
    return {"auto": {"family": "Auto-Select", "base_in": 0, "base_out": 0, "max_exp": 0}, **CONTROLLER_SPECS}


@app.post("/api/assemble")
async def api_assemble(req: AssembleRequest):
    try:
        config = assemble(req.modules, controller_model=req.controller_model,
                          equipment_family=req.equipment_family)
        inject_program_code(config)
    except ValueError as e:
        raise HTTPException(400, str(e))

    config_id = store_config(
        config.inputs, config.outputs, config.equipment_family,
        config.controller_model, "ahu",
        {
            "modules": req.modules,
            "controller_model": req.controller_model,
            "equipment_family": req.equipment_family,
        },
    )

    return {
        "config_id": config_id,
        "modules": config.selected_modules,
        "controller": {
            "model": config.controller_model,
            "expansion_count": config.expansion_count,
            "expansion_model": config.expansion_model,
            "highest_input_row": config.highest_input_row,
            "highest_output_row": config.highest_output_row,
            "display_max_input_row": getattr(config, 'display_max_input_row', config.highest_input_row),
            "display_max_output_row": getattr(config, 'display_max_output_row', config.highest_output_row),
        },
        "counts": {
            "inputs": len(config.inputs), "outputs": len(config.outputs),
            "values": len(config.values), "loops": len(config.loops),
            "tables": len(config.tables), "programs": len(config.programs),
            "schedules": len(config.schedules), "trends": len(config.trends),
            "system_groups": len(config.system_groups),
            "max_value_inst": max((v.instance for v in config.values), default=0),
            "max_loop_inst": max((l.instance for l in config.loops), default=0),
            "max_prg_inst": max((p.instance for p in config.programs), default=0),
        },
        "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "units": p.units, "range": p.range_code, "module": p.module} for p in config.inputs],
        "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module, "reverse": p.reverse, "min_v": p.min_v, "max_v": p.max_v} for p in config.outputs],
        "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description, "module": v.module} for v in config.values],
        "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "p": l.p_band, "i": l.integral, "action": l.action, "desc": l.description} for l in config.loops],
        "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "enabled": p.enabled, "desc": p.description, "has_code": bool(p.code and len(p.code) > 50), "code": p.code or ""} for p in sorted(config.programs, key=lambda x: x.exec_order)],
        "tables": [{"instance": t.instance, "name": t.name, "in_units": t.input_units, "out_units": t.output_units, "points": len(t.data_points), "desc": t.description} for t in config.tables],
        "trends": [{"instance": t.instance, "name": t.name, "monitored": t.monitored_point, "type": t.trend_type, "interval": t.interval, "cov_delta": t.cov_delta, "buffer": t.buffer_size} for t in config.trends],
        "schedules": [{"instance": s.instance, "name": s.name, "default": s.default_state, "states": "/".join(s.states), "priority": s.priority, "desc": s.description} for s in config.schedules],
        "system_groups": [{"name": g.name, "desc": g.description} for g in config.system_groups],
        "soo": config.soo_document,
        "warnings": getattr(config, 'warnings', []),
    }


class HWPAssembleRequest(BaseModel):
    params: dict
    controller_model: str = "MPS"


@app.post("/api/hwp-assemble")
async def api_hwp_assemble(req: HWPAssembleRequest):
    """Assemble HW plant from wizard parameters."""
    try:
        modules = hwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, f"HW Plant assembly error: {str(e)}")

    # Merge modules manually (HW plant uses its own merge, not the AHU assembler)
    from composition.modules.hw_plant.core import build as _hwp_core
    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas
    merged = merge_modules(modules)
    trends = generate_trends(merged)

    # Load .bas program code — skip if already has dynamic code from assembly
    prg_dir = Path(__file__).parent / "programs" / "hw_plant"
    for prg in merged['programs']:
        if prg.code and len(prg.code) > 50:
            continue  # Dynamic code already injected by hwp_assemble
        bas_path = prg_dir / prg.filename
        if bas_path.exists():
            prg.code = bas_path.read_text()

    # Build alarm program
    alarm_code = generate_alarm_bas(merged)
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])

    # Calculate highest I/O rows
    highest_in = max((p.row for p in merged['inputs']), default=0)
    highest_out = max((p.row for p in merged['outputs']), default=0)

    config_id = store_config(
        merged['inputs'], merged['outputs'], "HW-PLANT",
        req.controller_model or "MPS", "hwp", {"params": req.params},
    )

    return {
        "config_id": config_id,
        "modules": [m.id for m in modules],
        "controller": {
            "model": req.controller_model or "MPS",
            "expansion_count": max(0, (highest_in - 8 + 11) // 12) if highest_in > 8 else 0,
            "expansion_model": "MPP-IO-U",
            "highest_input_row": highest_in,
            "highest_output_row": highest_out,
        },
        "counts": {
            "inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
            "values": len(merged['values']), "loops": len(merged['loops']),
            "tables": len(merged['tables']), "programs": len(merged['programs']),
            "schedules": len(merged['schedules']), "trends": len(trends),
            "max_value_inst": max((v.instance for v in merged['values']), default=0),
            "max_loop_inst": max((l.instance for l in merged['loops']), default=0),
            "max_prg_inst": max((p.instance for p in merged['programs']), default=0),
            "system_groups": len(merged['system_groups']),
        },
        "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "units": p.units, "range": p.range_code, "module": p.module} for p in merged['inputs']],
        "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in merged['outputs']],
        "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description, "module": v.module} for v in merged['values']],
        "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "p": l.p_band, "i": l.integral, "action": l.action, "desc": l.description} for l in merged['loops']],
        "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "enabled": p.enabled, "desc": p.description, "has_code": bool(p.code and len(p.code) > 50), "code": p.code or ""} for p in sorted(merged['programs'], key=lambda x: x.exec_order)],
        "tables": [{"instance": t.instance, "name": t.name, "in_units": t.input_units, "out_units": t.output_units, "points": len(t.data_points), "desc": t.description} for t in merged.get('tables', [])],
        "trends": [{"instance": t.instance, "name": t.name, "monitored": t.monitored_point, "type": t.trend_type, "interval": t.interval, "cov_delta": t.cov_delta, "buffer": t.buffer_size} for t in trends],
        "schedules": [{"instance": s.instance, "name": s.name, "default": s.default_state, "states": "/".join(s.states), "priority": s.priority, "desc": s.description} for s in merged.get('schedules', [])],
        "system_groups": [{"name": g.name, "desc": g.description} for g in merged.get('system_groups', [])],
        "soo": '\n\n'.join(m.soo_paragraph for m in modules if m.soo_paragraph),
        "warnings": [],
        "hwp_params": req.params,
    }


@app.post("/api/hwp-generate")
async def api_hwp_generate(req: HWPAssembleRequest):
    """Generate HW plant Excel + .bas package from wizard parameters."""
    try:
        modules = hwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, str(e))

    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
    merged = merge_modules(modules)
    trends = generate_trends(merged)
    alarm_code = generate_alarm_bas(merged)

    config_name = req.params.get('config_name', 'HW-Plant')

    # Load .bas code and add line numbers
    prg_dir = Path(__file__).parent / "programs" / "hw_plant"
    for prg in merged['programs']:
        if not prg.code or len(prg.code) <= 50:
            bas_path = prg_dir / prg.filename
            if bas_path.exists():
                prg.code = bas_path.read_text()
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])

    wb = write_excel(merged, trends, alarm_code, config_name)

    # Compile .pan via temp dir
    import tempfile, shutil
    pan_data = b""
    tmp = tempfile.mkdtemp(prefix="sbs-hwp-")
    try:
        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_bytes = excel_buf.getvalue()
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_bytes)
        tmp_prg = os.path.join(tmp, "programs")
        os.makedirs(tmp_prg, exist_ok=True)
        for prg in merged['programs']:
            with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                f.write(prg.code or f"10 REM {prg.name}\n")
        try:
            from compile_from_excel import compile_package
            model = req.controller_model or "MPS"
            pan_data = compile_package(tmp, controller_model=model, verbose=False)
        except Exception as e:
            import traceback; traceback.print_exc()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    merged['trends'] = trends
    hw_config = _config_from_merged(merged, modules, family="HW-PLANT",
                                     model=req.controller_model or "MPS")
    report = _build_validation_report(hw_config)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("RC-Studio-Output.xlsx", excel_bytes)
        zf.writestr("SBS-Validation-Report.md", report)
        for prg in merged['programs']:
            zf.writestr(f"programs/{prg.filename}", prg.code or f"10 REM {prg.name}\n")

        if pan_data:
            zf.writestr(f"{config_name}.pan", pan_data)
        soo = '\n\n'.join(m.soo_paragraph for m in modules if m.soo_paragraph)
        zf.writestr("SOO.txt", soo)
        zf.writestr("summary.json", json.dumps({
            "name": config_name, "family": "HW-PLANT", "controller": "MPS",
            "params": req.params,
            "pan_included": bool(pan_data), "pan_size": len(pan_data),
            "counts": {"inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
                       "values": len(merged['values']), "programs": len(merged['programs'])+1,
                       "trends": len(trends)},
        }, indent=2, default=str))
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename={config_name}.zip"})


class CHWPAssembleRequest(BaseModel):
    params: dict
    controller_model: str = "MPS"


@app.post("/api/chwp-assemble")
async def api_chwp_assemble(req: CHWPAssembleRequest):
    """Assemble CHW plant from wizard parameters."""
    try:
        modules = chwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, f"CHW Plant assembly error: {str(e)}")

    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas
    merged = merge_modules(modules)
    trends = generate_trends(merged)

    prg_dir = Path(__file__).parent / "programs" / "chw_plant"
    for prg in merged['programs']:
        if prg.code and len(prg.code) > 50:
            continue  # Dynamic code already injected by chwp_assemble
        bas_path = prg_dir / prg.filename
        if bas_path.exists():
            prg.code = bas_path.read_text()

    alarm_code = generate_alarm_bas(merged)
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])

    highest_in = max((p.row for p in merged['inputs']), default=0)
    highest_out = max((p.row for p in merged['outputs']), default=0)

    family = "CHW-PLANT-TOWER" if req.params.get('num_towers') else "CHW-PLANT-AIR"
    config_id = store_config(
        merged['inputs'], merged['outputs'], family,
        req.controller_model or "MPS", "chwp", {"params": req.params},
    )

    return {
        "config_id": config_id,
        "modules": [m.id for m in modules],
        "controller": {
            "model": req.controller_model or "MPS",
            "expansion_count": max(0, (highest_in - 8 + 11) // 12) if highest_in > 8 else 0,
            "expansion_model": "MPP-IO-U",
            "highest_input_row": highest_in,
            "highest_output_row": highest_out,
        },
        "counts": {
            "inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
            "values": len(merged['values']), "loops": len(merged['loops']),
            "tables": len(merged['tables']), "programs": len(merged['programs']),
            "schedules": len(merged['schedules']), "trends": len(trends),
            "max_value_inst": max((v.instance for v in merged['values']), default=0),
            "max_loop_inst": max((l.instance for l in merged['loops']), default=0),
            "max_prg_inst": max((p.instance for p in merged['programs']), default=0),
            "system_groups": len(merged['system_groups']),
        },
        "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "units": p.units, "range": p.range_code, "module": p.module} for p in merged['inputs']],
        "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in merged['outputs']],
        "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description, "module": v.module} for v in merged['values']],
        "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "p": l.p_band, "i": l.integral, "action": l.action, "desc": l.description} for l in merged['loops']],
        "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "enabled": p.enabled, "desc": p.description, "has_code": bool(p.code and len(p.code) > 50), "code": p.code or ""} for p in sorted(merged['programs'], key=lambda x: x.exec_order)],
        "tables": [{"instance": t.instance, "name": t.name, "in_units": t.input_units, "out_units": t.output_units, "points": len(t.data_points), "desc": t.description} for t in merged.get('tables', [])],
        "trends": [{"instance": t.instance, "name": t.name, "monitored": t.monitored_point, "type": t.trend_type, "interval": t.interval, "cov_delta": t.cov_delta, "buffer": t.buffer_size} for t in trends],
        "schedules": [{"instance": s.instance, "name": s.name, "default": s.default_state, "states": "/".join(s.states), "priority": s.priority, "desc": s.description} for s in merged.get('schedules', [])],
        "system_groups": [{"name": g.name, "desc": g.description} for g in merged.get('system_groups', [])],
        "soo": '\n\n'.join(m.soo_paragraph for m in modules if m.soo_paragraph),
        "warnings": [],
        "chwp_params": req.params,
    }


@app.post("/api/chwp-generate")
async def api_chwp_generate(req: CHWPAssembleRequest):
    """Generate CHW plant Excel + .bas package from wizard parameters."""
    try:
        modules = chwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, str(e))

    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
    merged = merge_modules(modules)
    trends = generate_trends(merged)
    alarm_code = generate_alarm_bas(merged)

    config_name = req.params.get('config_name', 'CHW-Plant')

    # Load .bas code and add line numbers
    prg_dir = Path(__file__).parent / "programs" / "chw_plant"
    for prg in merged['programs']:
        if not prg.code or len(prg.code) <= 50:
            bas_path = prg_dir / prg.filename
            if bas_path.exists():
                prg.code = bas_path.read_text()
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])

    wb = write_excel(merged, trends, alarm_code, config_name)

    # Compile .pan via temp dir
    import tempfile, shutil
    pan_data = b""
    tmp = tempfile.mkdtemp(prefix="sbs-chwp-")
    try:
        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_bytes = excel_buf.getvalue()
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_bytes)
        tmp_prg = os.path.join(tmp, "programs")
        os.makedirs(tmp_prg, exist_ok=True)
        for prg in merged['programs']:
            with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                f.write(prg.code or f"10 REM {prg.name}\n")
        try:
            from compile_from_excel import compile_package
            model = req.controller_model or "MPS"
            pan_data = compile_package(tmp, controller_model=model, verbose=False)
        except Exception as e:
            import traceback; traceback.print_exc()
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    family = "CHW-PLANT-TOWER" if req.params.get('num_towers') else "CHW-PLANT-AIR"
    merged['trends'] = trends
    chw_config = _config_from_merged(merged, modules, family=family,
                                      model=req.controller_model or "MPS")
    report = _build_validation_report(chw_config)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("RC-Studio-Output.xlsx", excel_bytes)
        zf.writestr("SBS-Validation-Report.md", report)
        for prg in merged['programs']:
            zf.writestr(f"programs/{prg.filename}", prg.code or f"10 REM {prg.name}\n")
        if pan_data:
            zf.writestr(f"{config_name}.pan", pan_data)
        soo = '\n\n'.join(m.soo_paragraph for m in modules if m.soo_paragraph)
        zf.writestr("SOO.txt", soo)
        zf.writestr("summary.json", json.dumps({
            "name": config_name, "family": family, "controller": "MPS",
            "params": req.params,
            "pan_included": bool(pan_data), "pan_size": len(pan_data),
            "counts": {"inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
                       "values": len(merged['values']), "programs": len(merged['programs'])+1,
                       "trends": len(trends)},
        }, indent=2, default=str))
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename={config_name}.zip"})


@app.post("/api/hwp-generate-pan")
async def api_hwp_generate_pan(req: HWPAssembleRequest):
    """Generate HW plant .pan file only."""
    try:
        modules = hwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, str(e))
    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
    merged = merge_modules(modules)
    trends = generate_trends(merged)
    alarm_code = generate_alarm_bas(merged)
    config_name = req.params.get('config_name', 'HW-Plant')
    prg_dir = Path(__file__).parent / "programs" / "hw_plant"
    for prg in merged['programs']:
        if not prg.code or len(prg.code) <= 50:
            bas_path = prg_dir / prg.filename
            if bas_path.exists():
                prg.code = bas_path.read_text()
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])
    wb = write_excel(merged, trends, alarm_code, config_name)
    import tempfile, shutil
    tmp = tempfile.mkdtemp(prefix="sbs-hwp-pan-")
    try:
        excel_buf = io.BytesIO(); wb.save(excel_buf)
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_buf.getvalue())
        tmp_prg = os.path.join(tmp, "programs"); os.makedirs(tmp_prg, exist_ok=True)
        for prg in merged['programs']:
            with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                f.write(prg.code or f"10 REM {prg.name}\n")
        from compile_from_excel import compile_package
        model = req.controller_model or "MPS"
        pan_data = compile_package(tmp, controller_model=model, verbose=False)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    buf = io.BytesIO(pan_data)
    return StreamingResponse(buf, media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={config_name}.pan"})


@app.post("/api/chwp-generate-pan")
async def api_chwp_generate_pan(req: CHWPAssembleRequest):
    """Generate CHW plant .pan file only."""
    try:
        modules = chwp_assemble(req.params)
    except Exception as e:
        raise HTTPException(400, str(e))
    from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
    merged = merge_modules(modules)
    trends = generate_trends(merged)
    alarm_code = generate_alarm_bas(merged)
    config_name = req.params.get('config_name', 'CHW-Plant')
    prg_dir = Path(__file__).parent / "programs" / "chw_plant"
    for prg in merged['programs']:
        if not prg.code or len(prg.code) <= 50:
            bas_path = prg_dir / prg.filename
            if bas_path.exists():
                prg.code = bas_path.read_text()
    from composition.models import ProgramDef
    alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                           "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
    merged['programs'].append(alarm_prg)
    format_program_commas(merged['programs'])
    number_programs(merged['programs'])
    wb = write_excel(merged, trends, alarm_code, config_name)
    import tempfile, shutil
    tmp = tempfile.mkdtemp(prefix="sbs-chwp-pan-")
    try:
        excel_buf = io.BytesIO(); wb.save(excel_buf)
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_buf.getvalue())
        tmp_prg = os.path.join(tmp, "programs"); os.makedirs(tmp_prg, exist_ok=True)
        for prg in merged['programs']:
            with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                f.write(prg.code or f"10 REM {prg.name}\n")
        from compile_from_excel import compile_package
        model = req.controller_model or "MPS"
        pan_data = compile_package(tmp, controller_model=model, verbose=False)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    buf = io.BytesIO(pan_data)
    return StreamingResponse(buf, media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={config_name}.pan"})


class VVTZoneDef(BaseModel):
    address: int
    tag: str
    reheat: str = "none"
    stat: str = "hardwired"


class VVTAssembleRequest(BaseModel):
    system_tag: str = "RTU-1"
    htg_stages: int = 2
    clg_stages: int = 2
    has_bypass: bool = True
    zones: List[VVTZoneDef] = []


@app.post("/api/assemble-vvt")
async def api_assemble_vvt(req: VVTAssembleRequest):
    """Assemble a complete VVT system — MPV + bypass + all zone controllers."""
    import copy

    if len(req.zones) > 20:
        raise HTTPException(400, "Maximum 20 zones per VVT system.")
    if len(req.zones) < 1:
        raise HTTPException(400, "At least 1 zone is required.")

    zone_count = len(req.zones)
    warnings = []

    if not req.has_bypass:
        warnings.append("VVT system without bypass damper. Staged heating may cause duct overpressure.")

    # --- Build MPV controller ---
    from composition.modules.vvt.mpv_core import build as build_mpv
    mpv_mod = build_mpv(htg_stages=req.htg_stages, clg_stages=req.clg_stages, zone_count=zone_count)
    mpv_config = assemble(["vvt-mpv-core"], controller_model="auto", equipment_family="VVT-MPV")
    # Override with parameterized module (the registry has default params)
    mpv_config.inputs = copy.deepcopy(mpv_mod.inputs)
    mpv_config.outputs = copy.deepcopy(mpv_mod.outputs)
    mpv_config.values = copy.deepcopy(mpv_mod.values)
    mpv_config.loops = copy.deepcopy(mpv_mod.loops)
    mpv_config.arrays = copy.deepcopy(mpv_mod.arrays)
    mpv_config.programs = copy.deepcopy(mpv_mod.programs)
    mpv_config.schedules = copy.deepcopy(mpv_mod.schedules)
    mpv_config.system_groups = copy.deepcopy(mpv_mod.system_groups)
    mpv_config.soo_document = mpv_mod.soo_paragraph
    mpv_config.controller_model = "MACH-ProView LCD"
    prefix_local_points(mpv_config)
    format_program_commas(mpv_config.programs)
    number_programs(mpv_config.programs)

    result = {
        "system_tag": req.system_tag,
        "zone_count": zone_count,
        "htg_stages": req.htg_stages,
        "clg_stages": req.clg_stages,
        "has_bypass": req.has_bypass,
        "warnings": warnings,
        "controllers": {},
    }

    # MPV summary
    result["controllers"]["mpv"] = {
        "tag": f"{req.system_tag}-MPV",
        "controller": "MACH-ProView LCD",
        "family": "VVT-MPV",
        "counts": {
            "inputs": len(mpv_config.inputs),
            "outputs": len(mpv_config.outputs),
            "values": len(mpv_config.values),
            "arrays": len(mpv_config.arrays),
            "programs": len(mpv_config.programs),
        },
        "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in mpv_config.inputs],
        "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in mpv_config.outputs],
        "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description} for v in mpv_config.values],
        "arrays": [{"instance": a.instance, "name": a.name, "size": a.size, "desc": a.description} for a in mpv_config.arrays],
        "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "desc": p.description, "code": p.code or ""} for p in sorted(mpv_config.programs, key=lambda x: x.exec_order)],
    }

    # --- Build bypass controller (optional) ---
    if req.has_bypass:
        byp_config = assemble(["vvt-bypass-core"], controller_model="auto", equipment_family="VVT-BYPASS")
        inject_program_code(byp_config)
        result["controllers"]["bypass"] = {
            "tag": f"{req.system_tag}-BYP",
            "controller": byp_config.controller_model,
            "family": "VVT-BYPASS",
            "counts": {
                "inputs": len(byp_config.inputs),
                "outputs": len(byp_config.outputs),
                "values": len(byp_config.values),
                "loops": len(byp_config.loops),
                "programs": len(byp_config.programs),
            },
            "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in byp_config.inputs],
            "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in byp_config.outputs],
            "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description} for v in byp_config.values],
            "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "action": l.action, "desc": l.description} for l in byp_config.loops],
            "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "desc": p.description, "code": p.code or ""} for p in sorted(byp_config.programs, key=lambda x: x.exec_order)],
        }

    # --- Build each zone controller ---
    # Reheat module mapping
    REHEAT_MODULES = {
        "none": [],
        "hw-mod": ["vvt-rh-hw-mod"],
        "hw-flt": ["vvt-rh-hw-flt"],
        "elec-1": ["vvt-rh-elec-1"],
        "elec-2": ["vvt-rh-elec-2"],
    }
    # Stat module mapping (reuse VAV stat modules)
    STAT_MODULES = {
        "hardwired": ["vav-stat-hardwired"],
        "comm": ["vav-stat-comm"],
    }

    zones_result = []
    has_electric_reheat = False
    for zone in req.zones:
        zone_modules = ["vvt-zone-core"]
        rh_mods = REHEAT_MODULES.get(zone.reheat, [])
        stat_mods = STAT_MODULES.get(zone.stat, ["vav-stat-hardwired"])
        zone_modules.extend(rh_mods)
        zone_modules.extend(stat_mods)

        if zone.reheat in ("elec-1", "elec-2"):
            has_electric_reheat = True

        zn_config = assemble(zone_modules, controller_model="auto", equipment_family="VVT-ZONE")
        inject_program_code(zn_config)

        # Set zone address default
        for v in zn_config.values:
            if v.name == "CFG-ZONE-ADDR":
                v.default = float(zone.address)

        zones_result.append({
            "address": zone.address,
            "tag": zone.tag,
            "controller": zn_config.controller_model,
            "family": "VVT-ZONE",
            "reheat": zone.reheat,
            "stat": zone.stat,
            "counts": {
                "inputs": len(zn_config.inputs),
                "outputs": len(zn_config.outputs),
                "values": len(zn_config.values),
                "loops": len(zn_config.loops),
                "programs": len(zn_config.programs),
            },
            "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in zn_config.inputs],
            "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module} for p in zn_config.outputs],
            "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description} for v in zn_config.values],
            "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "action": l.action, "desc": l.description} for l in zn_config.loops],
            "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "desc": p.description, "code": p.code or ""} for p in sorted(zn_config.programs, key=lambda x: x.exec_order)],
        })

    result["controllers"]["zones"] = zones_result

    if has_electric_reheat:
        warnings.append("Electric reheat present. Verify DAT high limit interlock (87F) and SAT lockout (75F) at commissioning.")

    return result


def _build_vvt_configs(req: "VVTAssembleRequest"):
    """Assemble all controllers (MPV + bypass + per-zone) for a VVT system.

    Returns dict:
      {
        "warnings": [str, ...],
        "controllers": [
          {"folder": "MPV",       "filename_tag": "VVT-MPV",  "tag": "RTU-1-MPV",  "config": ControllerConfig},
          {"folder": "BYP",       "filename_tag": "VVT-BYP",  "tag": "RTU-1-BYP",  "config": ControllerConfig},
          {"folder": "ZONE-01",   "filename_tag": "VVT-ZN01", "tag": "VAV-1",      "config": ControllerConfig},
          ...
        ],
      }
    """
    import copy

    if len(req.zones) > 20:
        raise HTTPException(400, "Maximum 20 zones per VVT system.")
    if len(req.zones) < 1:
        raise HTTPException(400, "At least 1 zone is required.")

    zone_count = len(req.zones)
    warnings: List[str] = []
    if not req.has_bypass:
        warnings.append("VVT system without bypass damper. Staged heating may cause duct overpressure.")

    controllers = []

    # MPV (parameterized build → assembled config)
    from composition.modules.vvt.mpv_core import build as build_mpv
    mpv_mod = build_mpv(htg_stages=req.htg_stages, clg_stages=req.clg_stages, zone_count=zone_count)
    mpv_config = assemble(["vvt-mpv-core"], controller_model="auto", equipment_family="VVT-MPV")
    mpv_config.inputs = copy.deepcopy(mpv_mod.inputs)
    mpv_config.outputs = copy.deepcopy(mpv_mod.outputs)
    mpv_config.values = copy.deepcopy(mpv_mod.values)
    mpv_config.loops = copy.deepcopy(mpv_mod.loops)
    mpv_config.arrays = copy.deepcopy(mpv_mod.arrays)
    mpv_config.programs = copy.deepcopy(mpv_mod.programs)
    mpv_config.schedules = copy.deepcopy(mpv_mod.schedules)
    mpv_config.system_groups = copy.deepcopy(mpv_mod.system_groups)
    mpv_config.soo_document = mpv_mod.soo_paragraph
    prefix_local_points(mpv_config)
    format_program_commas(mpv_config.programs)
    number_programs(mpv_config.programs)
    controllers.append({
        "folder": "MPV",
        "filename_tag": "VVT-MPV",
        "tag": f"{req.system_tag}-MPV",
        "config": mpv_config,
    })

    # Bypass (optional)
    if req.has_bypass:
        byp_config = assemble(["vvt-bypass-core"], controller_model="auto", equipment_family="VVT-BYPASS")
        inject_program_code(byp_config)
        controllers.append({
            "folder": "BYP",
            "filename_tag": "VVT-BYP",
            "tag": f"{req.system_tag}-BYP",
            "config": byp_config,
        })

    # Zones
    REHEAT_MODULES = {
        "none":   [],
        "hw-mod": ["vvt-rh-hw-mod"],
        "hw-flt": ["vvt-rh-hw-flt"],
        "elec-1": ["vvt-rh-elec-1"],
        "elec-2": ["vvt-rh-elec-2"],
    }
    STAT_MODULES = {
        "hardwired": ["vav-stat-hardwired"],
        "comm":      ["vav-stat-comm"],
    }

    has_electric_reheat = False
    for idx, zone in enumerate(req.zones, start=1):
        zone_modules = ["vvt-zone-core"]
        zone_modules.extend(REHEAT_MODULES.get(zone.reheat, []))
        zone_modules.extend(STAT_MODULES.get(zone.stat, ["vav-stat-hardwired"]))
        if zone.reheat in ("elec-1", "elec-2"):
            has_electric_reheat = True

        zn_config = assemble(zone_modules, controller_model="auto", equipment_family="VVT-ZONE")
        inject_program_code(zn_config)
        for v in zn_config.values:
            if v.name == "CFG-ZONE-ADDR":
                v.default = float(zone.address)

        controllers.append({
            "folder": f"ZONE-{idx:02d}",
            "filename_tag": f"VVT-ZN{idx:02d}",
            "tag": zone.tag,
            "config": zn_config,
        })

    if has_electric_reheat:
        warnings.append("Electric reheat present. Verify DAT high limit interlock (87F) and SAT lockout (75F) at commissioning.")

    return {"warnings": warnings, "controllers": controllers}


def _write_vvt_controller_files(zf: zipfile.ZipFile, folder: str, filename_tag: str, tag: str, config):
    """Write Excel + .bas + .pan + SOO + validation report into zf under <folder>/."""
    from composition.alarm_gen import generate_alarm_bas

    excel_data = generate_excel(config)
    alarm_bas = generate_alarm_bas(config)
    model = config.controller_model or "MPS"
    pan_data = _compile_pan_from_config(config)
    report = _build_validation_report(config)

    zf.writestr(f"{folder}/SBS-Validation-Report.md", report)
    zf.writestr(f"{folder}/RC-Studio-Output.xlsx", excel_data)
    zf.writestr(f"{folder}/SBS-{filename_tag}-{model}.pan", pan_data)
    has_alarm_prg = any(p.filename == "PRG-ALARMS.bas" for p in config.programs)
    for prg in config.programs:
        zf.writestr(f"{folder}/programs/{prg.filename}", prg.code or "")
    if alarm_bas and not has_alarm_prg:
        zf.writestr(f"{folder}/programs/PRG-ALARMS.bas", alarm_bas)
    zf.writestr(f"{folder}/SOO.txt", config.soo_document or "")
    zf.writestr(f"{folder}/summary.json", json.dumps({
        "tag": tag,
        "controller": model,
        "family": config.equipment_family,
        "pan_size": len(pan_data),
        "counts": {
            "inputs":   len(config.inputs),
            "outputs":  len(config.outputs),
            "values":   len(config.values),
            "loops":    len(config.loops),
            "programs": len(config.programs),
            "trends":   len(config.trends),
        },
    }, indent=2))


@app.post("/api/vvt-generate-full")
async def api_vvt_generate_full(req: VVTAssembleRequest):
    """Generate complete VVT system package: per-controller Excel + .bas + .pan + SOO,
    bundled into one zip with subfolders MPV/, BYP/, ZONE-01/, ZONE-02/, ..."""
    try:
        built = _build_vvt_configs(req)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(400, str(e))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Per-controller subfolders
        controllers_summary = []
        for entry in built["controllers"]:
            _write_vvt_controller_files(zf, entry["folder"], entry["filename_tag"], entry["tag"], entry["config"])
            cfg = entry["config"]
            controllers_summary.append({
                "folder": entry["folder"],
                "tag": entry["tag"],
                "family": cfg.equipment_family,
                "controller": cfg.controller_model,
                "counts": {
                    "inputs":   len(cfg.inputs),
                    "outputs":  len(cfg.outputs),
                    "values":   len(cfg.values),
                    "loops":    len(cfg.loops),
                    "programs": len(cfg.programs),
                },
            })

        # Top-level system summary
        zf.writestr("system-summary.json", json.dumps({
            "system_tag": req.system_tag,
            "zone_count": len(req.zones),
            "htg_stages": req.htg_stages,
            "clg_stages": req.clg_stages,
            "has_bypass": req.has_bypass,
            "warnings": built["warnings"],
            "controllers": controllers_summary,
        }, indent=2))

    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=vvt-system-package.zip"})


@app.post("/api/vvt-generate-pan")
async def api_vvt_generate_pan(req: VVTAssembleRequest):
    """Generate .pan binaries for every controller in a VVT system, returned as one zip."""
    try:
        built = _build_vvt_configs(req)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(400, str(e))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for entry in built["controllers"]:
            cfg = entry["config"]
            model = cfg.controller_model or "MPS"
            pan_data = _compile_pan_from_config(cfg)
            zf.writestr(f"SBS-{entry['filename_tag']}-{model}.pan", pan_data)
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=vvt-system-pan.zip"})


@app.post("/api/generate")
async def api_generate(req: GenerateRequest):
    try:
        config = assemble(req.modules, controller_model=req.controller_model,
                          equipment_family=req.equipment_family)
        inject_program_code(config)
    except ValueError as e:
        raise HTTPException(400, str(e))

    excel_data = generate_excel(config)
    readme = _build_readme(config, include_pan=False)
    report = _build_validation_report(config)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", readme)
        zf.writestr("SBS-Validation-Report.md", report)
        zf.writestr("RC-Studio-Output.xlsx", excel_data)
        for prg in config.programs:
            zf.writestr(f"programs/{prg.filename}", prg.code or "")
        zf.writestr("SOO.txt", config.soo_document)
        zf.writestr("summary.json", json.dumps({
            "modules": config.selected_modules,
            "controller": config.controller_model,
            "expansion": f"{config.expansion_count}x {config.expansion_model}" if config.expansion_count else "none",
            "counts": {"inputs": len(config.inputs), "outputs": len(config.outputs),
                       "values": len(config.values), "loops": len(config.loops),
                       "programs": len(config.programs), "trends": len(config.trends)},
        }, indent=2))
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=composition-package.zip"})


def _config_from_merged(merged, modules, family="HW-PLANT", model="MPS"):
    """Build a minimal ControllerConfig from a merged dict for validation report."""
    from composition.models import ControllerConfig
    config = ControllerConfig()
    config.equipment_family = family
    config.controller_model = model
    config.selected_modules = [m.id for m in modules]
    config.inputs = merged.get('inputs', [])
    config.outputs = merged.get('outputs', [])
    config.values = merged.get('values', [])
    config.loops = merged.get('loops', [])
    config.tables = merged.get('tables', [])
    config.programs = merged.get('programs', [])
    config.schedules = merged.get('schedules', [])
    config.trends = merged.get('trends', [])
    config.system_groups = merged.get('system_groups', [])
    config.highest_input_row = max((p.row for p in config.inputs), default=0)
    config.highest_output_row = max((p.row for p in config.outputs), default=0)
    config.soo_document = '\n\n'.join(m.soo_paragraph for m in modules if m.soo_paragraph)
    return config


def _build_validation_report(config) -> str:
    """Build SBS-Validation-Report.md from assembled config."""
    from datetime import datetime
    model = config.controller_model or "MPS"
    exp = f"{config.expansion_count}x {config.expansion_model}" if config.expansion_count else "none"
    warnings = getattr(config, 'warnings', [])

    # Count by type
    ai = sum(1 for p in config.inputs if p.point_type == 'AI')
    bi = sum(1 for p in config.inputs if p.point_type == 'BI')
    ao = sum(1 for p in config.outputs if p.point_type == 'AO')
    bo = sum(1 for p in config.outputs if p.point_type == 'BO')
    av = sum(1 for v in config.values if v.point_type == 'AV')
    bv = sum(1 for v in config.values if v.point_type == 'BV')
    mv = sum(1 for v in config.values if v.point_type == 'MV')

    lines = [
        "# SBS Validation Report",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Equipment Family: {config.equipment_family}",
        "",
        "---",
        "",
        "## System Summary",
        "",
        f"- **Controller:** {model}" + (f" + {exp}" if config.expansion_count else ""),
        f"- **Inputs:** {len(config.inputs)} (AI: {ai}, BI: {bi}) — highest row: {config.highest_input_row}",
        f"- **Outputs:** {len(config.outputs)} (AO: {ao}, BO: {bo}) — highest row: {config.highest_output_row}",
        f"- **Values:** {len(config.values)} (AV: {av}, BV: {bv}, MV: {mv})",
        f"- **Loops:** {len(config.loops)}",
        f"- **Programs:** {len(config.programs)}",
        f"- **Trends:** {len(config.trends)}",
        f"- **Schedules:** {len(config.schedules)}",
        f"- **Tables:** {len(config.tables)}",
        f"- **System Groups:** {len(config.system_groups)}",
        "",
        "### Modules",
        "",
    ]
    for m in config.selected_modules:
        lines.append(f"- {m}")

    # I/O Summary
    lines += [
        "",
        "---",
        "",
        "## I/O Summary",
        "",
        "### Inputs",
        "",
        "| Row | Type | Name | Range | Units | Description |",
        "|-----|------|------|-------|-------|-------------|",
    ]
    for row in range(1, config.highest_input_row + 1):
        pt = next((p for p in config.inputs if p.row == row), None)
        if pt:
            lines.append(f"| {row} | {pt.point_type} | {{device-name}}-{pt.name} | {pt.range_code} | {getattr(pt, 'units', '')} | {pt.description} |")
        else:
            lines.append(f"| {row} | | *— unused —* | | | |")

    lines += [
        "",
        "### Outputs",
        "",
        "| Row | Type | Name | Range | Description |",
        "|-----|------|------|-------|-------------|",
    ]
    for row in range(1, config.highest_output_row + 1):
        pt = next((p for p in config.outputs if p.row == row), None)
        if pt:
            rev = " (REV)" if getattr(pt, 'reverse', False) else ""
            lines.append(f"| {row} | {pt.point_type} | {{device-name}}-{pt.name}{rev} | {pt.range_code} | {pt.description} |")
        else:
            lines.append(f"| {row} | | *— unused —* | | |")

    # Warnings
    lines += [
        "",
        "---",
        "",
        "## Warnings",
        "",
    ]
    if warnings:
        for w in warnings:
            lines.append(f"- {w}")
    else:
        lines.append("No warnings.")

    # Parent references
    lines += [
        "",
        "---",
        "",
        "## Parent References",
        "",
        "Programs containing `{parent}` — **requires parent device ID at commissioning:**",
        "",
    ]
    parent_prgs = [p for p in config.programs if p.code and '{parent}' in p.code]
    if parent_prgs:
        for p in parent_prgs:
            lines.append(f"- **PRG{p.instance}:** {p.name} ({p.filename})")
    else:
        lines.append("No programs reference `{parent}`.")

    # Commissioning notes
    lines += [
        "",
        "---",
        "",
        "## Commissioning Notes",
        "",
        "- All point names use `{device-name}` — set the controller's BACnet device name at commissioning.",
    ]
    if parent_prgs:
        lines.append("- Programs listed above reference `{parent}` — set the parent AHU device ID before downloading.")
    if config.schedules:
        lines.append(f"- {len(config.schedules)} schedule(s) defined — configure weekly schedule times in RC Studio after download.")
    if config.loops:
        lines.append(f"- {len(config.loops)} PID loop(s) — verify tuning parameters match field conditions.")
    if config.trends:
        lines.append(f"- {len(config.trends)} trend log(s) configured — verify buffer sizes are adequate for site requirements.")

    lines += [
        "",
        "---",
        "",
        "*Generated by SBS Controls — Ameresco*",
    ]

    return "\n".join(lines)


def _build_readme(config, include_pan=False):
    """Build instructions README for download package."""
    model = config.controller_model or "MPS"
    lines = [
        "SBS COMPOSITION ENGINE — PACKAGE INSTRUCTIONS",
        "=" * 55,
        f"Generated by SBS Composition Engine v2",
        f"Controller: {model}" + (f" + {config.expansion_count}x {config.expansion_model}" if config.expansion_count else ""),
        f"Family: {config.equipment_family}",
        f"Modules: {', '.join(config.selected_modules)}",
        "",
        "PACKAGE CONTENTS",
        "-" * 40,
        "  RC-Studio-Output.xlsx  — Full data reference (13 tabs)",
        "  programs/              — Control-BASIC .bas files",
        "  programs/PRG-ALARMS.bas — Auto-generated alarm program",
        "  SOO.txt                — Sequence of Operations",
    ]
    if include_pan:
        lines.append(f"  SBS-{config.equipment_family}-{model}.pan — Controller template file")
    lines += [
        "",
        "STEP-BY-STEP INSTRUCTIONS",
        "-" * 40,
    ]
    if include_pan:
        lines += [
            "",
            "1. OPEN .PAN FILE",
            "   Open the .pan file in RC Studio or PFU.",
            "   The file has all points with correct ranges and loop tuning.",
            "",
            "2. SET DEFAULT VALUES",
            "   Open the Excel file, go to the 'Values' tab.",
            "   For each AV point with a non-zero default:",
            "   - Find the point in RC Studio/PFU",
            "   - Set the Present Value to match the Excel 'Default Value' column",
            "",
            "3. CONFIGURE LOOPS",
            "   Open the Excel 'Loops' tab for full PID parameters.",
            "   In RC Studio, for each loop set:",
            "   - Input point (from 'Input' column)",
            "   - Setpoint point (from 'Setpoint' column)",
            "   - Output point (from 'Output' column)",
            "   - P-band, Integral, Action are already set in the .pan",
            "",
            "4. FILL TABLE SCALING DATA",
            "   Open the Excel 'Tables' tab.",
            "   In RC Studio, for each table enter the data points",
            "   from the 'Data Points' column (input -> output pairs).",
            "",
            "5. PASTE PROGRAM CODE",
            "   For each program in the programs/ folder:",
            "   - Open the .bas file in a text editor",
            "   - In RC Studio/PFU, go to the corresponding program",
            "   - Paste the code into the program editor",
            "   - Compile the program",
            "   Also paste PRG-ALARMS.bas as an additional alarm program.",
            "",
            "6. SET UP TRENDS",
            "   Open the Excel 'Trends' tab.",
            "   In RC Studio, create trend logs matching the list:",
            "   - STL name, monitored point, type (Polled/COV), interval",
            "",
            "7. VERIFY",
            "   Use the 'Commissioning' tab in Excel as a point checkout list.",
            "   Green columns are for field data entry (Actual, Pass, Notes).",
        ]
    else:
        lines += [
            "",
            "1. CREATE PROJECT IN RC STUDIO",
            "   Create a new project for the target controller ({model}).",
            "   Use the Excel 'Inputs' and 'Outputs' tabs to set up I/O points.",
            "",
            "2. IMPORT VALUES",
            "   Use the 'Values' tab to create AV/BV/MV points with defaults.",
            "   MV state text is in the Units/Description columns.",
            "",
            "3. PASTE PROGRAM CODE",
            "   For each program in the programs/ folder:",
            "   - Open the .bas file, copy all text",
            "   - Create the program in RC Studio",
            "   - Paste and compile",
            "   Also create PRG-ALARMS.bas as an additional alarm program.",
            "",
            "4. CONFIGURE LOOPS, TABLES, TRENDS",
            "   Use the Excel tabs as reference for all configuration.",
            "",
            "5. VERIFY WITH COMMISSIONING CHECKLIST",
            "   The 'Commissioning' tab has a point checkout list.",
        ]
    lines += [
        "",
        "EXCEL TABS REFERENCE",
        "-" * 40,
        "  Inputs        — AI/BI points with terminal rows and ranges",
        "  Outputs       — AO/BO points with voltage ranges",
        "  Values        — AV/BV/MV with defaults, units, state text",
        "  Loops         — PID tuning: P-band, integral, action, bindings",
        "  Tables        — Scaling table data points",
        "  Schedules     — Weekly schedule definitions",
        "  Trends        — Trend log definitions (STL polled/COV)",
        "  System Groups — Graphic page assignments",
        "  Programs      — Program list with descriptions",
        "  Custom Units  — SBS standard enumerations",
        "  Alarms        — Alarm definitions + BAS code preview",
        "  Commissioning — Field point checkout (green = fill in)",
        "",
        "Generated by SBS Controls — Ameresco",
    ]
    return "\n".join(lines)


# Store last assembled config for GET-based downloads
_last_config = {"modules": [], "controller_model": "auto"}


def _compile_pan_from_config(config):
    """Write config to temp dir and compile via compile_from_excel.py."""
    import tempfile, shutil
    from compile_from_excel import compile_package

    tmp = tempfile.mkdtemp(prefix="sbs-pan-")
    try:
        # Write Excel
        excel_data = generate_excel(config)
        with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
            f.write(excel_data)
        # Write .bas files
        prg_dir = os.path.join(tmp, "programs")
        os.makedirs(prg_dir, exist_ok=True)
        for prg in config.programs:
            if prg.code:
                with open(os.path.join(prg_dir, prg.filename), "w") as f:
                    f.write(prg.code)
        # Compile with correct blank for selected controller
        model = config.controller_model or "MPS"
        pan_data = compile_package(tmp, controller_model=model, verbose=False)
        return pan_data
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@app.post("/api/generate-pan")
async def api_generate_pan(req: GenerateRequest):
    """Generate a .pan binary file using compile_from_excel compiler."""
    _last_config["modules"] = req.modules
    _last_config["controller_model"] = req.controller_model
    try:
        config = assemble(req.modules, controller_model=req.controller_model,
                          equipment_family=req.equipment_family)
        inject_program_code(config)
    except ValueError as e:
        raise HTTPException(400, str(e))

    model = config.controller_model or "MPS"
    pan_data = _compile_pan_from_config(config)

    buf = io.BytesIO(pan_data)
    filename = f"SBS-{config.equipment_family}-{model}.pan"
    return StreamingResponse(buf, media_type="application/octet-stream",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/api/generate-full")
async def api_generate_full(req: GenerateRequest):
    """Generate complete package: Excel + .bas + .pan + SOO."""
    try:
        config = assemble(req.modules, controller_model=req.controller_model,
                          equipment_family=req.equipment_family)
        inject_program_code(config)
    except ValueError as e:
        raise HTTPException(400, str(e))

    from composition.alarm_gen import generate_alarm_bas
    model = config.controller_model or "MPS"
    excel_data = generate_excel(config)
    alarm_bas = generate_alarm_bas(config)
    pan_data = _compile_pan_from_config(config)

    report = _build_validation_report(config)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("SBS-Validation-Report.md", report)
        zf.writestr("RC-Studio-Output.xlsx", excel_data)
        zf.writestr(f"SBS-{config.equipment_family}-{model}.pan", pan_data)
        for prg in config.programs:
            zf.writestr(f"programs/{prg.filename}", prg.code or "")
        if alarm_bas:
            zf.writestr("programs/PRG-ALARMS.bas", alarm_bas)
        zf.writestr("SOO.txt", config.soo_document)
        zf.writestr("summary.json", json.dumps({
            "modules": config.selected_modules,
            "controller": model,
            "expansion": f"{config.expansion_count}x {config.expansion_model}" if config.expansion_count else "none",
            "pan_size": len(pan_data),
            "programs_compiled": sum(1 for p in config.programs if p.code),
            "counts": {"inputs": len(config.inputs), "outputs": len(config.outputs),
                       "values": len(config.values), "loops": len(config.loops),
                       "programs": len(config.programs), "trends": len(config.trends)},
        }, indent=2))
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=sbs-full-package.zip"})


@app.get("/api/download-pan")
async def api_download_pan(modules: str = "", controller: str = "auto"):
    """GET-based .pan download (browser-friendly)."""
    mod_list = [m.strip() for m in modules.split(",") if m.strip()]
    if not mod_list:
        raise HTTPException(400, "No modules specified")
    config = assemble(mod_list, controller_model=controller)
    inject_program_code(config)
    model = config.controller_model or "MPS"
    pan_data = _compile_pan_from_config(config)
    buf = io.BytesIO(pan_data)
    return StreamingResponse(buf, media_type="application/octet-stream",
                             headers={"Content-Disposition": f"attachment; filename=SBS-{config.equipment_family}-{model}.pan"})


@app.get("/api/download-full")
async def api_download_full(modules: str = "", controller: str = "auto"):
    """GET-based full package download (browser-friendly)."""
    from composition.alarm_gen import generate_alarm_bas
    mod_list = [m.strip() for m in modules.split(",") if m.strip()]
    if not mod_list:
        raise HTTPException(400, "No modules specified")
    config = assemble(mod_list, controller_model=controller)
    inject_program_code(config)
    model = config.controller_model or "MPS"
    excel_data = generate_excel(config)
    alarm_bas = generate_alarm_bas(config)
    pan_data = _compile_pan_from_config(config)
    readme = _build_readme(config, include_pan=True)
    report = _build_validation_report(config)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", readme)
        zf.writestr("SBS-Validation-Report.md", report)
        zf.writestr("RC-Studio-Output.xlsx", excel_data)
        zf.writestr(f"SBS-{config.equipment_family}-{model}.pan", pan_data)
        for prg in config.programs:
            zf.writestr(f"programs/{prg.filename}", prg.code or "")
        if alarm_bas:
            zf.writestr("programs/PRG-ALARMS.bas", alarm_bas)
        zf.writestr("SOO.txt", config.soo_document)
    zip_buf.seek(0)
    return StreamingResponse(zip_buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=sbs-full-package.zip"})


# --- Admin Auth ---

import hashlib, time

ADMIN_USERS_PATH = Path(__file__).parent / "admin_users.json"
_admin_sessions = {}  # token -> username


def _load_admin_users() -> dict:
    if ADMIN_USERS_PATH.exists():
        return json.loads(ADMIN_USERS_PATH.read_text())
    return {"Admin": "D@mo2142"}


def _save_admin_users(users: dict):
    ADMIN_USERS_PATH.write_text(json.dumps(users, indent=2))


class LoginRequest(BaseModel):
    username: str
    password: str


class AdminUserRequest(BaseModel):
    username: str
    password: str
    token: str = ""


class BasSaveRequest(BaseModel):
    filename: str
    code: str
    token: str = ""


def _check_admin(token: str):
    if token not in _admin_sessions:
        raise HTTPException(403, "Not authenticated — login required")


@app.post("/api/admin/login")
async def api_admin_login(req: LoginRequest):
    users = _load_admin_users()
    if req.username not in users or users[req.username] != req.password:
        raise HTTPException(403, "Invalid credentials")
    token = hashlib.sha256(f"{time.time()}-{req.username}".encode()).hexdigest()[:32]
    _admin_sessions[token] = req.username
    return {"ok": True, "token": token, "username": req.username}


@app.post("/api/admin/add-user")
async def api_admin_add_user(req: AdminUserRequest):
    _check_admin(req.token)
    if not req.username or not req.password:
        raise HTTPException(400, "Username and password required")
    users = _load_admin_users()
    users[req.username] = req.password
    _save_admin_users(users)
    return {"ok": True, "users": list(users.keys())}


@app.post("/api/admin/remove-user")
async def api_admin_remove_user(req: AdminUserRequest):
    _check_admin(req.token)
    users = _load_admin_users()
    if req.username not in users:
        raise HTTPException(404, f"User not found: {req.username}")
    if len(users) <= 1:
        raise HTTPException(400, "Cannot remove last admin user")
    del users[req.username]
    _save_admin_users(users)
    # Invalidate any sessions for removed user
    to_remove = [t for t, u in _admin_sessions.items() if u == req.username]
    for t in to_remove:
        del _admin_sessions[t]
    return {"ok": True, "users": list(users.keys())}


@app.get("/api/admin/list-users")
async def api_admin_list_users(token: str = ""):
    _check_admin(token)
    users = _load_admin_users()
    return {"users": list(users.keys())}


# --- .bas Editor API (admin-protected) ---

BAS_ROOT = Path(__file__).parent / "programs"


def _resolve_bas(filename: str) -> Path:
    """Resolve a .bas filename to its path, searching all subdirectories."""
    if ".." in filename or "\\" in filename:
        raise HTTPException(400, "Invalid filename")
    # If filename includes subfolder prefix like "hw_plant/HW-PRG-xxx.bas"
    if "/" in filename:
        path = BAS_ROOT / filename
    else:
        # Search all subdirectories
        for sub in sorted(BAS_ROOT.iterdir()):
            if sub.is_dir():
                candidate = sub / filename
                if candidate.exists() and candidate.is_file():
                    return candidate
        path = BAS_ROOT / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(404, f"File not found: {filename}")
    return path


@app.get("/api/bas/list")
async def api_bas_list(token: str = ""):
    """List all .bas template files from all program subdirectories."""
    _check_admin(token)
    files = []
    for sub in sorted(BAS_ROOT.iterdir()):
        if sub.is_dir():
            folder = sub.name
            for f in sorted(sub.glob("*.bas")):
                if f.name.endswith(".bak-20260323") or f.name.endswith(".editor-bak"):
                    continue
                files.append({"filename": f.name, "folder": folder, "path": f"{folder}/{f.name}"})
    return {"files": files}


@app.get("/api/bas/read")
async def api_bas_read(filename: str, token: str = ""):
    """Read a .bas template file."""
    _check_admin(token)
    path = _resolve_bas(filename)
    folder = path.parent.name
    return {"filename": path.name, "folder": folder, "code": path.read_text(encoding="utf-8", errors="replace")}


@app.post("/api/bas/save")
async def api_bas_save(req: BasSaveRequest):
    """Save a .bas template file (admin-protected)."""
    _check_admin(req.token)
    path = _resolve_bas(req.filename)
    import shutil
    bak = path.parent / (path.name + ".editor-bak")
    shutil.copy2(path, bak)
    path.write_text(req.code, encoding="utf-8")
    return {"ok": True, "filename": path.name, "folder": path.parent.name, "size": len(req.code)}


# --- .pan Intake ---

@app.post("/api/intake/upload")
async def api_intake_upload(file: UploadFile = File(...), token: str = ""):
    """Upload and decompile a .pan file."""
    _check_admin(token)
    from composition.pan_intake import decompile_pan
    content = await file.read()
    if len(content) < 0x0410:
        raise HTTPException(400, "File too small to be a valid .pan")
    result = decompile_pan(content)
    result["filename"] = file.filename
    return result


# --- Standard I/O Map Export/Import ---

IO_MAP_PATH = Path(__file__).parent / "standard_io_map.json"


@app.get("/api/io-map/export")
async def api_io_map_export(token: str = ""):
    """Export standard I/O map as Excel."""
    _check_admin(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not IO_MAP_PATH.exists():
        raise HTTPException(404, "Standard I/O map not found")

    data = json.loads(IO_MAP_PATH.read_text())
    wb = openpyxl.Workbook()

    # --- Inputs sheet ---
    ws = wb.active
    ws.title = "Inputs"
    headers = ["Row", "Name", "Type", "Range", "Units", "Description", "Module"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1a237e")
    thin = Side(style="thin", color="334155")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, pt in enumerate(data["inputs"], 2):
        ws.cell(row=i, column=1, value=pt["row"])
        ws.cell(row=i, column=2, value=pt["name"])
        ws.cell(row=i, column=3, value=pt["type"])
        ws.cell(row=i, column=4, value=pt["range"])
        ws.cell(row=i, column=5, value=pt.get("units", ""))
        ws.cell(row=i, column=6, value=pt["description"])
        ws.cell(row=i, column=7, value=pt["module"])
    for col in [("A", 6), ("B", 16), ("C", 6), ("D", 18), ("E", 8), ("F", 35), ("G", 16)]:
        ws.column_dimensions[col[0]].width = col[1]

    # --- Outputs sheet ---
    ws2 = wb.create_sheet("Outputs")
    headers2 = ["Row", "Name", "Type", "Range", "Description", "Module", "Min V", "Max V", "Reverse"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, pt in enumerate(data["outputs"], 2):
        ws2.cell(row=i, column=1, value=pt["row"])
        ws2.cell(row=i, column=2, value=pt["name"])
        ws2.cell(row=i, column=3, value=pt["type"])
        ws2.cell(row=i, column=4, value=pt["range"])
        ws2.cell(row=i, column=5, value=pt["description"])
        ws2.cell(row=i, column=6, value=pt["module"])
        ws2.cell(row=i, column=7, value=pt.get("min_v", ""))
        ws2.cell(row=i, column=8, value=pt.get("max_v", ""))
        ws2.cell(row=i, column=9, value="Yes" if pt.get("reverse") else "")
    for col in [("A", 6), ("B", 20), ("C", 6), ("D", 14), ("E", 35), ("F", 16), ("G", 8), ("H", 8), ("I", 8)]:
        ws2.column_dimensions[col[0]].width = col[1]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=SBS-Standard-IO-Map.xlsx"})


@app.post("/api/io-map/import")
async def api_io_map_import(file: UploadFile = File(...), token: str = ""):
    """Import standard I/O map from Excel. Updates the JSON reference."""
    _check_admin(token)
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    inputs = []
    if "Inputs" in wb.sheetnames:
        ws = wb["Inputs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            inputs.append({
                "row": int(row[0]),
                "name": str(row[1] or ""),
                "type": str(row[2] or ""),
                "range": str(row[3] or ""),
                "units": str(row[4] or "") if len(row) > 4 else "",
                "description": str(row[5] or "") if len(row) > 5 else "",
                "module": str(row[6] or "") if len(row) > 6 else "",
            })

    outputs = []
    if "Outputs" in wb.sheetnames:
        ws = wb["Outputs"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            outputs.append({
                "row": int(row[0]),
                "name": str(row[1] or ""),
                "type": str(row[2] or ""),
                "range": str(row[3] or ""),
                "description": str(row[4] or "") if len(row) > 4 else "",
                "module": str(row[5] or "") if len(row) > 5 else "",
                "min_v": float(row[6]) if len(row) > 6 and row[6] else 0.0,
                "max_v": float(row[7]) if len(row) > 7 and row[7] else 0.0,
                "reverse": (str(row[8] or "").strip().lower() in ("yes", "true", "1")) if len(row) > 8 else False,
            })

    wb.close()

    # Backup existing
    if IO_MAP_PATH.exists():
        import shutil
        shutil.copy2(IO_MAP_PATH, IO_MAP_PATH.with_suffix(".json.bak"))

    data = {"inputs": sorted(inputs, key=lambda x: x["row"]),
            "outputs": sorted(outputs, key=lambda x: x["row"])}
    IO_MAP_PATH.write_text(json.dumps(data, indent=2))

    return {"ok": True, "inputs": len(inputs), "outputs": len(outputs)}


@app.get("/api/io-map/json")
async def api_io_map_json(token: str = ""):
    """Return the standard I/O map as JSON (for UI display)."""
    _check_admin(token)
    if not IO_MAP_PATH.exists():
        raise HTTPException(404, "Standard I/O map not found")
    return json.loads(IO_MAP_PATH.read_text())


# --- IO Schedule Export/Import ---

@app.get("/composition/export-io/{config_id}")
async def api_export_io(config_id: str):
    """Export IO schedule Excel for a previously assembled config.

    Returns an Excel file with columns:
      Terminal | Point Name | Description | Type | Units | Controller | Notes

    Physical IO only (AI, AO, BI, BO). One tab per controller.
    Point Name column is locked — user can only edit Terminal column.
    """
    try:
        excel_data = export_io_schedule(config_id)
    except ValueError as e:
        raise HTTPException(404, str(e))

    buf = io.BytesIO(excel_data)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=IO-Schedule-{config_id}.xlsx"},
    )


@app.post("/composition/import-io/{config_id}")
async def api_import_io(config_id: str, file: UploadFile = File(...)):
    """Import modified IO schedule Excel for a previously assembled config.

    Reads the modified Terminal column and updates terminal assignments.
    Validates:
      - Point names unchanged (rejects with error if modified)
      - No terminal conflicts (two points on same terminal)
      - Terminal format valid (IN{n} or OUT{n})
    """
    try:
        get_stored_config(config_id)
    except ValueError as e:
        raise HTTPException(404, str(e))

    content = await file.read()
    try:
        result = import_io_schedule(config_id, content)
    except ValueError as e:
        raise HTTPException(400, str(e))

    return result


@app.get("/composition/preview-overrides/{config_id}")
async def api_preview_overrides(config_id: str):
    """Re-assemble with terminal overrides applied and return JSON for UI preview.

    Returns the same structure as /api/assemble but with overrides applied,
    plus a 'terminal_changes' list showing what moved.
    """
    try:
        cfg = get_stored_config(config_id)
    except ValueError as e:
        raise HTTPException(404, str(e))

    overrides = get_terminal_overrides(config_id)
    source = cfg["source"]
    params = cfg["params"]

    if source == "ahu":
        mod_list = params.get("modules", [])
        ctrl = params.get("controller_model", "auto")
        config = assemble(mod_list, controller_model=ctrl)
        inject_program_code(config)

        # Build original terminal map before overrides
        orig_inputs = {p.name: p.row for p in config.inputs}
        orig_outputs = {p.name: p.row for p in config.outputs}

        if overrides:
            apply_terminal_overrides(config_id, config.inputs, config.outputs)
            config.highest_input_row = max((p.row for p in config.inputs), default=0)
            config.highest_output_row = max((p.row for p in config.outputs), default=0)
            _select_controller(config, ctrl)

        # Build change list
        changes = []
        for p in config.inputs:
            orig = orig_inputs.get(p.name)
            if orig and orig != p.row:
                changes.append({"name": p.name, "type": p.point_type,
                                "old": f"IN{orig}", "new": f"IN{p.row}"})
        for p in config.outputs:
            orig = orig_outputs.get(p.name)
            if orig and orig != p.row:
                changes.append({"name": p.name, "type": p.point_type,
                                "old": f"OUT{orig}", "new": f"OUT{p.row}"})

        return {
            "config_id": config_id,
            "modules": config.selected_modules,
            "controller": {
                "model": config.controller_model,
                "expansion_count": config.expansion_count,
                "expansion_model": config.expansion_model,
                "highest_input_row": config.highest_input_row,
                "highest_output_row": config.highest_output_row,
            },
            "counts": {
                "inputs": len(config.inputs), "outputs": len(config.outputs),
                "values": len(config.values), "loops": len(config.loops),
                "tables": len(config.tables), "programs": len(config.programs),
                "schedules": len(config.schedules), "trends": len(config.trends),
                "system_groups": len(config.system_groups),
                "max_value_inst": max((v.instance for v in config.values), default=0),
                "max_loop_inst": max((l.instance for l in config.loops), default=0),
                "max_prg_inst": max((p.instance for p in config.programs), default=0),
            },
            "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "units": p.units, "range": p.range_code, "module": p.module} for p in config.inputs],
            "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module, "reverse": getattr(p, 'reverse', False), "min_v": getattr(p, 'min_v', 0), "max_v": getattr(p, 'max_v', 0)} for p in config.outputs],
            "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description, "module": v.module} for v in config.values],
            "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "p": l.p_band, "i": l.integral, "action": l.action, "desc": l.description} for l in config.loops],
            "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "enabled": p.enabled, "desc": p.description, "has_code": bool(p.code and len(p.code) > 50), "code": p.code or ""} for p in sorted(config.programs, key=lambda x: x.exec_order)],
            "soo": config.soo_document,
            "warnings": getattr(config, 'warnings', []),
            "terminal_changes": changes,
            "has_overrides": len(overrides) > 0,
        }

    elif source in ("hwp", "chwp"):
        plant_params = params.get("params", {})
        if source == "hwp":
            mods = hwp_assemble(plant_params)
        else:
            mods = chwp_assemble(plant_params)

        from composition.hw_plant_test import merge_modules
        merged = merge_modules(mods)

        # Load program code
        prg_dir = Path(__file__).parent / "programs" / ("hw_plant" if source == "hwp" else "chw_plant")
        for prg in merged['programs']:
            if prg.code and len(prg.code) > 50:
                continue
            bas_path = prg_dir / prg.filename
            if bas_path.exists():
                prg.code = bas_path.read_text()

        # Original terminals
        orig_inputs = {p.name: p.row for p in merged['inputs']}
        orig_outputs = {p.name: p.row for p in merged['outputs']}

        if overrides:
            apply_terminal_overrides(config_id, merged['inputs'], merged['outputs'])

        highest_in = max((p.row for p in merged['inputs']), default=0)
        highest_out = max((p.row for p in merged['outputs']), default=0)

        changes = []
        for p in merged['inputs']:
            orig = orig_inputs.get(p.name)
            if orig and orig != p.row:
                changes.append({"name": p.name, "type": p.point_type,
                                "old": f"IN{orig}", "new": f"IN{p.row}"})
        for p in merged['outputs']:
            orig = orig_outputs.get(p.name)
            if orig and orig != p.row:
                changes.append({"name": p.name, "type": p.point_type,
                                "old": f"OUT{orig}", "new": f"OUT{p.row}"})

        ctrl_model = cfg.get("controller", "MPS")
        return {
            "config_id": config_id,
            "modules": [m.id for m in mods],
            "controller": {
                "model": ctrl_model,
                "expansion_count": max(0, (highest_in - 12 + 11) // 12) if highest_in > 12 else 0,
                "expansion_model": "MPP-IO-U",
                "highest_input_row": highest_in,
                "highest_output_row": highest_out,
            },
            "counts": {
                "inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
                "values": len(merged['values']), "loops": len(merged['loops']),
                "tables": len(merged['tables']), "programs": len(merged['programs']),
                "schedules": len(merged['schedules']), "trends": 0,
                "system_groups": len(merged['system_groups']),
                "max_value_inst": max((v.instance for v in merged['values']), default=0),
                "max_loop_inst": max((l.instance for l in merged['loops']), default=0),
                "max_prg_inst": max((p.instance for p in merged['programs']), default=0),
            },
            "inputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "units": p.units, "range": p.range_code, "module": p.module} for p in merged['inputs']],
            "outputs": [{"row": p.row, "name": p.name, "type": p.point_type, "desc": p.description, "module": p.module, "reverse": False, "min_v": 0, "max_v": 0} for p in merged['outputs']],
            "values": [{"instance": v.instance, "name": v.name, "type": v.point_type, "default": str(v.default), "units": v.units, "desc": v.description, "module": v.module} for v in merged['values']],
            "loops": [{"instance": l.instance, "name": l.name, "input": l.input_ref, "setpoint": l.setpoint_ref, "p": l.p_band, "i": l.integral, "action": l.action, "desc": l.description} for l in merged['loops']],
            "programs": [{"instance": p.instance, "name": p.name, "filename": p.filename, "enabled": p.enabled, "desc": p.description, "has_code": bool(p.code and len(p.code) > 50), "code": p.code or ""} for p in sorted(merged['programs'], key=lambda x: x.exec_order)],
            "soo": '\n\n'.join(m.soo_paragraph for m in mods if m.soo_paragraph),
            "warnings": [],
            "terminal_changes": changes,
            "has_overrides": len(overrides) > 0,
        }

    raise HTTPException(400, f"Unknown config source: {source}")


class GenerateFromConfigRequest(BaseModel):
    config_id: str
    controller_model: str = "auto"


@app.post("/composition/generate-from-config")
async def api_generate_from_config(req: GenerateFromConfigRequest):
    """Generate a full package from a stored config, applying any terminal overrides from import.

    Re-assembles from original parameters, applies terminal overrides, then generates
    Excel + .bas + .pan + SOO as a ZIP package.
    """
    try:
        cfg = get_stored_config(req.config_id)
    except ValueError as e:
        raise HTTPException(404, str(e))

    overrides = get_terminal_overrides(req.config_id)
    source = cfg["source"]
    params = cfg["params"]

    if source == "ahu":
        # AHU path: re-assemble via standard assembler.
        # MUST pass the original equipment_family — otherwise the assembler
        # falls back to its "AHU-VAV" default, pulls in dsp-ctrl + fan-sf-vfd
        # as cores, and crashes with a conflict against the original
        # fan-sf-cs (CV families) or other non-VAV cores.
        mod_list = params.get("modules", [])
        ctrl = req.controller_model if req.controller_model != "auto" else params.get("controller_model", "auto")
        fam = cfg.get("family") or params.get("equipment_family") or "AHU-VAV"
        config = assemble(mod_list, controller_model=ctrl, equipment_family=fam)
        inject_program_code(config)

        # Apply terminal overrides
        if overrides:
            apply_terminal_overrides(req.config_id, config.inputs, config.outputs)
            config.highest_input_row = max((p.row for p in config.inputs), default=0)
            config.highest_output_row = max((p.row for p in config.outputs), default=0)
            _select_controller(config, ctrl)

        from composition.alarm_gen import generate_alarm_bas
        model = config.controller_model or "MPS"
        excel_data = generate_excel(config)
        pan_data = _compile_pan_from_config(config)

        report = _build_validation_report(config)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("SBS-Validation-Report.md", report)
            zf.writestr("RC-Studio-Output.xlsx", excel_data)
            zf.writestr(f"SBS-{config.equipment_family}-{model}.pan", pan_data)
            for prg in config.programs:
                zf.writestr(f"programs/{prg.filename}", prg.code or "")
            zf.writestr("SOO.txt", config.soo_document)
            zf.writestr("summary.json", json.dumps({
                "config_id": req.config_id,
                "modules": config.selected_modules,
                "controller": model,
                "terminal_overrides": overrides,
                "counts": {"inputs": len(config.inputs), "outputs": len(config.outputs),
                           "values": len(config.values), "programs": len(config.programs)},
            }, indent=2))
        zip_buf.seek(0)
        return StreamingResponse(zip_buf, media_type="application/zip",
                                 headers={"Content-Disposition": "attachment; filename=sbs-package.zip"})

    elif source in ("hwp", "chwp"):
        # Plant path: re-assemble via plant wizard
        plant_params = params.get("params", {})
        if source == "hwp":
            modules = hwp_assemble(plant_params)
            prg_dir = Path(__file__).parent / "programs" / "hw_plant"
            config_name = plant_params.get('config_name', 'HW-Plant')
        else:
            modules = chwp_assemble(plant_params)
            prg_dir = Path(__file__).parent / "programs" / "chw_plant"
            config_name = plant_params.get('config_name', 'CHW-Plant')

        from composition.hw_plant_test import merge_modules, generate_trends, generate_alarm_bas, write_excel
        merged = merge_modules(modules)
        trends = generate_trends(merged)
        alarm_code = generate_alarm_bas(merged)

        # Load .bas code and add line numbers
        for prg in merged['programs']:
            if not prg.code or len(prg.code) <= 50:
                bas_path = prg_dir / prg.filename
                if bas_path.exists():
                    prg.code = bas_path.read_text()
        from composition.models import ProgramDef
        alarm_prg = ProgramDef(50, "ALARMS-PRG", "PRG-ALARMS.bas", alarm_code, True,
                               "Auto-generated alarm definitions", "alarm-gen", exec_order=50)
        merged['programs'].append(alarm_prg)
        format_program_commas(merged['programs'])
        number_programs(merged['programs'])

        # Apply terminal overrides
        if overrides:
            apply_terminal_overrides(req.config_id, merged['inputs'], merged['outputs'])

        wb = write_excel(merged, trends, alarm_code, config_name)

        # Compile .pan
        import tempfile, shutil
        pan_data = b""
        tmp = tempfile.mkdtemp(prefix="sbs-iogen-")
        try:
            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_bytes = excel_buf.getvalue()
            with open(os.path.join(tmp, "RC-Studio-Output.xlsx"), "wb") as f:
                f.write(excel_bytes)
            tmp_prg = os.path.join(tmp, "programs")
            os.makedirs(tmp_prg, exist_ok=True)
            for prg in merged['programs']:
                with open(os.path.join(tmp_prg, prg.filename), "w") as f:
                    f.write(prg.code or f"10 REM {prg.name}\n")
            try:
                from compile_from_excel import compile_package
                model = req.controller_model or "MPS"
                pan_data = compile_package(tmp, controller_model=model, verbose=False)
            except Exception as e:
                import traceback
                tb = traceback.format_exc()
                print(f"[generate-from-config] .pan compile failed for config_id={req.config_id}: {e}\n{tb}", flush=True)
                raise HTTPException(500, f".pan compile failed: {type(e).__name__}: {e}")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("RC-Studio-Output.xlsx", excel_bytes)
            for prg in merged['programs']:
                zf.writestr(f"programs/{prg.filename}", prg.code or f"10 REM {prg.name}\n")
            if pan_data:
                zf.writestr(f"{config_name}.pan", pan_data)
            zf.writestr("summary.json", json.dumps({
                "config_id": req.config_id,
                "name": config_name, "family": cfg["family"],
                "controller": cfg["controller"],
                "terminal_overrides": overrides,
                "pan_included": bool(pan_data),
                "counts": {"inputs": len(merged['inputs']), "outputs": len(merged['outputs']),
                           "values": len(merged['values']), "programs": len(merged['programs'])},
            }, indent=2, default=str))
        zip_buf.seek(0)
        return StreamingResponse(zip_buf, media_type="application/zip",
                                 headers={"Content-Disposition": f"attachment; filename={config_name}.zip"})

    else:
        raise HTTPException(400, f"Unknown config source: {source}")


# --- UI ---

@app.get("/", response_class=HTMLResponse)
async def ui():
    return HTML_UI


HTML_UI = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SBS Composition Engine v2</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#0a0e17;color:#e0e6ed}
.hdr{background:linear-gradient(135deg,#1a237e,#0d47a1);padding:14px 24px;display:flex;justify-content:space-between;align-items:center}
.hdr h1{font-size:1.3em;font-weight:600}
.hdr .sub{color:#90caf9;font-size:0.8em}
.layout{display:grid;grid-template-columns:360px 1fr;height:calc(100vh - 56px)}
.side{background:#111827;border-right:1px solid #1e293b;overflow-y:auto;padding:14px}
.main{overflow-y:auto;padding:16px 20px}
.sec{margin-bottom:14px}
.sec-t{font-size:0.7em;text-transform:uppercase;letter-spacing:1px;color:#64748b;margin-bottom:6px;font-weight:600}
select,input{background:#1e293b;border:1px solid #334155;color:#e0e6ed;padding:7px 10px;border-radius:4px;width:100%;font-size:0.85em;margin-bottom:8px}
select:focus,input:focus{border-color:#3b82f6;outline:none}
.btn{padding:9px 18px;border:none;border-radius:5px;cursor:pointer;font-size:0.85em;font-weight:600;transition:background 0.15s}
.btn-p{background:#2563eb;color:#fff}.btn-p:hover{background:#1d4ed8}
.btn-s{background:#059669;color:#fff}.btn-s:hover{background:#047857}
.btn-o{background:#d97706;color:#fff}.btn-o:hover{background:#b45309}
.btn-grp{display:flex;gap:8px;margin:10px 0}
.cfg-card{background:#1e293b;border:1px solid #334155;border-radius:6px;padding:10px 14px;margin-bottom:6px;cursor:pointer;transition:border 0.15s}
.cfg-card:hover{border-color:#3b82f6}
.cfg-card.active{border-color:#3b82f6;background:#172554}
.cfg-card h4{color:#e0e6ed;font-size:0.85em}
.cfg-card p{color:#94a3b8;font-size:0.75em;margin-top:2px}
.mod-grp{margin-bottom:10px}
.mod-grp-t{font-size:0.8em;color:#94a3b8;margin-bottom:3px;font-weight:500}
.mod-item{display:flex;align-items:center;gap:6px;padding:3px 6px;border-radius:3px;font-size:0.8em}
.mod-item:hover{background:#1e293b}
.mod-item.on{color:#60a5fa}
.mod-item.core{opacity:0.5}
.mod-item input{width:auto;margin:0}
.stat-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:6px;margin:10px 0}
.stat{background:#1e293b;padding:8px;border-radius:5px;text-align:center}
.stat .v{font-size:1.4em;font-weight:700;color:#60a5fa}
.stat .l{font-size:0.7em;color:#94a3b8}
.tabs{display:flex;gap:3px;border-bottom:1px solid #1e293b;padding-bottom:6px;margin-bottom:10px;flex-wrap:wrap}
.tab{padding:5px 12px;border-radius:4px 4px 0 0;cursor:pointer;font-size:0.8em;color:#94a3b8;transition:all 0.15s}
.tab.act{background:#1e293b;color:#60a5fa;font-weight:600}
.tab:hover{color:#e0e6ed}
.tp{display:none}.tp.act{display:block}
table{width:100%;border-collapse:collapse;font-size:0.8em}
th{background:#1e293b;padding:6px 8px;text-align:left;color:#94a3b8;font-weight:500;position:sticky;top:0}
td{padding:5px 8px;border-bottom:1px solid #1e293b}
tr.unused td{color:#475569;font-style:italic}
tr.factory td{color:#64748b;font-style:italic;opacity:0.7}
tr.io-changed{background:#1a2e1a}
tr.io-changed td{color:#4ade80}
tr.io-changed td .old-term{text-decoration:line-through;color:#ef4444;font-size:0.85em;margin-left:6px}
.override-banner{background:#312e81;border:1px solid #6366f1;border-radius:6px;padding:10px 14px;margin:8px 0;color:#c7d2fe;font-size:0.85em;display:flex;align-items:center;gap:10px}
.override-banner b{color:#a5b4fc}
.tag{display:inline-block;padding:1px 5px;border-radius:3px;font-size:0.75em;font-weight:600}
.tag-ai{background:#1e3a5f;color:#60a5fa}.tag-bi{background:#1a3636;color:#5eead4}
.tag-ao{background:#3b1f1f;color:#fca5a5}.tag-bo{background:#3b2f1f;color:#fbbf24}.tag-mo{background:#2d1f3b;color:#c084fc}
.tag-av{background:#1e293b;color:#a78bfa}.tag-bv{background:#1e293b;color:#34d399}.tag-mv{background:#1e293b;color:#fb923c}
.soo{white-space:pre-wrap;font-family:'Courier New',monospace;font-size:0.78em;line-height:1.4;background:#0f172a;padding:14px;border-radius:5px;max-height:500px;overflow-y:auto}
#status{padding:6px 10px;background:#1e293b;border-radius:4px;font-size:0.8em;color:#94a3b8;margin-bottom:10px}
.row-label{font-size:0.7em;color:#475569;text-align:right;padding-right:4px}
.modal-bg{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,0.7);z-index:1000;justify-content:center;align-items:center}
.modal-bg.open{display:flex}
.modal{background:#111827;border:1px solid #334155;border-radius:8px;width:90%;max-width:1100px;height:85vh;display:flex;flex-direction:column;overflow:hidden}
.modal-hdr{display:flex;justify-content:space-between;align-items:center;padding:12px 16px;border-bottom:1px solid #1e293b}
.modal-hdr h3{color:#60a5fa;font-size:1em}
.modal-body{display:flex;flex:1;overflow:hidden}
.modal-side{width:220px;border-right:1px solid #1e293b;overflow-y:auto;padding:8px}
.modal-side .bf{padding:5px 8px;cursor:pointer;border-radius:4px;font-size:0.8em;color:#94a3b8}
.modal-side .bf:hover{background:#1e293b;color:#e0e6ed}
.modal-side .bf.sel{background:#172554;color:#60a5fa}
.modal-edit{flex:1;display:flex;flex-direction:column;padding:10px}
.modal-edit textarea{flex:1;background:#0a0e17;color:#e0e6ed;border:1px solid #334155;border-radius:4px;font-family:'Courier New',monospace;font-size:0.82em;line-height:1.5;padding:10px;resize:none;tab-size:4}
.modal-edit textarea:focus{border-color:#3b82f6;outline:none}
.modal-foot{display:flex;gap:8px;align-items:center;padding:8px 0 0 0}
.modal-foot input{width:180px}
.ed-status{font-size:0.8em;color:#94a3b8;margin-left:auto}
</style>
</head>
<body>
<div class="hdr">
  <div><h1>SBS Composition Engine v2</h1><div class="sub">Reliable Controls Output Tool</div></div>
  <div class="btn-grp">
    <button class="btn btn-p" onclick="doAssemble()">Assemble</button>
    <button class="btn btn-s btn-default-dl" onclick="doGenerate()">Download Excel + .bas</button>
    <button class="btn btn-s btn-default-dl" style="background:#1e40af" onclick="doGeneratePan()">Download .pan</button>
    <button class="btn btn-s btn-default-dl" style="background:#065f46" onclick="doGenerateFull()">Full Package</button>
    <button class="btn btn-s" id="btnExportIOSched" style="display:none;background:#7c3aed" onclick="exportIOSchedule()">Export IO Schedule</button>
    <button class="btn btn-s" id="btnImportIOSched" style="display:none;background:#5b21b6" onclick="document.getElementById('ioSchedFile').click()">Import IO Schedule</button>
    <button class="btn btn-s" id="btnGenFromConfig" style="display:none;background:#4338ca" onclick="generateFromConfig()">Generate (with overrides)</button>
    <input type="file" id="ioSchedFile" accept=".xlsx" style="display:none" onchange="importIOSchedule(this)">
    <button class="btn btn-o" id="btnAdmin" onclick="showAdminLogin()">Admin</button>
    <button class="btn btn-o" id="btnEditor" style="display:none" onclick="openEditor()">Edit .bas</button>
    <button class="btn btn-o" id="btnExportIO" style="display:none;background:#7c3aed" onclick="exportIOMap()">Export I/O Map</button>
    <button class="btn btn-o" id="btnImportIO" style="display:none;background:#5b21b6" onclick="document.getElementById('ioMapFile').click()">Import I/O Map</button>
    <button class="btn btn-o" id="btnIntake" style="display:none;background:#0e7490" onclick="openIntake()">Intake .pan</button>
    <button class="btn btn-o" id="btnUsers" style="display:none;background:#991b1b" onclick="openUsers()">Users</button>
    <input type="file" id="ioMapFile" accept=".xlsx" style="display:none" onchange="importIOMap(this)">
    <a id="hiddenDownload" style="display:none"></a>
  </div>
</div>
<div class="layout">
<div class="side">
  <div class="sec">
    <div class="sec-t">Equipment Family</div>
    <select id="selFamily" onchange="onFamilyChange()"><option value="">-- Select Family --</option></select>
    <div id="familyDesc" style="font-size:0.75em;color:#64748b;margin-bottom:8px"></div>
  </div>
  <div class="sec">
    <div class="sec-t">Standard Configuration</div>
    <div id="cfgList"></div>
  </div>
  <div class="sec" id="ctrlSection">
    <div class="sec-t">Controller Model</div>
    <select id="selCtrl"><option value="auto">Auto-Select (recommended)</option></select>
  </div>
  <div class="sec" id="vvtCtrlFixed" style="display:none">
    <div class="sec-t">RTU Controller</div>
    <div style="color:#94a3b8;font-size:0.85em;padding:4px 0">MACH-ProView LCD (fixed)</div>
  </div>
  <div class="sec" id="modToggles">
    <div class="sec-t">Module Toggles <span style="font-size:0.85em;color:#475569">(on/off from standard)</span></div>
    <div id="modList"></div>
  </div>
  <div class="sec" id="hwpWizard" style="display:none">
    <div class="sec-t">HW Plant Configuration</div>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Boiler Control</label>
    <select id="hwp_boiler_type" onchange="hwpUpdate()">
      <option value="cascade">Cascade (enable + setpoint)</option>
      <option value="full">Full Control (direct fire rate)</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Number of Boilers</label>
    <select id="hwp_num_boilers" onchange="hwpUpdate()">
      <option value="1">1</option><option value="2" selected>2</option>
      <option value="3">3</option><option value="4">4</option>
    </select>
    <div id="hwp_spt_row">
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Boiler SPT Output</label>
      <select id="hwp_spt_output" onchange="hwpUpdate()">
        <option value="analog">Analog AO (0-10V)</option>
        <option value="bacnet">BACnet (AV only)</option>
      </select>
    </div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:4px 0 8px">
      <input type="checkbox" id="hwp_monitor_temps" onchange="hwpUpdate()"> Monitor individual boiler supply + return temps
    </label>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Pump Configuration</label>
    <select id="hwp_pump_type" onchange="hwpUpdate()">
      <option value="cs">Constant Speed</option>
      <option value="vfd">VFD (with DP control)</option>
      <option value="pri-sec">Primary / Secondary</option>
    </select>
    <div id="hwp_pump_count_row">
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Number of Pumps</label>
      <select id="hwp_num_pumps" onchange="hwpUpdate()">
        <option value="1">1</option><option value="2" selected>2</option>
        <option value="3">3</option><option value="4">4</option>
      </select>
    </div>
    <div id="hwp_prisec_row" style="display:none">
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Primary Pumps</label>
      <select id="hwp_num_primary" onchange="hwpUpdate()">
        <option value="1">1</option><option value="2" selected>2</option>
        <option value="3">3</option><option value="4">4</option>
      </select>
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Secondary Pumps</label>
      <select id="hwp_num_secondary" onchange="hwpUpdate()">
        <option value="1">1</option><option value="2" selected>2</option>
        <option value="3">3</option><option value="4">4</option>
      </select>
    </div>
    <div class="sec-t" style="margin-top:10px">Optional Add-ons</div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_mixing_valve" onchange="hwpUpdate()"> Mixing Valve
    </label>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_iso_valves" onchange="hwpUpdate()"> Isolation Valves
    </label>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_comb_damper" onchange="hwpUpdate()"> Combustion Damper
    </label>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_heat_exchanger" onchange="hwpUpdate()"> Heat Exchanger
    </label>
    <div id="hwp_hx_row" style="display:none;margin-left:20px">
      <select id="hwp_hx_valve_type" onchange="hwpUpdate()">
        <option value="single_mod">Single Modulating</option>
        <option value="single_onoff">Single On/Off</option>
        <option value="third_twothird">1/3 + 2/3 Sequence</option>
      </select>
    </div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_ahu_integ" onchange="hwpUpdate()"> AHU Integration
    </label>
    <div id="hwp_ahu_row" style="display:none;margin-left:20px">
      <select id="hwp_num_ahus" onchange="hwpUpdate()">
        <option value="1">1 AHU</option><option value="2" selected>2 AHUs</option>
        <option value="3">3</option><option value="4">4</option>
        <option value="5">5</option><option value="6">6</option>
        <option value="7">7</option><option value="8">8</option>
      </select>
    </div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="hwp_makeup_water" onchange="hwpUpdate()"> Makeup Water Monitoring
    </label>
  </div>
  <div class="sec" id="chwpWizard" style="display:none">
    <div class="sec-t">CHW Plant Configuration</div>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Number of Chillers</label>
    <select id="chwp_num_chillers" onchange="chwpUpdate()">
      <option value="1">1</option><option value="2" selected>2</option>
      <option value="3">3</option><option value="4">4</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Primary CHW Pumps</label>
    <select id="chwp_num_pri" onchange="chwpUpdate()">
      <option value="1">1</option><option value="2" selected>2</option>
      <option value="3">3</option><option value="4">4</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Secondary CHW Pumps</label>
    <select id="chwp_num_sec" onchange="chwpUpdate()">
      <option value="1">1</option><option value="2" selected>2</option>
      <option value="3">3</option><option value="4">4</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">DP Sensors</label>
    <select id="chwp_num_dp" onchange="chwpUpdate()">
      <option value="1">1</option><option value="2" selected>2 (averaged)</option>
    </select>
    <div id="chwp_tower_section" style="display:none">
      <div class="sec-t" style="margin-top:10px">Condenser Water / Towers</div>
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">CW Pumps</label>
      <select id="chwp_num_cw" onchange="chwpUpdate()">
        <option value="1">1</option><option value="2" selected>2</option>
        <option value="3">3</option><option value="4">4</option>
      </select>
      <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Cooling Towers</label>
      <select id="chwp_num_towers" onchange="chwpUpdate()">
        <option value="1">1</option><option value="2" selected>2</option>
        <option value="3">3</option><option value="4">4</option>
      </select>
      <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
        <input type="checkbox" id="chwp_tower_bypass" checked onchange="chwpUpdate()"> Tower CW Bypass Valve
      </label>
    </div>
    <div class="sec-t" style="margin-top:10px">Optional Add-ons</div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="chwp_bypass_valve" onchange="chwpUpdate()"> CHW Bypass Valve
    </label>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="chwp_iso_valves" onchange="chwpUpdate()"> Isolation Valves
    </label>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="chwp_ahu_integ" onchange="chwpUpdate()"> AHU Integration
    </label>
    <div id="chwp_ahu_row" style="display:none;margin-left:20px">
      <select id="chwp_num_ahus" onchange="chwpUpdate()">
        <option value="1">1 AHU</option><option value="2" selected>2 AHUs</option>
        <option value="3">3</option><option value="4">4</option>
        <option value="5">5</option><option value="6">6</option>
        <option value="7">7</option><option value="8">8</option>
      </select>
    </div>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:3px 0">
      <input type="checkbox" id="chwp_makeup_water" onchange="chwpUpdate()"> Makeup Water Monitoring
    </label>
  </div>
  <div class="sec" id="vvtWizard" style="display:none">
    <div class="sec-t">VVT System Configuration</div>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">System Tag</label>
    <input type="text" id="vvt_system_tag" value="RTU-1" style="width:100%;padding:4px;background:#1e293b;color:#e2e8f0;border:1px solid #334155;border-radius:4px;margin-bottom:6px">
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Heating Stages</label>
    <select id="vvt_htg_stages">
      <option value="1">1</option><option value="2" selected>2</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;margin-bottom:2px;display:block">Cooling Stages</label>
    <select id="vvt_clg_stages">
      <option value="1">1</option><option value="2" selected>2</option>
    </select>
    <label style="font-size:0.75em;color:#94a3b8;display:flex;align-items:center;gap:6px;margin:6px 0">
      <input type="checkbox" id="vvt_has_bypass" checked> Include Bypass Damper
    </label>
    <div class="sec-t" style="margin-top:10px">Zone Table</div>
    <div id="vvtZoneTable" style="font-size:0.8em">
      <table style="width:100%;border-collapse:collapse">
        <thead><tr style="color:#94a3b8;text-align:left;border-bottom:1px solid #334155">
          <th style="padding:2px 4px">#</th><th style="padding:2px 4px">Tag</th>
          <th style="padding:2px 4px">Reheat</th><th style="padding:2px 4px">Stat</th>
          <th style="padding:2px 4px"></th>
        </tr></thead>
        <tbody id="vvtZoneRows"></tbody>
      </table>
    </div>
    <button onclick="vvtAddZone()" style="margin-top:6px;font-size:0.8em;padding:3px 10px;background:#334155;color:#e2e8f0;border:1px solid #475569;border-radius:4px;cursor:pointer">+ Add Zone</button>
  </div>

</div>
<div class="main">
  <div id="status">Select an equipment family to begin.</div>
  <div id="results" style="display:none">
    <div class="stat-grid" id="stats"></div>
    <div class="tabs" id="tabBar"></div>
    <div id="tabContents"></div>
  </div>
</div>
</div>
<script>
let families={}, standards={}, modules={}, controllers={};
let selected=new Set(), activeCfg='', activeFamily='';
let currentConfigId=''; // IO schedule export/import

async function init(){
  try{
    var resp=await Promise.all([
      fetch('api/families'),fetch('api/standards'),fetch('api/modules'),fetch('api/controllers')
    ]);
    families=await resp[0].json();
    standards=await resp[1].json();
    modules=await resp[2].json();
    controllers=await resp[3].json();
    // Family dropdown
    var sf=document.getElementById('selFamily');
    var fkeys=Object.keys(families);
    for(var i=0;i<fkeys.length;i++){
      var o=document.createElement('option');
      o.value=fkeys[i];
      o.textContent=families[fkeys[i]].name;
      sf.appendChild(o);
    }
    // Controller dropdown
    var sc=document.getElementById('selCtrl');
    var ckeys=Object.keys(controllers);
    for(var i=0;i<ckeys.length;i++){
      if(ckeys[i]==='auto')continue;
      var o=document.createElement('option');
      o.value=ckeys[i];
      var cc=controllers[ckeys[i]];
      o.textContent=ckeys[i]+' ('+cc.family+') - '+cc.base_in+'in/'+cc.base_out+'out, '+cc.max_exp+' exp';
      sc.appendChild(o);
    }
    document.getElementById('status').textContent='Loaded: '+fkeys.length+' families, '+Object.keys(standards).length+' configs, '+Object.keys(controllers).length+' controllers. Select a family to begin.';
  }catch(err){
    document.getElementById('status').textContent='ERROR loading: '+err.message;
    console.error('Init error:',err);
  }
}

function onFamilyChange(){
  activeFamily=document.getElementById('selFamily').value;
  const f=families[activeFamily];
  document.getElementById('familyDesc').textContent=f?f.description:'';
  activeCfg='';
  selected.clear();
  const isHWP=activeFamily==='HW-PLANT';
  const isCHWP=activeFamily==='CHW-PLANT-AIR'||activeFamily==='CHW-PLANT-TOWER';
  const isVVT=activeFamily.startsWith('VVT-');
  const isPlant=isHWP||isCHWP;
  const isWizard=isPlant||isVVT;
  const isVAV=activeFamily.startsWith('VAV-SD-')||activeFamily.startsWith('VAV-PF-')||activeFamily.startsWith('VAV-SF-')||activeFamily.startsWith('VAV-DD-');
  document.getElementById('modToggles').style.display=isWizard?'none':'';
  document.getElementById('ctrlSection').style.display=isVVT?'none':'';
  document.getElementById('vvtCtrlFixed').style.display=isVVT?'':'none';
  document.getElementById('hwpWizard').style.display=isHWP?'':'none';
  document.getElementById('chwpWizard').style.display=isCHWP?'':'none';
  document.getElementById('vvtWizard').style.display=isVVT?'':'none';
  if(isCHWP)chwpUpdate();
  if(isVVT&&document.getElementById('vvtZoneRows').children.length===0){vvtAddZone();vvtAddZone();vvtAddZone();}
  // VAV families: pre-select recommended controller
  if(isVAV){
    var reqMods=f?f.required_modules:[];
    var hasAO=false;
    for(var i=0;i<reqMods.length;i++){var mm=reqMods[i];if(mm==='vav-rh-hw-mod'||mm==='vav-rh-elec-scr')hasAO=true;}
    var hasFan=false;
    for(var i=0;i<reqMods.length;i++){if(reqMods[i].indexOf('fan')!==-1)hasFan=true;}
    var hasDD=false;
    for(var i=0;i<reqMods.length;i++){if(reqMods[i].indexOf('dd-')!==-1)hasDD=true;}
    var hasRH=false;
    for(var i=0;i<reqMods.length;i++){if(reqMods[i].indexOf('rh-')!==-1)hasRH=true;}
    var rec='RCFA-12';
    if(hasAO||hasRH||hasFan||hasDD)rec='RCFA-34';
    document.getElementById('selCtrl').value=rec;
  }else if(activeFamily.startsWith('FCU-')){
    document.getElementById('selCtrl').value='MPZ-88';
  }else{
    document.getElementById('selCtrl').value='auto';
  }
  renderConfigs();
  if(!isPlant)renderModules();
  document.getElementById('results').style.display='none';
  var statusMsg=f?(isPlant?'Configure plant options, then click Assemble.':'Select a standard configuration, then click Assemble.'):'Select an equipment family.';
  if(isVAV)statusMsg='Confirm controller model, then click Assemble.';
  if(activeFamily.startsWith('FCU-'))statusMsg='Select config or customize modules. MPZ-88 default, MPV-LCD available for display.';
  document.getElementById('status').textContent=statusMsg;
}

function hwpUpdate(){
  var bt=document.getElementById('hwp_boiler_type').value;
  document.getElementById('hwp_spt_row').style.display=bt==='cascade'?'':'none';
  var pt=document.getElementById('hwp_pump_type').value;
  document.getElementById('hwp_pump_count_row').style.display=pt!=='pri-sec'?'':'none';
  document.getElementById('hwp_prisec_row').style.display=pt==='pri-sec'?'':'none';
  document.getElementById('hwp_hx_row').style.display=document.getElementById('hwp_heat_exchanger').checked?'':'none';
  document.getElementById('hwp_ahu_row').style.display=document.getElementById('hwp_ahu_integ').checked?'':'none';
}

function hwpGetParams(){
  var p={};
  p.boiler_type=document.getElementById('hwp_boiler_type').value;
  p.num_boilers=parseInt(document.getElementById('hwp_num_boilers').value);
  if(p.boiler_type==='cascade') p.spt_output=document.getElementById('hwp_spt_output').value;
  p.monitor_boiler_temps=document.getElementById('hwp_monitor_temps').checked;
  p.pump_type=document.getElementById('hwp_pump_type').value;
  if(p.pump_type==='pri-sec'){
    p.num_primary=parseInt(document.getElementById('hwp_num_primary').value);
    p.num_secondary=parseInt(document.getElementById('hwp_num_secondary').value);
  }else{
    p.num_pumps=parseInt(document.getElementById('hwp_num_pumps').value);
  }
  p.mixing_valve=document.getElementById('hwp_mixing_valve').checked;
  p.iso_valves=document.getElementById('hwp_iso_valves').checked;
  p.comb_damper=document.getElementById('hwp_comb_damper').checked;
  p.heat_exchanger=document.getElementById('hwp_heat_exchanger').checked;
  if(p.heat_exchanger) p.hx_valve_type=document.getElementById('hwp_hx_valve_type').value;
  p.ahu_integration=document.getElementById('hwp_ahu_integ').checked;
  if(p.ahu_integration) p.num_ahus=parseInt(document.getElementById('hwp_num_ahus').value);
  p.makeup_water=document.getElementById('hwp_makeup_water').checked;
  return p;
}

function hwpLoadPreset(params){
  document.getElementById('hwp_boiler_type').value=params.boiler_type||'cascade';
  document.getElementById('hwp_num_boilers').value=params.num_boilers||2;
  document.getElementById('hwp_spt_output').value=params.spt_output||'analog';
  document.getElementById('hwp_monitor_temps').checked=!!params.monitor_boiler_temps;
  document.getElementById('hwp_pump_type').value=params.pump_type||'cs';
  document.getElementById('hwp_num_pumps').value=params.num_pumps||2;
  document.getElementById('hwp_num_primary').value=params.num_primary||2;
  document.getElementById('hwp_num_secondary').value=params.num_secondary||2;
  document.getElementById('hwp_mixing_valve').checked=!!params.mixing_valve;
  document.getElementById('hwp_iso_valves').checked=!!params.iso_valves;
  document.getElementById('hwp_comb_damper').checked=!!params.comb_damper;
  document.getElementById('hwp_heat_exchanger').checked=!!params.heat_exchanger;
  document.getElementById('hwp_hx_valve_type').value=params.hx_valve_type||'single_mod';
  document.getElementById('hwp_ahu_integ').checked=!!params.ahu_integration;
  document.getElementById('hwp_num_ahus').value=params.num_ahus||2;
  document.getElementById('hwp_makeup_water').checked=!!params.makeup_water;
  hwpUpdate();
}

function chwpUpdate(){
  var isTower=activeFamily==='CHW-PLANT-TOWER';
  document.getElementById('chwp_tower_section').style.display=isTower?'':'none';
  document.getElementById('chwp_ahu_row').style.display=document.getElementById('chwp_ahu_integ').checked?'':'none';
}

function chwpGetParams(){
  var p={};
  p.num_chillers=parseInt(document.getElementById('chwp_num_chillers').value);
  p.num_pri_pumps=parseInt(document.getElementById('chwp_num_pri').value);
  p.num_sec_pumps=parseInt(document.getElementById('chwp_num_sec').value);
  p.num_dp_sensors=parseInt(document.getElementById('chwp_num_dp').value);
  if(activeFamily==='CHW-PLANT-TOWER'){
    p.num_cw_pumps=parseInt(document.getElementById('chwp_num_cw').value);
    p.num_towers=parseInt(document.getElementById('chwp_num_towers').value);
    p.tower_bypass=document.getElementById('chwp_tower_bypass').checked;
  }
  p.bypass_valve=document.getElementById('chwp_bypass_valve').checked;
  p.iso_valves=document.getElementById('chwp_iso_valves').checked;
  p.ahu_integration=document.getElementById('chwp_ahu_integ').checked;
  if(p.ahu_integration) p.num_ahus=parseInt(document.getElementById('chwp_num_ahus').value);
  p.makeup_water=document.getElementById('chwp_makeup_water').checked;
  return p;
}

function chwpLoadPreset(params){
  document.getElementById('chwp_num_chillers').value=params.num_chillers||2;
  document.getElementById('chwp_num_pri').value=params.num_pri_pumps||2;
  document.getElementById('chwp_num_sec').value=params.num_sec_pumps||2;
  document.getElementById('chwp_num_dp').value=params.num_dp_sensors||2;
  if(params.num_cw_pumps) document.getElementById('chwp_num_cw').value=params.num_cw_pumps;
  if(params.num_towers) document.getElementById('chwp_num_towers').value=params.num_towers;
  document.getElementById('chwp_tower_bypass').checked=!!params.tower_bypass;
  document.getElementById('chwp_bypass_valve').checked=!!params.bypass_valve;
  document.getElementById('chwp_iso_valves').checked=!!params.iso_valves;
  document.getElementById('chwp_ahu_integ').checked=!!params.ahu_integration;
  if(params.num_ahus) document.getElementById('chwp_num_ahus').value=params.num_ahus;
  document.getElementById('chwp_makeup_water').checked=!!params.makeup_water;
  chwpUpdate();
}

function renderConfigs(){
  var el=document.getElementById('cfgList');
  if(!activeFamily){el.innerHTML='<div style="color:#475569;font-size:0.8em">Select a family first</div>';return;}
  var html='';
  var keys=Object.keys(standards).filter(function(k){return standards[k].family===activeFamily;}).sort();
  if(!keys.length){el.innerHTML='<div style="color:#475569;font-size:0.8em">No configs for this family</div>';return;}
  for(var i=0;i<keys.length;i++){
    var id=keys[i],cfg=standards[id];
    var act=id===activeCfg?' active':'';
    html+='<div class="cfg-card'+act+'" onclick="selectCfg(\\x27'+id+'\\x27)">';
    html+='<h4>'+id+': '+cfg.name+'</h4>';
    html+='<p>'+cfg.description+'</p></div>';
  }
  el.innerHTML=html;
}

function selectCfg(id){
  activeCfg=id;
  const cfg=standards[id];
  selected=new Set(cfg.modules);
  renderConfigs();
  if(cfg.hwp_params){
    hwpLoadPreset(cfg.hwp_params);
    document.getElementById('status').textContent='Loaded: '+id+' — '+cfg.name+' (HW Plant preset)';
  }else if(cfg.chwp_params){
    chwpLoadPreset(cfg.chwp_params);
    document.getElementById('status').textContent='Loaded: '+id+' — '+cfg.name+' (CHW Plant preset)';
  }else{
    renderModules();
    document.getElementById('status').textContent='Loaded: '+id+' — '+cfg.name;
  }
}

function renderModules(){
  var el=document.getElementById('modList');
  if(!activeFamily){el.innerHTML='';return;} var f=families[activeFamily];
  var html='';
  var catOrder=['core','fan','cooling','heating','preheat','economizer','energy-recovery','ventilation','humidity','pump','safety','optimum-start','boiler','plant','reheat','dual-duct','thermostat','thermostat-addon'];
  var allowedCats=f&&f.available_categories?f.available_categories:[];
  var reqMods=f&&f.required_modules?f.required_modules:[];
  // Show required modules as locked section first (for families with limited categories)
  if(allowedCats.length>0&&reqMods.length>0){
    var reqItems=[];
    for(var cat in modules){var ms=modules[cat];if(!ms)continue;for(var i=0;i<ms.length;i++){if(reqMods.indexOf(ms[i].id)!==-1)reqItems.push(ms[i]);}}
    if(reqItems.length>0){
      html+='<div class="mod-grp"><div class="mod-grp-t">SELECTED ('+reqItems.length+')</div>';
      for(var ri=0;ri<reqItems.length;ri++){
        var m=reqItems[ri];
        html+='<div class="mod-item core" style="display:flex;align-items:center;gap:6px"><input type="checkbox" checked disabled>';
        html+='<span style="flex:1;cursor:pointer" onclick="showModuleDetail(\\x27'+m.id+'\\x27)">'+m.name+'</span>';
        html+='<span style="font-size:0.7em;color:#64748b">'+m.programs+'P</span></div>';
      }
      html+='</div>';
    }
  }
  var noOptionals=(allowedCats.length===0);
  // Prefix filter: FCU families only show fcu-/vav-stat- modules, VVT only show vvt- modules
  var modPrefix='';
  if(activeFamily.startsWith('FCU-'))modPrefix='fcu-';
  else if(activeFamily.startsWith('UV-'))modPrefix='uv-';
  else if(activeFamily.startsWith('VVT-')||activeFamily==='VVT-SYSTEM')modPrefix='vvt-';
  for(var ci=0;ci<catOrder.length;ci++){
    var cat=catOrder[ci];
    var mods=modules[cat];if(!mods)continue;
    if(modPrefix==='fcu-')mods=mods.filter(function(m){return m.id.startsWith('fcu-')||m.id.startsWith('vav-stat-');});
    else if(modPrefix==='uv-')mods=mods.filter(function(m){return m.id.startsWith('uv-')||m.id.startsWith('vav-stat-');});
    else if(modPrefix)mods=mods.filter(function(m){return m.id.startsWith(modPrefix);});
    if(noOptionals){mods=mods.filter(function(m){return reqMods.indexOf(m.id)!==-1;});if(!mods.length)continue;}
    else if(allowedCats.indexOf(cat)===-1)continue;
    html+='<div class="mod-grp"><div class="mod-grp-t">'+cat.toUpperCase()+' ('+mods.length+')</div>';
    for(var mi=0;mi<mods.length;mi++){
      var m=mods[mi];
      var isCore=m.is_core;
      var checked=(isCore||selected.has(m.id))?'checked':'';
      var cls=isCore?'mod-item core':(selected.has(m.id)?'mod-item on':'mod-item');
      html+='<div class="'+cls+'" style="display:flex;align-items:center;gap:6px"><input type="checkbox" '+checked+' '+(isCore?'disabled':'')+' onchange="toggleMod(\\x27'+m.id+'\\x27,this.checked)">';
      html+='<span style="flex:1;cursor:pointer" onclick="showModuleDetail(\\x27'+m.id+'\\x27)">'+m.name+'</span>';
      html+='<span style="font-size:0.7em;color:#64748b">'+m.programs+'P</span></div>';
    }
    html+='</div>';
  }
  el.innerHTML=html;
}

function toggleMod(id,on){
  if(on){
    // Enforce mutual exclusion — deselect conflicting modules in same group
    var thisGroup=null;
    for(var cat in modules){
      var mods=modules[cat];if(!mods)continue;
      for(var i=0;i<mods.length;i++){
        if(mods[i].id===id){thisGroup=mods[i].mutually_exclusive_group;break;}
      }
      if(thisGroup)break;
    }
    if(thisGroup){
      for(var cat in modules){
        var mods=modules[cat];if(!mods)continue;
        for(var i=0;i<mods.length;i++){
          if(mods[i].id!==id && mods[i].mutually_exclusive_group===thisGroup){
            selected.delete(mods[i].id);
          }
        }
      }
    }
    selected.add(id);
  }else{
    selected.delete(id);
  }
  renderModules();
}

async function showModuleDetail(modId){
  var panel=document.getElementById('modDetailPanel');
  if(!panel){
    var d=document.createElement('div');
    d.id='modDetailPanel';
    d.style.cssText='position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#0f172a;border:2px solid #3b82f6;border-radius:8px;padding:0;z-index:1000;min-width:500px;max-width:700px;max-height:80vh;overflow:auto;box-shadow:0 8px 32px rgba(0,0,0,0.5)';
    document.body.appendChild(d);
    panel=d;
    var overlay=document.createElement('div');
    overlay.id='modDetailOverlay';
    overlay.style.cssText='position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:999';
    overlay.onclick=function(){closeModuleDetail();};
    document.body.appendChild(overlay);
  }
  document.getElementById('modDetailOverlay').style.display='block';
  panel.style.display='block';
  panel.innerHTML='<div style="padding:16px;color:#94a3b8">Loading...</div>';
  try{
    var res=await fetch('api/modules/'+modId);
    var m=await res.json();
    var html='<div style="padding:12px 16px;background:#1e293b;border-bottom:1px solid #334155;display:flex;justify-content:space-between;align-items:center">';
    html+='<div><h3 style="margin:0;color:#60a5fa;font-size:1em">'+m.name+'</h3><div style="font-size:0.75em;color:#94a3b8;margin-top:2px">'+m.category.toUpperCase()+' module — '+m.id+'</div></div>';
    html+='<button class="btn btn-o" style="padding:4px 12px;font-size:0.8em" onclick="closeModuleDetail()">Close</button></div>';
    html+='<div style="padding:12px 16px">';
    if(m.description)html+='<div style="font-size:0.8em;color:#cbd5e1;margin-bottom:10px">'+m.description+'</div>';
    var ps=m.point_summary;
    var counts=[];
    if(ps.AI)counts.push(ps.AI+' AI');if(ps.AO)counts.push(ps.AO+' AO');if(ps.DI)counts.push(ps.DI+' DI');if(ps.DO)counts.push(ps.DO+' DO');
    if(ps.AV)counts.push(ps.AV+' AV');if(ps.BV)counts.push(ps.BV+' BV');if(ps.MV)counts.push(ps.MV+' MV');
    if(counts.length)html+='<div style="font-size:0.75em;color:#64748b;margin-bottom:10px">Points: '+counts.join(' | ')+'</div>';
    if(m.programs.length){
      html+='<div style="font-size:0.8em;font-weight:700;color:#a78bfa;margin-bottom:6px">Programs ('+m.programs.length+')</div>';
      html+='<table style="width:100%;border-collapse:collapse;font-size:0.75em"><tr style="background:#1e293b"><th style="padding:4px 8px;text-align:left;color:#60a5fa">PRG#</th><th style="padding:4px 8px;text-align:left;color:#60a5fa">Name</th><th style="padding:4px 8px;text-align:left;color:#60a5fa">File</th><th style="padding:4px 8px;text-align:left;color:#60a5fa">Description</th></tr>';
      m.programs.sort(function(a,b){return a.exec_order-b.exec_order;});
      for(var i=0;i<m.programs.length;i++){
        var p=m.programs[i];
        html+='<tr style="border-bottom:1px solid #1e293b"><td style="padding:4px 8px;color:#e2e8f0">PRG'+p.instance+'</td><td style="padding:4px 8px;color:#e2e8f0">'+p.name+'</td>';
        html+='<td style="padding:4px 8px"><span style="color:#34d399;cursor:pointer;text-decoration:underline" onclick="closeModuleDetail();openBasFromModule(\\x27'+p.filename+'\\x27)">'+p.filename+'</span></td>';
        html+='<td style="padding:4px 8px;color:#94a3b8">'+p.description+'</td></tr>';
      }
      html+='</table>';
    }else{
      html+='<div style="font-size:0.8em;color:#64748b;font-style:italic">No programs in this module</div>';
    }
    html+='</div>';
    panel.innerHTML=html;
  }catch(e){panel.innerHTML='<div style="padding:16px;color:#ef4444">Error: '+e.message+'</div>';}
}

function closeModuleDetail(){
  var p=document.getElementById('modDetailPanel');if(p)p.style.display='none';
  var o=document.getElementById('modDetailOverlay');if(o)o.style.display='none';
}

function openBasFromModule(filename){
  if(!adminToken){alert('Login to admin to view .bas files');return;}
  openEditor();
  loadBasFile(filename);
}

async function doAssemble(){
  // VAV families require explicit controller selection
  var isVAV=activeFamily.startsWith('VAV-SD-')||activeFamily.startsWith('VAV-PF-')||activeFamily.startsWith('VAV-SF-')||activeFamily.startsWith('VAV-DD-');
  if(isVAV&&document.getElementById('selCtrl').value==='auto'){
    document.getElementById('status').textContent='⚠ Select a controller model before assembling VAV families.';return;
  }
  hasOverrides=false;
  currentConfigId='';
  document.getElementById('btnGenFromConfig').style.display='none';
  document.querySelectorAll('.btn-default-dl').forEach(function(b){b.style.opacity='1';b.title='';});
  document.getElementById('status').textContent='Assembling...';
  try{
    var res;
    if(activeFamily==='HW-PLANT'){
      var params=hwpGetParams();
      res=await fetch('api/hwp-assemble',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    }else if(activeFamily==='CHW-PLANT-AIR'||activeFamily==='CHW-PLANT-TOWER'){
      var params=chwpGetParams();
      res=await fetch('api/chwp-assemble',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    }else if(activeFamily.startsWith('VVT-')){
      var vvtData=vvtGetParams();
      res=await fetch('api/assemble-vvt',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(vvtData)});
      if(!res.ok){var e=await res.json();throw new Error(e.detail);}
      var r=await res.json();
      vvtRenderResults(r);
      return;
    }else{
      var mods=Array.from(selected);
      var plantCoreCats={'hw-core':1,'chw-core':1};
      var cats=Object.keys(modules);
      for(var ci=0;ci<cats.length;ci++){
        if(plantCoreCats[cats[ci]])continue;
        var ms=modules[cats[ci]];
        for(var mi=0;mi<ms.length;mi++){
          if(ms[mi].is_core&&mods.indexOf(ms[mi].id)===-1)mods.push(ms[mi].id);
        }
      }
      res=await fetch('api/assemble',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({modules:mods,controller_model:document.getElementById('selCtrl').value,equipment_family:activeFamily})});
    }
    if(!res.ok){var e=await res.json();throw new Error(e.detail);}
    var r=await res.json();
    renderResults(r);
  }catch(e){document.getElementById('status').textContent='Error: '+e.message;}
}

function renderResults(r){
  document.getElementById('results').style.display='block';
  // Store config_id for IO schedule export/import
  if(r.config_id){
    currentConfigId=r.config_id;
    var isTerminal=activeFamily.startsWith('VAV-SD-')||activeFamily.startsWith('VAV-PF-')||activeFamily.startsWith('VAV-SF-')||activeFamily.startsWith('VAV-DD-')||activeFamily.startsWith('VVT-');
    var btnExp=document.getElementById('btnExportIOSched');
    var btnImp=document.getElementById('btnImportIOSched');
    btnExp.style.display='';
    btnImp.style.display='';
    if(isTerminal){
      btnExp.disabled=true;btnExp.style.opacity='0.4';btnExp.title='IO import/export not available for VAV terminal units \u2014 factory points are firmware-controlled.';
      btnImp.disabled=true;btnImp.style.opacity='0.4';btnImp.title='IO import/export not available for VAV terminal units \u2014 factory points are firmware-controlled.';
    }else{
      btnExp.disabled=false;btnExp.style.opacity='1';btnExp.title='';
      btnImp.disabled=false;btnImp.style.opacity='1';btnImp.title='';
    }
  }
  const c=r.counts,ctrl=r.controller;
  const exp=ctrl.expansion_count?ctrl.expansion_count+'x '+ctrl.expansion_model:'none';
  var statusText=ctrl.model+(ctrl.expansion_count?' + '+exp:'')+' | '+r.modules.length+' modules | '+c.inputs+' inputs, '+c.outputs+' outputs, '+c.programs+' programs';
  if(r.warnings&&r.warnings.length>0){statusText+=' | ⚠ '+r.warnings.length+' warning(s)';}
  document.getElementById('status').textContent=statusText;
  // Build set of changed point names for highlighting
  var changedPoints={};
  if(r.terminal_changes&&r.terminal_changes.length>0){
    r.terminal_changes.forEach(function(c){changedPoints[c.name]={old:c.old,new_term:c.new};});
  }
  // Clear previous banners before rendering new ones
  document.querySelectorAll('.override-banner,.warning-banner').forEach(function(el){el.remove();});
  // Override banner
  if(r.has_overrides&&r.terminal_changes&&r.terminal_changes.length>0){
    var bhtml='<div class="override-banner"><b>'+r.terminal_changes.length+' terminal override(s) active</b> — green rows below show moved points. Use <b>Generate (with overrides)</b> to download.';
    bhtml+='<button class="btn" style="padding:4px 12px;font-size:0.75em;background:#4338ca;color:#fff;margin-left:auto" onclick="generateFromConfig()">Generate (with overrides)</button></div>';
    document.getElementById('stats').insertAdjacentHTML('afterend',bhtml);
  }
  if(r.warnings&&r.warnings.length>0){
    var whtml='<div class="warning-banner" style="background:#78350f;border:1px solid #f59e0b;border-radius:6px;padding:12px;margin:8px 0;color:#fef3c7;font-size:13px"><b>⚠ Warnings ('+r.warnings.length+'):</b><ul style="margin:6px 0 0 16px;padding:0">';
    r.warnings.forEach(function(w){whtml+='<li style="margin:2px 0">'+w+'</li>';});
    whtml+='</ul></div>';
    document.getElementById('stats').insertAdjacentHTML('afterend',whtml);
  }

  document.getElementById('stats').innerHTML=
    '<div class="stat"><div class="v">'+ctrl.model+'</div><div class="l">Controller</div></div>'+
    '<div class="stat"><div class="v">'+(ctrl.expansion_count||0)+'</div><div class="l">Expansion</div></div>'+
    '<div class="stat"><div class="v">'+c.inputs+'</div><div class="l">Inputs (r'+ctrl.highest_input_row+')</div></div>'+
    '<div class="stat"><div class="v">'+c.outputs+'</div><div class="l">Outputs (r'+ctrl.highest_output_row+')</div></div>'+
    '<div class="stat"><div class="v">'+c.values+'</div><div class="l">Values</div></div>'+
    '<div class="stat"><div class="v">'+c.loops+'</div><div class="l">PID Loops</div></div>'+
    '<div class="stat"><div class="v">'+c.programs+'</div><div class="l">Programs</div></div>'+
    '<div class="stat"><div class="v">'+c.trends+'</div><div class="l">Trends</div></div>';

  const tabs=['Inputs','Outputs','Values','Loops','Programs','Tables','Trends','Schedules','Sys Groups','SOO'];
  document.getElementById('tabBar').innerHTML=tabs.map((t,i)=>'<div class="tab'+(i===0?' act':'')+'" onclick="showTab('+i+')">'+t+'</div>').join('');

  let tc='';
  // Inputs — use display_max to include factory reserved rows
  var dispMaxIn=ctrl.display_max_input_row||ctrl.highest_input_row;
  tc+='<div class="tp act" id="t0"><table><tr><th>Row</th><th>Type</th><th>Name</th><th>Range</th><th>Units</th><th>Description</th><th>Module</th></tr>';
  for(let row=1;row<=dispMaxIn;row++){
    const pt=r.inputs.find(p=>p.row===row);
    if(pt){
      var ch=changedPoints[pt.name];
      var isFact=pt.module==='FACTORY';
      var cls=isFact?'factory':(ch?'io-changed':'');
      var oldTag=ch?'<span class="old-term">was '+ch.old+'</span>':'';
      tc+='<tr class="'+cls+'"><td>IN'+row+oldTag+'</td><td><span class="tag tag-'+pt.type.toLowerCase()+'">'+pt.type+'</span></td><td>{device-name}-'+pt.name+'</td><td>'+(pt.range||'')+'</td><td>'+(pt.units||'')+'</td><td>'+pt.desc+'</td><td>'+(isFact?'Factory Reserved':pt.module)+'</td></tr>';
    }else{
      tc+='<tr class="unused"><td>IN'+row+'</td><td></td><td colspan="5">--- unused ---</td></tr>';
    }
  }
  tc+='</table></div>';

  // Outputs — use display_max to include factory reserved rows
  var dispMaxOut=ctrl.display_max_output_row||ctrl.highest_output_row;
  tc+='<div class="tp" id="t1"><table><tr><th>Row</th><th>Type</th><th>Name</th><th>Min V</th><th>Max V</th><th>Description</th><th>Module</th></tr>';
  for(let row=1;row<=dispMaxOut;row++){
    const pt=r.outputs.find(p=>p.row===row);
    if(pt){
      var ch=changedPoints[pt.name];
      var isFact=pt.module==='FACTORY';
      var cls=isFact?'factory':(ch?'io-changed':'');
      var oldTag=ch?'<span class="old-term">was '+ch.old+'</span>':'';
      tc+='<tr class="'+cls+'"><td>OUT'+row+oldTag+'</td><td><span class="tag tag-'+pt.type.toLowerCase()+'">'+pt.type+'</span></td><td>{device-name}-'+pt.name+(pt.reverse?' (REV)':'')+'</td><td>'+(pt.min_v||'')+'</td><td>'+(pt.max_v||'')+'</td><td>'+(pt.desc||'')+'</td><td>'+(isFact?'Factory Reserved':pt.module)+'</td></tr>';
    }else{
      tc+='<tr class="unused"><td>OUT'+row+'</td><td></td><td colspan="5">--- unused ---</td></tr>';
    }
  }
  tc+='</table></div>';

  // Values — skip unused rows, show only real values
  tc+='<div class="tp" id="t2"><table><tr><th>Instance</th><th>Type</th><th>Name</th><th>Default</th><th>Units</th><th>Description</th><th>Module</th></tr>';
  var valMap={};r.values.forEach(function(v){valMap[v.instance]=v;});
  for(var vi=1;vi<=c.max_value_inst;vi++){
    var v=valMap[vi];
    if(v){
      var pre={AV:'AV',BV:'BV',MV:'MV'}[v.type]||'AV';
      var isFact=v.module==='FACTORY';
      tc+='<tr class="'+(isFact?'factory':'')+'"><td>'+pre+vi+'</td><td><span class="tag tag-'+v.type.toLowerCase()+'">'+v.type+'</span></td><td>{device-name}-'+v.name+'</td><td>'+v.default+'</td><td>'+(v.units||'')+'</td><td>'+v.desc+'</td><td>'+(isFact?'Factory Reserved':v.module)+'</td></tr>';
    }
  }
  tc+='</table></div>';

  // Loops — skip unused rows, show only real loops
  tc+='<div class="tp" id="t3"><table><tr><th>Loop</th><th>Name</th><th>Input</th><th>Setpoint</th><th>Action</th><th>P Band</th><th>Integral</th><th>Description</th></tr>';
  var loopMap={};r.loops.forEach(function(l){loopMap[l.instance]=l;});
  for(var li=1;li<=c.max_loop_inst;li++){
    var l=loopMap[li];
    if(l){
      tc+='<tr><td>LOOP'+li+'</td><td>'+l.name+'</td><td>{device-name}-'+l.input+'</td><td>{device-name}-'+l.setpoint+'</td><td>'+(l.action==='direct'?'+':'-')+'</td><td>'+l.p+'</td><td>'+l.i+'</td><td>'+l.desc+'</td></tr>';
    }
  }
  tc+='</table></div>';

  // Programs — skip unused rows, show only real programs
  window._programs=r.programs;
  var prgMap={};r.programs.forEach(function(p,i){prgMap[p.instance]={p:p,i:i};});
  tc+='<div class="tp" id="t4"><table><tr><th>PRG#</th><th>Name</th><th>Filename</th><th>Enabled</th><th>Status</th><th>Description</th><th>View</th></tr>';
  for(var pi=1;pi<=c.max_prg_inst;pi++){
    var pe=prgMap[pi];
    if(pe){
      var p=pe.p;
      tc+='<tr><td>PRG'+pi+'</td><td>{device-name}-'+p.name+'</td><td>'+p.filename+'</td><td>'+(p.enabled?'Yes':'No')+'</td><td>'+(p.has_code?'OK':'STUB')+'</td><td>'+p.desc+'</td>';
      tc+='<td><button class="btn btn-p" style="padding:3px 10px;font-size:0.75em" onclick="viewProgram('+pe.i+')">View</button></td></tr>';
    }
  }
  tc+='</table><div id="prgViewer" style="display:none;margin-top:12px"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px"><h4 id="prgViewerTitle" style="color:#60a5fa;font-size:0.9em"></h4><button class="btn btn-o" style="padding:3px 10px;font-size:0.75em" onclick="document.getElementById(\\x27prgViewer\\x27).style.display=\\x27none\\x27">Close</button></div><pre class="soo" id="prgViewerCode" style="max-height:400px"></pre></div></div>';

  // Tables
  tc+='<div class="tp" id="t5"><table><tr><th>Instance</th><th>Name</th><th>Input Units</th><th>Output Units</th><th>Data Points</th><th>Description</th></tr>';
  (r.tables||[]).forEach(function(t){tc+='<tr><td>TBL'+t.instance+'</td><td>{device-name}-'+t.name+'</td><td>'+t.in_units+'</td><td>'+t.out_units+'</td><td>'+t.points+'</td><td>'+t.desc+'</td></tr>';});
  tc+='</table></div>';

  // Trends
  tc+='<div class="tp" id="t6"><table><tr><th>Instance</th><th>Name</th><th>Monitored Point</th><th>Type</th><th>Interval</th><th>COV Delta</th><th>Buffer</th></tr>';
  (r.trends||[]).forEach(function(t){tc+='<tr><td>STL'+t.instance+'</td><td>{device-name}-'+t.name+'</td><td>{device-name}-'+t.monitored+'</td><td>'+t.type+'</td><td>'+t.interval+'</td><td>'+t.cov_delta+'</td><td>'+t.buffer+'</td></tr>';});
  tc+='</table></div>';

  // Schedules
  tc+='<div class="tp" id="t7"><table><tr><th>Instance</th><th>Name</th><th>Default State</th><th>States</th><th>Priority</th><th>Description</th></tr>';
  (r.schedules||[]).forEach(function(s){tc+='<tr><td>SCHED'+s.instance+'</td><td>{device-name}-'+s.name+'</td><td>'+s.default+'</td><td>'+s.states+'</td><td>'+s.priority+'</td><td>'+s.desc+'</td></tr>';});
  tc+='</table></div>';

  // System Groups
  tc+='<div class="tp" id="t8"><table><tr><th>Name</th><th>Description</th></tr>';
  (r.system_groups||[]).forEach(function(g){tc+='<tr><td>'+g.name+'</td><td>'+g.desc+'</td></tr>';});
  tc+='</table></div>';

  // SOO
  tc+='<div class="tp" id="t9"><div class="soo">'+r.soo.replace(/</g,'&lt;')+'</div></div>';

  document.getElementById('tabContents').innerHTML=tc;
}

function viewProgram(idx){
  var p=window._programs[idx];
  document.getElementById('prgViewer').style.display='block';
  document.getElementById('prgViewerTitle').textContent='PRG'+p.instance+': '+p.filename;
  document.getElementById('prgViewerCode').textContent=p.code||'(no code)';
  showTab(4);
}

function showTab(i){
  document.querySelectorAll('.tab').forEach((t,j)=>t.classList.toggle('act',j===i));
  document.querySelectorAll('.tp').forEach((t,j)=>t.classList.toggle('act',j===i));
}

async function doGenerate(){
  document.getElementById('status').textContent='Generating package...';
  try{
    var res;
    if(activeFamily==='HW-PLANT'){
      var params=hwpGetParams();
      params.config_name='SBS-HW-Plant';
      res=await fetch('api/hwp-generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    }else if(activeFamily==='CHW-PLANT-AIR'||activeFamily==='CHW-PLANT-TOWER'){
      var params=chwpGetParams();
      params.config_name='SBS-CHW-Plant';
      res=await fetch('api/chwp-generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    }else{
      var mods=Array.from(selected);
      var plantCoreCats={'hw-core':1,'chw-core':1};
      var cats=Object.keys(modules);
      for(var ci=0;ci<cats.length;ci++){
        if(plantCoreCats[cats[ci]])continue;
        var ms=modules[cats[ci]];
        for(var mi=0;mi<ms.length;mi++){
          if(ms[mi].is_core&&mods.indexOf(ms[mi].id)===-1)mods.push(ms[mi].id);
        }
      }
      res=await fetch('api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({modules:mods,controller_model:document.getElementById('selCtrl').value,equipment_family:activeFamily})});
    }
    if(!res.ok){document.getElementById('status').textContent='Error generating';return;}
    var blob=await res.blob();
    var url=URL.createObjectURL(blob);
    var fname=activeFamily==='HW-PLANT'?'hw-plant-package.zip':(activeFamily.startsWith('CHW')?'chw-plant-package.zip':'composition-package.zip');
    var a=document.createElement('a');a.href=url;a.download=fname;a.click();
    document.getElementById('status').textContent='Package downloaded!';
  }catch(e){document.getElementById('status').textContent='Error: '+e.message;}
}

function getModList(){
  var mods=Array.from(selected);
  var plantCoreCats={'hw-core':1,'chw-core':1};
  var cats=Object.keys(modules);
  for(var ci=0;ci<cats.length;ci++){
    if(plantCoreCats[cats[ci]])continue;
    var ms=modules[cats[ci]];
    for(var mi=0;mi<ms.length;mi++){
      if(ms[mi].is_core&&mods.indexOf(ms[mi].id)===-1)mods.push(ms[mi].id);
    }
  }
  return mods;
}

async function doGeneratePan(){
  if(activeFamily==='HW-PLANT'){
    var params=hwpGetParams();params.config_name='SBS-HW-Plant';
    document.getElementById('status').textContent='Compiling HW Plant .pan...';
    try{
      var res=await fetch('api/hwp-generate-pan',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
      if(!res.ok){document.getElementById('status').textContent='Error compiling .pan';return;}
      var blob=await res.blob();var url=URL.createObjectURL(blob);
      var a=document.createElement('a');a.href=url;a.download='SBS-HW-Plant.pan';a.click();
      document.getElementById('status').textContent='.pan downloaded ('+Math.round(blob.size/1024)+'KB)';
    }catch(e){document.getElementById('status').textContent='Error: '+e;}
    return;
  }
  if(activeFamily==='CHW-PLANT-AIR'||activeFamily==='CHW-PLANT-TOWER'){
    var params=chwpGetParams();params.config_name='SBS-CHW-Plant';
    document.getElementById('status').textContent='Compiling CHW Plant .pan...';
    try{
      var res=await fetch('api/chwp-generate-pan',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
      if(!res.ok){document.getElementById('status').textContent='Error compiling .pan';return;}
      var blob=await res.blob();var url=URL.createObjectURL(blob);
      var a=document.createElement('a');a.href=url;a.download='SBS-CHW-Plant.pan';a.click();
      document.getElementById('status').textContent='.pan downloaded ('+Math.round(blob.size/1024)+'KB)';
    }catch(e){document.getElementById('status').textContent='Error: '+e;}
    return;
  }
  if(activeFamily.startsWith('VVT-')){
    var params=vvtGetParams();
    document.getElementById('status').textContent='Compiling VVT system .pan files (MPV + bypass + zones)...';
    try{
      var res=await fetch('api/vvt-generate-pan',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(params)});
      if(!res.ok){document.getElementById('status').textContent='Error compiling VVT .pan';return;}
      var blob=await res.blob();var url=URL.createObjectURL(blob);
      var a=document.createElement('a');a.href=url;a.download='vvt-system-pan.zip';
      document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
      document.getElementById('status').textContent='VVT .pan zip downloaded ('+Math.round(blob.size/1024)+'KB)';
    }catch(e){document.getElementById('status').textContent='Error: '+e;}
    return;
  }
  var mods=getModList();
  if(mods.length===0){document.getElementById('status').textContent='Assemble first';return;}
  var body=JSON.stringify({modules:mods,controller_model:document.getElementById('selCtrl').value,equipment_family:activeFamily});
  document.getElementById('status').textContent='Generating .pan...';
  try{
    var res=await fetch('api/generate-pan',{method:'POST',headers:{'Content-Type':'application/json'},body:body});
    if(!res.ok){document.getElementById('status').textContent='Error';return;}
    var blob=await res.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');a.href=url;a.download='SBS-controller.pan';
    document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
    document.getElementById('status').textContent='.pan downloaded ('+Math.round(blob.size/1024)+'KB)';
  }catch(e){document.getElementById('status').textContent='Error: '+e;}
}

async function doGenerateFull(){
  if(activeFamily==='HW-PLANT'){
    var params=hwpGetParams();params.config_name='SBS-HW-Plant';
    var res=await fetch('api/hwp-generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    if(!res.ok){document.getElementById('status').textContent='Error generating';return;}
    var blob=await res.blob();var url=URL.createObjectURL(blob);
    var a=document.createElement('a');a.href=url;a.download='hw-plant-package.zip';a.click();
    document.getElementById('status').textContent='HW Plant package downloaded!';return;
  }
  if(activeFamily==='CHW-PLANT-AIR'||activeFamily==='CHW-PLANT-TOWER'){
    var params=chwpGetParams();params.config_name='SBS-CHW-Plant';
    var res=await fetch('api/chwp-generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({params:params,controller_model:document.getElementById('selCtrl').value})});
    if(!res.ok){document.getElementById('status').textContent='Error generating';return;}
    var blob=await res.blob();var url=URL.createObjectURL(blob);
    var a=document.createElement('a');a.href=url;a.download='chw-plant-package.zip';a.click();
    document.getElementById('status').textContent='CHW Plant package downloaded!';return;
  }
  if(activeFamily.startsWith('VVT-')){
    var params=vvtGetParams();
    document.getElementById('status').textContent='Generating VVT system package (MPV + bypass + zones)...';
    try{
      var res=await fetch('api/vvt-generate-full',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(params)});
      if(!res.ok){document.getElementById('status').textContent='Error generating VVT package';return;}
      var blob=await res.blob();var url=URL.createObjectURL(blob);
      var a=document.createElement('a');a.href=url;a.download='vvt-system-package.zip';
      document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
      document.getElementById('status').textContent='VVT system package downloaded!';
    }catch(e){document.getElementById('status').textContent='Error: '+e;}
    return;
  }
  var mods=getModList();
  if(mods.length===0){document.getElementById('status').textContent='Assemble first';return;}
  var body=JSON.stringify({modules:mods,controller_model:document.getElementById('selCtrl').value,equipment_family:activeFamily});
  document.getElementById('status').textContent='Generating full package...';
  try{
    var res=await fetch('api/generate-full',{method:'POST',headers:{'Content-Type':'application/json'},body:body});
    if(!res.ok){document.getElementById('status').textContent='Error';return;}
    var blob=await res.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');a.href=url;a.download='sbs-full-package.zip';
    document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
    document.getElementById('status').textContent='Full package downloaded!';
  }catch(e){document.getElementById('status').textContent='Error: '+e;}
}

// --- Admin Auth ---
var adminToken='';

function showAdminLogin(){
  document.getElementById('loginModal').classList.add('open');
  document.getElementById('loginUser').focus();
}
async function doAdminLogin(){
  var u=document.getElementById('loginUser').value;
  var p=document.getElementById('loginPass').value;
  try{
    var res=await fetch('api/admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:u,password:p})});
    if(!res.ok){document.getElementById('loginError').textContent='Invalid credentials';return;}
    var d=await res.json();
    adminToken=d.token;
    document.getElementById('loginModal').classList.remove('open');
    document.getElementById('btnAdmin').style.display='none';
    document.getElementById('btnEditor').style.display='';
    document.getElementById('btnExportIO').style.display='';
    document.getElementById('btnImportIO').style.display='';
    document.getElementById('btnUsers').style.display='';
    document.getElementById('btnIntake').style.display='';
    document.getElementById('status').textContent='Admin access granted';
  }catch(e){document.getElementById('loginError').textContent='Error: '+e;}
}

// --- I/O Map Export/Import ---
async function exportIOMap(){
  var a=document.createElement('a');
  a.href='api/io-map/export?token='+adminToken;a.download='SBS-Standard-IO-Map.xlsx';
  document.body.appendChild(a);a.click();a.remove();
  document.getElementById('status').textContent='I/O Map exported';
}
function importIOMap(input){
  if(!input.files.length)return;
  var fd=new FormData();
  fd.append('file',input.files[0]);
  document.getElementById('status').textContent='Importing I/O map...';
  fetch('api/io-map/import?token='+adminToken,{method:'POST',body:fd}).then(r=>r.json()).then(d=>{
    if(d.ok)document.getElementById('status').textContent='I/O Map imported: '+d.inputs+' inputs, '+d.outputs+' outputs';
    else document.getElementById('status').textContent='Import error: '+JSON.stringify(d);
  }).catch(e=>{document.getElementById('status').textContent='Import error: '+e;});
  input.value='';
}

// --- IO Schedule Export/Import ---
var hasOverrides=false;
async function exportIOSchedule(){
  if(!currentConfigId){document.getElementById('status').textContent='Assemble first to get a config';return;}
  document.getElementById('status').textContent='Exporting IO schedule...';
  try{
    var a=document.createElement('a');
    a.href='composition/export-io/'+currentConfigId;
    a.download='IO-Schedule-'+currentConfigId+'.xlsx';
    document.body.appendChild(a);a.click();a.remove();
    document.getElementById('status').textContent='IO schedule exported — edit Terminal column (green), leave Point Name (gray) unchanged, then Import';
  }catch(e){document.getElementById('status').textContent='Export error: '+e;}
}
function importIOSchedule(input){
  if(!input.files.length)return;
  if(!currentConfigId){document.getElementById('status').textContent='Assemble first';input.value='';return;}
  var fd=new FormData();
  fd.append('file',input.files[0]);
  document.getElementById('status').textContent='Importing IO schedule...';
  fetch('composition/import-io/'+currentConfigId,{method:'POST',body:fd}).then(function(r){
    if(!r.ok)return r.json().then(function(e){throw new Error(e.detail);});
    return r.json();
  }).then(async function(d){
    if(d.ok){
      if(d.overrides_applied>0){
        // Fetch preview with overrides applied and re-render
        document.getElementById('status').textContent='Applying '+d.overrides_applied+' terminal overrides...';
        var pres=await fetch('composition/preview-overrides/'+currentConfigId);
        if(pres.ok){
          var preview=await pres.json();
          hasOverrides=true;
          renderResults(preview);
          updateDownloadButtons();
        }
      }else{
        document.getElementById('status').textContent='IO schedule imported — no terminal changes detected';
      }
    }
  }).catch(function(e){document.getElementById('status').textContent='Import error: '+e.message;});
  input.value='';
}
function updateDownloadButtons(){
  if(hasOverrides){
    // Hide default downloads, show override generate
    document.getElementById('btnGenFromConfig').style.display='';
    document.querySelectorAll('.btn-default-dl').forEach(function(b){b.style.opacity='0.4';b.title='Terminals modified — use Generate (with overrides)';});
  }else{
    document.getElementById('btnGenFromConfig').style.display='none';
    document.querySelectorAll('.btn-default-dl').forEach(function(b){b.style.opacity='1';b.title='';});
  }
}
async function generateFromConfig(){
  if(!currentConfigId){document.getElementById('status').textContent='No config';return;}
  document.getElementById('status').textContent='Generating package with terminal overrides...';
  try{
    var res=await fetch('composition/generate-from-config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({config_id:currentConfigId,controller_model:document.getElementById('selCtrl').value})});
    if(!res.ok){var e=await res.json();document.getElementById('status').textContent='Error: '+e.detail;return;}
    var blob=await res.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');a.href=url;a.download='sbs-package-modified.zip';
    document.body.appendChild(a);a.click();a.remove();URL.revokeObjectURL(url);
    document.getElementById('status').textContent='Package with terminal overrides downloaded';
  }catch(e){document.getElementById('status').textContent='Error: '+e.message;}
}

// --- .pan Intake ---
var intakeData=null;
function openIntake(){document.getElementById('intakeModal').classList.add('open');}
function closeIntake(){document.getElementById('intakeModal').classList.remove('open');intakeData=null;}
async function uploadPan(){
  var f=document.getElementById('panFile').files[0];
  if(!f){document.getElementById('intakeStatus').textContent='Select a .pan file';return;}
  var fd=new FormData();fd.append('file',f);
  document.getElementById('intakeStatus').textContent='Analyzing...';
  try{
    var res=await fetch('api/intake/upload?token='+adminToken,{method:'POST',body:fd});
    if(!res.ok){var e=await res.json();document.getElementById('intakeStatus').textContent='Error: '+e.detail;return;}
    intakeData=await res.json();
    renderIntake(intakeData);
  }catch(e){document.getElementById('intakeStatus').textContent='Error: '+e;}
}
function renderIntake(d){
  document.getElementById('intakeStatus').textContent=d.filename+' — '+d.file_size.toLocaleString()+' bytes, '+d.total_blocks+' blocks';
  // Summary
  var bc=d.block_counts;
  var summary='<div class="stat-grid" style="margin:8px 0">';
  var keys=Object.keys(bc);
  for(var i=0;i<keys.length;i++){
    if(keys[i]==='PRG'||keys[i]==='DEVICE'||keys[i]==='NOTIF_CLS'||keys[i]==='NC_GROUP')continue;
    summary+='<div class="stat"><div class="v">'+bc[keys[i]]+'</div><div class="l">'+keys[i]+'</div></div>';
  }
  summary+='</div>';
  document.getElementById('intakeSummary').innerHTML=summary;
  // Tabs
  var tabs=['Inputs','Outputs','Values','Loops','Schedules','Tables','Trends'];
  document.getElementById('intakeTabBar').innerHTML=tabs.map(function(t,i){return '<div class="tab'+(i===0?' act':'')+'" onclick="showIntakeTab('+i+')">'+t+'</div>';}).join('');
  var tc='';
  // Inputs
  var inp=d.inputs.filter(function(x){return x.name;});
  tc+='<div class="tp act" id="it0"><table><tr><th>Row</th><th>Type</th><th>Name</th><th>Units</th><th>Range</th><th>Description</th></tr>';
  inp.forEach(function(p){tc+='<tr><td>'+p.instance+'</td><td><span class="tag tag-'+p.type.toLowerCase()+'">'+p.type+'</span></td><td>'+p.name+'</td><td>'+p.units+'</td><td>'+p.range+'</td><td>'+p.desc+'</td></tr>';});
  tc+='</table></div>';
  // Outputs
  var out=d.outputs.filter(function(x){return x.name;});
  tc+='<div class="tp" id="it1"><table><tr><th>Row</th><th>Type</th><th>Name</th><th>Units</th><th>Range</th><th>Description</th></tr>';
  out.forEach(function(p){tc+='<tr><td>'+p.instance+'</td><td><span class="tag tag-'+p.type.toLowerCase()+'">'+p.type+'</span></td><td>'+p.name+'</td><td>'+p.units+'</td><td>'+p.range+'</td><td>'+p.desc+'</td></tr>';});
  tc+='</table></div>';
  // Values
  var vals=d.values.filter(function(x){return x.name;});
  tc+='<div class="tp" id="it2"><table><tr><th>Instance</th><th>Type</th><th>Name</th><th>Default</th><th>Units</th><th>Description</th></tr>';
  vals.forEach(function(v){tc+='<tr><td>'+v.type+v.instance+'</td><td><span class="tag tag-'+v.type.toLowerCase()+'">'+v.type+'</span></td><td>'+v.name+'</td><td>'+(v.default||'')+'</td><td>'+v.units+'</td><td>'+(v.desc||(v.states||''))+'</td></tr>';});
  tc+='</table></div>';
  // Loops
  var lps=d.loops.filter(function(x){return x.name;});
  tc+='<div class="tp" id="it3"><table><tr><th>Loop</th><th>Name</th><th>Action</th><th>P Band</th><th>Integral</th><th>Derivative</th></tr>';
  lps.forEach(function(l){tc+='<tr><td>LOOP'+l.instance+'</td><td>'+l.name+'</td><td>'+l.action+'</td><td>'+l.p_band+'</td><td>'+l.integral+'</td><td>'+l.derivative+'</td></tr>';});
  tc+='</table></div>';
  // Schedules
  tc+='<div class="tp" id="it4"><table><tr><th>Schedule</th><th>Name</th></tr>';
  d.schedules.forEach(function(s){tc+='<tr><td>SCHED'+s.instance+'</td><td>'+s.name+'</td></tr>';});
  tc+='</table></div>';
  // Tables
  tc+='<div class="tp" id="it5"><table><tr><th>Table</th><th>Name</th><th>Units</th><th>Description</th></tr>';
  d.tables.forEach(function(t){tc+='<tr><td>TBL'+t.instance+'</td><td>'+t.name+'</td><td>'+t.units+'</td><td>'+t.desc+'</td></tr>';});
  tc+='</table></div>';
  // Trends
  tc+='<div class="tp" id="it6"><table><tr><th>STL</th><th>Name</th><th>Monitored</th><th>Type</th><th>Interval</th><th>Buffer</th></tr>';
  d.trends.forEach(function(t){tc+='<tr><td>STL'+t.instance+'</td><td>'+t.name+'</td><td>'+t.monitored+'</td><td>'+t.type+'</td><td>'+t.interval+'</td><td>'+t.buffer+'</td></tr>';});
  tc+='</table></div>';
  document.getElementById('intakeContents').innerHTML=tc;
}
function showIntakeTab(i){
  document.querySelectorAll('#intakeTabBar .tab').forEach(function(t,j){t.classList.toggle('act',j===i);});
  for(var j=0;j<7;j++){var el=document.getElementById('it'+j);if(el)el.classList.toggle('act',j===i);}
}

// --- User Management ---
function openUsers(){
  document.getElementById('usersModal').classList.add('open');
  loadUserList();
}
function closeUsers(){document.getElementById('usersModal').classList.remove('open');}
async function loadUserList(){
  var res=await fetch('api/admin/list-users?token='+adminToken);
  var d=await res.json();
  var el=document.getElementById('userList');
  el.innerHTML=d.users.map(function(u){
    return '<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 8px;border-bottom:1px solid #1e293b"><span style="color:#e0e6ed;font-size:0.85em">'+u+'</span>'+(d.users.length>1?'<button class="btn btn-o" style="padding:2px 8px;font-size:0.7em" onclick="removeUser(\\x27'+u+'\\x27)">Remove</button>':'')+'</div>';
  }).join('');
}
async function addUser(){
  var u=document.getElementById('newUser').value.trim();
  var p=document.getElementById('newPass').value;
  if(!u||!p){document.getElementById('userStatus').textContent='Enter username and password';return;}
  var res=await fetch('api/admin/add-user',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:u,password:p,token:adminToken})});
  if(!res.ok){var e=await res.json();document.getElementById('userStatus').textContent='Error: '+e.detail;return;}
  document.getElementById('newUser').value='';document.getElementById('newPass').value='';
  document.getElementById('userStatus').textContent='User added: '+u;
  loadUserList();
}
async function removeUser(u){
  if(!confirm('Remove user: '+u+'?'))return;
  var res=await fetch('api/admin/remove-user',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:u,password:'',token:adminToken})});
  if(!res.ok){var e=await res.json();document.getElementById('userStatus').textContent='Error: '+e.detail;return;}
  document.getElementById('userStatus').textContent='Removed: '+u;
  loadUserList();
}

// --- .bas Editor ---
var editorDirty=false, editorFile='';

function openEditor(){
  document.getElementById('editorModal').classList.add('open');
  loadBasList();
}
function closeEditor(){
  if(editorDirty&&!confirm('Unsaved changes. Close anyway?'))return;
  document.getElementById('editorModal').classList.remove('open');
  editorDirty=false;
}
async function loadBasList(){
  var res=await fetch('api/bas/list?token='+adminToken);
  var d=await res.json();
  var el=document.getElementById('basFileList');
  var folders={};
  d.files.forEach(function(f){if(!folders[f.folder])folders[f.folder]=[];folders[f.folder].push(f);});
  var html='';
  Object.keys(folders).sort().forEach(function(folder){
    html+='<div style="padding:4px 8px;font-weight:700;color:#7c3aed;font-size:11px;background:#1e293b;border-bottom:1px solid #334155;text-transform:uppercase;letter-spacing:0.05em">'+folder+' ('+folders[folder].length+')</div>';
    folders[folder].forEach(function(f){
      html+='<div class="bf'+(f.path===editorFile?' sel':'')+'" onclick="loadBasFile(\\x27'+f.path+'\\x27)">'+f.filename+'</div>';
    });
  });
  el.innerHTML=html;
}
async function loadBasFile(fn){
  if(editorDirty&&!confirm('Unsaved changes in '+editorFile+'. Discard?'))return;
  var res=await fetch('api/bas/read?filename='+encodeURIComponent(fn)+'&token='+adminToken);
  if(!res.ok){document.getElementById('edStatus').textContent='Error loading';return;}
  var d=await res.json();
  editorFile=d.folder+'/'+d.filename;
  document.getElementById('basEditor').value=d.code;
  document.getElementById('edStatus').textContent='Loaded: '+editorFile+' ('+d.code.length+' chars)';
  editorDirty=false;
  loadBasList();
}
function onEditorChange(){editorDirty=true;document.getElementById('edStatus').textContent=editorFile+' (modified)';}
async function saveBasFile(){
  if(!editorFile){document.getElementById('edStatus').textContent='No file selected';return;}
  var code=document.getElementById('basEditor').value;
  var res=await fetch('api/bas/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({filename:editorFile,code:code,token:adminToken})});
  if(!res.ok){var e=await res.json();document.getElementById('edStatus').textContent='Error: '+e.detail;return;}
  editorDirty=false;
  document.getElementById('edStatus').textContent='Saved: '+editorFile;
}

// VVT Builder Functions
var vvtZoneCounter=0;
function vvtAddZone(){
  vvtZoneCounter++;
  var tbody=document.getElementById('vvtZoneRows');
  var tr=document.createElement('tr');
  tr.id='vvtZone_'+vvtZoneCounter;
  tr.style.borderBottom='1px solid #1e293b';
  tr.innerHTML='<td style="padding:2px 4px">'+vvtZoneCounter+'</td>'+
    '<td style="padding:2px 4px"><input type="text" value="VAV-'+vvtZoneCounter+'" style="width:60px;padding:2px;background:#1e293b;color:#e2e8f0;border:1px solid #334155;border-radius:3px" class="vvt-tag"></td>'+
    '<td style="padding:2px 4px"><select class="vvt-rh" style="font-size:0.85em;background:#1e293b;color:#e2e8f0;border:1px solid #334155;border-radius:3px">'+
    '<option value="none">None</option><option value="hw-mod">HW Mod</option><option value="hw-flt">HW Float</option>'+
    '<option value="elec-1">Elec 1</option><option value="elec-2">Elec 2</option></select></td>'+
    '<td style="padding:2px 4px"><select class="vvt-stat" style="font-size:0.85em;background:#1e293b;color:#e2e8f0;border:1px solid #334155;border-radius:3px">'+
    '<option value="hardwired">HW</option><option value="comm">Comm</option></select></td>'+
    '<td style="padding:2px 4px;cursor:pointer;color:#ef4444" onclick="this.parentElement.remove()">x</td>';
  tbody.appendChild(tr);
}
function vvtGetParams(){
  var zones=[];
  var rows=document.getElementById('vvtZoneRows').children;
  for(var i=0;i<rows.length;i++){
    zones.push({
      address:i+1,
      tag:rows[i].querySelector('.vvt-tag').value,
      reheat:rows[i].querySelector('.vvt-rh').value,
      stat:rows[i].querySelector('.vvt-stat').value
    });
  }
  return {
    system_tag:document.getElementById('vvt_system_tag').value,
    htg_stages:parseInt(document.getElementById('vvt_htg_stages').value),
    clg_stages:parseInt(document.getElementById('vvt_clg_stages').value),
    has_bypass:document.getElementById('vvt_has_bypass').checked,
    zones:zones
  };
}
function vvtRenderResults(r){
  document.getElementById('results').style.display='block';
  var html='<div style="color:#94a3b8;font-size:0.9em">';
  html+='<h3 style="color:#e2e8f0">VVT System: '+r.system_tag+' ('+r.zone_count+' zones, '+r.htg_stages+'H/'+r.clg_stages+'C)</h3>';
  if(r.warnings&&r.warnings.length>0){
    r.warnings.forEach(function(w){html+='<div style="color:#fbbf24;margin:4px 0">&#9888; '+w+'</div>';});
  }
  // MPV
  var mpv=r.controllers.mpv;
  html+='<h4 style="color:#38bdf8;margin-top:12px">MPV: '+mpv.tag+' ('+mpv.controller+')</h4>';
  html+='<div>I:'+mpv.counts.inputs+' O:'+mpv.counts.outputs+' V:'+mpv.counts.values+' ARR:'+mpv.counts.arrays+' PRG:'+mpv.counts.programs+'</div>';
  html+='<details><summary style="cursor:pointer;color:#94a3b8">Programs ('+mpv.programs.length+')</summary><table style="width:100%;font-size:0.85em;border-collapse:collapse">';
  mpv.programs.forEach(function(p){html+='<tr style="border-bottom:1px solid #1e293b"><td style="padding:2px">'+p.instance+'</td><td>'+p.name+'</td><td style="color:#64748b">'+p.desc+'</td></tr>';});
  html+='</table></details>';
  // Bypass
  if(r.controllers.bypass){
    var byp=r.controllers.bypass;
    html+='<h4 style="color:#38bdf8;margin-top:12px">Bypass: '+byp.tag+' ('+byp.controller+')</h4>';
    html+='<div>I:'+byp.counts.inputs+' O:'+byp.counts.outputs+' V:'+byp.counts.values+' L:'+byp.counts.loops+' PRG:'+byp.counts.programs+'</div>';
  }
  // Zones
  html+='<h4 style="color:#38bdf8;margin-top:12px">Zone Controllers</h4>';
  html+='<table style="width:100%;font-size:0.85em;border-collapse:collapse"><tr style="color:#94a3b8;border-bottom:1px solid #334155"><th style="text-align:left;padding:2px 4px">#</th><th style="text-align:left;padding:2px 4px">Tag</th><th style="text-align:left;padding:2px 4px">Controller</th><th style="text-align:left;padding:2px 4px">Reheat</th><th style="text-align:left;padding:2px 4px">Stat</th><th style="text-align:left;padding:2px 4px">I/O/V/PRG</th></tr>';
  r.controllers.zones.forEach(function(z){
    html+='<tr style="border-bottom:1px solid #1e293b"><td style="padding:2px 4px">'+z.address+'</td><td style="padding:2px 4px">'+z.tag+'</td><td style="padding:2px 4px">'+z.controller+'</td><td style="padding:2px 4px">'+z.reheat+'</td><td style="padding:2px 4px">'+z.stat+'</td><td style="padding:2px 4px">'+z.counts.inputs+'/'+z.counts.outputs+'/'+z.counts.values+'/'+z.counts.programs+'</td></tr>';
  });
  html+='</table></div>';
  document.getElementById('tabContents').innerHTML=html;
  document.getElementById('tabBar').innerHTML='';
  document.getElementById('stats').innerHTML='';
  document.getElementById('status').textContent='VVT System assembled: '+r.zone_count+' zones + MPV'+(r.has_bypass?' + bypass':'');
}
init();
</script>
<div class="modal-bg" id="intakeModal">
  <div class="modal" style="height:80vh">
    <div class="modal-hdr">
      <h3>.pan Intake Tool</h3>
      <button class="btn btn-o" style="padding:4px 12px;font-size:0.8em" onclick="closeIntake()">Close</button>
    </div>
    <div style="padding:12px 16px;border-bottom:1px solid #1e293b;display:flex;gap:8px;align-items:center">
      <input type="file" id="panFile" accept=".pan,.panx" style="width:auto;margin:0">
      <button class="btn btn-p" style="padding:5px 14px" onclick="uploadPan()">Upload &amp; Analyze</button>
      <span id="intakeStatus" style="color:#94a3b8;font-size:0.8em;margin-left:8px">Select a .pan file</span>
    </div>
    <div id="intakeSummary" style="padding:0 16px"></div>
    <div class="tabs" id="intakeTabBar" style="padding:0 16px"></div>
    <div id="intakeContents" style="flex:1;overflow-y:auto;padding:0 16px"></div>
  </div>
</div>
<div class="modal-bg" id="usersModal">
  <div style="background:#111827;border:1px solid #334155;border-radius:8px;padding:24px;width:400px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <h3 style="color:#60a5fa">Admin Users</h3>
      <button class="btn btn-o" style="padding:4px 12px;font-size:0.8em" onclick="closeUsers()">Close</button>
    </div>
    <div id="userList" style="border:1px solid #1e293b;border-radius:4px;margin-bottom:16px;max-height:200px;overflow-y:auto"></div>
    <div style="font-size:0.8em;color:#94a3b8;margin-bottom:8px">Add new admin user:</div>
    <input type="text" id="newUser" placeholder="Username" style="margin-bottom:6px">
    <input type="password" id="newPass" placeholder="Password" style="margin-bottom:10px" onkeydown="if(event.key==='Enter')addUser()">
    <div style="display:flex;gap:8px;align-items:center">
      <button class="btn btn-p" style="padding:5px 14px" onclick="addUser()">Add User</button>
      <span id="userStatus" style="color:#94a3b8;font-size:0.8em"></span>
    </div>
  </div>
</div>
<div class="modal-bg" id="loginModal">
  <div style="background:#111827;border:1px solid #334155;border-radius:8px;padding:24px;width:340px">
    <h3 style="color:#60a5fa;margin-bottom:16px">Admin Login</h3>
    <input type="text" id="loginUser" placeholder="Username" style="margin-bottom:8px" onkeydown="if(event.key==='Enter')document.getElementById('loginPass').focus()">
    <input type="password" id="loginPass" placeholder="Password" style="margin-bottom:12px" onkeydown="if(event.key==='Enter')doAdminLogin()">
    <div id="loginError" style="color:#f87171;font-size:0.8em;margin-bottom:8px"></div>
    <div style="display:flex;gap:8px">
      <button class="btn btn-p" onclick="doAdminLogin()">Login</button>
      <button class="btn btn-o" onclick="document.getElementById('loginModal').classList.remove('open')">Cancel</button>
    </div>
  </div>
</div>
<div class="modal-bg" id="editorModal">
  <div class="modal">
    <div class="modal-hdr">
      <h3>.bas Template Editor</h3>
      <button class="btn btn-o" style="padding:4px 12px;font-size:0.8em" onclick="closeEditor()">Close</button>
    </div>
    <div class="modal-body">
      <div class="modal-side" id="basFileList"></div>
      <div class="modal-edit">
        <textarea id="basEditor" placeholder="Select a .bas file to edit..." oninput="onEditorChange()"></textarea>
        <div class="modal-foot">
          <button class="btn btn-p" style="padding:5px 14px" onclick="saveBasFile()">Save</button>
          <span class="ed-status" id="edStatus">Select a file</span>
        </div>
      </div>
    </div>
  </div>
</div>
</body>
</html>"""


SUMMARY_TABLE_HTML = Path(__file__).parent / "SBS-Summary-Table-Generator.html"


@app.get("/summary-table", response_class=HTMLResponse)
async def summary_table():
    return HTMLResponse(SUMMARY_TABLE_HTML.read_text())


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8087, log_level="info")
