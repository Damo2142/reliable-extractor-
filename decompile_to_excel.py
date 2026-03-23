#!/usr/bin/env python3
"""
Decompiler: Read .pan binary → write Excel in RC-Studio-Output.xlsx format.
Same sheets, same columns, same data.
"""
import struct, sys, re
from openpyxl import Workbook

PAN = sys.argv[1] if len(sys.argv) > 1 else '/srv/dfa/drops/SBS-AHU-FINAL-r04.pan'
OUT = sys.argv[2] if len(sys.argv) > 2 else '/srv/dfa/drops/RC-Studio-OutputsRev.xlsx'

pan = open(PAN, 'rb').read()

UNIT_NAMES = {
    64: '°F', 98: '%', 96: 'ppm', 84: 'CFM', 58: 'WC', 117: 'BTU/lb',
    72: 'Min.', 73: 'Sec', 77: 'FPM', 95: '', 29: '%RH', 5: 'Volts',
    42: 'kW', 85: 'GPM', 71: 'hours', 0: '',
}

RANGE_NAMES = {
    0: 'Off/On', 1: 'Close/Open', 2: 'Stop/Start', 3: '0.0 ->100%',
    4: 'Normal/Alarm', 7: 'WC', 15: 'CFM', 17: 'min', 21: 'Clean/Dirty',
    22: '0 ->100% (0-5V)', 43: 'ppm', 58: 'BTU/lb',
}

AI_RANGE_NAMES = {
    0: 'Off/On', 1: 'Close/Open', 2: 'Stop/Start', 3: '10K -40 ->250',
    4: 'Normal/Alarm', 7: 'WC', 15: 'CFM', 17: 'min', 21: 'Clean/Dirty',
    22: '0 ->100% (0-5V)', 43: 'ppm', 58: 'BTU/lb',
}

TYPE_NAMES = {0: 'AI', 1: 'AO', 2: 'AV', 3: 'BI', 4: 'BO', 5: 'BV',
              8: 'DEV', 12: 'LOOP', 15: 'NC', 16: 'PRG', 17: 'SCHED',
              19: 'MV', 20: 'STL', 26: 'NOTIF', 141: 'TBL'}

# ── Parse index ──
def ri(d, o):
    v = []
    for i in range(4):
        h = struct.unpack_from('<H', d, o + i * 4)[0]
        l = struct.unpack_from('<H', d, o + i * 4 + 2)[0]
        v.append((h << 16) | l)
    return v

first = ri(pan, 0x400)
all_blocks = {}  # type_id -> {instance: block_bytes}
ents = [(first[1], first[2], first[3])]
for i in range(1, first[0]):
    ents.append(tuple(ri(pan, 0x440 + (i - 1) * 0x40)[1:]))

for tid, off, cnt in ents:
    d = {}
    pos = off
    for _ in range(cnt):
        if pos + 2 > len(pan): break
        sz = struct.unpack_from('<H', pan, pos)[0]
        blk = pan[pos:pos + sz + 2]
        inst = struct.unpack_from('>I', blk, 6)[0] & 0x3FFFFF
        d[inst] = blk
        pos += sz + 2
    all_blocks[tid] = d


# ── Property extraction ──
def get_props(blk):
    p = blk[12:]
    props = []
    r = 3
    while r < len(p):
        rl = p[r]
        if rl == 0: break
        rec = p[r:min(r + rl, len(p))]
        is_last = (r + rl >= len(p))
        if len(rec) >= 6:
            ctx = (rec[2] << 8) | rec[3]
            prop = rec[4]
            tag = rec[5]
            val = rec[6:]
            # Strip trailing 0x00 terminator, but NOT for last record
            # (last record has no terminator — it was stripped by _seed_payload)
            if not is_last and val and val[-1] == 0x00:
                val = val[:-1]
            props.append((ctx, prop, tag, val))
        r += rl
    return props


def get_str(props, ctx, pid):
    for c, p, t, v in props:
        if c == ctx and p == pid and t == 0x75 and len(v) > 2:
            slen = v[0]
            return bytes(v[2:2 + slen - 1]).decode('latin-1', errors='replace')
    return ''


def get_float(props, ctx, pid):
    for c, p, t, v in props:
        if c == ctx and p == pid and t == 0x44 and len(v) >= 4:
            return struct.unpack('>f', v[:4])[0]
    return None


def get_u8(props, ctx, pid):
    for c, p, t, v in props:
        if c == ctx and p == pid and t in (0x91, 0x21) and len(v) >= 1:
            return v[0]
    return None


def get_u16(props, ctx, pid):
    for c, p, t, v in props:
        if c == ctx and p == pid and t == 0x22 and len(v) >= 2:
            return struct.unpack('>H', v[:2])[0]
    return None


def get_objref(props, ctx, pid):
    for c, p, t, v in props:
        if c == ctx and p == pid and t == 0x0C and len(v) >= 4:
            oid = struct.unpack('>I', v[:4])[0]
            return ((oid >> 22) & 0x3FF, oid & 0x3FFFFF)
    return None


# ── Build reverse name lookup from all blocks ──
# instance→name for resolving ObjID refs back to point names
obj_to_name = {}
for tid in (0, 1, 2, 3, 4, 5, 19):
    for inst, blk in all_blocks.get(tid, {}).items():
        props = get_props(blk)
        name = get_str(props, 0, 0x4D)
        if name:
            obj_to_name[(tid, inst)] = name


def resolve_name(ref_tuple):
    if ref_tuple and ref_tuple in obj_to_name:
        return obj_to_name[ref_tuple]
    return ''


# ── Create workbook ──
wb = Workbook()
wb.remove(wb.active)


def add_title(ws, title, subtitle):
    ws.append([title])
    ws.append([subtitle])
    ws.append([])
    return 4  # header row (1-indexed)


# ── Build AI-to-Table mapping ──
# AIs with range_code=3 and non-temperature units use table scaling
# Map them 1:1 in order to non-filler tables
ai_table_map = {}  # ai_instance -> table_instance
_table_ais = []
for inst in sorted(all_blocks.get(0, {}).keys()):
    blk = all_blocks[0][inst]
    props = get_props(blk)
    name = get_str(props, 0, 0x4D)
    if not name or name.startswith('---'):
        continue
    rc = get_u8(props, 4, 0x1D)
    uc = get_u8(props, 0, 0x75)
    if rc == 3 and uc not in (64, 0):
        _table_ais.append(inst)
_real_tables = []
for inst in sorted(all_blocks.get(141, {}).keys()):
    blk = all_blocks[141][inst]
    props = get_props(blk)
    name = get_str(props, 0, 0x4D)
    short = name.replace('{device-name}-', '') if name else ''
    if short:
        _real_tables.append(inst)
for ai_inst, tbl_inst in zip(_table_ais, _real_tables):
    ai_table_map[ai_inst] = tbl_inst


# ═══════════════════════════════════════════
# INPUTS
# ═══════════════════════════════════════════
ws = wb.create_sheet("Inputs")
add_title(ws, "INPUTS", "{device-name}")
ws.append(['Terminal', 'Name', 'Type', 'Value', 'Units', 'Range', 'Description', 'Module'])

# Merge AI and BI by terminal number
all_inputs = {}
for tid, type_name in [(0, 'AI'), (3, 'BI')]:
    for inst, blk in all_blocks.get(tid, {}).items():
        all_inputs[inst] = (type_name, blk)

if all_inputs:
    max_term = max(all_inputs.keys())
    for t in range(1, max_term + 1):
        if t in all_inputs:
            typ, blk = all_inputs[t]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name or name.startswith('---'):
                ws.append([t, name if name else f'--- IN{t} unused ---'])
                continue
            units_code = get_u8(props, 0, 0x75)
            units = UNIT_NAMES.get(units_code, '') if typ == 'AI' else ''
            rc = get_u8(props, 4, 0x1D)
            if typ == 'AI' and t in ai_table_map:
                range_str = f'Table{ai_table_map[t]}'
            elif typ == 'AI':
                range_str = AI_RANGE_NAMES.get(rc, '10K -40 ->250')
            else:
                range_str = RANGE_NAMES.get(rc, '')
            desc = get_str(props, 0, 0x1C)
            ws.append([t, name, typ, None, units if units else None, range_str if range_str else None, desc, ''])
        else:
            ws.append([t, f'--- IN{t} unused ---'])


# ═══════════════════════════════════════════
# OUTPUTS
# ═══════════════════════════════════════════
ws = wb.create_sheet("Outputs")
add_title(ws, "OUTPUTS", "{device-name}")
ws.append(['Terminal', 'Name', 'Type', 'Value', 'Units', 'Range', 'Min V', 'Max V', 'Description', 'Module'])

all_outputs = {}
for tid, type_name in [(1, 'AO'), (4, 'BO')]:
    for inst, blk in all_blocks.get(tid, {}).items():
        all_outputs[inst] = (type_name, blk)

if all_outputs:
    max_term = max(all_outputs.keys())
    for t in range(1, max_term + 1):
        if t in all_outputs:
            typ, blk = all_outputs[t]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name or name.startswith('---'):
                ws.append([t, name if name else f'--- OUT{t} unused ---'])
                continue
            units_code = get_u8(props, 0, 0x75)
            units = UNIT_NAMES.get(units_code, '%')
            rc = get_u8(props, 4, 0x1D)
            range_str = RANGE_NAMES.get(rc, '0.0 ->100%') if typ == 'AO' else RANGE_NAMES.get(rc, 'Stop/Start')
            if typ == 'AO':
                minv = get_float(props, 4, 0x22)
                maxv = get_float(props, 4, 0x23)
            else:
                # BO: standard convention 2V/10V
                minv = 2
                maxv = 10
            desc = get_str(props, 0, 0x1C)
            ws.append([t, name, typ, None, units, range_str,
                       round(minv, 1) if minv is not None else None,
                       round(maxv, 1) if maxv is not None else None,
                       desc, ''])
        else:
            ws.append([t, f'--- OUT{t} unused ---'])


# ═══════════════════════════════════════════
# VALUES
# ═══════════════════════════════════════════
ws = wb.create_sheet("Values")
add_title(ws, "VALUES (AV / BV / MV)", "{device-name}")
ws.append(['Instance', 'Name', 'Type', 'Default Value', 'Units', 'Range / States', 'Min', 'Max', 'Description', 'Module'])

# Merge AV, BV, MV by instance
all_values = {}
for tid, prefix in [(2, 'AV'), (5, 'BV'), (19, 'MV')]:
    for inst, blk in all_blocks.get(tid, {}).items():
        all_values[inst] = (prefix, blk)

if all_values:
    max_inst = max(all_values.keys())
    for i in range(1, max_inst + 1):
        if i in all_values:
            prefix, blk = all_values[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'AV{i}'])
                continue

            if prefix == 'AV':
                pv = get_float(props, 0, 0x55)
                units_code = get_u8(props, 0, 0x75)
                units = UNIT_NAMES.get(units_code, '')
                pv_str = str(round(pv, 1)) if pv is not None else '0.0'
                # Time format for units=72 and pv=0
                if units_code == 72 and pv is not None and pv == 0.0:
                    pv_str = '00:00:00'
                    units = 'Time'
                desc = get_str(props, 0, 0x1C)
                ws.append([f'AV{i}', name, 'AV', pv_str, units, None, None, None, desc, ''])

            elif prefix == 'BV':
                pv = get_u8(props, 0, 0x55)
                pv_str = 'On' if pv == 1 else 'Off'
                desc = get_str(props, 0, 0x1C)
                ws.append([f'BV{i}', name, 'BV', pv_str, 'Off/On', 'Off/On', None, None, desc, ''])

            elif prefix == 'MV':
                pv = get_u8(props, 0, 0x55)
                states = get_str(props, 0, 0x1C)  # states in description
                # Find state name for pv
                pv_name = str(pv)
                if states:
                    for part in states.split('/'):
                        part = part.strip()
                        if '-' in part:
                            num, nm = part.split('-', 1)
                            try:
                                if int(num.strip()) == pv:
                                    pv_name = nm.strip()
                            except:
                                pass
                ws.append([f'MV{i}', name, 'MV', pv_name, states, states, None, None, states, ''])
        else:
            ws.append([f'AV{i}'])


# ═══════════════════════════════════════════
# LOOPS
# ═══════════════════════════════════════════
ws = wb.create_sheet("Loops")
add_title(ws, "PID LOOPS", "{device-name}")
ws.append(['Loop', 'Name', 'Input', 'Setpoint', 'Action', 'P Band', 'Integral', 'D', 'Bias', 'Description', 'Module'])

loop_blocks = all_blocks.get(12, {})
if loop_blocks:
    max_l = max(loop_blocks.keys())
    for i in range(1, max_l + 1):
        if i in loop_blocks:
            blk = loop_blocks[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'LOOP{i}'])
                continue
            # Strip {device-name}- prefix for short name
            short = name.replace('{device-name}-', '')
            # Detect filler rows: name is just '{device-name}-' (short='')
            if not short:
                ws.append([f'LOOP{i}'])
                continue
            act = get_u8(props, 0, 0x02)
            act_s = '-' if act == 0 else '+'
            pb = get_float(props, 4, 0x4D)
            integral = get_float(props, 0, 0x6C)
            deriv = get_float(props, 0, 0x0E)
            # Input ref
            inp_ref = get_objref(props, 0, 0x13)
            inp_name = resolve_name(inp_ref) if inp_ref else ''
            # Setpoint ref — in 0x6D record
            sp_name = ''
            for c, p, t, v in props:
                if c == 0 and p == 0x6D and t == 0x0E and len(v) >= 5:
                    if v[0] == 0x0C and len(v) >= 5:
                        oid = struct.unpack('>I', v[1:5])[0]
                        sp_type = (oid >> 22) & 0x3FF
                        sp_inst = oid & 0x3FFFFF
                        sp_name = resolve_name((sp_type, sp_inst))
            desc = get_str(props, 0, 0x1C)
            ws.append([f'LOOP{i}', short, inp_name, sp_name, act_s,
                       round(pb, 1) if pb else 0, round(integral, 1) if integral else 0,
                       round(deriv, 1) if deriv else 0, 'H', desc, ''])
        else:
            ws.append([f'LOOP{i}'])


# ═══════════════════════════════════════════
# TABLES
# ═══════════════════════════════════════════
ws = wb.create_sheet("Tables")
add_title(ws, "TABLES", "{device-name}")
ws.append(['Table', 'Name', 'Input Units', 'Output Units', 'Description', 'Module',
           'X1', 'Y1', 'X2', 'Y2', 'X3', 'Y3'])

tbl_blocks = all_blocks.get(141, {})
if tbl_blocks:
    max_t = max(tbl_blocks.keys())
    for i in range(1, max_t + 1):
        if i in tbl_blocks:
            blk = tbl_blocks[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'TBL{i}'])
                continue
            short = name.replace('{device-name}-', '')
            # Detect filler rows: name is just '{device-name}-' (short='')
            if not short:
                ws.append([f'TBL{i}'])
                continue
            units_code = get_u8(props, 0, 0x75)
            out_units = UNIT_NAMES.get(units_code, '')
            # XY data from 0x0428
            xy = []
            for c, p, t, v in props:
                if c == 4 and p == 0x28:
                    # Parse XY pairs: [00 19 00] then [0C float][1C float] pairs
                    j = 0
                    while j < len(v):
                        if v[j] == 0x0C and j + 4 < len(v):
                            x = struct.unpack('>f', v[j + 1:j + 5])[0]
                            j += 5
                            if j < len(v) and v[j] == 0x1C and j + 4 < len(v):
                                y = struct.unpack('>f', v[j + 1:j + 5])[0]
                                xy.append((round(x, 2), round(y, 2)))
                                j += 5
                            else:
                                j += 1
                        else:
                            j += 1
            desc = get_str(props, 0, 0x1C)
            row = [f'TBL{i}', short, 'Volts', out_units, desc, '']
            for x, y in xy[:3]:
                row.extend([x, y])
            ws.append(row)
        else:
            ws.append([f'TBL{i}'])


# ═══════════════════════════════════════════
# SCHEDULES
# ═══════════════════════════════════════════
ws = wb.create_sheet("Schedules")
add_title(ws, "SCHEDULES", "{device-name}")
ws.append(['Schedule', 'Name', 'Default State', 'States', 'Priority', 'Description', 'Module'])

sched_blocks = all_blocks.get(17, {})
if sched_blocks:
    max_s = max(sched_blocks.keys())
    for i in range(1, max_s + 1):
        if i in sched_blocks:
            blk = sched_blocks[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'SCHED{i}'])
                continue
            default = get_u8(props, 0, 0xAE)
            default_str = 'Unoccupied' if default == 0 else 'Occupied'
            ws.append([f'SCHED{i}', name, default_str, 'Unoccupied / Occupied', 10, '', ''])
        else:
            ws.append([f'SCHED{i}'])


# ═══════════════════════════════════════════
# TRENDS
# ═══════════════════════════════════════════
ws = wb.create_sheet("Trends")
add_title(ws, "SINGLE POINT TREND LOGS (STL)", "{device-name}")
ws.append(['STL', 'Name', 'Monitored Point', 'Type', 'Interval/COV Delta', 'Buffer'])

stl_blocks = all_blocks.get(20, {})
if stl_blocks:
    max_stl = max(stl_blocks.keys())
    for i in range(1, max_stl + 1):
        if i in stl_blocks:
            blk = stl_blocks[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'STL{i}'])
                continue
            lt = get_u8(props, 0, 0x48)
            type_str = 'Polled' if lt == 1 else 'Cov'
            buf = get_u16(props, 0, 0x7E) or 512
            mon_ref = get_objref(props, 0, 0x84)
            mon_name = resolve_name(mon_ref) if mon_ref else ''
            if lt == 1:
                interval = get_u16(props, 0, 0x80) or 900
                h = interval // 3600
                m = (interval % 3600) // 60
                s = interval % 60
                interval_str = f'{h:02d}:{m:02d}:{s:02d}'
            else:
                cov = get_float(props, 0, 0x7F)
                interval_str = str(round(cov, 1)) if cov else '0.2'
            ws.append([f'STL{i}', name, mon_name, type_str, interval_str, buf])
        else:
            ws.append([f'STL{i}'])


# ═══════════════════════════════════════════
# SYSTEM GROUPS
# ═══════════════════════════════════════════
ws = wb.create_sheet("System Groups")
add_title(ws, "SYSTEM GROUPS (Graphic Pages)", "{device-name}")
ws.append(['Group Name', 'Description', 'Module'])

notif_blocks = all_blocks.get(26, {})
if notif_blocks:
    for inst in sorted(notif_blocks.keys()):
        blk = notif_blocks[inst]
        props = get_props(blk)
        name = get_str(props, 0, 0x4D)
        if name:
            ws.append([name, '', ''])


# ═══════════════════════════════════════════
# PROGRAMS
# ═══════════════════════════════════════════
ws = wb.create_sheet("Programs")
add_title(ws, "PROGRAMS", "{device-name}")
ws.append(['PRG#', 'Name', 'Filename', 'Enabled', 'Exec Order', 'Description', 'Module'])

prg_blocks = all_blocks.get(16, {})
if prg_blocks:
    max_p = max(prg_blocks.keys())
    for i in range(1, max_p + 1):
        if i in prg_blocks:
            blk = prg_blocks[i]
            props = get_props(blk)
            name = get_str(props, 0, 0x4D)
            if not name:
                ws.append([f'PRG{i}'])
                continue
            en = get_u8(props, 4, 0x0A)
            en_str = 'Yes' if en == 1 else 'No'
            has_bc = len(blk) > 100
            ws.append([f'PRG{i}', name, '', en_str, i, '', ''])
        else:
            ws.append([f'PRG{i}'])


# ═══════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
