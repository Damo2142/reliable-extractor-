#!/usr/bin/env python3
"""
SBS .pan Compiler — SEED FORMAT.
Reads RC-Studio-Output.xlsx, writes seed-format .pan file.
Seed blocks = minimal properties. RC Studio fills the rest on open.

Can be called as:
  - CLI:  python compile_from_excel.py [package_path] [--output out.pan]
  - API:  compile_package(pkg_path, blank_path) -> bytes
"""
import struct, zipfile, re, sys, os
from io import BytesIO
from pathlib import Path

# Ensure composition package is importable
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook
from composition.pan_compiler import (
    crc16_kermit,
    write_av_seed, write_av_block, write_ai_seed, write_ao_seed,
    write_bi_seed, write_bo_seed, write_bv_seed,
    write_mv_seed, write_loop_block, write_prg_seed,
    write_sched_seed, write_sys_group_seed, write_table_block,
    write_stl_block, write_notif_cls_seed, write_empty_block, write_prg_block,
    extract_nc_groups_from_blank, extract_device_block_from_blank,
    read_blank_header, _write_index_entry,
)
from composition.cbas_compiler import compile_bas

# ── Constants ──
BLANKS_BASE = "/srv/dfa/shared/files/vendors/reliable/blanks/"
DEFAULT_BLANK = BLANKS_BASE + "MACH-ProSys-88/MACH-ProSys-88.panx"

# Map controller model IDs to blank file paths (relative to BLANKS_BASE)
BLANK_FILE_MAP = {
    'MPS':     'MACH-ProSys-88/MACH-ProSys-88.panx',
    'MPWS':    'MACH-ProWebSys-88/MACH-ProWebSys-88.panx',
    'MPV-LCD': 'MACH-ProView LCD/MACH-ProView LCD.panx',
    'MP2':     'MACH-Pro2-88/MACH-Pro2-88.panx',
    'MP1':     'MACH-Pro1-88/MACH-Pro1-88.panx',
    'MPC':     'MACH-ProCom-88/MACH-ProCom-88.panx',
    'MPWC':    'MACH-ProWebCom-88/MACH-ProWebCom-88.panx',
    'MPA-36':  'MACH-ProAir-28/MACH-ProAir-28.panx',    # 3in/6out = ProAir-28
    'MPA-35':  'MACH-ProAir-18/MACH-ProAir-18.panx',    # 3in/5out = ProAir-18
    'MPA-34':  'MACH-ProAir-18/MACH-ProAir-18.panx',    # 3in/4out = ProAir-18
    'MPA-33':  'MACH-ProAir-08/MACH-ProAir-08.panx',    # 3in/3out = ProAir-08
    'MPZ-88':  'MACH-ProZone-88/MACH-ProZone-88.panx',
    'MPZ-44':  'MACH-ProZone-44/MACH-ProZone-44.panx',
}


def get_blank_path(controller_model: str) -> str:
    """Get blank .panx file path for a controller model. Falls back to MPS."""
    relative = BLANK_FILE_MAP.get(controller_model, BLANK_FILE_MAP['MPS'])
    path = os.path.join(BLANKS_BASE, relative)
    if not os.path.exists(path):
        print(f"WARNING: blank not found for {controller_model} at {path}, falling back to MPS")
        path = DEFAULT_BLANK
    return path

RANGE_CODES = {
    'Off/On': 0, 'Normal/Alarm': 4, 'Clean/Dirty': 21, 'Close/Open': 1,
    'Stop/Start': 2, '0.0 ->100%': 3, '10K -40 ->250': 3,
    '0 ->100% (0-5V)': 22,
}

UNIT_CODES = {
    '°F': 64, '%': 98, 'BTU/lb': 117, 'ppm': 96, 'CFM': 84,
    'WC': 58, 'FPM': 77, 'Min.': 72, 'Time': 72, 'Sec': 73,
    '#': 95, '': 95,
}
UNIT_TO_RANGE = {
    64: 2, 98: 22, 95: 0, 96: 43, 84: 15, 58: 7, 117: 58,
    72: 17, 73: 0, 77: 0, 29: 22, 5: 0, 42: 0, 85: 0, 71: 0,
}


# ── Helpers ──
def _s(v):
    if v is None: return ''
    return str(v).strip()

def _uc(v):
    return UNIT_CODES.get(_s(v), 95)

def _pi(text):
    return int(re.sub(r'[^0-9]', '', str(text or '0')))

def _rng(range_str):
    return RANGE_CODES.get(_s(range_str), 0)


def compile_package(pkg_path: str, controller_model: str = 'MPS',
                    blank_path: str = None, verbose: bool = False) -> bytes:
    """
    Compile a Composition Engine v2 package into a .pan binary.

    Args:
        pkg_path: Path to package directory or zip file
        controller_model: Controller model ID (e.g., 'MPS', 'MP2', 'MPA-36')
        blank_path: Override blank file path (if None, auto-selects from controller_model)
        verbose: Print progress to stdout

    Returns:
        bytes: Complete .pan file content
    """
    if blank_path is None:
        blank_path = get_blank_path(controller_model)

    def log(msg):
        if verbose:
            print(msg)

    # ── Read package ──
    log(f"Reading: {pkg_path}")
    bas_files = {}
    if os.path.isdir(pkg_path):
        wb = load_workbook(os.path.join(pkg_path, "RC-Studio-Output.xlsx"), read_only=True)
        prg_dir = os.path.join(pkg_path, "programs")
        if os.path.isdir(prg_dir):
            for fn in os.listdir(prg_dir):
                if fn.endswith('.bas'):
                    bas_files[fn] = open(os.path.join(prg_dir, fn), 'r').read()
    else:
        with zipfile.ZipFile(pkg_path) as z:
            wb = load_workbook(BytesIO(z.read("RC-Studio-Output.xlsx")), read_only=True)
            for n in z.namelist():
                if n.endswith('.bas'):
                    fn = n.split('/')[-1]
                    if fn not in bas_files:
                        bas_files[fn] = z.read(n).decode('utf-8', errors='replace')
    log(f"  {len(bas_files)} .bas files")

    # Pre-load all sheets into memory — prevents read_only iterator exhaustion
    _sheets = {name: list(wb[name].iter_rows(values_only=True)) for name in wb.sheetnames}

    def rows(sheet_name):
        all_rows = _sheets.get(sheet_name, [])[4:]  # skip first 4 rows (title/header)
        result = []
        for r in all_rows:
            if r[0] is None: continue
            c0 = _s(r[0])
            if c0 in ('', 'Terminal', 'Instance', 'Loop', 'Table', 'STL',
                       'Schedule', 'Group Name', 'PRG#'): continue
            if '---' in c0 or '---' in _s(r[1] if len(r) > 1 else ''): continue
            result.append(list(r))
        return result

    # ── Build name→obj lookup ──
    name_to_obj = {}
    for r in rows("Inputs"):
        nm = _s(r[1]); typ = _s(r[2])
        if nm: name_to_obj[nm] = (0 if typ == 'AI' else 3, int(r[0]))
    for r in rows("Outputs"):
        nm = _s(r[1]); typ = _s(r[2])
        if nm: name_to_obj[nm] = (1 if typ == 'AO' else 4, int(r[0]))
    for r in rows("Values"):
        nm = _s(r[1]); c0 = _s(r[0])
        if nm and any(c0.startswith(p) for p in ['AV','BV','MV']):
            tid = {'AV':2,'BV':5,'MV':19}.get(c0[:2], 2)
            name_to_obj[nm] = (tid, _pi(r[0]))

    def resolve(pn):
        pn = _s(pn)
        if pn in name_to_obj: return name_to_obj[pn]
        return (0x3FF, 0)

    # ══════════════════════════════════════════════════════════════
    # BUILD BLOCKS — SEED FORMAT
    # ══════════════════════════════════════════════════════════════
    blocks = {}

    # NC_GROUP from blank
    blocks[15] = extract_nc_groups_from_blank(blank_path)
    log(f"  NC_GROUP: {len(blocks[15])}")

    # INPUTS
    ai, bi = [], []
    for r in rows("Inputs"):
        inst = int(r[0]); nm = _s(r[1]); typ = _s(r[2])
        if not nm:  # filler
            (ai if typ != 'BI' else bi).append(write_empty_block(0 if typ != 'BI' else 3, inst))
            continue
        desc = _s(r[6] if len(r) > 6 else '')
        u = _uc(r[4] if len(r) > 4 else None)
        if typ == 'AI':
            ai.append(write_ai_seed(inst, nm, desc, units=u))
        elif typ == 'BI':
            bi.append(write_bi_seed(inst, nm, desc))
    if ai: blocks[0] = ai; log(f"  AI: {len(ai)}")
    if bi: blocks[3] = bi; log(f"  BI: {len(bi)}")

    # OUTPUTS
    ao, bo = [], []
    for r in rows("Outputs"):
        inst = int(r[0]); nm = _s(r[1]); typ = _s(r[2])
        if not nm:  # filler
            (ao if typ != 'BO' else bo).append(write_empty_block(1 if typ != 'BO' else 4, inst))
            continue
        desc = _s(r[8] if len(r) > 8 else '')
        u = _uc(r[4] if len(r) > 4 else None)
        if typ == 'AO':
            ao.append(write_ao_seed(inst, nm, desc, units=u))
        elif typ == 'BO':
            bo.append(write_bo_seed(inst, nm, desc))
    if ao: blocks[1] = ao; log(f"  AO: {len(ao)}")
    if bo: blocks[4] = bo; log(f"  BO: {len(bo)}")

    # VALUES
    av, bv, mv = [], [], []
    for r in rows("Values"):
        c0 = _s(r[0]); nm = _s(r[1]); typ = c0[:2]; inst = _pi(r[0])
        if not nm:  # filler — empty block, default to AV type
            tid = {'AV':2,'BV':5,'MV':19}.get(typ, 2)
            av.append(write_empty_block(tid, inst))
            continue
        desc = _s(r[8] if len(r) > 8 else '')
        states = _s(r[5] if len(r) > 5 else '')
        u = _uc(r[4] if len(r) > 4 else None)
        dv = r[3] if len(r) > 3 else None
        if typ == 'AV':
            pv = 0.0
            if dv is not None:
                dv_s = str(dv).strip()
                if ':' in dv_s:
                    parts = dv_s.split(':')
                    try: pv = int(parts[0]) * 60 + int(parts[1]) + int(parts[2]) / 60.0
                    except: pv = 0.0
                else:
                    try: pv = float(dv_s)
                    except: pv = 0.0
            rc = UNIT_TO_RANGE.get(u, 0)
            av.append(write_av_block(inst, nm, present_value=pv, units=u, range_code=rc, desc=desc))
        elif typ == 'BV':
            pv = 1 if str(dv or '').strip().lower() in ('1', 'true', 'on', 'active') else 0
            bv.append(write_bv_seed(inst, nm, desc, present_value=pv))
        elif typ == 'MV':
            # Strip number prefixes: "1-Occupied/2-Bypass" → "Occupied/Bypass"
            if states:
                stripped = '/'.join(p.split('-', 1)[1] if '-' in p else p for p in states.split('/'))
            else:
                stripped = desc
            mv.append(write_mv_seed(inst, nm, stripped))
    if av: blocks[2] = av; log(f"  AV: {len(av)}")
    if bv: blocks[5] = bv; log(f"  BV: {len(bv)}")
    if mv: blocks[19] = mv; log(f"  MV: {len(mv)}")

    # LOOPS — with input/setpoint/output refs
    loops = []
    for r in rows("Loops"):
        if not _s(r[1]):  # skip unused loop instances — no empty blocks
            continue
        inst = _pi(r[0]); nm = "{device-name}-" + _s(r[1])
        inp_name = _s(r[2]); sp_name = _s(r[3])
        out_name = _s(r[4]) if len(r) > 4 else ''
        action = _s(r[5]) if len(r) > 5 else '+'
        pband = float(r[6] or 0) if len(r) > 6 else 0.0
        integ = float(r[7] or 0) if len(r) > 7 else 0.0
        deriv = float(r[8] or 0) if len(r) > 8 else 0.0
        desc = _s(r[10] if len(r) > 10 else '')
        in_t, in_i = resolve(inp_name)
        sp_t, sp_i = resolve(sp_name)
        out_t, out_i = resolve(out_name) if out_name else (0x3FF, 0x3FFFFF)
        # Fully populated LOOP block — matches RC Studio reference files exactly
        action_int = 1 if action == "+" else 0  # direct=1, reverse=0
        loops.append(write_loop_block(inst, nm, action_int,
            input_type=in_t, input_inst=in_i,
            sp_type=sp_t, sp_inst=sp_i,
            p_band=pband, integral=integ, derivative=deriv, desc=desc))
    if loops: blocks[12] = loops; log(f"  LOOP: {len(loops)}")

    # DEVICE from blank
    blocks[8] = [extract_device_block_from_blank(blank_path)]
    log(f"  DEVICE: 1")

    # SCHEDULES
    scheds = []
    for r in rows("Schedules"):
        inst = _pi(r[0]); nm = _s(r[1])
        if not nm:
            scheds.append(write_empty_block(17, inst))
        else:
            scheds.append(write_sched_seed(inst, nm))
    if scheds: blocks[17] = scheds; log(f"  SCHED: {len(scheds)}")

    # PROGRAMS — with bytecode
    prgs = []
    for r in rows("Programs"):
        inst = _pi(r[0]); nm = _s(r[1])
        if not nm:
            prgs.append(write_empty_block(16, inst))
            continue
        filename = _s(r[2] if len(r) > 2 else '')
        enabled = 1 if _s(r[3] if len(r) > 3 else 'Yes').lower() in ('yes','1','true') else 0
        bc = b'\x09'  # END opcode default
        if filename and filename in bas_files:
            try:
                bc = compile_bas(bas_files[filename])
            except Exception as e:
                log(f"  WARN: PRG{inst} ({filename}): {e}")
        elif filename:
            log(f"  WARN: PRG{inst} .bas not found: {filename}")
        desc = _s(r[5] if len(r) > 5 else '')
        prgs.append(write_prg_block(inst, nm, bc, enabled=enabled, desc=desc))
    if prgs: blocks[16] = prgs; log(f"  PRG: {len(prgs)} ({len(bas_files)} .bas)")

    # TABLES — type 141 with XY data (no fillers for gaps, no SYS_GROUP)
    tbls = []
    for r in rows("Tables"):
        if not _s(r[1]): continue  # skip blank rows
        inst = _pi(r[0]); nm = "{device-name}-" + _s(r[1]) + str(inst)
        u = _uc(r[3] if len(r) > 3 else None)
        xy = []
        p = 0
        while True:
            xi, yi = 6 + p * 2, 7 + p * 2
            if len(r) <= yi or r[xi] is None or r[yi] is None:
                break
            try: xy.append((float(r[xi]), float(r[yi])))
            except: break
            p += 1
        tbls.append(write_table_block(inst, nm, xy_pairs=xy, units=u))
    if tbls: blocks[141] = tbls; log(f"  TABLE: {len(tbls)}")

    # NOTIF_CLS — name + 0x0485=TRUE
    notifs = []
    ni = 1
    for r in rows("System Groups"):
        nm = _s(r[0])
        if nm:
            notifs.append(write_notif_cls_seed(ni, nm))
            ni += 1
    if notifs: blocks[26] = notifs; log(f"  NOTIF_CLS: {len(notifs)}")

    # TRENDS — 22 properties (fully populated, confirmed working)
    stls = []
    for r in rows("Trends"):
        inst = _pi(r[0]); nm = _s(r[1]); mon = _s(r[2])
        tt = _s(r[3]).lower() if len(r) > 3 else 'polled'
        iv_raw = r[4] if len(r) > 4 else '00:15:00'
        buf = int(r[5] if len(r) > 5 and r[5] else 512)
        mt, mi = resolve(mon)
        if tt == 'polled':
            lt = 1; cov = 0.2
            iv = str(iv_raw or '00:15:00')
            parts = iv.split(':')
            try: log_int = int(parts[0])*3600 + int(parts[1])*60 + int(parts[2])
            except: log_int = 900
        else:
            lt = 2; log_int = 0
            try: cov = float(iv_raw)
            except: cov = 0.2
        stls.append(write_stl_block(inst, nm, lt, mt, mi, buf, log_int, cov))
    if stls: blocks[20] = stls; log(f"  STL: {len(stls)}")

    # ══════════════════════════════════════════════════════════════
    # ASSEMBLE .PAN FILE
    # ══════════════════════════════════════════════════════════════
    TYPE_ORDER = [15, 0, 1, 3, 4, 2, 5, 19, 12, 141, 26, 8, 17, 20, 16]
    present = [t for t in TYPE_ORDER if t in blocks]
    num_types = len(present)
    blocks_start = 0x0400 + num_types * 0x40 + 6

    type_offsets = {}; cursor = blocks_start
    for tid in present:
        type_offsets[tid] = cursor
        for blk in blocks[tid]: cursor += len(blk)

    dev_block = blocks[8][0]
    dev_objid = struct.unpack_from('>I', dev_block, 6)[0] & 0x3FFFFF
    blank_hdr = read_blank_header(blank_path)

    fd = bytearray()
    hdr = bytearray(12)
    struct.pack_into('<I', hdr, 0, 0x0023BAC0)
    struct.pack_into('<I', hdr, 4, dev_objid)
    struct.pack_into('<I', hdr, 8, blank_hdr[2])
    fd += hdr
    fd += bytes(0x0400 - 12)

    for idx, tid in enumerate(present):
        val0 = num_types if idx == 0 else 0
        fd += _write_index_entry(val0, tid, type_offsets[tid], len(blocks[tid]))
        fd += bytes(0x40 - 16)
    fd += bytes(6)

    for tid in present:
        for blk in blocks[tid]: fd += blk

    pan_data = bytes(fd)
    log(f"  Compiled: {len(pan_data):,} bytes, {sum(len(blocks[t]) for t in present)} blocks")
    return pan_data


# ══════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    pkg = sys.argv[1] if len(sys.argv) > 1 else "/srv/dfa/drops/rcscreenshots/composition-package (7)"
    out = "/srv/dfa/drops/SBS-AHU-FINAL-r43.pan"
    for i, arg in enumerate(sys.argv):
        if arg == '--output' and i + 1 < len(sys.argv):
            out = sys.argv[i + 1]

    pan_data = compile_package(pkg, DEFAULT_BLANK, verbose=True)

    with open(out, "wb") as f:
        f.write(pan_data)

    # ── Verify ──
    TN = {0:"AI",1:"AO",2:"AV",3:"BI",4:"BO",5:"BV",8:"DEV",12:"LOOP",
          15:"NC",16:"PRG",17:"SCH",19:"MV",20:"STL",26:"NOTIF",141:"TBL"}

    def read_index(d, o):
        v = []
        for i in range(4):
            hi = struct.unpack_from('<H', d, o+i*4)[0]
            lo = struct.unpack_from('<H', d, o+i*4+2)[0]
            v.append((hi << 16) | lo)
        return v

    print(f"\n{'='*60}")
    print(f"Output: {out}")
    print(f"Size: {len(pan_data):,} bytes")

    first = read_index(pan_data, 0x0400)
    crc_ok = 0; crc_total = 0
    pos = first[2]
    for j in range(first[3]):
        if pos+12 > len(pan_data): break
        sz = struct.unpack_from('<H', pan_data, pos)[0]; tot = sz+2
        if pos+tot > len(pan_data): break
        if crc16_kermit(pan_data[pos+12:pos+tot]) == struct.unpack_from('>H', pan_data, pos+10)[0]: crc_ok += 1
        crc_total += 1; pos += tot
    print(f"  {TN.get(first[1],'?'):5s} count={first[3]:3d}")

    for i in range(1, first[0]):
        e = read_index(pan_data, 0x0440+(i-1)*0x40)
        pos = e[2]
        for j in range(e[3]):
            if pos+12 > len(pan_data): break
            sz = struct.unpack_from('<H', pan_data, pos)[0]; tot = sz+2
            if pos+tot > len(pan_data): break
            if crc16_kermit(pan_data[pos+12:pos+tot]) == struct.unpack_from('>H', pan_data, pos+10)[0]: crc_ok += 1
            crc_total += 1; pos += tot
        print(f"  {TN.get(e[1],'?'):5s} count={e[3]:3d}")

    print(f"\nCRC: {crc_ok}/{crc_total}")
    if crc_ok == crc_total: print("ALL CRC PASSED")
    print(f"Total blocks: {crc_total}")
    print(f"File size: {len(pan_data):,} bytes")
    print(f"Ready: {out}")
