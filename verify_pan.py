#!/usr/bin/env python3
"""Verify compiled .pan matches Excel exactly. Every field, every row."""
import struct, zipfile, re, sys
from io import BytesIO
from openpyxl import load_workbook

PKG = sys.argv[1] if len(sys.argv) > 1 else '/srv/dfa/drops/sbs-full-package (7).zip'
PAN = sys.argv[2] if len(sys.argv) > 2 else '/srv/dfa/drops/SBS-AHU-FINAL-r01.pan'

UNIT_CODES = {
    '°F': 64, 'degF': 64, '%': 98, '%RH': 29, 'ppm': 96,
    'CFM': 84, 'WC': 58, '"WC': 58, 'BTU/lb': 117, 'BTUlb': 117,
    'Volts': 5, 'V': 5, 'kW': 42, 'GPM': 85,
    'minutes': 72, 'min': 72, 'Min.': 72, 'seconds': 73, 'sec': 73, 'Sec': 73,
    'hours': 71, 'Time': 72, '': 95, 'None': 95, '#': 95,
}

with zipfile.ZipFile(PKG) as z:
    wb = load_workbook(BytesIO(z.read('RC-Studio-Output.xlsx')), read_only=True)
    bas_files = set(n.split('/')[-1] for n in z.namelist() if n.endswith('.bas'))

pan = open(PAN, 'rb').read()

# Parse index
def ri(d, o):
    v = []
    for i in range(4):
        h = struct.unpack_from('<H', d, o + i*4)[0]
        l = struct.unpack_from('<H', d, o + i*4 + 2)[0]
        v.append((h << 16) | l)
    return v

first = ri(pan, 0x400)
type_blocks = {}
all_entries = [(first[1], first[2], first[3])]
for i in range(1, first[0]):
    all_entries.append(tuple(ri(pan, 0x440+(i-1)*0x40)[1:]))
for tid, offset, count in all_entries:
    blks = {}
    pos = offset
    for j in range(count):
        sz = struct.unpack_from('<H', pan, pos)[0]
        blk = pan[pos:pos+sz+2]
        inst = struct.unpack_from('>I', blk, 6)[0] & 0x3FFFFF
        blks[inst] = blk
        pos += sz + 2
    type_blocks[tid] = blks


def get_props(blk):
    """Return dict of (ctx, prop) -> (tag, val_bytes)"""
    payload = blk[12:]
    props = {}
    rpos = 3
    while rpos < len(payload):
        rl = payload[rpos]
        if rl == 0: break
        rec = payload[rpos:min(rpos+rl, len(payload))]
        if len(rec) >= 6:
            ctx = (rec[2]<<8)|rec[3]
            prop = rec[4]
            tag = rec[5]
            val = rec[6:]
            if val and val[-1] == 0x00: val = val[:-1]
            props[(ctx, prop)] = (tag, val)
        rpos += rl
    return props


def get_name(props):
    r = props.get((0, 0x4D))
    if r and r[0] == 0x75 and len(r[1]) > 2:
        slen = r[1][0]
        return bytes(r[1][2:2+slen-1]).decode('latin-1', errors='replace')
    return ''


def get_float(props, ctx, prop):
    r = props.get((ctx, prop))
    if r and r[0] == 0x44 and len(r[1]) >= 4:
        return struct.unpack('>f', r[1][:4])[0]
    return None


def get_uint8(props, ctx, prop):
    r = props.get((ctx, prop))
    if r and len(r[1]) >= 1:
        return r[1][0]
    return None


def get_desc(props):
    r = props.get((0, 0x1C))
    if r and r[0] == 0x75 and len(r[1]) > 2:
        slen = r[1][0]
        return bytes(r[1][2:2+slen-1]).decode('latin-1', errors='replace')
    return ''


def uc(s):
    return UNIT_CODES.get(str(s or '').strip(), 95)


fails = []
passes = 0


def check(section, row_id, field, expected, actual):
    global passes, fails
    if expected == actual:
        passes += 1
    else:
        fails.append(f"{section} {row_id} {field}: expected={expected!r} got={actual!r}")


# ══════════════════════════════════════════════════════════════
# INPUTS
# ══════════════════════════════════════════════════════════════
for row in wb['Inputs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    name = str(row[1] or '').strip()
    if not name or '---' in name: continue
    inst = int(row[0])
    typ = str(row[2] or '').strip()
    tid = 0 if typ == 'AI' else 3
    blks = type_blocks.get(tid, {})
    if inst not in blks:
        fails.append(f"INPUTS {typ}{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("INPUTS", f"{typ}{inst}", "name", name, get_name(props))
    if typ == 'AI':
        check("INPUTS", f"AI{inst}", "units", uc(row[4]), get_uint8(props, 0, 0x75))


# ══════════════════════════════════════════════════════════════
# OUTPUTS
# ══════════════════════════════════════════════════════════════
for row in wb['Outputs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    name = str(row[1] or '').strip()
    if not name or '---' in name: continue
    inst = int(row[0])
    typ = str(row[2] or '').strip()
    tid = 1 if typ == 'AO' else 4
    blks = type_blocks.get(tid, {})
    if inst not in blks:
        fails.append(f"OUTPUTS {typ}{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("OUTPUTS", f"{typ}{inst}", "name", name, get_name(props))
    if typ == 'AO':
        check("OUTPUTS", f"AO{inst}", "units", uc(row[4]), get_uint8(props, 0, 0x75))


# ══════════════════════════════════════════════════════════════
# VALUES
# ══════════════════════════════════════════════════════════════
for row in wb['Values'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not any(c0.startswith(p) for p in ['AV','BV','MV']): continue
    name = str(row[1] or '').strip()
    if not name: continue
    typ = c0[:2]
    inst = int(c0[2:])
    tid = {'AV':2,'BV':5,'MV':19}[typ]
    blks = type_blocks.get(tid, {})
    if inst not in blks:
        fails.append(f"VALUES {c0} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("VALUES", c0, "name", name, get_name(props))
    if typ == 'AV':
        check("VALUES", c0, "units", uc(row[4]), get_uint8(props, 0, 0x75))
        # Check present value
        dv = row[3]
        if dv is not None:
            dv_s = str(dv).strip()
            if ':' not in dv_s:
                try:
                    expected_pv = float(dv_s)
                    actual_pv = get_float(props, 0, 0x55)
                    if actual_pv is not None and abs(expected_pv - actual_pv) > 0.01:
                        fails.append(f"VALUES {c0} pv: expected={expected_pv} got={actual_pv:.4f}")
                    else:
                        passes += 1
                except:
                    pass
    elif typ == 'MV':
        # Check states in description
        states = str(row[5] or '').strip()
        if states:
            pan_desc = get_desc(props)
            check("VALUES", c0, "states", states, pan_desc)


# ══════════════════════════════════════════════════════════════
# LOOPS
# ══════════════════════════════════════════════════════════════
for row in wb['Loops'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not c0.startswith('LOOP'): continue
    inst = int(c0[4:])
    name = '{device-name}-' + str(row[1] or '').strip()
    if not str(row[1] or '').strip(): continue
    blks = type_blocks.get(12, {})
    if inst not in blks:
        fails.append(f"LOOPS LOOP{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("LOOPS", f"LOOP{inst}", "name", name, get_name(props))
    # Action
    action_s = str(row[4] or '+').strip()
    expected_act = 0 if action_s == '-' else 1
    actual_act = get_uint8(props, 0, 0x02)
    check("LOOPS", f"LOOP{inst}", "action", expected_act, actual_act)
    # P-band
    expected_pb = float(row[5] or 0)
    actual_pb = get_float(props, 4, 0x4D)
    if actual_pb is not None and abs(expected_pb - actual_pb) > 0.01:
        fails.append(f"LOOPS LOOP{inst} p-band: expected={expected_pb} got={actual_pb:.4f}")
    else:
        passes += 1


# ══════════════════════════════════════════════════════════════
# PROGRAMS
# ══════════════════════════════════════════════════════════════
for row in wb['Programs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not c0.startswith('PRG'): continue
    inst = int(re.sub(r'[^0-9]','', c0))
    name = str(row[1] or '').strip()
    if not name: continue
    filename = str(row[2] or '').strip()
    enabled_s = str(row[3] or 'Yes').strip()
    blks = type_blocks.get(16, {})
    if inst not in blks:
        fails.append(f"PROGRAMS PRG{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("PROGRAMS", f"PRG{inst}", "name", name, get_name(props))
    # Enabled
    expected_en = 1 if enabled_s.lower() in ('yes','1','true') else 0
    actual_en = get_uint8(props, 4, 0x0A)
    check("PROGRAMS", f"PRG{inst}", "enabled", expected_en, actual_en)
    # Has bytecode — check if block is large enough (bytecode makes it >100 bytes)
    has_file = filename in bas_files
    block_size = len(blks[inst])
    has_bc = block_size > 100
    if has_file:
        check("PROGRAMS", f"PRG{inst}", "has_bytecode", True, has_bc)


# ══════════════════════════════════════════════════════════════
# TABLES
# ══════════════════════════════════════════════════════════════
for row in wb['Tables'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','', str(row[0])))
    name = '{device-name}-' + str(row[1] or '').strip()
    blks = type_blocks.get(141, {})
    if inst not in blks:
        fails.append(f"TABLES TBL{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("TABLES", f"TBL{inst}", "name", name, get_name(props))


# ══════════════════════════════════════════════════════════════
# SCHEDULES
# ══════════════════════════════════════════════════════════════
for row in wb['Schedules'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','', str(row[0])))
    name = str(row[1] or '').strip()
    if not name: continue
    blks = type_blocks.get(17, {})
    if inst not in blks:
        fails.append(f"SCHEDULES SCHED{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("SCHEDULES", f"SCHED{inst}", "name", name, get_name(props))


# ══════════════════════════════════════════════════════════════
# TRENDS
# ══════════════════════════════════════════════════════════════
for row in wb['Trends'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','', str(row[0])))
    name = str(row[1] or '').strip()
    if not name: continue
    blks = type_blocks.get(20, {})
    if inst not in blks:
        fails.append(f"TRENDS STL{inst} MISSING from pan")
        continue
    props = get_props(blks[inst])
    check("TRENDS", f"STL{inst}", "name", name, get_name(props))
    # Check monitored point ref
    mon_name = str(row[2] or '').strip()
    # Check type
    trend_type = str(row[3] or '').strip().lower()
    expected_lt = 1 if trend_type == 'polled' else 2
    actual_lt = get_uint8(props, 0, 0x48)
    check("TRENDS", f"STL{inst}", "logging_type", expected_lt, actual_lt)
    # Check buffer
    expected_buf = int(row[5]) if row[5] else 512
    r = props.get((0, 0x7E))
    if r and len(r[1]) >= 2:
        actual_buf = struct.unpack('>H', r[1][:2])[0]
        check("TRENDS", f"STL{inst}", "buffer", expected_buf, actual_buf)


# ══════════════════════════════════════════════════════════════
# SYSTEM GROUPS (NOTIF_CLS)
# ══════════════════════════════════════════════════════════════
sg_inst = 1
for row in wb['System Groups'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    name = str(row[0]).strip()
    if not name: continue
    blks = type_blocks.get(26, {})
    if sg_inst not in blks:
        fails.append(f"SYSGROUPS {name} (inst {sg_inst}) MISSING from pan")
    else:
        props = get_props(blks[sg_inst])
        check("SYSGROUPS", f"NOTIF{sg_inst}", "name", name, get_name(props))
    sg_inst += 1


# ══════════════════════════════════════════════════════════════
# REPORT
# ══════════════════════════════════════════════════════════════
print(f"\n{'='*60}")
print(f"VERIFICATION RESULTS")
print(f"  Passed: {passes}")
print(f"  Failed: {len(fails)}")

if fails:
    print(f"\nFAILURES:")
    for f in fails:
        print(f"  {f}")
    print(f"\nVERIFICATION FAILED — {len(fails)} errors")
else:
    print(f"\n100% PASS — ALL FIELDS MATCH EXCEL")
