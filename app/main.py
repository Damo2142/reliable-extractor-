"""
Reliable Controls Extraction Tool
FastAPI backend — serves the web UI and REST API for DFA integration
"""

import asyncio
import json
import mimetypes
import os
import struct
import uuid
from pathlib import Path
from typing import Optional
import logging

from fastapi import FastAPI, BackgroundTasks, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from app.extractor import ExtractionEngine
from app.composer import Composer
from app.config import Config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Reliable Controls Extraction Tool",
    description="BAS programming library extractor and viewer for Reliable Controls .panx/.pan files",
    version="1.1.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

cfg = Config()
engine = ExtractionEngine(cfg)
composer = Composer(cfg)

# In-memory job state
jobs: dict[str, dict] = {}


# ─── Static Files & UI ────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root():
    html_path = Path(__file__).parent.parent / "static" / "index.html"
    return HTMLResponse(content=html_path.read_text())


# ─── Variant Discovery ────────────────────────────────────────────────────────

@app.get("/api/variants")
async def list_variants():
    """List all discovered variants grouped by category."""
    return engine.discover_variants()


@app.get("/api/variants/{category}/{variant_id}")
async def get_variant(category: str, variant_id: str):
    """Get full extracted data for a single variant."""
    data = engine.load_library_entry(category, variant_id)
    if data is None:
        raise HTTPException(404, f"Variant {variant_id} not yet processed or not found")
    return data


# ─── Processing ───────────────────────────────────────────────────────────────

@app.post("/api/process")
async def process_all(background_tasks: BackgroundTasks, category: Optional[str] = Query(None)):
    """Trigger extraction pipeline. Optionally filter by category."""
    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {"status": "queued", "progress": 0, "total": 0, "current": "", "errors": [], "done": []}
    background_tasks.add_task(run_extraction, job_id, category, None)
    return {"job_id": job_id, "message": "Extraction started"}


@app.post("/api/process/selected")
async def process_selected(background_tasks: BackgroundTasks, body: dict = None):
    """Process specific variants. Body: { "variants": ["VAV/VAV-IS10001", "RTU/RTU-ISA11110E", ...] }"""
    if not body or "variants" not in body:
        raise HTTPException(400, "Must provide 'variants' list")
    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {"status": "queued", "progress": 0, "total": 0, "current": "", "errors": [], "done": []}
    background_tasks.add_task(run_extraction, job_id, None, body["variants"])
    return {"job_id": job_id, "message": f"Processing {len(body['variants'])} variants"}


@app.get("/api/process/{job_id}/status")
async def job_status(job_id: str):
    """Poll extraction job status."""
    if job_id not in jobs:
        raise HTTPException(404, "Job not found")
    return jobs[job_id]


@app.get("/api/process/{job_id}/stream")
async def job_stream(job_id: str):
    """SSE stream of job progress events."""
    async def event_gen():
        last_progress = -1
        for _ in range(3600):  # max 1hr
            job = jobs.get(job_id, {})
            if job.get("progress") != last_progress or job.get("status") in ("done", "error"):
                last_progress = job.get("progress", 0)
                yield f"data: {json.dumps(job)}\n\n"
            if job.get("status") in ("done", "error"):
                break
            await asyncio.sleep(1)
        yield "data: {\"status\":\"timeout\"}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream")


# ─── Library ──────────────────────────────────────────────────────────────────

@app.get("/api/library")
async def library_summary():
    """Summary of all processed variants with object counts — main dashboard feed."""
    return engine.library_summary()


@app.get("/api/library/compare")
async def compare_variants(a: str = Query(...), b: str = Query(...)):
    """Compare two variants side-by-side."""
    cat_a, var_a = a.split("/", 1)
    cat_b, var_b = b.split("/", 1)
    data_a = engine.load_library_entry(cat_a, var_a)
    data_b = engine.load_library_entry(cat_b, var_b)
    if data_a is None or data_b is None:
        raise HTTPException(404, "One or both variants not found in library")
    return {"a": data_a, "b": data_b}


@app.get("/api/library/export")
async def export_library():
    """Export entire library as a single JSON bundle (for DFA import)."""
    return engine.full_library_export()


# ─── Settings / Categories ────────────────────────────────────────────────────

@app.get("/api/settings/categories")
async def get_categories():
    """Get all category folder mappings (built-in + custom)."""
    # Separate built-in from custom
    builtin = {
        "RC VAV Programming": "VAV", "RC RTU Programming": "RTU",
        "RC FCU Programming": "FCU", "RC AHU Programming": "AHU",
        "RC G36AHU Programming": "G36AHU", "RC G36VAV Programming": "G36VAV",
        "RC UH Programming": "UH", "RC VVT Programming": "VVT",
        "RC WSHP Programming": "WSHP",
    }
    custom = {}
    if engine.cfg.custom_categories_file.exists():
        try:
            custom = json.loads(engine.cfg.custom_categories_file.read_text())
        except Exception:
            pass
    # List existing folders in upload root
    existing_folders = []
    if engine.cfg.upload_root.exists():
        for d in sorted(engine.cfg.upload_root.iterdir()):
            if d.is_dir():
                existing_folders.append(d.name)
    return {
        "builtin": builtin,
        "custom": custom,
        "all": engine.cfg.CATEGORIES,
        "upload_root": str(engine.cfg.upload_root),
        "existing_folders": existing_folders,
    }


@app.post("/api/settings/categories")
async def add_category(body: dict = None):
    """Add a custom category mapping. Body: { "folder": "My Custom AHU", "key": "CUSTOM_AHU" }"""
    if not body or "folder" not in body or "key" not in body:
        raise HTTPException(400, "Must provide 'folder' and 'key'")
    folder = body["folder"].strip()
    key = body["key"].strip().upper()
    if not folder or not key:
        raise HTTPException(400, "Folder and key cannot be empty")

    # Create the folder if it doesn't exist
    folder_path = engine.cfg.upload_root / folder
    folder_path.mkdir(parents=True, exist_ok=True)

    # Load existing custom categories and add new one
    custom = {}
    if engine.cfg.custom_categories_file.exists():
        try:
            custom = json.loads(engine.cfg.custom_categories_file.read_text())
        except Exception:
            pass
    custom[folder] = key
    engine.cfg.save_custom_categories(custom)

    return {"status": "added", "folder": folder, "key": key, "path": str(folder_path)}


@app.delete("/api/settings/categories/{key}")
async def remove_category(key: str):
    """Remove a custom category mapping by key."""
    custom = {}
    if engine.cfg.custom_categories_file.exists():
        try:
            custom = json.loads(engine.cfg.custom_categories_file.read_text())
        except Exception:
            pass
    # Find and remove by key
    to_remove = [k for k, v in custom.items() if v == key.upper()]
    if not to_remove:
        raise HTTPException(404, f"Custom category '{key}' not found")
    for k in to_remove:
        del custom[k]
        # Also remove from live config
        engine.cfg.CATEGORIES.pop(k, None)
    engine.cfg.custom_categories_file.write_text(json.dumps(custom, indent=2))
    return {"status": "removed", "key": key}


# ─── File Serving ─────────────────────────────────────────────────────────────

@app.get("/api/files/assets/{category}/{variant_id}")
async def list_asset_files(category: str, variant_id: str):
    """List all graphics/asset files for a variant."""
    asset_dir = engine.cfg.assets_root / category / variant_id
    if not asset_dir.exists():
        return {"files": []}
    files = []
    for f in sorted(asset_dir.iterdir()):
        if f.is_file():
            mime = mimetypes.guess_type(f.name)[0] or "application/octet-stream"
            files.append({
                "name": f.name,
                "size": f.stat().st_size,
                "mime": mime,
                "is_image": mime.startswith("image/"),
                "url": f"/api/files/assets/{category}/{variant_id}/{f.name}",
            })
    return {"files": files, "count": len(files)}


@app.get("/api/files/assets/{category}/{variant_id}/download")
async def download_variant_assets(category: str, variant_id: str):
    """Download all assets for a variant as a zip.
    Sources (in priority order):
    1. Variant-specific asset folder (from .panx extraction)
    2. Required images parsed from GRP JSON files in the library entry
    3. Shared asset library as fallback
    """
    import zipfile
    import io
    import re as _re

    asset_dir = engine.cfg.assets_root / category / variant_id
    shared_dir = engine.cfg.assets_root / "_shared"

    buf = io.BytesIO()
    added = set()  # Track by full relative path to allow same filename in different dirs

    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Add variant-specific assets if they exist
        if asset_dir.exists():
            for root, dirs, files in os.walk(asset_dir):
                for fname in files:
                    fpath = Path(root) / fname
                    arcname = str(fpath.relative_to(asset_dir))
                    zf.write(fpath, arcname)
                    added.add(arcname.lower().replace('\\', '/'))

        # 2. Parse GRP JSONs + meta GroupAssets for ALL required files
        lib_entry = engine.load_library_entry(category, variant_id)
        required_files = set()
        required_dirs = set()  # Animation directories to include entirely
        if lib_entry:
            grp_files = lib_entry.get("grp_files", {})
            for grp_name, grp_data in grp_files.items():
                text = json.dumps(grp_data)
                for m in _re.finditer(r'"(?:external_file|gel_filename|image|background_image)"\s*:\s*"([^"]+)"', text):
                    val = m.group(1).replace('\\\\', '/').replace('\\', '/').replace('pic/', '')
                    if val and '.' in val:
                        required_files.add(val)

            # GroupAssets — includes animation subdirectories
            for ga in lib_entry.get("meta", {}).get("GroupAssets", []):
                job_path = ga.get("JobPath", "").replace('\\', '/').replace('pic/', '')
                if job_path and '.' in job_path:
                    required_files.add(job_path)
                    # If it's in an Animation subdir, mark the whole dir
                    if 'Animation' in job_path:
                        parts = job_path.replace('\\', '/').split('/')
                        # Animation/DirName/ — include all files in that dir
                        if len(parts) >= 2:
                            anim_dir = '/'.join(parts[:2])  # e.g. "Animation/Damper-Vert-02-V02"
                            required_dirs.add(anim_dir)

        # 3. Find required files in shared library
        if shared_dir.exists():
            # First: copy entire animation directories
            for anim_dir in required_dirs:
                src_dir = shared_dir / anim_dir
                if src_dir.exists():
                    for root, dirs, files in os.walk(src_dir):
                        for fname in files:
                            fpath = Path(root) / fname
                            arcname = str(fpath.relative_to(shared_dir)).replace('\\', '/')
                            if arcname.lower().replace('\\', '/') not in added:
                                zf.write(fpath, arcname)
                                added.add(arcname.lower().replace('\\', '/'))

            # Then: individual files (images, icons)
            for req in required_files:
                req_norm = req.replace('\\', '/')
                if req_norm.lower() in added:
                    continue
                # Try exact relative path first
                exact = shared_dir / req_norm
                if exact.exists():
                    zf.write(exact, req_norm)
                    added.add(req_norm.lower())
                    continue
                # Fallback: search by filename
                matches = list(shared_dir.rglob(Path(req_norm).name))
                if matches:
                    zf.write(matches[0], req_norm)
                    added.add(req_norm.lower())

    if not added:
        raise HTTPException(404, "No assets found for this variant")

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{variant_id}_assets.zip"'}
    )


@app.post("/api/files/assets/{category}/{variant_id}/upload")
async def upload_variant_assets(category: str, variant_id: str, files: list[UploadFile] = File(...)):
    """Upload asset files (images, animations) to a variant's asset folder.
    Supports individual files or a zip of files.
    """
    import zipfile
    import io

    asset_dir = engine.cfg.assets_root / category / variant_id
    asset_dir.mkdir(parents=True, exist_ok=True)

    uploaded = []
    for f in files:
        data = await f.read()

        # If it's a zip, extract its contents
        if f.filename and f.filename.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        # Strip leading folder if all files share one
                        arcname = info.filename
                        dest = asset_dir / arcname
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        dest.write_bytes(zf.read(info))
                        uploaded.append(arcname)
            except zipfile.BadZipFile:
                raise HTTPException(400, f"{f.filename} is not a valid zip")
        else:
            # Save individual file
            dest = asset_dir / f.filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(data)
            uploaded.append(f.filename)

    # Also update the library JSON's GroupAssets if it exists
    lib_entry = engine.load_library_entry(category, variant_id)
    if lib_entry:
        existing_assets = set()
        for ga in lib_entry.get("meta", {}).get("GroupAssets", []):
            existing_assets.add(ga.get("Asset", ""))

        new_assets = []
        for fname in uploaded:
            # Normalize to forward slashes for cross-platform compatibility
            fname_norm = fname.replace('\\', '/')
            asset_path = f"group_assets/{fname_norm}"
            if asset_path not in existing_assets:
                new_assets.append({
                    "Asset": asset_path,
                    "JobPath": fname_norm,
                })

        if new_assets:
            lib_entry.setdefault("meta", {}).setdefault("GroupAssets", []).extend(new_assets)
            engine._save_library_entry(category, variant_id, lib_entry)

    return {"status": "uploaded", "count": len(uploaded), "files": uploaded}


@app.get("/api/binary/report/{category}/{variant_id}")
async def binary_point_report(category: str, variant_id: str):
    """Generate a complete point report from the .pan binary.

    Includes data that PFG XML export misses: loop bindings, verified present
    values, trend references, unit mappings.
    """
    from app.pan_binary import PanBinary

    # Find the source .panx or .pan
    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name

    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    report = pan.generate_point_report()
    return StreamingResponse(
        __import__('io').BytesIO(report.encode()),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{variant_id}_binary_report.txt"'}
    )


@app.get("/api/binary/data/{category}/{variant_id}")
async def binary_data(category: str, variant_id: str):
    """Get structured binary data as JSON for a variant.

    Returns loops with actual input/setpoint bindings, points with verified
    present values, and trend references — all from the .pan binary.
    """
    from app.pan_binary import PanBinary

    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name

    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    # Get arrays and tables, stripping raw data_region (not JSON-serializable)
    arrays = []
    for arr in pan.get_arrays():
        arrays.append({
            'name': arr['name'],
            'values': arr['values'],
            'value_count': arr['value_count'],
        })
    tables = []
    for tbl in pan.get_tables():
        tables.append({
            'name': tbl['name'],
            'rows': tbl['rows'],
            'values': tbl['values'],
            'value_count': tbl['value_count'],
        })

    return {
        "variant_id": variant_id,
        "object_count": len(pan.objects),
        "loops": pan.get_loops(),
        "trends": pan.get_trends(),
        "points": pan.get_point_details(),
        "arrays": arrays,
        "tables": tables,
        "summary": {cat: len(items) for cat, items in pan.get_all_objects().items()},
    }


@app.post("/api/binary/diff")
async def binary_diff(body: dict = None):
    """Compare two controllers and return differences.

    Body: {
        "first": {"category": "SBS_AHU", "variant_id": "1003"},
        "second": {"category": "SBS_AHU", "variant_id": "PS-AHU-ERW-0100"}
    }
    """
    from app.pan_binary import diff_controllers

    if not body or 'first' not in body or 'second' not in body:
        raise HTTPException(400, "Must provide 'first' and 'second' controller references")

    paths = []
    for key in ['first', 'second']:
        ref = body[key]
        folder = engine._cat_folder(ref['category'])
        cat_dir = engine.cfg.upload_root / folder
        src = next(cat_dir.rglob(f"{ref['variant_id']}.panx"), None) or \
              next(cat_dir.rglob(f"{ref['variant_id']}.pan"), None)
        if not src:
            raise HTTPException(404, f"Controller not found: {ref['variant_id']}")
        paths.append(src)

    try:
        result = diff_controllers(paths[0], paths[1])
        return result
    except Exception as e:
        raise HTTPException(500, f"Diff failed: {e}")


@app.post("/api/binary/copy")
async def binary_copy(body: dict = None):
    """Copy a controller with new device name and ID.

    Body: {
        "source": {"category": "SBS_AHU", "variant_id": "1003"},
        "new_device_id": 900,
        "new_device_name": "AHU-TEST",    // must be same byte length as source
        "values": {"RMT-SP": 72.0}        // optional value overrides
    }
    """
    from app.pan_binary import PanWriter

    if not body or 'source' not in body:
        raise HTTPException(400, "Must provide 'source' controller reference")

    ref = body['source']
    folder = engine._cat_folder(ref['category'])
    cat_dir = engine.cfg.upload_root / folder
    src = next(cat_dir.rglob(f"{ref['variant_id']}.panx"), None) or \
          next(cat_dir.rglob(f"{ref['variant_id']}.pan"), None)
    if not src:
        raise HTTPException(404, f"Controller not found: {ref['variant_id']}")

    try:
        if str(src).endswith('.panx'):
            writer = PanWriter.from_panx(src)
        else:
            writer = PanWriter.from_file(src)

        if 'new_device_id' in body:
            writer.set_device_id(body['new_device_id'])

        if 'new_device_name' in body:
            # Auto-detect source name
            from app.pan_binary import PanBinary as _PB
            parser = _PB(bytes(writer.data))
            # Find the most common multi-segment prefix (e.g., "MYS-AHU3")
            # by looking at object names that have at least 2 dashes
            prefixes = {}
            for obj in parser.objects:
                name = obj['name'].strip().strip('\x00')
                parts = name.split('-')
                if len(parts) >= 3:
                    # Try 2-segment prefix: "MYS-AHU3" from "MYS-AHU3-DAT"
                    prefix = f"{parts[0]}-{parts[1]}"
                    if prefix and len(prefix) > 2:
                        prefixes[prefix] = prefixes.get(prefix, 0) + 1
            source_name = max(prefixes, key=prefixes.get) if prefixes else ""
            new_name = body['new_device_name']

            if source_name and len(source_name) == len(new_name):
                count = writer.rename_device(source_name, new_name)
            elif source_name:
                return JSONResponse(
                    status_code=400,
                    content={"detail": f"Name length mismatch: source='{source_name}' ({len(source_name)} chars), new='{new_name}' ({len(new_name)} chars). Must be same length."}
                )

        out_path = Path(f"/tmp/binary_copy_{body.get('new_device_id', 'output')}.pan")
        writer.save(out_path)

        return FileResponse(
            out_path,
            filename=f"{body.get('new_device_name', 'controller')}.pan",
            media_type="application/octet-stream",
        )
    except Exception as e:
        raise HTTPException(500, f"Copy failed: {e}")


# ─── #15: Binary Create (bypass PFG) ─────────────────────────────────────────

@app.post("/api/binary/create")
async def binary_create(body: dict = None):
    """Create a .pan file from a source template WITHOUT using PFG/Wine.

    Takes a source .pan/.panx (by category/variant), applies device_id,
    device_name, and optional value overrides. Returns a modified .pan file.

    Body: {
        "source": {"category": "SBS_AHU", "variant_id": "1003"},
        "device_id": 900,
        "device_name": "AHU-01",
        "values": {"RMT-SP": 72.0, "OCC-CMD": 1.0}
    }
    """
    from app.pan_binary import PanWriter, PanBinary as _PB

    if not body or 'source' not in body:
        raise HTTPException(400, "Must provide 'source' with category and variant_id")

    ref = body['source']
    folder = engine._cat_folder(ref['category'])
    cat_dir = engine.cfg.upload_root / folder
    src = next(cat_dir.rglob(f"{ref['variant_id']}.panx"), None) or \
          next(cat_dir.rglob(f"{ref['variant_id']}.pan"), None)
    if not src:
        raise HTTPException(404, f"Source not found: {ref['variant_id']}")

    try:
        if str(src).endswith('.panx'):
            writer = PanWriter.from_panx(src)
        else:
            writer = PanWriter.from_file(src)

        # Set device ID
        new_id = body.get('device_id')
        if new_id is not None:
            writer.set_device_id(int(new_id))

        # Rename device
        new_name = body.get('device_name')
        rename_count = 0
        source_name = ""
        if new_name:
            parser = _PB(bytes(writer.data))
            prefixes = {}
            for obj in parser.objects:
                name = obj['name'].strip().strip('\x00')
                parts = name.split('-')
                if len(parts) >= 3:
                    prefix = f"{parts[0]}-{parts[1]}"
                    if prefix and len(prefix) > 2:
                        prefixes[prefix] = prefixes.get(prefix, 0) + 1
            source_name = max(prefixes, key=prefixes.get) if prefixes else ""

            if source_name:
                if len(source_name) == len(new_name):
                    rename_count = writer.rename_device(source_name, new_name)
                else:
                    return JSONResponse(
                        status_code=400,
                        content={
                            "detail": f"Name length mismatch: source='{source_name}' "
                                      f"({len(source_name)} chars), new='{new_name}' "
                                      f"({len(new_name)} chars). Must be same byte length.",
                            "source_name": source_name,
                        }
                    )

        # Set present values
        values_set = {}
        for point_name, value in body.get('values', {}).items():
            ok = writer.set_present_value(point_name, float(value))
            values_set[point_name] = "set" if ok else "not found"

        out_path = Path(f"/tmp/binary_create_{body.get('device_id', 'output')}.pan")
        writer.save(out_path)

        filename = f"{new_name or ref['variant_id']}.pan"
        return FileResponse(
            out_path,
            filename=filename,
            media_type="application/octet-stream",
            headers={
                "X-Source-Name": source_name,
                "X-Rename-Count": str(rename_count),
                "X-Values-Status": json.dumps(values_set),
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Binary create failed: {e}")


# ─── #17: Bulk Programming (improved copy) ───────────────────────────────────

@app.post("/api/binary/bulk")
async def binary_bulk(body: dict = None):
    """Generate multiple .pan files from a single source template.

    Returns a zip containing all generated .pan files plus a summary CSV.

    Body: {
        "source": {"category": "SBS_VAV", "variant_id": "2001"},
        "devices": [
            {"device_id": 901, "device_name": "VAV-01", "values": {"RMT-SP": 72.0}},
            {"device_id": 902, "device_name": "VAV-02", "values": {"RMT-SP": 74.0}},
            {"device_id": 903, "device_name": "VAV-03", "values": {}}
        ]
    }
    """
    import zipfile
    import csv
    import io
    from app.pan_binary import PanWriter, PanBinary as _PB

    if not body or 'source' not in body or 'devices' not in body:
        raise HTTPException(400, "Must provide 'source' and 'devices' list")

    devices = body['devices']
    if not devices:
        raise HTTPException(400, "devices list is empty")

    ref = body['source']
    folder = engine._cat_folder(ref['category'])
    cat_dir = engine.cfg.upload_root / folder
    src = next(cat_dir.rglob(f"{ref['variant_id']}.panx"), None) or \
          next(cat_dir.rglob(f"{ref['variant_id']}.pan"), None)
    if not src:
        raise HTTPException(404, f"Source not found: {ref['variant_id']}")

    # Read source binary once
    if str(src).endswith('.panx'):
        import zipfile as _zf
        with _zf.ZipFile(src) as z:
            pan_name = [n for n in z.namelist() if n.endswith('.pan')][0]
            source_data = z.read(pan_name)
    else:
        source_data = src.read_bytes()

    # Auto-detect source device name
    parser = _PB(source_data)
    prefixes = {}
    for obj in parser.objects:
        name = obj['name'].strip().strip('\x00')
        parts = name.split('-')
        if len(parts) >= 3:
            prefix = f"{parts[0]}-{parts[1]}"
            if prefix and len(prefix) > 2:
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
    source_name = max(prefixes, key=prefixes.get) if prefixes else ""

    # Build output zip
    zip_buffer = io.BytesIO()
    csv_buffer = io.StringIO()
    csv_writer = csv.writer(csv_buffer)
    csv_writer.writerow(['device_id', 'device_name', 'status', 'rename_count', 'values_set', 'errors'])

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dev in devices:
            dev_id = dev.get('device_id', 0)
            dev_name = dev.get('device_name', f'device_{dev_id}')
            errors = []

            try:
                writer = PanWriter(bytes(source_data))

                # Set device ID
                if dev_id:
                    writer.set_device_id(int(dev_id))

                # Rename
                rename_count = 0
                if dev_name and source_name:
                    if len(dev_name) == len(source_name):
                        rename_count = writer.rename_device(source_name, dev_name)
                    else:
                        errors.append(
                            f"Name length mismatch: '{source_name}'({len(source_name)}) "
                            f"vs '{dev_name}'({len(dev_name)})"
                        )

                # Set values
                values_ok = 0
                for pt, val in dev.get('values', {}).items():
                    if writer.set_present_value(pt, float(val)):
                        values_ok += 1
                    else:
                        errors.append(f"Point not found: {pt}")

                pan_bytes = bytes(writer.data)
                filename = f"{dev_name}.pan"
                zf.writestr(filename, pan_bytes)

                status = "ok" if not errors else "partial"
                csv_writer.writerow([dev_id, dev_name, status, rename_count,
                                     values_ok, '; '.join(errors) if errors else ''])
            except Exception as e:
                csv_writer.writerow([dev_id, dev_name, 'error', 0, 0, str(e)])

        # Add summary CSV
        zf.writestr('summary.csv', csv_buffer.getvalue())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="bulk_{ref["variant_id"]}_{len(devices)}devices.zip"'
        }
    )


# ─── #24: Change Tracking / Snapshots ─────────────────────────────────────────

SNAPSHOT_ROOT = Path("/srv/dfa/shared/files/vendors/reliable/library-dev/.snapshots")


@app.post("/api/binary/snapshot")
async def binary_snapshot(body: dict = None):
    """Save a named snapshot of a variant's current binary state.

    Body: {
        "category": "VAV",
        "variant_id": "VAV-IS10001",
        "notes": "Before tuning loop gains",
        "created_by": "dave"
    }
    """
    import hashlib
    from app.pan_binary import PanBinary

    if not body or 'category' not in body or 'variant_id' not in body:
        raise HTTPException(400, "Must provide 'category' and 'variant_id'")

    category = body['category']
    variant_id = body['variant_id']

    # Find the source .pan/.panx
    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name
    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    # Compute file hash
    file_hash = hashlib.md5(bytes(pan.data)).hexdigest()

    # Gather metrics
    all_objs = pan.get_all_objects()
    object_count = len(pan.objects)
    loop_count = len(all_objs.get('LOOP', []))
    point_count = sum(len(v) for k, v in all_objs.items()
                      if k in ('AI', 'AO', 'AV', 'BI', 'BO', 'BV', 'MO', 'MV'))

    # Loop summary
    loops = pan.get_loops()
    loop_summary = []
    for lp in loops:
        loop_summary.append({
            "name": lp.get('name', ''),
            "input": lp.get('input_ref', ''),
            "setpoint": lp.get('setpoint_ref', '') or str(lp.get('setpoint_value', '')),
            "output": lp.get('output_ref', ''),
        })

    # Build snapshot
    from datetime import datetime
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    snapshot_id = f"{variant_id}_{ts}"

    snapshot = {
        "snapshot_id": snapshot_id,
        "variant_id": variant_id,
        "category": category,
        "file_hash": file_hash,
        "object_count": object_count,
        "loop_count": loop_count,
        "point_count": point_count,
        "loop_summary": loop_summary,
        "created_at": datetime.now().isoformat(),
        "created_by": body.get('created_by', ''),
        "notes": body.get('notes', ''),
    }

    # Save snapshot
    snap_dir = SNAPSHOT_ROOT / category / variant_id
    snap_dir.mkdir(parents=True, exist_ok=True)
    snap_file = snap_dir / f"{ts}.json"
    snap_file.write_text(json.dumps(snapshot, indent=2))

    return snapshot


@app.get("/api/binary/history/{category}/{variant_id}")
async def binary_history(category: str, variant_id: str):
    """List all snapshots for a variant, newest first."""
    snap_dir = SNAPSHOT_ROOT / category / variant_id
    if not snap_dir.exists():
        return {"variant_id": variant_id, "category": category, "snapshots": []}

    snapshots = []
    for f in sorted(snap_dir.glob("*.json"), reverse=True):
        try:
            data = json.loads(f.read_text())
            snapshots.append(data)
        except Exception:
            continue

    return {
        "variant_id": variant_id,
        "category": category,
        "snapshot_count": len(snapshots),
        "snapshots": snapshots,
    }


@app.get("/api/binary/history/compare/{snapshot1}/{snapshot2}")
async def binary_history_compare(snapshot1: str, snapshot2: str):
    """Compare two snapshots by their snapshot_id.

    snapshot_id format: {variant_id}_{YYYYMMDD-HHMMSS}
    Searches all snapshot directories to find the matching files.
    """
    def find_snapshot(snap_id: str) -> dict | None:
        for cat_dir in SNAPSHOT_ROOT.iterdir():
            if not cat_dir.is_dir():
                continue
            for var_dir in cat_dir.iterdir():
                if not var_dir.is_dir():
                    continue
                for f in var_dir.glob("*.json"):
                    try:
                        data = json.loads(f.read_text())
                        if data.get('snapshot_id') == snap_id:
                            return data
                    except Exception:
                        continue
        return None

    s1 = find_snapshot(snapshot1)
    s2 = find_snapshot(snapshot2)

    if not s1:
        raise HTTPException(404, f"Snapshot not found: {snapshot1}")
    if not s2:
        raise HTTPException(404, f"Snapshot not found: {snapshot2}")

    # Compare metrics
    comparison = {
        "snapshot1": s1,
        "snapshot2": s2,
        "changes": {
            "file_hash_changed": s1['file_hash'] != s2['file_hash'],
            "object_count_delta": s2['object_count'] - s1['object_count'],
            "loop_count_delta": s2['loop_count'] - s1['loop_count'],
            "point_count_delta": s2['point_count'] - s1['point_count'],
        },
        "loop_changes": [],
    }

    # Compare loop summaries
    loops1 = {lp['name']: lp for lp in s1.get('loop_summary', [])}
    loops2 = {lp['name']: lp for lp in s2.get('loop_summary', [])}

    all_loop_names = set(loops1.keys()) | set(loops2.keys())
    for name in sorted(all_loop_names):
        l1 = loops1.get(name)
        l2 = loops2.get(name)
        if l1 is None:
            comparison['loop_changes'].append({"name": name, "change": "added"})
        elif l2 is None:
            comparison['loop_changes'].append({"name": name, "change": "removed"})
        elif l1 != l2:
            comparison['loop_changes'].append({
                "name": name, "change": "modified",
                "before": l1, "after": l2,
            })

    # If both snapshots are from the same variant and we have the .pan files,
    # try a full binary diff
    if (s1['category'] == s2['category'] and
        s1['variant_id'] == s2['variant_id'] and
        s1['file_hash'] != s2['file_hash']):
        comparison['changes']['note'] = (
            "Binary content differs. Use the Controller Diff tool for full object-level comparison."
        )

    return comparison


# ─── #22: Field Diagnostic ───────────────────────────────────────────────────

@app.post("/api/binary/diagnose")
async def binary_diagnose(file: UploadFile = File(...)):
    """Upload a .pan/.panx file and receive a full diagnostic report.

    Returns device info, object summary, loop configs, point details,
    trend configs, schedule summaries, and warnings.
    """
    from app.pan_binary import PanBinary

    content = await file.read()
    filename = file.filename or "upload.pan"

    try:
        if filename.endswith('.panx'):
            import zipfile as _zf
            import io as _io
            with _zf.ZipFile(_io.BytesIO(content)) as z:
                pan_name = [n for n in z.namelist() if n.endswith('.pan')][0]
                pan_data = z.read(pan_name)
        else:
            pan_data = content

        pan = PanBinary(pan_data)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    # Device info
    device_id = struct.unpack('<I', pan.data[4:8])[0] if len(pan.data) >= 8 else None

    # Find device name from most common prefix
    prefixes = {}
    for obj in pan.objects:
        name = obj['name'].strip().strip('\x00')
        parts = name.split('-')
        if len(parts) >= 3:
            prefix = f"{parts[0]}-{parts[1]}"
            if prefix and len(prefix) > 2:
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
    device_name = max(prefixes, key=prefixes.get) if prefixes else "unknown"

    # Object summary
    summary = {cat: len(items) for cat, items in pan.get_all_objects().items()}

    # Points with details
    points = pan.get_point_details()

    # Loops
    loops = pan.get_loops()

    # Trends
    trends = pan.get_trends()

    # Schedules
    schedules = pan.get_schedules()

    # Programs
    programs = pan.get_programs()

    # --- Generate warnings ---
    warnings = []

    # Out-of-service points
    oos_points = [p['name'] for p in points if p.get('out_of_service')]
    if oos_points:
        warnings.append({
            "type": "out_of_service",
            "severity": "warning",
            "message": f"{len(oos_points)} point(s) are out-of-service",
            "details": oos_points,
        })

    # Points with unusual present values (analog points at exactly 0 or very large)
    suspect_values = []
    for p in points:
        pv = p.get('present_value')
        if pv is not None:
            if abs(pv) > 10000:
                suspect_values.append({"name": p['name'], "value": pv, "reason": "unusually large"})
    if suspect_values:
        warnings.append({
            "type": "unusual_values",
            "severity": "info",
            "message": f"{len(suspect_values)} point(s) have unusually large values",
            "details": suspect_values,
        })

    # Loops missing bindings
    incomplete_loops = []
    for loop in loops:
        missing = []
        if not loop.get('input_ref'):
            missing.append('input')
        if not loop.get('output_ref'):
            missing.append('output')
        if not loop.get('setpoint_ref') and loop.get('setpoint_value') is None:
            missing.append('setpoint')
        if missing:
            incomplete_loops.append({"name": loop['name'], "missing": missing})
    if incomplete_loops:
        warnings.append({
            "type": "incomplete_loops",
            "severity": "warning",
            "message": f"{len(incomplete_loops)} loop(s) have missing bindings",
            "details": incomplete_loops,
        })

    # Empty trends (no point references)
    empty_trends = [t['name'] for t in trends if not t.get('refs')]
    if empty_trends:
        warnings.append({
            "type": "empty_trends",
            "severity": "info",
            "message": f"{len(empty_trends)} trend(s) have no point references",
            "details": empty_trends,
        })

    # Disabled programs
    disabled_progs = [p['name'] for p in programs if p.get('enabled') is False]
    if disabled_progs:
        warnings.append({
            "type": "disabled_programs",
            "severity": "info",
            "message": f"{len(disabled_progs)} program(s) are disabled",
            "details": disabled_progs,
        })

    return {
        "filename": filename,
        "device": {
            "id": device_id,
            "name": device_name,
            "object_count": len(pan.objects),
            "file_size": len(pan_data),
        },
        "object_summary": summary,
        "loops": loops,
        "points": points,
        "trends": trends,
        "schedules": schedules,
        "programs": programs,
        "warnings": warnings,
        "warning_count": len(warnings),
    }


# ─── #19: Schedule Management ────────────────────────────────────────────────

@app.get("/api/binary/schedules/{category}/{variant_id}")
async def binary_schedules(category: str, variant_id: str):
    """Get all schedule objects with decoded weekly schedules from the .pan binary.

    Returns schedule names, default values, and weekly schedule byte pairs.
    """
    from app.pan_binary import PanBinary

    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name

    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    schedules = pan.get_schedules()
    return {
        "variant_id": variant_id,
        "category": category,
        "schedule_count": len(schedules),
        "schedules": schedules,
    }


# ─── #20: Trend Info ─────────────────────────────────────────────────────────

@app.get("/api/binary/trends/{category}/{variant_id}")
async def binary_trends(category: str, variant_id: str):
    """Get all trend objects with point references and log intervals from .pan binary.

    Returns trend names, referenced points, and configuration details.
    """
    from app.pan_binary import PanBinary

    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name

    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    trends = pan.get_trends()

    # Enrich with additional data from the raw objects
    trend_details = []
    for obj in pan.objects:
        if obj['category'] != 'TREND':
            continue

        props = obj.get('properties', {})
        data = obj['data_region']

        detail = {
            'name': obj['name'],
            'refs': [r['ref'] for r in obj['refs']
                     if r['type'] != 'NULL' and r['instance'] < 300],
            'present_value': obj.get('present_value'),
        }

        # Try to extract log interval (property 0x6c = 108 for trends stores interval)
        for i in range(min(len(data) - 6, 200)):
            if data[i] == 0x6c and data[i+1] == 0x44:
                import struct as _st
                val = _st.unpack('>f', data[i+2:i+6])[0]
                if 0 < val < 86400:  # Reasonable interval (up to 24h in seconds)
                    detail['log_interval_seconds'] = round(val, 1)
                    break

        # Check enabled state
        for tag, val in props.get(0x0A, []):
            if tag == 'uint8':
                detail['enabled'] = val == 1

        trend_details.append(detail)

    return {
        "variant_id": variant_id,
        "category": category,
        "trend_count": len(trend_details),
        "trends": trend_details,
    }


@app.get("/api/files/assets/shared")
async def list_shared_assets():
    """List all files in the shared asset library."""
    shared_dir = engine.cfg.assets_root / "_shared"
    if not shared_dir.exists():
        return {"files": [], "count": 0}
    files = []
    for root, dirs, fnames in os.walk(shared_dir):
        for fname in sorted(fnames):
            fpath = Path(root) / fname
            rel = str(fpath.relative_to(shared_dir))
            mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
            files.append({
                "name": fname,
                "path": rel,
                "size": fpath.stat().st_size,
                "mime": mime,
                "is_image": mime.startswith("image/"),
            })
    return {"files": files, "count": len(files)}


@app.post("/api/files/assets/shared/upload")
async def upload_shared_assets(files: list[UploadFile] = File(...)):
    """Upload files to the shared asset library. Supports files or zips."""
    import zipfile
    import io

    shared_dir = engine.cfg.assets_root / "_shared"
    shared_dir.mkdir(parents=True, exist_ok=True)
    uploaded = []

    for f in files:
        data = await f.read()
        if f.filename and f.filename.lower().endswith('.zip'):
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        dest = shared_dir / info.filename
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        dest.write_bytes(zf.read(info))
                        uploaded.append(info.filename)
            except zipfile.BadZipFile:
                raise HTTPException(400, f"{f.filename} is not a valid zip")
        else:
            dest = shared_dir / f.filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(data)
            uploaded.append(f.filename)

    return {"status": "uploaded", "count": len(uploaded), "files": uploaded}


@app.get("/api/files/assets/shared/download")
async def download_shared_assets():
    """Download the entire shared asset library as a zip."""
    import zipfile
    import io

    shared_dir = engine.cfg.assets_root / "_shared"
    if not shared_dir.exists():
        raise HTTPException(404, "No shared assets found")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(shared_dir):
            for fname in files:
                fpath = Path(root) / fname
                arcname = str(fpath.relative_to(shared_dir))
                zf.write(fpath, arcname)

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="shared_assets.zip"'}
    )


@app.get("/api/files/assets/{category}/{variant_id}/{filename}")
async def serve_asset_file(category: str, variant_id: str, filename: str):
    """Serve a single asset file (image, etc.)."""
    file_path = engine.cfg.assets_root / category / variant_id / filename
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(404, "File not found")
    return FileResponse(file_path)


@app.get("/api/files/source/{category}/{variant_id}")
async def list_source_files(category: str, variant_id: str):
    """List original source files (.pan, .bas, .pdf, etc.) for a variant."""
    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name
    if not cat_dir.exists():
        return {"files": []}

    files = []
    # Find the variant's source directory
    for src in cat_dir.rglob(f"{variant_id}.*"):
        if src.is_file():
            mime = mimetypes.guess_type(src.name)[0] or "application/octet-stream"
            files.append({
                "name": src.name,
                "size": src.stat().st_size,
                "mime": mime,
                "is_image": mime.startswith("image/"),
                "url": f"/api/files/source/{category}/{variant_id}/{src.name}",
            })

    # Also find .bas files and other files in the same directory
    for pattern in [f"{variant_id}*.bas", f"{variant_id}*.pdf", f"{variant_id}*.txt", f"{variant_id}*.doc*"]:
        for src in cat_dir.rglob(pattern):
            if src.is_file() and not any(f["name"] == src.name for f in files):
                mime = mimetypes.guess_type(src.name)[0] or "application/octet-stream"
                files.append({
                    "name": src.name,
                    "size": src.stat().st_size,
                    "mime": mime,
                    "url": f"/api/files/source/{category}/{variant_id}/{src.name}",
                })

    # Check the variant's parent folder for any extra files
    for panx in cat_dir.rglob(f"{variant_id}.panx"):
        parent = panx.parent
        for f in parent.iterdir():
            if f.is_file() and not any(ef["name"] == f.name for ef in files):
                mime = mimetypes.guess_type(f.name)[0] or "application/octet-stream"
                files.append({
                    "name": f.name,
                    "size": f.stat().st_size,
                    "mime": mime,
                    "url": f"/api/files/source/{category}/{variant_id}/{f.name}",
                })

    for pan in cat_dir.rglob(f"{variant_id}.pan"):
        parent = pan.parent
        for f in parent.iterdir():
            if f.is_file() and not any(ef["name"] == f.name for ef in files):
                mime = mimetypes.guess_type(f.name)[0] or "application/octet-stream"
                files.append({
                    "name": f.name,
                    "size": f.stat().st_size,
                    "mime": mime,
                    "url": f"/api/files/source/{category}/{variant_id}/{f.name}",
                })

    return {"files": files, "count": len(files)}


@app.get("/api/files/source/{category}/{variant_id}/{filename}")
async def serve_source_file(category: str, variant_id: str, filename: str):
    """Serve a source file."""
    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name
    # Search for the file
    for f in cat_dir.rglob(filename):
        if f.is_file():
            return FileResponse(f, filename=filename)
    raise HTTPException(404, "File not found")


@app.post("/api/library/{category}/{variant_id}/save")
async def save_variant(category: str, variant_id: str, body: dict = None):
    """Save edits to a library entry (point names, code, trend toggles, etc.)."""
    if not body:
        raise HTTPException(400, "No data to save")
    existing = engine.load_library_entry(category, variant_id)
    if existing is None:
        raise HTTPException(404, "Variant not found")

    # Merge edits into existing record
    if "objects" in body:
        existing["objects"] = body["objects"]
        existing["counts"] = {k: len(v) for k, v in existing["objects"].items() if isinstance(v, list) and v}
    if "meta" in body:
        existing["meta"].update(body["meta"])
    if "description" in body:
        existing["description"] = body["description"]
        # Also update master_descriptions.json so it persists across re-extractions
        _update_master_description(variant_id, body["description"])

    engine._save_library_entry(category, variant_id, existing)
    return {"status": "saved", "id": variant_id}


def _update_master_description(variant_id: str, description: str):
    """Update a single variant description in master_descriptions.json."""
    descs = {}
    if cfg.master_descriptions.exists():
        try:
            descs = json.loads(cfg.master_descriptions.read_text())
        except Exception:
            pass
    descs[variant_id] = description
    cfg.master_descriptions.write_text(json.dumps(descs, indent=2))
    # Also update the extractor's in-memory cache
    engine.descriptions[variant_id] = description


@app.get("/api/library/descriptions")
async def get_all_descriptions():
    """Get all variant descriptions from master_descriptions.json."""
    if cfg.master_descriptions.exists():
        try:
            return json.loads(cfg.master_descriptions.read_text())
        except Exception:
            return {}
    return {}


@app.put("/api/library/{category}/{variant_id}/description")
async def set_variant_description(category: str, variant_id: str, body: dict = None):
    """Set the description for a single variant.
    Body: { "description": "My custom controller description" }
    """
    if not body or "description" not in body:
        raise HTTPException(400, "Must provide 'description' in body")

    description = body["description"]

    # Update master_descriptions.json
    _update_master_description(variant_id, description)

    # Update the library JSON if it exists
    existing = engine.load_library_entry(category, variant_id)
    if existing:
        existing["description"] = description
        engine._save_library_entry(category, variant_id, existing)

    return {"status": "updated", "id": variant_id, "description": description}


@app.put("/api/library/descriptions")
async def set_bulk_descriptions(body: dict = None):
    """Set descriptions for multiple variants at once.
    Body: { "VAV-IS10001": "Single Duct, Floating HW Reheat", "MY-CUSTOM": "My desc" }
    """
    if not body:
        raise HTTPException(400, "Must provide variant_id: description map")

    descs = {}
    if cfg.master_descriptions.exists():
        try:
            descs = json.loads(cfg.master_descriptions.read_text())
        except Exception:
            pass

    descs.update(body)
    cfg.master_descriptions.write_text(json.dumps(descs, indent=2))

    # Update extractor in-memory cache
    engine.descriptions.update(body)

    # Update any existing library JSONs
    updated = []
    for vid, desc in body.items():
        for cat_dir in cfg.library_root.iterdir():
            if not cat_dir.is_dir() or cat_dir.name.startswith("_"):
                continue
            entry_path = cat_dir / f"{vid}.json"
            if entry_path.exists():
                try:
                    entry = json.loads(entry_path.read_text())
                    entry["description"] = desc
                    entry_path.write_text(json.dumps(entry, indent=2))
                    updated.append(vid)
                except Exception:
                    pass

    return {"status": "updated", "count": len(body), "library_updated": updated}


# ─── Composer ────────────────────────────────────────────────────────────────

@app.get("/api/composer/programs")
async def composer_program_index():
    """Get flat index of every program across the entire library with dependencies."""
    return composer.build_program_index()


@app.post("/api/composer/compose")
async def composer_compose(body: dict = None):
    """Compose a new controller from selected programs.

    Body:
    {
        "selections": [
            {"category": "VAV", "variant_id": "VAV-IS10001", "program_instance": "1"},
            {"category": "RTU", "variant_id": "RTU-ISA11110E", "program_instance": "4"},
            ...
        ],
        "device_name": "{device-name}",  (optional, default "{device-name}")
        "device_id": "900"               (optional, default "900")
    }
    """
    if not body or "selections" not in body:
        raise HTTPException(400, "Must provide 'selections' list")

    selections = body["selections"]
    if not selections:
        raise HTTPException(400, "Must select at least one program")

    device_name = body.get("device_name", "{device-name}")
    device_id = body.get("device_id", "900")
    primary_variant = body.get("primary_variant", None)

    try:
        result = composer.compose(selections, device_name, device_id, primary_variant=primary_variant)
        # Cache for Excel export
        composer._last_composition = result
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.exception("Composition failed")
        raise HTTPException(500, f"Composition failed: {e}")


@app.post("/api/composer/save")
async def composer_save(body: dict = None):
    """Save a composed controller to the library.

    Body:
    {
        "name": "My-Custom-VAV",
        "composition": { ... composed controller JSON ... }
    }
    """
    if not body or "name" not in body or "composition" not in body:
        raise HTTPException(400, "Must provide 'name' and 'composition'")

    try:
        path = composer.save_composition(body["name"], body["composition"])
        return {"status": "saved", "path": str(path), "id": body["name"]}
    except Exception as e:
        raise HTTPException(500, f"Save failed: {e}")


@app.get("/api/composer/compositions")
async def composer_list():
    """List all saved compositions."""
    return composer.list_compositions()


@app.get("/api/composer/compositions/{name}")
async def composer_get(name: str):
    """Load a saved composition."""
    data = composer.load_composition(name)
    if data is None:
        raise HTTPException(404, f"Composition '{name}' not found")
    return data


@app.delete("/api/composer/compositions/{name}")
async def composer_delete(name: str):
    """Delete a saved composition."""
    if composer.delete_composition(name):
        return {"status": "deleted", "id": name}
    raise HTTPException(404, f"Composition '{name}' not found")


@app.post("/api/composer/generate-xml")
async def composer_generate_xml(body: dict = None):
    """Generate PFG-compatible XML from a composed controller.

    Body: the composed controller JSON (from /api/composer/compose or /api/composer/compositions/{name})
    OR: {"name": "saved-composition-name"} to load from saved
    """
    from generator import generate_xml

    if not body:
        raise HTTPException(400, "Must provide composition data")

    # If a name is provided, load the saved composition
    if "name" in body and "objects" not in body:
        data = composer.load_composition(body["name"])
        if data is None:
            raise HTTPException(404, f"Composition '{body['name']}' not found")
    else:
        data = body

    device_id = data.get("meta", {}).get("device_id", "900")
    device_name = data.get("meta", {}).get("device_name", "{device-name}")

    try:
        xml = generate_xml(data, device_id=device_id, device_name=device_name)
        return JSONResponse(content={
            "xml": xml,
            "device_id": device_id,
            "device_name": device_name,
        })
    except Exception as e:
        logger.exception("XML generation failed")
        raise HTTPException(500, f"XML generation failed: {e}")


@app.get("/api/composer/descriptions")
async def composer_variant_descriptions():
    """Get auto-generated human-readable descriptions for all variant IDs.
    Analyzes program code dependencies (AO vs BO on reheat) to determine
    floating vs modulating control. Non-VAV types use manual overrides."""
    return composer.build_variant_descriptions()


@app.get("/api/composer/variant-metadata")
async def composer_variant_metadata():
    """Get enriched metadata for all variants: friendly labels, function tags,
    compatible equipment types, and descriptions."""
    return composer.build_variant_metadata()


@app.get("/api/composer/blanks")
async def composer_list_blanks():
    """List available blank controller model templates for .pan/.panx generation."""
    return composer.list_blank_panels()


def _resolve_composition(body: dict) -> dict:
    """Helper: resolve body to composition data (load by name if needed)."""
    if "name" in body and "objects" not in body:
        data = composer.load_composition(body["name"])
        if data is None:
            raise HTTPException(404, f"Composition '{body['name']}' not found")
        return data
    return body


@app.post("/api/composer/generate-pan")
async def composer_generate_pan(body: dict = None):
    """Generate a .pan file from a composed controller.

    Body: {
        ...composition data or {"name": "saved-name"},
        "blank_model": "RC-FLEXair-34-A-F"  (optional, selects controller template)
    }
    Returns the .pan file as a download.
    """
    if not body:
        raise HTTPException(400, "Must provide composition data")

    blank_model = body.pop("blank_model", None)
    data = _resolve_composition(body)

    try:
        pan_path = await asyncio.get_event_loop().run_in_executor(
            None, composer.generate_pan, data, blank_model
        )
        comp_id = data.get("id", "composed")
        # Check if values document was generated alongside
        values_path = pan_path.parent / f"{comp_id}_values.txt"
        headers = {}
        if values_path.exists():
            headers["X-Values-Document"] = "true"
        return FileResponse(
            pan_path,
            filename=f"{comp_id}.pan",
            media_type="application/octet-stream",
            headers=headers,
        )
    except Exception as e:
        logger.exception("PAN generation failed")
        raise HTTPException(500, f"PAN generation failed: {e}")


@app.post("/api/composer/generate-panx")
async def composer_generate_panx(body: dict = None):
    """Generate a .panx file from a composed controller.

    Body: {
        ...composition data or {"name": "saved-name"},
        "blank_model": "RC-FLEXair-34-A-F"  (optional, selects controller template)
    }
    Returns the .panx file as a download.
    """
    if not body:
        raise HTTPException(400, "Must provide composition data")

    blank_model = body.pop("blank_model", None)
    data = _resolve_composition(body)

    try:
        panx_path = await asyncio.get_event_loop().run_in_executor(
            None, composer.generate_panx, data, blank_model
        )
        comp_id = data.get("id", "composed")
        return FileResponse(
            panx_path,
            filename=f"{comp_id}.panx",
            media_type="application/octet-stream",
        )
    except Exception as e:
        logger.exception("PANX generation failed")
        raise HTTPException(500, f"PANX generation failed: {e}")


def _generate_excel_from_composition(comp_data):
    """Generate the grouped Excel point schedule from a composition — same format as the standalone endpoint."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    objects = comp_data.get("objects", {})
    device_name = comp_data.get("meta", {}).get("device_name", "{device-name}")

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    empty_fill = PatternFill("solid", fgColor="F2F2F2")

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        if not rows:
            return
        all_int = all(isinstance(row[0], int) for row in rows)
        if all_int:
            by_inst = {}
            for row in rows:
                by_inst[row[0]] = row
            max_inst = max(by_inst.keys())
            r = 2
            for inst in range(1, max_inst + 1):
                if inst in by_inst:
                    for c, val in enumerate(by_inst[inst], 1):
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.border = thin_border
                else:
                    cell = ws.cell(row=r, column=1, value=inst)
                    cell.border = thin_border
                    cell.fill = empty_fill
                    for c in range(2, len(headers) + 1):
                        cell = ws.cell(row=r, column=c, value="")
                        cell.border = thin_border
                        cell.fill = empty_fill
                r += 1
        else:
            r = 2
            for row in rows:
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                r += 1
        for c in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row, c).value or '')) for row in range(1, r))
            ws.column_dimensions[ws.cell(1, c).column_letter].width = min(max_len + 3, 50)

    for ptype, sheet_name, extra_cols in [
        ("AV", "Analog Values", ["present_value", "range", "unit", "increment", "description"]),
        ("AI", "Analog Inputs", ["present_value", "range", "unit", "description"]),
        ("AO", "Analog Outputs", ["present_value", "range", "unit", "description"]),
        ("BV", "Binary Values", ["present_value", "description"]),
        ("BI", "Binary Inputs", ["present_value", "description"]),
        ("BO", "Binary Outputs", ["present_value", "description"]),
        ("MV", "Multistate Values", ["present_value", "description"]),
        ("MO", "Multistate Outputs", ["present_value", "description"]),
    ]:
        pts = objects.get(ptype, [])
        if not pts:
            continue
        headers = ["Instance", "Name"] + [c.replace("_", " ").title() for c in extra_cols]
        rows = []
        for p in sorted(pts, key=lambda x: int(x.get("instance", 0))):
            name = p.get("name", "").replace("{device-name}", device_name)
            row = [int(p.get("instance", 0)), name]
            for col in extra_cols:
                row.append(p.get(col, ""))
            rows.append(row)
        if rows:
            add_sheet(sheet_name, headers, rows)

    # Loops sheet — infer input/setpoint/output from point names if not in data
    loops = objects.get("LOOP", [])
    if loops:
        # Build point name lookup for matching
        point_lookup = {}  # mnemonic -> "TYPE:instance (name)"
        for ptype in ['AI', 'AO', 'AV', 'BI', 'BO', 'BV', 'MO', 'MV']:
            for p in objects.get(ptype, []):
                pname = p.get("name", "")
                mnem = pname.replace("{device-name}-", "").replace(device_name + "-", "") if pname else ""
                if mnem:
                    point_lookup[mnem.upper()] = f"{ptype}:{p.get('instance','')} ({pname.replace('{device-name}', device_name)})"

        headers = ["Instance", "Name", "P", "I", "D", "Bias", "Deadband", "Action", "I-Units",
                   "Input (Controlled Variable)", "Setpoint", "Output (Manipulated Variable)"]
        rows = []
        for l in sorted(loops, key=lambda x: int(x.get("instance", 0))):
            name = l.get("name", "").replace("{device-name}", device_name)
            raw_name = l.get("name", "").replace("{device-name}-", "")

            # Try to get refs from data first, then infer from name
            input_ref = l.get("input_ref", l.get("suggested_input", ""))
            setpoint_ref = l.get("setpoint_ref", l.get("suggested_setpoint", ""))
            output_ref = l.get("output_ref", l.get("suggested_output", ""))

            # Infer from loop name if refs are empty
            if not input_ref or not output_ref:
                import re as _re_loop
                # Extract mnemonic from loop name: e.g. "DSP-LOOP1" -> "DSP"
                loop_mnem_match = _re_loop.match(r'(.+?)-?LOOP\d*$', raw_name, _re_loop.I)
                if loop_mnem_match:
                    loop_mnem = loop_mnem_match.group(1).upper()
                    # Search for matching points
                    for pt_mnem, pt_ref in point_lookup.items():
                        if not input_ref and pt_mnem == loop_mnem and pt_ref.startswith('AI:'):
                            input_ref = pt_ref
                        elif not input_ref and loop_mnem in pt_mnem and pt_ref.startswith('AI:'):
                            input_ref = pt_ref
                        elif not setpoint_ref and (loop_mnem + '-SPT') in pt_mnem and pt_ref.startswith('AV:'):
                            setpoint_ref = pt_ref
                        elif not setpoint_ref and loop_mnem in pt_mnem and 'SPT' in pt_mnem and pt_ref.startswith('AV:'):
                            setpoint_ref = pt_ref
                        elif not output_ref and loop_mnem in pt_mnem and pt_ref.startswith('AO:'):
                            output_ref = pt_ref

            action_text = "Direct" if str(l.get("action", "")) == "1" else "Reverse" if str(l.get("action", "")) == "0" else str(l.get("action", ""))

            rows.append([int(l.get("instance", 0)), name,
                        l.get("proportional", ""), l.get("integral", ""), l.get("derivative", ""),
                        l.get("bias", ""), l.get("deadband", ""), action_text,
                        l.get("integralunits", ""),
                        input_ref, setpoint_ref, output_ref])
        add_sheet("Loops", headers, rows)

    # Arrays sheet — each value on its own row for clarity
    arrays = objects.get("ARRAY", [])
    if arrays:
        headers = ["Instance", "Name", "Index", "Value"]
        rows = []
        for a in sorted(arrays, key=lambda x: int(x.get("instance", 0))):
            name = a.get("name", "").replace("{device-name}", device_name)
            vals = a.get("values", [])
            if vals:
                for idx, v in enumerate(vals):
                    rows.append([int(a.get("instance", 0)) if idx == 0 else "", name if idx == 0 else "", idx, v])
            else:
                rows.append([int(a.get("instance", 0)), name, "", ""])
        add_sheet("Arrays", headers, rows)

    # Tables sheet — show in/out pairs
    tables = objects.get("TABLE", [])
    if tables:
        headers = ["Instance", "Name", "Unit", "Input Unit", "Row #", "Input", "Output"]
        rows = []
        for t in sorted(tables, key=lambda x: int(x.get("instance", 0))):
            name = t.get("name", "").replace("{device-name}", device_name)
            tbl_rows = t.get("rows", [])
            if tbl_rows:
                for ri, r in enumerate(tbl_rows):
                    rows.append([
                        int(t.get("instance", 0)) if ri == 0 else "",
                        name if ri == 0 else "",
                        t.get("unit", "") if ri == 0 else "",
                        t.get("inunit", "") if ri == 0 else "",
                        ri + 1,
                        r.get("in", ""),
                        r.get("out", ""),
                    ])
            else:
                rows.append([int(t.get("instance", 0)), name, t.get("unit", ""), t.get("inunit", ""), "", "", ""])
        add_sheet("Tables", headers, rows)

    # Programs sheet
    progs = objects.get("PROGRAM", [])
    if progs:
        headers = ["Instance", "Name", "Lines", "Description"]
        rows = []
        for p in sorted(progs, key=lambda x: int(x.get("instance", 0))):
            name = p.get("name", "").replace("{device-name}", device_name)
            code = p.get("code", "")
            rows.append([int(p.get("instance", 0)), name, len(code.split("\n")), p.get("description", "")])
        add_sheet("Programs", headers, rows)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.post("/api/composer/generate-template-package")
async def composer_generate_template_package(body: dict = None):
    """Generate a complete template package zip containing:
    - .pan file (with {device-name} template names)
    - values reference document
    - assets folder (images + animations from primary variant + shared library)

    Uses the first selected variant as the primary source for assets.
    """
    import zipfile as _zf
    import io
    import re as _re_pkg

    if not body:
        raise HTTPException(400, "Must provide composition data")

    blank_model = body.pop("blank_model", None)
    data = _resolve_composition(body)
    comp_id = data.get("id", "composed")

    try:
        # 1. Generate .pan
        pan_path = await asyncio.get_event_loop().run_in_executor(
            None, composer.generate_pan, data, blank_model
        )

        # 2. Get values doc
        values_path = pan_path.parent / f"{comp_id}_values.txt"

        # 3. Collect assets from primary variant + shared library
        meta = data.get("meta", {})
        graphics_sources = meta.get("graphics_sources", [])

        # Primary variant = first graphics source
        primary_cat = graphics_sources[0].get("from_category", "") if graphics_sources else ""
        primary_var = graphics_sources[0].get("from_variant", "") if graphics_sources else ""

        # Gather required files from GRP JSONs + GroupAssets
        required_files = set()
        required_dirs = set()
        grp_files = data.get("grp_files", {})
        for grp_name, grp_data in grp_files.items():
            text = json.dumps(grp_data)
            for m in _re_pkg.finditer(r'"(?:external_file|gel_filename|image|background_image)"\s*:\s*"([^"]+)"', text):
                val = m.group(1).replace('\\\\', '/').replace('\\', '/').replace('pic/', '')
                if val and '.' in val:
                    required_files.add(val)

        for ga in meta.get("GroupAssets", []):
            job_path = ga.get("JobPath", "").replace('\\', '/').replace('pic/', '')
            if job_path and '.' in job_path:
                required_files.add(job_path)
                if 'Animation' in job_path:
                    parts = job_path.replace('\\', '/').split('/')
                    if len(parts) >= 2:
                        required_dirs.add('/'.join(parts[:2]))

        # Build zip
        buf = io.BytesIO()
        with _zf.ZipFile(buf, 'w', _zf.ZIP_DEFLATED) as zf:
            # Add .pan
            zf.write(pan_path, f"{comp_id}.pan")

            # Add values doc
            if values_path.exists():
                zf.write(values_path, f"{comp_id}_values.txt")

            # Add assets
            added_assets = set()
            asset_sources = []

            # Primary variant assets first
            if primary_cat and primary_var:
                pdir = engine.cfg.assets_root / primary_cat / primary_var
                if pdir.exists():
                    asset_sources.append(pdir)

            # Other variant assets
            for gs in graphics_sources[1:]:
                d = engine.cfg.assets_root / gs.get("from_category", "") / gs.get("from_variant", "")
                if d.exists() and d not in asset_sources:
                    asset_sources.append(d)

            # Shared library last
            shared = engine.cfg.assets_root / "_shared"
            if shared.exists():
                asset_sources.append(shared)

            # Copy from variant asset dirs
            for adir in asset_sources:
                if adir.name == "_shared":
                    continue  # Handle shared separately
                for root, dirs, files in os.walk(adir):
                    for fname in files:
                        fpath = Path(root) / fname
                        arcname = "assets/" + str(fpath.relative_to(adir)).replace('\\', '/')
                        if arcname.lower() not in added_assets:
                            zf.write(fpath, arcname)
                            added_assets.add(arcname.lower())

            # Copy animation dirs from shared
            if shared.exists():
                for anim_dir in required_dirs:
                    src_dir = shared / anim_dir
                    if src_dir.exists():
                        for root, dirs, files in os.walk(src_dir):
                            for fname in files:
                                fpath = Path(root) / fname
                                arcname = "assets/" + str(fpath.relative_to(shared)).replace('\\', '/')
                                if arcname.lower() not in added_assets:
                                    zf.write(fpath, arcname)
                                    added_assets.add(arcname.lower())

                # Individual files from shared
                for req in required_files:
                    req_norm = req.replace('\\', '/')
                    arcname = "assets/" + req_norm
                    if arcname.lower() in added_assets:
                        continue
                    exact = shared / req_norm
                    if exact.exists():
                        zf.write(exact, arcname)
                        added_assets.add(arcname.lower())
                    else:
                        matches = list(shared.rglob(Path(req_norm).name))
                        if matches:
                            zf.write(matches[0], arcname)
                            added_assets.add(arcname.lower())

                # 5. Generate and add Excel point schedule (same format as standalone endpoint)
                try:
                    # Cache composition for the Excel endpoint to use
                    composer._last_composition = data
                    # Call the Excel generation inline using the same logic
                    excel_buf = _generate_excel_from_composition(data)
                    if excel_buf:
                        zf.writestr(f"{comp_id}_point_schedule.xlsx", excel_buf.read())
                except Exception as exc:
                    logger.warning(f"Could not add Excel to template package: {exc}")

        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{comp_id}_template_package.zip"'}
        )
    except Exception as e:
        logger.exception("Template package generation failed")
        raise HTTPException(500, f"Template package generation failed: {e}")


@app.get("/api/composer/values-document/{comp_id}")
async def composer_values_document(comp_id: str):
    """Download the companion values reference document for a generated controller.

    This text file lists all configured values (present values, PID settings,
    ranges, units) from the source library — useful for verifying or manually
    entering values that PFG may not preserve in the .pan binary.
    """
    output_dir = cfg.library_root / "COMPOSED" / "_output"
    values_path = output_dir / f"{comp_id}_values.txt"
    if not values_path.exists():
        raise HTTPException(404, f"Values document not found for '{comp_id}'")
    return FileResponse(
        values_path,
        filename=f"{comp_id}_values.txt",
        media_type="text/plain",
    )


@app.get("/api/composer/values-document/{comp_id}/excel")
async def composer_values_excel(comp_id: str):
    """Download values as an Excel file for easy copy/paste into RC Studio."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output_dir = cfg.library_root / "COMPOSED" / "_output"

    # Load the composition data to build Excel
    comp_path = output_dir / f"{comp_id}.pan"
    if not comp_path.exists():
        raise HTTPException(404, f"Composition not found for '{comp_id}'")

    # Try to load from the last compose result
    last_comp = getattr(composer, '_last_composition', None)
    if not last_comp or last_comp.get("id") != comp_id:
        raise HTTPException(404, "Composition data expired — recompose first")

    objects = last_comp.get("objects", {})
    device_name = last_comp.get("meta", {}).get("device_name", "{device-name}")

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    empty_fill = PatternFill("solid", fgColor="F2F2F2")

    def add_sheet(name, headers, rows):
        """Add a sheet with rows. First column must be instance number.
        Fills empty rows for gaps so the spreadsheet is contiguous for copy/paste."""
        ws = wb.create_sheet(name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border

        if not rows:
            return

        # Check if all row[0] values are integers (point-style with gap filling)
        all_int = all(isinstance(row[0], int) for row in rows)

        if all_int:
            # Build dict by instance for gap filling
            by_inst = {}
            for row in rows:
                by_inst[row[0]] = row

            max_inst = max(by_inst.keys())
            r = 2
            for inst in range(1, max_inst + 1):
                if inst in by_inst:
                    for c, val in enumerate(by_inst[inst], 1):
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.border = thin_border
                else:
                    # Empty row — just instance number, greyed out
                    cell = ws.cell(row=r, column=1, value=inst)
                    cell.border = thin_border
                    cell.fill = empty_fill
                    for c in range(2, len(headers) + 1):
                        cell = ws.cell(row=r, column=c, value="")
                        cell.border = thin_border
                        cell.fill = empty_fill
                r += 1
        else:
            # Direct row writing (arrays, tables, loops with continuation rows)
            r = 2
            for row in rows:
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                r += 1

        # Auto-width
        for c in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row, c).value or '')) for row in range(1, r))
            ws.column_dimensions[ws.cell(1, c).column_letter].width = min(max_len + 3, 50)

    # Analog Values
    av_rows = []
    for p in sorted(objects.get("AV", []), key=lambda x: int(x.get("instance", 0))):
        name = p.get("name", "").replace("{device-name}", device_name)
        av_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                       p.get("range", ""), p.get("unit", ""), p.get("increment", ""),
                       p.get("description", "")])
    if av_rows:
        add_sheet("Analog Values", ["Instance", "Name", "Value", "Range", "Unit", "Increment", "Description"], av_rows)

    # Analog Inputs
    ai_rows = []
    for p in sorted(objects.get("AI", []), key=lambda x: int(x.get("instance", 0))):
        name = p.get("name", "").replace("{device-name}", device_name)
        ai_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                       p.get("range", ""), p.get("unit", ""), p.get("increment", "")])
    if ai_rows:
        add_sheet("Analog Inputs", ["Instance", "Name", "Value", "Range", "Unit", "Increment"], ai_rows)

    # Analog Outputs
    ao_rows = []
    for p in sorted(objects.get("AO", []), key=lambda x: int(x.get("instance", 0))):
        name = p.get("name", "").replace("{device-name}", device_name)
        ao_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                       p.get("range", ""), p.get("unit", ""), p.get("increment", "")])
    if ao_rows:
        add_sheet("Analog Outputs", ["Instance", "Name", "Value", "Range", "Unit", "Increment"], ao_rows)

    # Binary Values
    bv_rows = []
    for p in sorted(objects.get("BV", []), key=lambda x: int(x.get("instance", 0))):
        name = p.get("name", "").replace("{device-name}", device_name)
        bv_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                       p.get("range", ""), p.get("unit", "")])
    if bv_rows:
        add_sheet("Binary Values", ["Instance", "Name", "Value", "Range", "Unit"], bv_rows)

    # Binary Inputs/Outputs
    for btype, sheet_name in [("BI", "Binary Inputs"), ("BO", "Binary Outputs")]:
        b_rows = []
        for p in sorted(objects.get(btype, []), key=lambda x: int(x.get("instance", 0))):
            name = p.get("name", "").replace("{device-name}", device_name)
            b_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                          p.get("range", ""), p.get("unit", "")])
        if b_rows:
            add_sheet(sheet_name, ["Instance", "Name", "Value", "Range", "Unit"], b_rows)

    # Multistate
    for mtype, sheet_name in [("MV", "Multistate Values"), ("MO", "Multistate Outputs")]:
        m_rows = []
        for p in sorted(objects.get(mtype, []), key=lambda x: int(x.get("instance", 0))):
            name = p.get("name", "").replace("{device-name}", device_name)
            m_rows.append([int(p.get("instance", 0)), name, p.get("present_value", ""),
                          p.get("range", "")])
        if m_rows:
            add_sheet(sheet_name, ["Instance", "Name", "Value", "Range"], m_rows)

    # Loops — include suggested input/setpoint
    import re as _re_xl
    # Build point name lookup
    all_pts = {}
    for ptype in ["AV", "AI", "AO", "BV", "BI", "BO", "MV", "MO"]:
        for p in objects.get(ptype, []):
            pname = p.get("name", "").replace("{device-name}", device_name)
            all_pts[pname] = f"{ptype}{p.get('instance','')}"

    # Find LOOP output assignments from code
    loop_outputs = {}
    for prog in objects.get("PROGRAM", []):
        code = prog.get("code", "").replace("{device-name}", device_name)
        for line in code.split('\n'):
            # AO4 = LOOP8 pattern
            m = _re_xl.search(r'(AO\d+)\s*=\s*LOOP(\d+)', line)
            if m:
                loop_outputs.setdefault(m.group(2), []).append(m.group(1))
            # F = LOOP1 then AO10 = ... pattern
            m2 = _re_xl.search(r'\w+\s*=\s*LOOP(\d+)', line)

    loop_rows = []
    for l in sorted(objects.get("LOOP", []), key=lambda x: int(x.get("instance", 0))):
        name = l.get("name", "").replace("{device-name}", device_name)
        inst = l.get("instance", "")

        # Use binary-extracted bindings first, fall back to inferred suggestions
        input_pt = ""
        setpoint_pt = ""
        output_pt = ""

        # Binary extraction (accurate)
        ir = l.get("input_ref", "")
        irn = l.get("input_name", "").replace("{device-name}", device_name)
        if ir:
            input_pt = f"{ir} ({irn})" if irn else ir
        else:
            # Fallback: inference from name matching
            si = l.get("suggested_input", "")
            sin = l.get("suggested_input_name", "").replace("{device-name}", device_name)
            if si:
                input_pt = f"{si} ({sin}) [inferred]" if sin else f"{si} [inferred]"

        sr = l.get("setpoint_ref", "")
        srn = l.get("setpoint_name", "").replace("{device-name}", device_name)
        if sr:
            setpoint_pt = f"{sr} ({srn})" if srn else sr
        else:
            sp = l.get("suggested_setpoint", "")
            spn = l.get("suggested_setpoint_name", "").replace("{device-name}", device_name)
            if sp:
                setpoint_pt = f"{sp} ({spn}) [inferred]" if spn else f"{sp} [inferred]"

        so = l.get("suggested_output", "")
        son = l.get("suggested_output_name", "").replace("{device-name}", device_name)
        if so:
            output_pt = f"{so} ({son})" if son else so

        loop_rows.append([int(inst), name,
                         l.get("proportional", ""), l.get("integral", ""),
                         l.get("derivative", ""), l.get("bias", ""),
                         l.get("deadband", ""), l.get("action", ""),
                         l.get("integralunits", ""),
                         input_pt, setpoint_pt, output_pt])
    if loop_rows:
        # Force direct write (no gap filling) — loop instances may skip
        loop_headers = ["Instance", "Name", "P", "I", "D", "Bias", "Deadband", "Action", "I-Units",
                       "Input (suggested)", "Setpoint (suggested)", "Output"]
        ws_loops = wb.create_sheet("Loops")
        for c, h in enumerate(loop_headers, 1):
            cell = ws_loops.cell(row=1, column=c, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for ri, row in enumerate(loop_rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws_loops.cell(row=ri, column=c, value=val)
                cell.border = thin_border
        for c in range(1, len(loop_headers) + 1):
            max_len = max(len(str(ws_loops.cell(r, c).value or '')) for r in range(1, len(loop_rows) + 2))
            ws_loops.column_dimensions[ws_loops.cell(1, c).column_letter].width = min(max_len + 3, 50)

    # Programs
    prog_rows = []
    for p in sorted(objects.get("PROGRAM", []), key=lambda x: int(x.get("instance", 0))):
        name = p.get("name", "").replace("{device-name}", device_name)
        pv = p.get("present_value", "1")
        prog_rows.append([int(p.get("instance", 0)), name,
                         "Yes" if str(pv) == "1" else "No",
                         p.get("code", "").replace("{device-name}", device_name)])
    if prog_rows:
        add_sheet("Programs", ["Instance", "Name", "Enabled", "Code"], prog_rows)

    # Arrays
    array_rows = []
    for a in sorted(objects.get("ARRAY", []), key=lambda x: int(x.get("instance", 0))):
        name = a.get("name", "").replace("{device-name}", device_name)
        vals = a.get("values", [])
        if vals:
            for idx, v in enumerate(vals):
                array_rows.append([int(a.get("instance", 0)) if idx == 0 else "", name if idx == 0 else "", idx, v])
        else:
            array_rows.append([int(a.get("instance", 0)), name, "", ""])
    if array_rows:
        add_sheet("Arrays", ["Instance", "Name", "Index", "Value"], array_rows)

    # Tables
    table_rows = []
    for t in sorted(objects.get("TABLE", []), key=lambda x: int(x.get("instance", 0))):
        name = t.get("name", "").replace("{device-name}", device_name)
        tbl_data = t.get("rows", [])
        if tbl_data:
            for ri, r in enumerate(tbl_data):
                table_rows.append([int(t.get("instance", 0)) if ri == 0 else "", name if ri == 0 else "",
                                  t.get("unit", "") if ri == 0 else "", ri + 1, r.get("in", ""), r.get("out", "")])
        else:
            table_rows.append([int(t.get("instance", 0)), name, t.get("unit", ""), "", "", ""])
    if table_rows:
        add_sheet("Tables", ["Instance", "Name", "Unit", "Row #", "Input", "Output"], table_rows)

    # Trends
    trend_rows = []
    for t in sorted(objects.get("TREND", []), key=lambda x: int(x.get("instance", 0))):
        name = t.get("name", "").replace("{device-name}", device_name)
        refs = ", ".join(r for r in t.get("references", []) if r)
        trend_rows.append([int(t.get("instance", 0)), name, t.get("type", ""),
                          t.get("interval", ""), refs])
    if trend_rows:
        add_sheet("Trends", ["Instance", "Name", "Type", "Interval", "References"], trend_rows)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{comp_id}_values.xlsx"'}
    )


# ─── #16 + #23: Auto-Commissioning & Wiring Verification ─────────────────────

def _load_pan_from_ref(ref: dict):
    """Helper: load a PanBinary from a category/variant_id reference."""
    from app.pan_binary import PanBinary
    folder = engine._cat_folder(ref['category'])
    cat_dir = engine.cfg.upload_root / folder
    src = next(cat_dir.rglob(f"{ref['variant_id']}.panx"), None) or \
          next(cat_dir.rglob(f"{ref['variant_id']}.pan"), None)
    if not src:
        raise HTTPException(404, f"Controller not found: {ref['variant_id']}")
    if str(src).endswith('.panx'):
        return PanBinary.from_panx(src)
    return PanBinary.from_file(src)


def _load_pan_from_upload(data: bytes, filename: str):
    """Helper: load a PanBinary from uploaded file bytes."""
    from app.pan_binary import PanBinary
    if filename.endswith('.panx'):
        import zipfile as _zf
        import io as _io
        with _zf.ZipFile(_io.BytesIO(data)) as z:
            pan_name = [n for n in z.namelist() if n.endswith('.pan')][0]
            return PanBinary(z.read(pan_name))
    return PanBinary(data)


def _get_device_name(pan) -> str:
    """Helper: detect device name prefix from a PanBinary."""
    prefixes = {}
    for obj in pan.objects:
        name = obj['name'].strip().strip('\x00')
        parts = name.split('-')
        if len(parts) >= 3:
            prefix = f"{parts[0]}-{parts[1]}"
            if prefix and len(prefix) > 2:
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
    return max(prefixes, key=prefixes.get) if prefixes else "unknown"


def _commission_check_logic(design_pan, asbuilt_pan, tolerance: float = 0.5) -> dict:
    """Core commissioning comparison logic used by the endpoint.

    Compares a design (template) .pan against an as-built (field) .pan and
    returns a comprehensive commissioning verification report including
    wiring verification (#23).
    """
    result = {
        'device_id_match': True,
        'device_id': {},
        'object_summary': {},
        'missing_in_asbuilt': [],
        'extra_in_asbuilt': [],
        'present_value_differences': [],
        'loop_binding_differences': [],
        'pid_differences': [],
        'out_of_service_issues': [],
        'unusual_values': [],
        'wiring_verification': [],
        'wiring_summary': {},
        'action_items': [],
        'score': 0.0,
    }

    # --- Device ID comparison ---
    design_id = struct.unpack('<I', design_pan.data[4:8])[0] if len(design_pan.data) >= 8 else 0
    asbuilt_id = struct.unpack('<I', asbuilt_pan.data[4:8])[0] if len(asbuilt_pan.data) >= 8 else 0
    result['device_id'] = {'design': design_id, 'as_built': asbuilt_id}
    if design_id != asbuilt_id:
        result['device_id_match'] = False

    design_name = _get_device_name(design_pan)
    asbuilt_name = _get_device_name(asbuilt_pan)
    result['device_id']['design_name'] = design_name
    result['device_id']['as_built_name'] = asbuilt_name

    # --- Build object maps ---
    # Use the last part of the name (after device prefix) for matching
    def strip_prefix(name, prefix):
        if name.startswith(prefix + '-'):
            return name[len(prefix) + 1:]
        return name

    design_points = {}
    for p in design_pan.get_point_details():
        key = strip_prefix(p['name'], design_name)
        design_points[key] = p

    asbuilt_points = {}
    for p in asbuilt_pan.get_point_details():
        key = strip_prefix(p['name'], asbuilt_name)
        asbuilt_points[key] = p

    # All design objects by stripped name
    design_all = {}
    for obj in design_pan.objects:
        key = strip_prefix(obj['name'], design_name)
        design_all[key] = obj

    asbuilt_all = {}
    for obj in asbuilt_pan.objects:
        key = strip_prefix(obj['name'], asbuilt_name)
        asbuilt_all[key] = obj

    # --- Object count summary ---
    design_cats = {}
    for obj in design_pan.objects:
        design_cats[obj['category']] = design_cats.get(obj['category'], 0) + 1
    asbuilt_cats = {}
    for obj in asbuilt_pan.objects:
        asbuilt_cats[obj['category']] = asbuilt_cats.get(obj['category'], 0) + 1
    result['object_summary'] = {'design': design_cats, 'as_built': asbuilt_cats}

    # --- Missing / Extra objects ---
    design_keys = set(design_all.keys())
    asbuilt_keys = set(asbuilt_all.keys())
    result['missing_in_asbuilt'] = sorted(design_keys - asbuilt_keys)
    result['extra_in_asbuilt'] = sorted(asbuilt_keys - design_keys)

    if result['missing_in_asbuilt']:
        result['action_items'].append({
            'severity': 'error',
            'message': f"{len(result['missing_in_asbuilt'])} object(s) from design are missing in as-built",
            'details': result['missing_in_asbuilt'][:10],
        })
    if result['extra_in_asbuilt']:
        result['action_items'].append({
            'severity': 'info',
            'message': f"{len(result['extra_in_asbuilt'])} extra object(s) in as-built not in design",
            'details': result['extra_in_asbuilt'][:10],
        })

    # --- Present value differences ---
    total_points = 0
    matching_points = 0
    common_keys = set(design_points.keys()) & set(asbuilt_points.keys())

    for key in sorted(common_keys):
        dp = design_points[key]
        ap = asbuilt_points[key]
        dpv = dp.get('present_value')
        apv = ap.get('present_value')

        if dpv is not None and apv is not None:
            total_points += 1
            diff = abs(dpv - apv)
            if diff <= tolerance:
                matching_points += 1
            else:
                result['present_value_differences'].append({
                    'name': key,
                    'design_value': round(dpv, 4),
                    'as_built_value': round(apv, 4),
                    'difference': round(diff, 4),
                })

    # --- Out-of-service checks ---
    design_oos = set()
    for key, p in design_points.items():
        if p.get('out_of_service'):
            design_oos.add(key)

    for key, p in asbuilt_points.items():
        if p.get('out_of_service') and key not in design_oos:
            result['out_of_service_issues'].append({
                'name': key,
                'message': 'Out-of-service in as-built but not in design',
                'present_value': p.get('present_value'),
            })

    if result['out_of_service_issues']:
        result['action_items'].append({
            'severity': 'warning',
            'message': f"{len(result['out_of_service_issues'])} point(s) out-of-service in field that should be active",
            'details': [i['name'] for i in result['out_of_service_issues']],
        })

    # --- Unusual values check ---
    for key, p in asbuilt_points.items():
        pv = p.get('present_value')
        if pv is None:
            continue
        units = p.get('units_name', '') or ''
        issues = []
        # Negative temperatures
        if 'deg' in units.lower() and pv < -40:
            issues.append(f"Negative temperature: {pv}")
        # Output > 100%
        if ('percent' in units.lower() or 'open' in units.lower()) and pv > 100:
            issues.append(f"Output exceeds 100%: {pv}")
        # Negative output percentage
        if ('percent' in units.lower() or 'open' in units.lower()) and pv < 0:
            issues.append(f"Negative output: {pv}")
        # Extremely large values
        if abs(pv) > 10000:
            issues.append(f"Unusually large value: {pv}")

        if issues:
            result['unusual_values'].append({
                'name': key,
                'present_value': round(pv, 4),
                'units': units,
                'issues': issues,
            })

    if result['unusual_values']:
        result['action_items'].append({
            'severity': 'warning',
            'message': f"{len(result['unusual_values'])} point(s) with unusual values need investigation",
            'details': [f"{i['name']}: {', '.join(i['issues'])}" for i in result['unusual_values']],
        })

    # --- Loop binding + PID differences ---
    design_loops = {l.get('instance', ''): l for l in design_pan.get_loops()}
    asbuilt_loops = {l.get('instance', ''): l for l in asbuilt_pan.get_loops()}

    all_loop_insts = set(design_loops.keys()) | set(asbuilt_loops.keys())
    for inst in sorted(all_loop_insts):
        dl = design_loops.get(inst, {})
        al = asbuilt_loops.get(inst, {})

        # Binding differences
        for field in ['input_ref', 'setpoint_ref', 'output_ref']:
            dv = dl.get(field, '')
            av = al.get(field, '')
            if dv != av:
                result['loop_binding_differences'].append({
                    'loop': f"LOOP{inst}",
                    'field': field.replace('_ref', ''),
                    'design': dv or 'not set',
                    'as_built': av or 'not set',
                })

        # PID parameter differences
        for param in ['proportional', 'integral', 'derivative']:
            dp_val = dl.get(param)
            ap_val = al.get(param)
            if dp_val is not None and ap_val is not None:
                if abs(dp_val - ap_val) > 0.001:
                    result['pid_differences'].append({
                        'loop': f"LOOP{inst}",
                        'parameter': param,
                        'design': dp_val,
                        'as_built': ap_val,
                    })

    if result['loop_binding_differences']:
        result['action_items'].append({
            'severity': 'error',
            'message': f"{len(result['loop_binding_differences'])} loop binding difference(s) - verify wiring",
            'details': [f"{d['loop']} {d['field']}: design={d['design']}, field={d['as_built']}"
                       for d in result['loop_binding_differences']],
        })

    if result['pid_differences']:
        result['action_items'].append({
            'severity': 'warning',
            'message': f"{len(result['pid_differences'])} PID parameter difference(s)",
            'details': [f"{d['loop']} {d['parameter']}: design={d['design']}, field={d['as_built']}"
                       for d in result['pid_differences']],
        })

    # --- #23: Wiring Verification ---
    # For each loop, check if input ref points to a valid object and whether
    # it's a hardware point (AI/AO) vs software point (AV)
    # Build instance lookup for as-built objects
    asbuilt_obj_instances = set()
    for obj in asbuilt_pan.objects:
        for ref in obj.get('refs', []):
            if ref['type'] != 'NULL':
                asbuilt_obj_instances.add(f"{ref['type']}{ref['instance']}")
    # Also add objects by category name matching
    for obj in asbuilt_pan.objects:
        name = obj['name'].upper()
        cat = obj['category']
        # Try to extract type+instance from name patterns
        import re as _re_wv
        m = _re_wv.search(r'(AI|AO|AV|BI|BO|BV|MO|MV|LOOP|TREND|SCHEDULE|PROGRAM)(\d+)', name)
        if m:
            asbuilt_obj_instances.add(f"{m.group(1)}{m.group(2)}")

    loops_verified = 0
    loops_need_attention = 0
    total_loops = len(asbuilt_loops)

    for inst, al in asbuilt_loops.items():
        wiring_entry = {
            'loop': f"LOOP{inst}",
            'name': al.get('name', ''),
            'status': 'ok',
            'issues': [],
        }

        input_ref = al.get('input_ref', '')
        input_type = al.get('input_type', '')
        setpoint_ref = al.get('setpoint_ref', '')
        output_ref = al.get('output_ref', '')

        # Check if input is AV (software) vs AI (hardware sensor)
        if input_type == 'AV':
            wiring_entry['issues'].append(
                f"Input {input_ref} is an Analog Value (software point) - "
                f"field wiring should use AI (hardware sensor)"
            )
            wiring_entry['status'] = 'attention'

        # Check if setpoint ref is a valid object
        if setpoint_ref:
            setpoint_type = al.get('setpoint_type', '')
            # Setpoints are typically AV which is fine
            pass

        # Check for references to non-existent instances
        for ref_field, ref_val in [('input', input_ref), ('setpoint', setpoint_ref), ('output', output_ref)]:
            if ref_val and ref_val != 'NULL':
                # Check if the referenced object exists in the as-built
                found = False
                for obj in asbuilt_pan.objects:
                    obj_name = obj['name'].upper()
                    if ref_val.upper() in obj_name or ref_val.upper().replace(' ', '') in obj_name.replace(' ', ''):
                        found = True
                        break
                # Also check refs
                if not found and ref_val in asbuilt_obj_instances:
                    found = True
                # Relax: if we can't find it by name, it may still be valid
                # (instance numbers don't always appear in names)
                if not found:
                    wiring_entry['issues'].append(
                        f"{ref_field.capitalize()} reference {ref_val} may not exist in as-built controller"
                    )
                    wiring_entry['status'] = 'attention'

        if wiring_entry['issues']:
            loops_need_attention += 1
        else:
            loops_verified += 1

        result['wiring_verification'].append(wiring_entry)

    result['wiring_summary'] = {
        'total_loops': total_loops,
        'verified': loops_verified,
        'need_attention': loops_need_attention,
        'message': f"{loops_verified} of {total_loops} loops verified, {loops_need_attention} need field attention",
    }

    if loops_need_attention > 0:
        result['action_items'].append({
            'severity': 'warning',
            'message': result['wiring_summary']['message'],
            'details': [f"{w['loop']}: {'; '.join(w['issues'])}"
                       for w in result['wiring_verification'] if w['status'] == 'attention'],
        })

    # --- Commissioning Score ---
    # Score components: object presence, value matching, loop bindings, OOS
    score_components = []

    # Object presence score
    total_design_objs = len(design_keys)
    if total_design_objs > 0:
        present_count = len(design_keys & asbuilt_keys)
        score_components.append(present_count / total_design_objs)

    # Value matching score
    if total_points > 0:
        score_components.append(matching_points / total_points)

    # Loop binding score
    total_loop_fields = len(all_loop_insts) * 3  # input, setpoint, output per loop
    if total_loop_fields > 0:
        binding_issues = len(result['loop_binding_differences'])
        score_components.append(max(0, (total_loop_fields - binding_issues) / total_loop_fields))

    # OOS penalty
    if len(asbuilt_points) > 0:
        oos_count = len(result['out_of_service_issues'])
        score_components.append(max(0, (len(asbuilt_points) - oos_count) / len(asbuilt_points)))

    if score_components:
        result['score'] = round(sum(score_components) / len(score_components) * 100, 1)
    else:
        result['score'] = 0.0

    # Score grade
    if result['score'] >= 95:
        result['grade'] = 'PASS'
    elif result['score'] >= 80:
        result['grade'] = 'CONDITIONAL'
    else:
        result['grade'] = 'FAIL'

    result['action_items'].append({
        'severity': 'info',
        'message': f"Commissioning score: {result['score']}% ({result['grade']})",
        'details': [],
    })

    return result


@app.post("/api/binary/commission-check")
async def binary_commission_check(body: dict = None):
    """Compare a design (template) .pan against an as-built (field) .pan.

    JSON body with category/variant pairs:
    {
        "design": {"category": "SBS_AHU", "variant_id": "1003"},
        "as_built": {"category": "SBS_AHU", "variant_id": "PS-AHU-ERW-0100"},
        "tolerance": 0.5
    }

    Returns comprehensive commissioning report with wiring verification.
    """
    from app.pan_binary import PanBinary

    if not body or 'design' not in body or 'as_built' not in body:
        raise HTTPException(400, "Must provide 'design' and 'as_built' controller references")

    tolerance = body.get('tolerance', 0.5)
    try:
        design_pan = _load_pan_from_ref(body['design'])
        asbuilt_pan = _load_pan_from_ref(body['as_built'])
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to load controllers: {e}")

    try:
        report = _commission_check_logic(design_pan, asbuilt_pan, tolerance)
        return report
    except Exception as e:
        logger.exception("Commission check failed")
        raise HTTPException(500, f"Commission check failed: {e}")


@app.post("/api/binary/commission-check/upload")
async def binary_commission_check_upload(
    design_file: UploadFile = File(...),
    asbuilt_file: UploadFile = File(...),
):
    """Compare uploaded design and as-built .pan/.panx files.

    Upload two files via multipart form: design_file and asbuilt_file.
    Returns comprehensive commissioning report with wiring verification.
    """
    try:
        design_data = await design_file.read()
        asbuilt_data = await asbuilt_file.read()
        design_pan = _load_pan_from_upload(design_data, design_file.filename or "design.pan")
        asbuilt_pan = _load_pan_from_upload(asbuilt_data, asbuilt_file.filename or "asbuilt.pan")
    except Exception as e:
        raise HTTPException(400, f"Failed to parse uploaded files: {e}")

    try:
        report = _commission_check_logic(design_pan, asbuilt_pan, tolerance=0.5)
        return report
    except Exception as e:
        logger.exception("Commission check failed")
        raise HTTPException(500, f"Commission check failed: {e}")


# ─── #21: Program State Management ───────────────────────────────────────────

@app.get("/api/binary/programs/{category}/{variant_id}")
async def binary_programs_list(category: str, variant_id: str):
    """List all programs with their enabled state from the .pan binary.

    Returns program names, enabled/disabled state, data sizes, and ref counts.
    """
    from app.pan_binary import PanBinary

    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name

    src_panx = next(cat_dir.rglob(f"{variant_id}.panx"), None)
    src_pan = next(cat_dir.rglob(f"{variant_id}.pan"), None)

    if src_panx:
        pan = PanBinary.from_panx(src_panx)
    elif src_pan:
        pan = PanBinary.from_file(src_pan)
    else:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    programs = pan.get_programs()

    enabled_count = sum(1 for p in programs if p.get('enabled', False))
    disabled_count = sum(1 for p in programs if p.get('enabled') is False)

    return {
        "variant_id": variant_id,
        "category": category,
        "program_count": len(programs),
        "enabled_count": enabled_count,
        "disabled_count": disabled_count,
        "programs": programs,
    }


@app.post("/api/binary/programs/toggle")
async def binary_program_toggle(body: dict = None):
    """Toggle a program's enabled/disabled state in the .pan binary.

    Body: {
        "category": "SBS_AHU",
        "variant_id": "1003",
        "program_name": "MYS-AHU3-PRG1",
        "enabled": true
    }

    Returns the modified .pan file as a download.
    Uses PanWriter to modify the program-enabled property (0x0A).
    """
    from app.pan_binary import PanBinary, PanWriter

    if not body or 'category' not in body or 'variant_id' not in body or 'program_name' not in body:
        raise HTTPException(400, "Must provide 'category', 'variant_id', and 'program_name'")

    category = body['category']
    variant_id = body['variant_id']
    program_name = body['program_name']
    target_enabled = body.get('enabled', None)  # None = toggle, True/False = set explicitly

    folder_name = engine._cat_folder(category)
    cat_dir = engine.cfg.upload_root / folder_name
    src = next(cat_dir.rglob(f"{variant_id}.panx"), None) or \
          next(cat_dir.rglob(f"{variant_id}.pan"), None)
    if not src:
        raise HTTPException(404, f"No .pan/.panx found for {variant_id}")

    try:
        if str(src).endswith('.panx'):
            writer = PanWriter.from_panx(src)
        else:
            writer = PanWriter.from_file(src)

        # Find the program object in the binary
        import re as _re_prg
        found = False
        program_name_bytes = program_name.encode('utf-8')

        for m in _re_prg.finditer(b'\x4d\x75(.)', bytes(writer.data)):
            name_len = m.group(1)[0]
            if name_len < 2 or name_len > 200:
                continue
            name_start = m.start() + 3
            name_bytes = writer.data[name_start:name_start + name_len]
            try:
                name = name_bytes.decode('utf-8').strip('\x00')
            except (UnicodeDecodeError, ValueError):
                continue

            if name.strip() != program_name.strip():
                continue

            # Found the program - now find property 0x0A in its data region
            name_end = name_start + name_len
            next_mu = bytes(writer.data).find(b'\x4d\x75', name_end + 10)
            if next_mu < 0:
                next_mu = min(name_end + 500, len(writer.data))

            data_region_start = name_end
            data_region_end = next_mu

            # Search for property 0x0A (program-enabled)
            for i in range(data_region_start, min(data_region_end, data_region_start + 300)):
                if i + 2 < len(writer.data) and writer.data[i] == 0x0A and writer.data[i + 1] == 0x91:
                    current_state = writer.data[i + 2]
                    if target_enabled is None:
                        # Toggle
                        new_state = 0 if current_state == 1 else 1
                    else:
                        new_state = 1 if target_enabled else 0

                    writer.data[i + 2] = new_state
                    found = True
                    break

            if found:
                break

        if not found:
            raise HTTPException(404, f"Program '{program_name}' not found or has no enable property")

        out_path = Path(f"/tmp/program_toggle_{variant_id}.pan")
        writer.save(out_path)

        new_state_str = "enabled" if (target_enabled if target_enabled is not None else new_state == 1) else "disabled"

        return FileResponse(
            out_path,
            filename=f"{variant_id}_program_{new_state_str}.pan",
            media_type="application/octet-stream",
            headers={
                "X-Program-Name": program_name,
                "X-New-State": new_state_str,
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Program toggle failed")
        raise HTTPException(500, f"Program toggle failed: {e}")


# ─── Background Task ─────────────────────────────────────────────────────────

async def run_extraction(job_id: str, category_filter: Optional[str], selected_keys: Optional[list]):
    job = jobs[job_id]
    try:
        job["status"] = "running"
        variants = engine.discover_variants()

        flat = []
        if selected_keys:
            # Process only the specified variants
            for key in selected_keys:
                cat, vid = key.split("/", 1)
                for cat_name, items in variants.items():
                    if cat_name.upper() == cat.upper():
                        for item in items:
                            if item["id"] == vid:
                                flat.append((cat_name, item))
                                break
        else:
            for cat, items in variants.items():
                if category_filter and cat.upper() != category_filter.upper():
                    continue
                for item in items:
                    flat.append((cat, item))

        job["total"] = len(flat)
        for i, (cat, item) in enumerate(flat):
            job["current"] = f"{cat}/{item['id']}"
            job["progress"] = i + 1
            try:
                await asyncio.get_event_loop().run_in_executor(
                    None, engine.process_variant, cat, item
                )
                job["done"].append(f"{cat}/{item['id']}")
            except Exception as e:
                logger.error(f"Failed {cat}/{item['id']}: {e}")
                job["errors"].append({"variant": f"{cat}/{item['id']}", "error": str(e)})

        job["status"] = "done"
        job["current"] = ""
    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        logger.exception("Extraction job failed")


# ─── Door Label PDF Generation ──────────────────────────────────────────────

# Controller terminal definitions — maps controller model to I/O terminal names
# Instance numbers in BACnet map 1:1 to terminal positions
CONTROLLER_TERMINALS = {
    # RC-FLEXair family (VAV/terminal units)
    'RCFA-M-36': {'inputs': ['IN:01', 'IN:02', 'IN:03'], 'analog_out': ['UO:01', 'UO:02', 'UO:03'], 'binary_out': ['BO:01', 'BO:02', 'BO:03'], 'label': 'RC-FLEXair 36'},
    'RCFA-M-35': {'inputs': ['IN:01', 'IN:02', 'IN:03'], 'analog_out': ['UO:01', 'UO:02', 'UO:03', 'UO:04', 'UO:05'], 'binary_out': [], 'label': 'RC-FLEXair 35'},
    'RCFA-M-34': {'inputs': ['IN:01', 'IN:02', 'IN:03'], 'analog_out': ['UO:01', 'UO:02', 'UO:03', 'UO:04'], 'binary_out': [], 'label': 'RC-FLEXair 34'},
    'RCFA-M-33': {'inputs': ['IN:01', 'IN:02', 'IN:03'], 'analog_out': ['UO:01', 'UO:02', 'UO:03'], 'binary_out': [], 'label': 'RC-FLEXair 33'},
    'RCFA-M-12': {'inputs': ['IN:01'], 'analog_out': [], 'binary_out': ['BO:01', 'BO:02'], 'label': 'RC-FLEXair 12'},
    # RC-FLEXone family (equipment/room control)
    'RCFO-36': {'inputs': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06', 'IN:01', 'IN:02', 'IN:03', 'IN:04', 'IN:05', 'IN:06'], 'analog_out': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06'], 'binary_out': [], 'label': 'RC-FLEXone 36'},
    'RCFO-M-36': {'inputs': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06', 'IN:01', 'IN:02', 'IN:03', 'IN:04', 'IN:05', 'IN:06'], 'analog_out': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06'], 'binary_out': [], 'label': 'RC-FLEXone M-36'},
    'RCFO-M-16': {'inputs': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06', 'IN:01'], 'analog_out': ['FX:01', 'FX:02', 'FX:03', 'FX:04', 'FX:05', 'FX:06'], 'binary_out': [], 'label': 'RC-FLEXone M-16'},
    # MACH-ProZone (VAV)
    'MPZ-44': {'inputs': ['IN:01+/IN:01-', 'IN:02+/IN:02-', 'IN:03+/IN:03-', 'IN:04+/IN:04-'], 'analog_out': ['OUT:01+/OUT:01-', 'OUT:02+/OUT:02-', 'OUT:03+/OUT:03-', 'OUT:04+/OUT:04-'], 'binary_out': [], 'label': 'MACH-ProZone 44'},
    'MPZ-88': {'inputs': ['IN:01+/IN:01-', 'IN:02+/IN:02-', 'IN:03+/IN:03-', 'IN:04+/IN:04-', 'IN:05+/IN:05-', 'IN:06+/IN:06-', 'IN:07+/IN:07-', 'IN:08+/IN:08-'], 'analog_out': ['OUT:01+/OUT:01-', 'OUT:02+/OUT:02-', 'OUT:03+/OUT:03-', 'OUT:04+/OUT:04-', 'OUT:05+/OUT:05-', 'OUT:06+/OUT:06-', 'OUT:07+/OUT:07-', 'OUT:08+/OUT:08-'], 'binary_out': [], 'label': 'MACH-ProZone 88'},
    # MACH-ProAir
    'MPA-33': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'BI:01', 'BI:02', 'BI:03'], 'analog_out': ['AO:01', 'AO:02', 'AO:03'], 'binary_out': ['BO:01', 'BO:02', 'BO:03'], 'label': 'MACH-ProAir 33'},
    'MPA-34': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'BI:01', 'BI:02', 'BI:03'], 'analog_out': ['AO:01', 'AO:02', 'AO:03', 'AO:04'], 'binary_out': ['BO:01', 'BO:02', 'BO:03'], 'label': 'MACH-ProAir 34'},
    'MPA-35': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'BI:01', 'BI:02', 'BI:03', 'BI:04', 'BI:05'], 'analog_out': ['AO:01', 'AO:02', 'AO:03', 'AO:04', 'AO:05'], 'binary_out': ['BO:01', 'BO:02', 'BO:03', 'BO:04', 'BO:05'], 'label': 'MACH-ProAir 35'},
    'MPA-36': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'BI:01', 'BI:02', 'BI:03', 'BI:04', 'BI:05', 'BI:06', 'BI:07', 'BI:08'], 'analog_out': ['AO:01', 'AO:02', 'AO:03', 'AO:04', 'AO:05', 'AO:06'], 'binary_out': ['BO:01', 'BO:02', 'BO:03', 'BO:04', 'BO:05', 'BO:06', 'BO:07', 'BO:08'], 'label': 'MACH-ProAir 36'},
    # MACH-ProSys
    'MPS': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'UI:04', 'UI:05', 'UI:06', 'UI:07', 'UI:08', 'BI:01', 'BI:02', 'BI:03', 'BI:04', 'BI:05', 'BI:06', 'BI:07', 'BI:08'], 'analog_out': ['AO:01', 'AO:02', 'AO:03', 'AO:04', 'AO:05', 'AO:06', 'AO:07', 'AO:08'], 'binary_out': ['BO:01', 'BO:02', 'BO:03', 'BO:04', 'BO:05', 'BO:06', 'BO:07', 'BO:08'], 'label': 'MACH-ProSys'},
    # MACH-Pro2
    'MP2': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'UI:04', 'UI:05', 'UI:06', 'UI:07', 'UI:08', 'UI:09', 'UI:10', 'UI:11', 'UI:12'], 'analog_out': ['UO:01', 'UO:02', 'UO:03', 'UO:04', 'UO:05', 'UO:06', 'UO:07', 'UO:08'], 'binary_out': [], 'label': 'MACH-Pro2'},
    # MACH-Pro1
    'MP1': {'inputs': ['UI:01', 'UI:02', 'UI:03', 'UI:04', 'UI:05', 'UI:06', 'UI:07', 'UI:08'], 'analog_out': ['UO:01', 'UO:02', 'UO:03', 'UO:04', 'UO:05', 'UO:06', 'UO:07', 'UO:08'], 'binary_out': [], 'label': 'MACH-Pro1'},
}

# Map library Model field to controller model ID
MODEL_NUM_TO_CONTROLLER = {
    47: 'RCFA-M-36',  # FLEXair — hardpoint config determines actual variant
    49: 'RCFO-M-36',  # FLEXone
    19: 'MPZ-44',     # ProZone
    28: 'MPS',        # ProSys (or MPA depending on I/O)
    32: 'MPA-36',     # ProAir
    40: 'MP2',        # Pro2
    44: 'MP1',        # Pro1
}

# HardPointConfig to FLEXair variant
HPC_TO_RCFA = {
    '12-F': 'RCFA-M-12', '12-M': 'RCFA-M-12',
    '33-F': 'RCFA-M-33', '33-M': 'RCFA-M-33',
    '34-F': 'RCFA-M-34', '34-M': 'RCFA-M-34',
    '35-F': 'RCFA-M-35', '35-M': 'RCFA-M-35',
    '36-F': 'RCFA-M-36', '36-M': 'RCFA-M-36',
}


def _resolve_controller(meta: dict) -> str:
    """Given variant meta, figure out the controller model."""
    model_num = meta.get('Model')
    hpc = meta.get('HardPointConfig', '')
    if model_num == 47 and hpc:
        return HPC_TO_RCFA.get(hpc, 'RCFA-M-36')
    if model_num in MODEL_NUM_TO_CONTROLLER:
        return MODEL_NUM_TO_CONTROLLER[model_num]
    return 'MPS'  # default fallback


def _build_io_table(variant_data: dict, device_name: str = "") -> list:
    """Build I/O terminal mapping table from variant JSON data.
    Returns list of dicts: {terminal, io_type, point_name, description}
    """
    meta = variant_data.get('meta', {})
    objs = variant_data.get('objects', {})
    controller_id = _resolve_controller(meta)
    ctrl = CONTROLLER_TERMINALS.get(controller_id, CONTROLLER_TERMINALS.get('MPS'))

    rows = []

    # Map AI/BI objects to input terminals
    ai_objs = sorted(objs.get('AI', []), key=lambda o: int(o.get('instance', 0)))
    bi_objs = sorted(objs.get('BI', []), key=lambda o: int(o.get('instance', 0)))
    ao_objs = sorted(objs.get('AO', []), key=lambda o: int(o.get('instance', 0)))
    bo_objs = sorted(objs.get('BO', []), key=lambda o: int(o.get('instance', 0)))

    # For FLEXair/FLEXone: AI maps to IN terminals, BI also maps to IN terminals
    # For MPA/MPS: AI maps to UI terminals, BI maps to BI terminals
    input_terminals = ctrl.get('inputs', [])
    analog_out_terminals = ctrl.get('analog_out', [])
    binary_out_terminals = ctrl.get('binary_out', [])

    # Inputs: AI points go to input terminals by instance
    for obj in ai_objs:
        inst = int(obj.get('instance', 0))
        term_idx = inst - 1
        terminal = input_terminals[term_idx] if term_idx < len(input_terminals) else f'IN:{inst:02d}'
        name = obj.get('name', '').replace('{device-name}', device_name) if device_name else obj.get('name', '')
        rows.append({
            'terminal': terminal,
            'io_type': 'AI',
            'point_name': name,
            'description': obj.get('description', ''),
        })

    # Binary Inputs
    for obj in bi_objs:
        inst = int(obj.get('instance', 0))
        # For MPA/MPS: BI terminals are separate, numbered BI:01+
        if controller_id.startswith('MPA') or controller_id in ('MPS', 'MP2', 'MP1'):
            # BI terminals come after UI terminals in the inputs list
            ui_count = len([t for t in input_terminals if t.startswith('UI')])
            term_idx = ui_count + inst - 1
            terminal = input_terminals[term_idx] if term_idx < len(input_terminals) else f'BI:{inst:02d}'
        else:
            # FLEXair: BI shares IN terminals
            term_idx = inst - 1
            terminal = input_terminals[term_idx] if term_idx < len(input_terminals) else f'IN:{inst:02d}'
        name = obj.get('name', '').replace('{device-name}', device_name) if device_name else obj.get('name', '')
        rows.append({
            'terminal': terminal,
            'io_type': 'BI',
            'point_name': name,
            'description': obj.get('description', ''),
        })

    # Analog Outputs
    for obj in ao_objs:
        inst = int(obj.get('instance', 0))
        term_idx = inst - 1
        terminal = analog_out_terminals[term_idx] if term_idx < len(analog_out_terminals) else f'AO:{inst:02d}'
        name = obj.get('name', '').replace('{device-name}', device_name) if device_name else obj.get('name', '')
        rows.append({
            'terminal': terminal,
            'io_type': 'AO',
            'point_name': name,
            'description': obj.get('description', ''),
        })

    # Binary Outputs
    for obj in bo_objs:
        inst = int(obj.get('instance', 0))
        term_idx = inst - 1
        terminal = binary_out_terminals[term_idx] if term_idx < len(binary_out_terminals) else f'BO:{inst:02d}'
        name = obj.get('name', '').replace('{device-name}', device_name) if device_name else obj.get('name', '')
        rows.append({
            'terminal': terminal,
            'io_type': 'BO',
            'point_name': name,
            'description': obj.get('description', ''),
        })

    return rows


def _get_label_template_file(controller_id: str) -> Optional[Path]:
    """Find the matching RC label template Word doc for a controller model."""
    templates_dir = Path('/srv/dfa/shared/files/vendors/reliable/label-templates')
    if not templates_dir.exists():
        return None

    # Map controller models to template files
    TEMPLATE_MAP = {
        # RC-FLEXone/FLEXair → RCFO templates based on I/O config
        'RCFO-36': 'RCFO-848-Door-Label-Template.docx',   # 8 IN, 4 FX, 8 OUT
        'RCFO-M-36': 'RCFO-848-Door-Label-Template.docx',
        'RCFO-M-16': 'RCFO-646-Door-Label-Template.docx', # 6 IN, 4 FX, 6 OUT
        'RCFA-M-36': 'RCFO-444-Door-Label-Template.docx', # 4 fields (3UI+3UO+3BO = 9 pts, fits 444)
        'RCFA-M-35': 'RCFO-444-Door-Label-Template.docx',
        'RCFA-M-34': 'RCFO-444-Door-Label-Template.docx',
        'RCFA-M-33': 'RCFO-242-Door-Label-Template.docx', # smaller configs
        'RCFA-M-12': 'RCFO-242-Door-Label-Template.docx',
        # MACH-Pro family → MP door label
        'MPZ-44': 'MPZ_Door_Label_Template.docx',
        'MPZ-88': 'MPZ_Door_Label_Template.docx',
        'MPA-33': 'MP-DL_door_label_ver-E.doc',
        'MPA-34': 'MP-DL_door_label_ver-E.doc',
        'MPA-35': 'MP-DL_door_label_ver-E.doc',
        'MPA-36': 'MP-DL_door_label_ver-E.doc',
        'MPS': 'MP-DL_door_label_ver-E.doc',
        'MP2': 'MP-DL_door_label_ver-E.doc',
        'MP1': 'MP-DL_door_label_ver-E.doc',
    }

    filename = TEMPLATE_MAP.get(controller_id)
    if filename:
        path = templates_dir / filename
        if path.exists():
            return path
    return None


def _fill_label_template(template_path: Path, point_names: list, device_name: str = "") -> bytes:
    """Fill an RC label template .docx by replacing ABCDEFGH placeholders with point names.
    Returns the modified docx as bytes.

    RC label templates have 4 labels per sheet (4 columns). Each label has N text boxes
    (e.g., 12 for RCFO-444, 20 for RCFO-848) in a fixed layout. All text boxes contain
    'ABCDEFGH' placeholder text. The Word format uses mc:AlternateContent with both
    DrawingML (wp:anchor) and VML (v:textbox) representations — both contain w:t text.

    Strategy: find all mc:AlternateContent blocks, parse positions from the DrawingML
    anchor, sort by column then row, then replace ABCDEFGH in both representations.
    """
    import shutil
    import tempfile
    import zipfile
    import re

    tmp_dir = tempfile.mkdtemp()
    try:
        extract_dir = Path(tmp_dir) / 'extracted'
        with zipfile.ZipFile(template_path, 'r') as z:
            z.extractall(extract_dir)

        doc_xml = extract_dir / 'word' / 'document.xml'
        xml_content = doc_xml.read_text(encoding='utf-8')

        # Find all mc:AlternateContent blocks that contain ABCDEFGH
        alt_pattern = re.compile(r'(<mc:AlternateContent\b[^>]*>.*?</mc:AlternateContent>)', re.DOTALL)
        pos_h_pattern = re.compile(r'<wp:posOffset>(\d+)</wp:posOffset>.*?</wp:positionH>', re.DOTALL)
        pos_v_pattern = re.compile(r'<wp:posOffset>(\d+)</wp:posOffset>.*?</wp:positionV>', re.DOTALL)
        posH_block = re.compile(r'<wp:positionH[^>]*>(.*?)</wp:positionH>', re.DOTALL)
        posV_block = re.compile(r'<wp:positionV[^>]*>(.*?)</wp:positionV>', re.DOTALL)

        blocks = []
        for m in alt_pattern.finditer(xml_content):
            block = m.group(1)
            if 'ABCDEFGH' not in block:
                continue
            # Get position from the DrawingML anchor
            ph_block = posH_block.search(block)
            pv_block = posV_block.search(block)
            if not ph_block or not pv_block:
                continue
            ph_off = re.search(r'<wp:posOffset>(\d+)</wp:posOffset>', ph_block.group(1))
            pv_off = re.search(r'<wp:posOffset>(\d+)</wp:posOffset>', pv_block.group(1))
            if not ph_off or not pv_off:
                continue
            x = int(ph_off.group(1))
            y = int(pv_off.group(1))
            blocks.append((x, y, m.start(), m.end(), block))

        if not blocks:
            logger.warning("No ABCDEFGH text boxes found in template")
            return template_path.read_bytes()

        # Sort by column (x) then row (y)
        blocks.sort(key=lambda b: (round(b[0] / (2.3 * 914400)), b[1]))

        # Group into labels (columns)
        labels = []
        last_col_key = -1
        for x, y, start, end, block in blocks:
            col_key = round(x / (2.3 * 914400))
            if col_key != last_col_key:
                labels.append([])
                last_col_key = col_key
            labels[-1].append((x, y, start, end, block))

        # Build replacement list — process in reverse order to preserve positions
        edits = []
        for label_blocks in labels:
            for i, (x, y, start, end, block) in enumerate(label_blocks):
                new_text = point_names[i] if i < len(point_names) else ''
                # Replace ALL <w:t>ABCDEFGH</w:t> in both Choice and Fallback
                new_block = block.replace('>ABCDEFGH</w:t>', f'>{new_text}</w:t>')
                if new_block != block:
                    edits.append((start, end, new_block))

        # Apply in reverse order
        edits.sort(key=lambda e: e[0], reverse=True)
        for start, end, new_block in edits:
            xml_content = xml_content[:start] + new_block + xml_content[end:]

        doc_xml.write_text(xml_content, encoding='utf-8')

        # Repackage
        output = Path(tmp_dir) / 'output.docx'
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
            for walk_root, dirs, files in os.walk(extract_dir):
                for f in files:
                    file_path = Path(walk_root) / f
                    arcname = file_path.relative_to(extract_dir)
                    zout.write(file_path, arcname)

        return output.read_bytes()
    finally:
        shutil.rmtree(tmp_dir)


def _generate_label_pdf(variant_data: dict, device_name: str = "", title: str = "") -> bytes:
    """Generate a printable door label PDF for a controller variant (fallback when no RC template)."""
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent / 'lib' / 'fpdf2_pkg'))
    from fpdf import FPDF

    meta = variant_data.get('meta', {})
    controller_id = _resolve_controller(meta)
    ctrl_info = CONTROLLER_TERMINALS.get(controller_id, {})
    ctrl_label = ctrl_info.get('label', controller_id)

    io_rows = _build_io_table(variant_data, device_name)
    if not io_rows:
        raise HTTPException(400, "No I/O points found in this variant")

    inputs = [r for r in io_rows if r['io_type'] in ('AI', 'BI')]
    outputs = [r for r in io_rows if r['io_type'] in ('AO', 'BO')]

    pdf = FPDF(orientation='P', unit='mm', format='Letter')
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    # Header
    pdf.set_fill_color(25, 55, 95)
    pdf.rect(10, 10, 196, 18, 'F')
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(14, 12)
    pdf.cell(0, 7, f'DOOR LABEL  --  {ctrl_label}', ln=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_xy(14, 19)
    display_name = device_name if device_name else title
    pdf.cell(0, 5, f'Device: {display_name}' if display_name else 'Template Label', ln=True)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_xy(160, 13)
    pdf.cell(40, 5, 'RELIABLE CONTROLS', align='R')
    pdf.set_text_color(0, 0, 0)
    y = 34

    def draw_section(section_title, rows, start_y, fill_color):
        sy = start_y
        r, g, b = fill_color
        pdf.set_fill_color(r, g, b)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(10, sy)
        pdf.cell(196, 7, f'  {section_title}', fill=True, ln=True)
        sy += 8
        pdf.set_fill_color(230, 235, 240)
        pdf.set_text_color(60, 60, 60)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_xy(10, sy)
        for header, w in zip(['TERMINAL', 'TYPE', 'POINT NAME', 'DESCRIPTION'], [28, 16, 65, 87]):
            pdf.cell(w, 6, f' {header}', border=1, fill=True)
        pdf.ln()
        sy += 6
        pdf.set_text_color(0, 0, 0)
        for ri, row in enumerate(rows):
            if sy > 260:
                pdf.add_page()
                sy = 15
            pdf.set_fill_color(248, 250, 252) if ri % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            pdf.set_xy(10, sy)
            pdf.set_font('Courier', 'B', 8)
            pdf.cell(28, 5.5, f' {row["terminal"]}', border='LTB', fill=True)
            pdf.set_font('Courier', '', 7)
            pdf.cell(16, 5.5, f' {row["io_type"]}', border='TB', fill=True)
            pdf.set_font('Courier', 'B', 8)
            pdf.cell(65, 5.5, f' {row["point_name"]}', border='TB', fill=True)
            pdf.set_font('Courier', '', 7)
            pdf.cell(87, 5.5, f' {row["description"]}', border='TBR', fill=True)
            pdf.ln()
            sy += 5.5
        return sy + 4

    if inputs:
        y = draw_section('INPUTS', inputs, y, (34, 120, 60))
    if outputs:
        y = draw_section('OUTPUTS', outputs, y, (30, 80, 160))

    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(140, 140, 140)
    pdf.set_xy(10, 268)
    pdf.cell(196, 4, f'Generated by DFA Platform  --  SBS Controls  --  {ctrl_label}  --  Print and affix to controller door', align='C')

    return pdf.output()


@app.get("/api/composer/label/{category}/{variant_id}")
async def get_label(category: str, variant_id: str, device_name: str = "", format: str = "auto"):
    """Generate a printable door label for a library variant.

    format=auto: uses RC Word template if available, otherwise PDF
    format=pdf: always generates PDF
    format=docx: always generates filled Word template (404 if no template)
    """
    cat_path = cfg.library_root / category
    json_path = cat_path / f"{variant_id}.json"
    if not json_path.exists():
        raise HTTPException(404, f"Variant {category}/{variant_id} not found")

    variant_data = json.loads(json_path.read_text())
    meta = variant_data.get('meta', {})
    controller_id = _resolve_controller(meta)
    title = variant_data.get('description', variant_id)

    io_rows = _build_io_table(variant_data, device_name)
    point_names = [r['point_name'] for r in io_rows]

    template_path = _get_label_template_file(controller_id)

    use_docx = (format == 'docx') or (format == 'auto' and template_path is not None and str(template_path).endswith('.docx'))

    if use_docx:
        if template_path is None:
            raise HTTPException(404, f"No RC label template found for controller {controller_id}")
        docx_bytes = _fill_label_template(template_path, point_names, device_name)
        fname = f"Door_Label_{device_name or variant_id}.docx"
        return StreamingResponse(
            iter([docx_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )
    else:
        pdf_bytes = _generate_label_pdf(variant_data, device_name=device_name, title=title)
        fname = f"Door_Label_{device_name or variant_id}.pdf"
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )


@app.get("/api/composer/label-data/{category}/{variant_id}")
async def get_label_data(category: str, variant_id: str, device_name: str = ""):
    """Return I/O terminal mapping data for a variant (JSON, for UI display)."""
    cat_path = cfg.library_root / category
    json_path = cat_path / f"{variant_id}.json"
    if not json_path.exists():
        raise HTTPException(404, f"Variant {category}/{variant_id} not found")

    variant_data = json.loads(json_path.read_text())
    meta = variant_data.get('meta', {})
    controller_id = _resolve_controller(meta)
    ctrl_info = CONTROLLER_TERMINALS.get(controller_id, {})

    io_rows = _build_io_table(variant_data, device_name)
    return {
        "controller": controller_id,
        "controller_label": ctrl_info.get('label', controller_id),
        "variant_id": variant_id,
        "category": category,
        "device_name": device_name,
        "io_points": io_rows,
        "summary": {
            "inputs": len([r for r in io_rows if r['io_type'] in ('AI', 'BI')]),
            "outputs": len([r for r in io_rows if r['io_type'] in ('AO', 'BO')]),
            "total": len(io_rows),
        }
    }


# ─── Point-to-Point Verification PDF ────────────────────────────────────────

# Map DFA equipment keys to library categories
EQUIP_TO_LIB_CAT = {
    'ahu_standard': 'AHU', 'ahu_doas': 'AHU', 'ahu_erw': 'AHU', 'ahu_rtu': 'RTU', 'ahu_other': 'AHU',
    'rtu_sf_rf': 'RTU', 'rtu_sf_only': 'RTU', 'rtu_hp': 'RTU', 'rtu_doas': 'RTU',
    'vav_hw_std': 'VAV', 'vav_cool': 'VAV', 'vav_dual': 'VAV', 'vav_elec': 'VAV', 'vav_hw_fp': 'VAV', 'vav_other': 'VAV',
    'fcu': 'FCU', 'fcu_2pipe': 'FCU', 'fcu_4pipe': 'FCU',
    'unit_heater': 'UH', 'wshp': 'WSHP', 'vvt': 'VVT', 'vvt_bypass': 'VVT',
    'boiler': 'SBS_PLANTS', 'boiler_cascade_leader': 'SBS_PLANTS', 'boiler_cascade_follower': 'SBS_PLANTS',
    'chiller': 'SBS_PLANTS', 'hwp_primary': 'SBS_PLANTS', 'hwp_secondary': 'SBS_PLANTS',
    'chwp_primary': 'SBS_PLANTS', 'chwp_secondary': 'SBS_PLANTS', 'cwp': 'SBS_PLANTS',
    'cooling_tower': 'SBS_PLANTS', 'hx': 'SBS_PLANTS',
}

TERMINAL_EQUIP = {'vav_hw_std', 'vav_cool', 'vav_dual', 'vav_elec', 'vav_hw_fp', 'vav_other',
                  'fcu', 'fcu_2pipe', 'fcu_4pipe', 'unit_heater', 'vvt', 'vvt_bypass',
                  'wshp', 'wshp_2pipe', 'wshp_4pipe', 'radiant'}

# BACnet unit codes to display strings
UNIT_LABELS = {
    '2': 'deg-F', '62': 'deg-C', '7': '"WC', '15': 'CFM', '45': 'in',
    '23': 'psi', '0': '', '3': 'deg-F', '98': '%RH',
}


def _get_variant_for_category(lib_cat: str) -> Optional[dict]:
    """Load the first variant JSON for a library category."""
    cat_path = cfg.library_root / lib_cat
    if not cat_path.is_dir():
        return None
    jsons = [f for f in os.listdir(cat_path) if f.endswith('.json')]
    if not jsons:
        return None
    return json.loads((cat_path / jsons[0]).read_text())


@app.get("/api/composer/p2p-pdf")
async def p2p_verification_pdf(
    equipment: str = "",
    device_name: str = "",
    project_name: str = "",
    project_number: str = "",
    qty: int = 1,
):
    """Generate a Point-to-Point Verification PDF for an equipment type.

    Uses library I/O data to create a verification sheet with programmed points
    and blank columns for field readings.

    For terminal units (VAV, FCU, etc.) with qty > 1, generates a bulk grid.
    """
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent / 'lib' / 'fpdf2_pkg'))
    from fpdf import FPDF

    def _c(s):
        if not isinstance(s, str): return str(s) if s else ''
        return s.replace('\u2014', '-').replace('\u2013', '-').replace('\u2018', "'").replace('\u2019', "'")

    lib_cat = EQUIP_TO_LIB_CAT.get(equipment)
    if not lib_cat:
        raise HTTPException(400, f"No library mapping for equipment type: {equipment}")

    variant_data = _get_variant_for_category(lib_cat)
    if not variant_data:
        raise HTTPException(404, f"No library variant found for category: {lib_cat}")

    dn = device_name or '{device-name}'
    io_rows = _build_io_table(variant_data, dn)
    if not io_rows:
        raise HTTPException(400, "No I/O points found")

    meta = variant_data.get('meta', {})
    controller_id = _resolve_controller(meta)
    ctrl_info = CONTROLLER_TERMINALS.get(controller_id, {})
    ctrl_label = ctrl_info.get('label', controller_id)

    # Get unit labels from variant objects
    objs = variant_data.get('objects', {})
    unit_map = {}
    for otype in ['AI', 'AO', 'BI', 'BO']:
        for o in objs.get(otype, []):
            name = o.get('name', '').replace('{device-name}', dn)
            unit_code = o.get('unit', '')
            unit_map[name] = UNIT_LABELS.get(str(unit_code), '')

    is_terminal = equipment in TERMINAL_EQUIP
    equip_label = _c(equipment.replace('_', ' ').title())
    proj = _c(project_name)
    pnum = _c(project_number)

    pdf = FPDF(orientation='L' if (is_terminal and qty > 3) else 'P', unit='mm', format='Letter')
    pdf.set_auto_page_break(auto=True, margin=15)

    def header(title, subtitle=''):
        pdf.set_fill_color(25, 55, 95)
        w = pdf.w - 20
        pdf.rect(10, 10, w, 16, 'F')
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(14, 11)
        pdf.cell(0, 6, _c(title))
        pdf.set_font('Helvetica', '', 8)
        pdf.set_xy(14, 18)
        pdf.cell(0, 5, _c(subtitle))
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_xy(pdf.w - 60, 11)
        pdf.set_text_color(200, 220, 255)
        pdf.cell(50, 6, pnum, align='R')
        pdf.set_text_color(0, 0, 0)

    if is_terminal and qty > 1:
        # ═══ BULK TERMINAL P2P: grid with units as columns ═══
        pdf.add_page()
        header('POINT-TO-POINT VERIFICATION - TERMINAL UNITS',
               f'{proj}  |  {equip_label}  |  {qty} units  |  Controller: {ctrl_label}')

        y = 32
        pdf.set_font('Helvetica', '', 8)
        pdf.set_xy(14, y)
        pdf.cell(0, 4, f'Date: ___________  Tech: ___________  Instructions: For each unit, verify point reads correctly. Write measured value or check off.')
        y += 8

        # How many units fit per page? Each unit column ~12mm wide
        page_w = pdf.w - 20
        label_col = 55  # point name column
        type_col = 12
        units_space = page_w - label_col - type_col
        units_per_page = max(1, int(units_space / 12))

        for page_start in range(0, qty, units_per_page):
            if page_start > 0:
                pdf.add_page()
                header('POINT-TO-POINT VERIFICATION (cont.)',
                       f'{proj}  |  {equip_label}  |  Units {page_start + 1}-{min(page_start + units_per_page, qty)}')
                y = 32

            page_units = min(units_per_page, qty - page_start)
            unit_col_w = min(12, units_space / page_units)

            # Table header
            pdf.set_fill_color(230, 235, 240)
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_xy(10, y)
            pdf.cell(label_col, 6, ' Point Name', border=1, fill=True)
            pdf.cell(type_col, 6, ' Type', border=1, fill=True)
            for u in range(page_units):
                pdf.cell(unit_col_w, 6, f' #{page_start + u + 1}', border=1, fill=True, align='C')
            pdf.ln()
            y += 6.5

            # Data rows
            pdf.set_font('Courier', '', 7)
            for ri, row in enumerate(io_rows):
                if y > (pdf.h - 20):
                    pdf.add_page()
                    header('POINT-TO-POINT VERIFICATION (cont.)', f'{equip_label}')
                    y = 32
                    pdf.set_fill_color(230, 235, 240)
                    pdf.set_font('Helvetica', 'B', 7)
                    pdf.set_xy(10, y)
                    pdf.cell(label_col, 6, ' Point Name', border=1, fill=True)
                    pdf.cell(type_col, 6, ' Type', border=1, fill=True)
                    for u in range(page_units):
                        pdf.cell(unit_col_w, 6, f' #{page_start + u + 1}', border=1, fill=True, align='C')
                    pdf.ln()
                    y += 6.5
                    pdf.set_font('Courier', '', 7)

                bg = (248, 250, 252) if ri % 2 == 0 else (255, 255, 255)
                pdf.set_fill_color(*bg)
                pdf.set_xy(10, y)
                pdf.set_font('Courier', 'B', 7)
                pdf.cell(label_col, 5, f' {_c(row["point_name"])}', border=1, fill=True)
                pdf.set_font('Courier', '', 6)
                pdf.cell(type_col, 5, f' {row["io_type"]}', border=1, fill=True)
                for u in range(page_units):
                    pdf.cell(unit_col_w, 5, '', border=1, fill=True)
                pdf.ln()
                y += 5

        # Notes
        y += 6
        if y > (pdf.h - 30):
            pdf.add_page()
            y = 20
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(25, 55, 95)
        pdf.set_xy(14, y)
        pdf.cell(0, 5, 'Notes / Discrepancies:')
        y += 6
        pdf.set_text_color(0, 0, 0)
        for _ in range(3):
            pdf.line(14, y + 4, pdf.w - 10, y + 4)
            y += 7

    else:
        # ═══ INDIVIDUAL EQUIPMENT P2P ═══
        pdf.add_page()
        header('POINT-TO-POINT VERIFICATION',
               f'{proj}  |  {equip_label}  |  Device: {dn}  |  Controller: {ctrl_label}')

        y = 32
        pdf.set_font('Helvetica', '', 8)
        pdf.set_xy(14, y)
        pdf.cell(0, 4, 'Date: ___________  Tech: ___________')
        y += 8

        # Table header
        col_w = [22, 12, 48, 38, 26, 26, 24]
        headers = ['Terminal', 'Type', 'Point Name', 'Description', 'Units', 'Field Reading', 'Pass/Fail']
        pdf.set_fill_color(230, 235, 240)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_xy(10, y)
        for h, w in zip(headers, col_w):
            pdf.cell(w, 6, f' {h}', border=1, fill=True)
        pdf.ln()
        y += 6.5

        # Data rows
        for ri, row in enumerate(io_rows):
            bg = (248, 250, 252) if ri % 2 == 0 else (255, 255, 255)
            pdf.set_fill_color(*bg)
            pdf.set_xy(10, y)
            pdf.set_font('Courier', 'B', 8)
            pdf.cell(col_w[0], 6, f' {_c(row["terminal"])}', border=1, fill=True)
            pdf.set_font('Courier', '', 7)
            pdf.cell(col_w[1], 6, f' {row["io_type"]}', border=1, fill=True)
            pdf.set_font('Courier', 'B', 8)
            pdf.cell(col_w[2], 6, f' {_c(row["point_name"])}', border=1, fill=True)
            pdf.set_font('Helvetica', '', 7)
            desc = _c(row.get('description', ''))[:22]
            pdf.cell(col_w[3], 6, f' {desc}', border=1, fill=True)
            units = unit_map.get(row['point_name'], '')
            pdf.cell(col_w[4], 6, f' {units}', border=1, fill=True)
            # Blank field reading column
            pdf.cell(col_w[5], 6, '', border=1, fill=True)
            # Pass/Fail checkboxes
            pdf.cell(col_w[6], 6, '  [ ] P  [ ] F', border=1, fill=True)
            pdf.ln()
            y += 6

        # Software points (AV/BV) — values to verify from RC Studio
        av_objs = objs.get('AV', [])
        bv_objs = objs.get('BV', [])
        sw_points = []
        for o in av_objs:
            name = o.get('name', '').replace('{device-name}', dn)
            desc = o.get('description', '')
            pv = o.get('present_value', '')
            sw_points.append((name, 'AV', desc, pv))
        for o in bv_objs:
            name = o.get('name', '').replace('{device-name}', dn)
            desc = o.get('description', '')
            pv = o.get('present_value', '')
            sw_points.append((name, 'BV', desc, pv))

        if sw_points:
            y += 6
            pdf.set_font('Helvetica', 'B', 10)
            pdf.set_fill_color(60, 100, 160)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(10, y)
            pdf.cell(196, 7, '  SOFTWARE POINTS (verify in RC Studio / Tridium)', fill=True)
            y += 8
            pdf.set_text_color(0, 0, 0)

            sw_cols = [48, 12, 50, 30, 30, 26]
            sw_headers = ['Point Name', 'Type', 'Description', 'Design Value', 'Actual Value', 'Pass/Fail']
            pdf.set_fill_color(230, 235, 240)
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_xy(10, y)
            for h, w in zip(sw_headers, sw_cols):
                pdf.cell(w, 5.5, f' {h}', border=1, fill=True)
            pdf.ln()
            y += 6

            pdf.set_font('Courier', '', 7)
            for ri, (name, stype, desc, pv) in enumerate(sw_points[:30]):  # Limit to 30
                if y > 260:
                    pdf.add_page()
                    header('SOFTWARE POINTS (cont.)', f'{equip_label} - {dn}')
                    y = 32
                bg = (248, 250, 252) if ri % 2 == 0 else (255, 255, 255)
                pdf.set_fill_color(*bg)
                pdf.set_xy(10, y)
                pdf.set_font('Courier', 'B', 7)
                pdf.cell(sw_cols[0], 5, f' {_c(name)[:28]}', border=1, fill=True)
                pdf.set_font('Courier', '', 6)
                pdf.cell(sw_cols[1], 5, f' {stype}', border=1, fill=True)
                pdf.set_font('Helvetica', '', 6)
                pdf.cell(sw_cols[2], 5, f' {_c(desc)[:30]}', border=1, fill=True)
                pdf.cell(sw_cols[3], 5, f' {_c(str(pv))[:18]}', border=1, fill=True)
                pdf.cell(sw_cols[4], 5, '', border=1, fill=True)
                pdf.cell(sw_cols[5], 5, '  [ ] P  [ ] F', border=1, fill=True)
                pdf.ln()
                y += 5

        # Notes + signature
        y += 8
        if y > 245:
            pdf.add_page()
            y = 20
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(25, 55, 95)
        pdf.set_xy(14, y)
        pdf.cell(0, 5, 'Notes / Discrepancies:')
        y += 6
        pdf.set_text_color(0, 0, 0)
        for _ in range(4):
            pdf.line(14, y + 4, 200, y + 4)
            y += 7
        y += 4
        pdf.set_font('Helvetica', '', 9)
        pdf.set_xy(14, y)
        pdf.cell(0, 5, 'All points verified:  [ ] Yes  [ ] No     Technician: ___________________________     Date: _______________')

    # Footer
    for pn in range(1, pdf.pages_count + 1):
        pdf.page = pn
        pdf.set_font('Helvetica', 'I', 7)
        pdf.set_text_color(150, 150, 150)
        pdf.set_xy(10, pdf.h - 10)
        pdf.cell(pdf.w - 20, 4, f'DFA Platform  |  SBS Controls  |  {pnum} {proj}  |  P2P Verification  |  Page {pn}/{pdf.pages_count}', align='C')
    pdf.set_text_color(0, 0, 0)

    pdf_bytes = bytes(pdf.output())
    fname = f"P2P_{_c(equip_label)}_{_c(dn)}.pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )
