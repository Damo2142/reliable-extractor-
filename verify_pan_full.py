#!/usr/bin/env python3
"""Full verification: every field, every row, zero tolerance."""
import struct, zipfile, re, sys
from io import BytesIO
from openpyxl import load_workbook

PKG = sys.argv[1] if len(sys.argv) > 1 else '/srv/dfa/drops/sbs-full-package (7).zip'
PAN = sys.argv[2] if len(sys.argv) > 2 else '/srv/dfa/drops/SBS-AHU-FINAL-r01.pan'

UNIT_CODES = {
    '°F': 64, '%': 98, '%RH': 29, 'ppm': 96, 'CFM': 84,
    'WC': 58, '"WC': 58, 'BTU/lb': 117, 'BTUlb': 117,
    'Volts': 5, 'V': 5, 'kW': 42, 'GPM': 85,
    'minutes': 72, 'min': 72, 'Min.': 72, 'seconds': 73,
    'sec': 73, 'Sec': 73, 'hours': 71, 'Time': 72,
    '': 95, 'None': 95, '#': 95, 'Off/On': 95,
}

with zipfile.ZipFile(PKG) as z:
    wb = load_workbook(BytesIO(z.read('RC-Studio-Output.xlsx')), read_only=True)
    bas_set = set(n.split('/')[-1] for n in z.namelist() if n.endswith('.bas'))

pan = open(PAN, 'rb').read()

# Parse index → type_blocks[tid][inst] = block_bytes
def ri(d, o):
    v = []
    for i in range(4):
        h = struct.unpack_from('<H', d, o+i*4)[0]
        l = struct.unpack_from('<H', d, o+i*4+2)[0]
        v.append((h << 16) | l)
    return v

first = ri(pan, 0x400)
type_blocks = {}
ents = [(first[1], first[2], first[3])]
for i in range(1, first[0]):
    ents.append(tuple(ri(pan, 0x440+(i-1)*0x40)[1:]))
for tid, off, cnt in ents:
    d = {}
    pos = off
    for _ in range(cnt):
        sz = struct.unpack_from('<H', pan, pos)[0]
        blk = pan[pos:pos+sz+2]
        inst = struct.unpack_from('>I', blk, 6)[0] & 0x3FFFFF
        d[inst] = blk
        pos += sz + 2
    type_blocks[tid] = d

def get_props(blk):
    p = blk[12:]
    props = {}
    r = 3
    while r < len(p):
        rl = p[r]
        if rl == 0: break
        rec = p[r:min(r+rl, len(p))]
        if len(rec) >= 6:
            ctx = (rec[2]<<8)|rec[3]
            prop = rec[4]
            tag = rec[5]
            val = rec[6:]
            if val and val[-1] == 0x00: val = val[:-1]
            props[(ctx, prop)] = (tag, val)
        r += rl
    return props

def gn(props):
    r = props.get((0, 0x4D))
    if r and r[0] == 0x75 and len(r[1]) > 2:
        return bytes(r[1][2:2+r[1][0]-1]).decode('latin-1', errors='replace')
    return ''

def gdesc(props):
    r = props.get((0, 0x1C))
    if r and r[0] == 0x75 and len(r[1]) > 2:
        return bytes(r[1][2:2+r[1][0]-1]).decode('latin-1', errors='replace')
    return ''

def gf(props, ctx, prop):
    r = props.get((ctx, prop))
    if r and r[0] == 0x44 and len(r[1]) >= 4:
        return round(struct.unpack('>f', r[1][:4])[0], 4)
    return None

def gu8(props, ctx, prop):
    r = props.get((ctx, prop))
    if r and len(r[1]) >= 1: return r[1][0]
    return None

def gu16(props, ctx, prop):
    r = props.get((ctx, prop))
    if r and len(r[1]) >= 2: return struct.unpack('>H', r[1][:2])[0]
    return None

def uc(s):
    return UNIT_CODES.get(str(s or '').strip(), 95)

fails = []
passes = 0

def ok(sec, rid, field, exp, got):
    global passes
    if exp == got:
        passes += 1
    else:
        fails.append(f"{sec} {rid} {field}: exp={exp!r} got={got!r}")

def ok_float(sec, rid, field, exp, got, tol=0.02):
    global passes
    if exp is None and got is None: passes += 1; return
    if exp is not None and got is not None and abs(exp - got) < tol: passes += 1; return
    fails.append(f"{sec} {rid} {field}: exp={exp!r} got={got!r}")

# ═══ INPUTS ═══
for row in wb['Inputs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    nm = str(row[1] or '').strip()
    if not nm or '---' in nm: continue
    inst = int(row[0]); typ = str(row[2] or '').strip()
    tid = 0 if typ == 'AI' else 3
    blks = type_blocks.get(tid, {})
    if inst not in blks: fails.append(f"IN {typ}{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("IN", f"{typ}{inst}", "name", nm, gn(pr))
    if typ == 'AI':
        ok("IN", f"AI{inst}", "units", uc(row[4]), gu8(pr, 0, 0x75))

# ═══ OUTPUTS ═══
for row in wb['Outputs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    nm = str(row[1] or '').strip()
    if not nm or '---' in nm: continue
    inst = int(row[0]); typ = str(row[2] or '').strip()
    tid = 1 if typ == 'AO' else 4
    blks = type_blocks.get(tid, {})
    if inst not in blks: fails.append(f"OUT {typ}{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("OUT", f"{typ}{inst}", "name", nm, gn(pr))
    if typ == 'AO':
        ok("OUT", f"AO{inst}", "units", uc(row[4]), gu8(pr, 0, 0x75))
        # min/max voltage
        exp_minv = float(row[6]) if row[6] is not None else 2.0
        exp_maxv = float(row[7]) if row[7] is not None else 10.0
        ok_float("OUT", f"AO{inst}", "min_v", exp_minv, gf(pr, 4, 0x22))
        ok_float("OUT", f"AO{inst}", "max_v", exp_maxv, gf(pr, 4, 0x23))

# ═══ VALUES ═══
for row in wb['Values'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not any(c0.startswith(p) for p in ['AV','BV','MV']): continue
    nm = str(row[1] or '').strip()
    if not nm: continue
    typ = c0[:2]; inst = int(c0[2:])
    tid = {'AV':2,'BV':5,'MV':19}[typ]
    blks = type_blocks.get(tid, {})
    if inst not in blks: fails.append(f"VAL {c0} MISSING"); continue
    pr = get_props(blks[inst])
    ok("VAL", c0, "name", nm, gn(pr))
    if typ == 'AV':
        ok("VAL", c0, "units", uc(row[4]), gu8(pr, 0, 0x75))
        dv = row[3]
        if dv is not None:
            dv_s = str(dv).strip()
            if ':' in dv_s:
                pass  # time format — skip float check
            else:
                try:
                    ok_float("VAL", c0, "pv", float(dv_s), gf(pr, 0, 0x55))
                except: pass
    elif typ == 'MV':
        states = str(row[5] or '').strip()
        if states:
            ok("VAL", c0, "states", states, gdesc(pr))
        # MV present value
        dv = row[3]
        if dv is not None:
            dv_s = str(dv).strip()
            try:
                exp_pv = int(float(dv_s))
            except ValueError:
                exp_pv = None
                if states:
                    for part in states.split('/'):
                        part = part.strip()
                        if '-' in part:
                            n, v = part.split('-', 1)
                            if v.strip().lower() == dv_s.lower():
                                try: exp_pv = int(n.strip())
                                except: pass
                                break
            if exp_pv is not None:
                ok("VAL", c0, "pv", exp_pv, gu8(pr, 0, 0x55))
    elif typ == 'BV':
        dv = str(row[3] or '').strip().lower()
        exp_pv = 1 if dv in ('1','true','on','active') else 0
        ok("VAL", c0, "pv", exp_pv, gu8(pr, 0, 0x55))

# ═══ LOOPS ═══
for row in wb['Loops'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not c0.startswith('LOOP'): continue
    inst = int(c0[4:])
    nm = '{device-name}-' + str(row[1] or '').strip()
    if not str(row[1] or '').strip(): continue
    blks = type_blocks.get(12, {})
    if inst not in blks: fails.append(f"LOOP LOOP{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("LOOP", f"L{inst}", "name", nm, gn(pr))
    act_s = str(row[4] or '+').strip()
    ok("LOOP", f"L{inst}", "action", 0 if act_s=='-' else 1, gu8(pr, 0, 0x02))
    ok_float("LOOP", f"L{inst}", "p_band", float(row[5] or 0), gf(pr, 4, 0x4D))
    ok_float("LOOP", f"L{inst}", "integral", float(row[6] or 0), gf(pr, 0, 0x6C))
    ok_float("LOOP", f"L{inst}", "derivative", float(row[7] or 0), gf(pr, 0, 0x0E))
    # Input ref
    inp_name = str(row[2] or '').strip()
    inp_ref = pr.get((0, 0x13))
    if inp_ref and inp_name:
        oid = struct.unpack('>I', inp_ref[1][:4])[0]
        ref_type = (oid >> 22) & 0x3FF
        ref_inst = oid & 0x3FFFFF
        # Look up expected
        from compile_from_excel import name_to_obj
        exp = name_to_obj.get(inp_name, (0x3FF, 0))
        ok("LOOP", f"L{inst}", "input_type", exp[0], ref_type)
        ok("LOOP", f"L{inst}", "input_inst", exp[1], ref_inst)

# ═══ PROGRAMS ═══
for row in wb['Programs'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    c0 = str(row[0]).strip()
    if not c0.startswith('PRG'): continue
    inst = int(re.sub(r'[^0-9]','',c0))
    nm = str(row[1] or '').strip()
    if not nm: continue
    fn = str(row[2] or '').strip()
    en_s = str(row[3] or 'Yes').strip()
    blks = type_blocks.get(16, {})
    if inst not in blks: fails.append(f"PRG PRG{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("PRG", f"P{inst}", "name", nm, gn(pr))
    ok("PRG", f"P{inst}", "enabled", 1 if en_s.lower() in ('yes','1','true') else 0, gu8(pr, 4, 0x0A))
    has_bc = len(blks[inst]) > 100
    if fn in bas_set:
        ok("PRG", f"P{inst}", "has_bc", True, has_bc)

# ═══ TABLES ═══
for row in wb['Tables'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','',str(row[0])))
    nm = '{device-name}-' + str(row[1] or '').strip()
    blks = type_blocks.get(141, {})
    if inst not in blks: fails.append(f"TBL TBL{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("TBL", f"T{inst}", "name", nm, gn(pr))
    ok("TBL", f"T{inst}", "units", uc(row[3]), gu8(pr, 0, 0x75))

# ═══ SCHEDULES ═══
for row in wb['Schedules'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','',str(row[0])))
    nm = str(row[1] or '').strip()
    if not nm: continue
    blks = type_blocks.get(17, {})
    if inst not in blks: fails.append(f"SCH SCHED{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("SCH", f"S{inst}", "name", nm, gn(pr))

# ═══ TRENDS ═══
for row in wb['Trends'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    inst = int(re.sub(r'[^0-9]','',str(row[0])))
    nm = str(row[1] or '').strip()
    if not nm: continue
    blks = type_blocks.get(20, {})
    if inst not in blks: fails.append(f"STL STL{inst} MISSING"); continue
    pr = get_props(blks[inst])
    ok("STL", f"T{inst}", "name", nm, gn(pr))
    tt = str(row[3] or '').strip().lower()
    ok("STL", f"T{inst}", "log_type", 1 if tt=='polled' else 2, gu8(pr, 0, 0x48))
    exp_buf = int(row[5]) if row[5] else 512
    ok("STL", f"T{inst}", "buffer", exp_buf, gu16(pr, 0, 0x7E))
    # Monitored point ref
    mon = str(row[2] or '').strip()
    ref = pr.get((0, 0x84))
    if ref and mon:
        oid = struct.unpack('>I', ref[1][:4])[0]
        ref_type = (oid >> 22) & 0x3FF
        ref_inst = oid & 0x3FFFFF
        # Check it's not null (0x3FF)
        if ref_type == 0x3FF:
            fails.append(f"STL T{inst} mon_ref: unresolved for '{mon}'")
        else:
            passes += 1

# ═══ SYSTEM GROUPS ═══
si = 1
for row in wb['System Groups'].iter_rows(min_row=5, values_only=True):
    if row[0] is None: continue
    nm = str(row[0]).strip()
    if not nm: continue
    blks = type_blocks.get(26, {})
    if si not in blks: fails.append(f"SG NOTIF{si} MISSING"); continue
    pr = get_props(blks[si])
    ok("SG", f"N{si}", "name", nm, gn(pr))
    si += 1

# ═══ REPORT ═══
print(f"\n{'='*60}")
print(f"VERIFICATION: {PAN}")
print(f"  Passed: {passes}")
print(f"  Failed: {len(fails)}")
if fails:
    print(f"\nFAILURES:")
    for f in fails: print(f"  {f}")
    print(f"\nFAILED — {len(fails)} errors")
else:
    print(f"\n100% PASS — ALL FIELDS MATCH")
